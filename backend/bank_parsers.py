"""
UAE Bank and PSP Statement Parsers
Supports: 
- Banks: Emirates NBD, ADCB, FAB, Mashreq, RAK Bank, DIB, CBD, and more
- PSPs: PayTabs, Telr, Network International, PayFort, Checkout.com, Stripe, PayPal, and more
"""

import re
import io
import logging
from typing import List, Dict, Optional, Tuple
from datetime import datetime

logger = logging.getLogger(__name__)


# Bank detection patterns
BANK_PATTERNS = {
    "emirates_nbd": [
        r"emirates\s*nbd",
        r"emiratesnbd",
        r"enbd",
        r"ﻲﺑد\s*تارﺎﻣﻹا\s*ﻚﻨﺑ",  # Arabic: Emirates NBD Bank
    ],
    "adcb": [
        r"abu\s*dhabi\s*commercial\s*bank",
        r"adcb",
        r"ﻲﺑﻮﻈﺑأ\s*يرﺎﺠﺘﻟا\s*ﻚﻨﺒﻟا",  # Arabic
    ],
    "fab": [
        r"first\s*abu\s*dhabi\s*bank",
        r"fab\s*bank",
        r"\bfab\b",
        r"nbad",  # National Bank of Abu Dhabi (merged into FAB)
    ],
    "mashreq": [
        r"mashreq\s*bank",
        r"mashreq",
        r"ﻖﺷﺮﻤﻟا\s*ﻚﻨﺑ",  # Arabic
    ],
    "rakbank": [
        r"rak\s*bank",
        r"rakbank",
        r"national\s*bank\s*of\s*ras\s*al\s*khaimah",
    ],
    "dib": [
        r"dubai\s*islamic\s*bank",
        r"\bdib\b",
        r"ﻲﺑد\s*ﻲﻣﻼﺳﻹا\s*ﻚﻨﺒﻟا",  # Arabic
    ],
    "cbd": [
        r"commercial\s*bank\s*of\s*dubai",
        r"\bcbd\b",
        r"ﻲﺑد\s*يرﺎﺠﺘﻟا\s*ﻚﻨﺒﻟا",  # Arabic
    ],
    "cbi": [
        r"commercial\s*bank\s*international",
        r"\bcbi\b",
    ],
    "nbf": [
        r"national\s*bank\s*of\s*fujairah",
        r"\bnbf\b",
    ],
    "ajman_bank": [
        r"ajman\s*bank",
    ],
}

# PSP (Payment Service Provider) detection patterns
PSP_PATTERNS = {
    "paytabs": [
        r"paytabs",
        r"pay\s*tabs",
        r"paytabs\.com",
    ],
    "telr": [
        r"telr",
        r"telr\.com",
        r"telrpay",
    ],
    "network_international": [
        r"network\s*international",
        r"network\s*int",
        r"n-genius",
        r"ngenius",
    ],
    "payfort": [
        r"payfort",
        r"pay\s*fort",
        r"amazon\s*payment\s*services",
    ],
    "checkout_com": [
        r"checkout\.com",
        r"checkout\s*com",
        r"cko\-",
    ],
    "stripe": [
        r"\bstripe\b",
        r"stripe\.com",
        r"stripe\s*inc",
    ],
    "paypal": [
        r"paypal",
        r"pay\s*pal",
        r"paypal\.com",
    ],
    "wise": [
        r"\bwise\b",
        r"transferwise",
        r"wise\.com",
    ],
    "tap_payments": [
        r"tap\s*payments",
        r"tap\.company",
        r"tappay",
    ],
    "hyperpay": [
        r"hyperpay",
        r"hyper\s*pay",
    ],
    "payby": [
        r"payby",
        r"pay\s*by",
    ],
    "magnati": [
        r"magnati",
        r"first\s*abu\s*dhabi\s*bank.*payment",
    ],
    "myfatoorah": [
        r"myfatoorah",
        r"my\s*fatoorah",
        r"fatoorah",
    ],
}


def detect_bank(text: str, filename: str = "") -> str:
    """Detect bank from text content or filename"""
    combined = (text + " " + filename).lower()
    
    for bank_name, patterns in BANK_PATTERNS.items():
        for pattern in patterns:
            if re.search(pattern, combined, re.IGNORECASE):
                logger.info(f"Detected bank: {bank_name}")
                return bank_name
    
    logger.info("Bank not detected, using generic parser")
    return "generic"


def detect_psp(text: str, filename: str = "") -> str:
    """Detect PSP from text content or filename"""
    combined = (text + " " + filename).lower()
    
    for psp_name, patterns in PSP_PATTERNS.items():
        for pattern in patterns:
            if re.search(pattern, combined, re.IGNORECASE):
                logger.info(f"Detected PSP: {psp_name}")
                return psp_name
    
    logger.info("PSP not detected, using generic parser")
    return "generic"


def detect_statement_type(text: str, filename: str = "") -> Tuple[str, str]:
    """
    Detect if statement is from a bank or PSP, and identify which one.
    Returns: (type: 'bank' or 'psp', name: detected name)
    """
    combined = (text + " " + filename).lower()
    
    # Check for PSP first (more specific patterns)
    for psp_name, patterns in PSP_PATTERNS.items():
        for pattern in patterns:
            if re.search(pattern, combined, re.IGNORECASE):
                logger.info(f"Detected PSP: {psp_name}")
                return ("psp", psp_name)
    
    # Then check for banks
    for bank_name, patterns in BANK_PATTERNS.items():
        for pattern in patterns:
            if re.search(pattern, combined, re.IGNORECASE):
                logger.info(f"Detected bank: {bank_name}")
                return ("bank", bank_name)
    
    logger.info("Statement type not detected, using generic parser")
    return ("unknown", "generic")


def parse_date(date_str: str, formats: List[str] = None) -> Optional[str]:
    """Parse date string to ISO format (YYYY-MM-DD)"""
    if not date_str:
        return None
    
    date_str = date_str.strip()
    
    # Default formats to try
    if formats is None:
        formats = [
            "%d/%m/%Y",      # DD/MM/YYYY (common in UAE)
            "%d-%m-%Y",      # DD-MM-YYYY
            "%Y-%m-%d",      # YYYY-MM-DD (ISO)
            "%d %b %Y",      # DD Mon YYYY
            "%d %B %Y",      # DD Month YYYY
            "%m/%d/%Y",      # MM/DD/YYYY (US format, less common)
            "%d.%m.%Y",      # DD.MM.YYYY
        ]
    
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    
    # Try regex extraction for common patterns
    # DD/MM/YYYY
    match = re.search(r'(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})', date_str)
    if match:
        day, month, year = match.groups()
        return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    
    return None


def parse_amount(amount_str: str) -> float:
    """Parse amount string to float"""
    if not amount_str:
        return 0.0
    
    # Clean the string
    cleaned = str(amount_str).strip()
    
    # Remove currency symbols and text
    cleaned = re.sub(r'[A-Za-z\s]', '', cleaned)
    
    # Handle negative indicators
    is_negative = False
    if cleaned.startswith('(') and cleaned.endswith(')'):
        is_negative = True
        cleaned = cleaned[1:-1]
    elif cleaned.startswith('-'):
        is_negative = True
        cleaned = cleaned[1:]
    elif 'DR' in str(amount_str).upper() or 'DEBIT' in str(amount_str).upper():
        is_negative = True
    
    # Remove commas and spaces
    cleaned = cleaned.replace(',', '').replace(' ', '')
    
    try:
        amount = float(cleaned)
        return -amount if is_negative else amount
    except ValueError:
        return 0.0


class BankStatementParser:
    """Base class for bank statement parsing"""
    
    def __init__(self, bank_name: str = "generic"):
        self.bank_name = bank_name
        self.entries = []
    
    def parse_ocr_text(self, text: str, fallback_date: str = None) -> List[Dict]:
        """Parse OCR text using bank-specific rules"""
        self.entries = []
        
        # Get bank-specific parser
        parser_method = getattr(self, f"_parse_{self.bank_name}", self._parse_generic)
        return parser_method(text, fallback_date)
    
    def _parse_generic(self, text: str, fallback_date: str = None) -> List[Dict]:
        """Generic parser using flexible pattern matching — works even without dates"""
        entries = []
        entry_id = 0
        
        lines = text.split('\n')
        
        # Date patterns: DD/MM/YYYY, DD-MM-YYYY, DD.MM.YYYY, YYYY-MM-DD, Mon DD YYYY
        date_patterns = [
            r'(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
            r'(\d{4}[/\-\.]\d{1,2}[/\-\.]\d{1,2})',
            r'(\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\s+\d{4})',
        ]
        # Amount pattern: numbers with commas and decimals
        amount_pattern = r'([\d,]+\.\d{2})'
        
        last_found_date = fallback_date
        
        for line in lines:
            if not line.strip() or len(line.strip()) < 5:
                continue
            
            # Look for date in line
            parsed_date = None
            for dp in date_patterns:
                date_match = re.search(dp, line, re.IGNORECASE)
                if date_match:
                    parsed_date = parse_date(date_match.group(1))
                    if parsed_date:
                        last_found_date = parsed_date
                        break
            
            # Find amounts in line
            amounts = re.findall(amount_pattern, line)
            if not amounts:
                continue
            
            # Skip header/footer lines (common keywords)
            line_lower = line.lower()
            if any(x in line_lower for x in ['page ', 'statement of', 'account number', 'opening balance', 'closing balance', 'total', 'balance brought', 'balance carried', 'tax registration', 'registered details', 'tel:', 'po box', 'head office', 'licensed by', 'electronically generated', 'www.', 'confirmation of', 'notice of disagreement']):
                if not any(x in line_lower for x in ['transfer', 'payment', 'deposit', 'withdrawal']):
                    continue
            
            entry_id += 1
            use_date = parsed_date or last_found_date or fallback_date
            
            # Determine amount and type
            first_amount = parse_amount(amounts[0])
            amount = first_amount
            
            # Debit indicators
            is_debit = any(x in line_lower for x in [
                'debit', 'dr', 'withdrawal', 'payment', 'purchase', 'transfer out',
                'atm', 'pos', 'card no', 'ddr', 'dds', 'loan'
            ])
            # Credit indicators  
            is_credit = any(x in line_lower for x in [
                'credit', 'cr', 'deposit', 'salary', 'transfer in',
                'refund', 'reversal', 'incoming', 'rma'
            ])
            
            # Check for Cr/Dr suffix after amount
            after_amounts = line[line.rfind(amounts[-1]) + len(amounts[-1]):].strip().lower()
            if after_amounts.startswith('cr'):
                is_credit = True
            elif after_amounts.startswith('dr'):
                is_debit = True
            
            if is_debit and not is_credit:
                amount = -abs(first_amount)
            elif is_credit:
                amount = abs(first_amount)
            
            # If multiple amounts, first is likely the transaction, last is balance
            # Use first amount for the entry
            
            # Extract description
            desc_parts = line
            if parsed_date and date_match:
                desc_parts = desc_parts[desc_parts.find(date_match.group(1)) + len(date_match.group(1)):]
            # Remove amounts from description
            for a in amounts:
                desc_parts = desc_parts.replace(a, '', 1)
            description = re.sub(r'\s+', ' ', desc_parts).strip()
            description = re.sub(r'[,\s]*$', '', description)[:200]
            
            if amount != 0:
                entries.append({
                    "id": f"stmt_{entry_id}",
                    "date": use_date,
                    "description": description,
                    "reference": "",
                    "amount": amount,
                    "bank": self.bank_name
                })
        
        return entries
    
    def _parse_emirates_nbd(self, text: str, fallback_date: str = None) -> List[Dict]:
        """Parser for Emirates NBD statements — handles bilingual Arabic/English"""
        entries = []
        entry_id = 0
        
        lines = text.split('\n')
        date_patterns = [
            r'(\d{2}/\d{2}/\d{4})',
            r'(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})',
        ]
        amount_pattern = r'([\d,]+\.\d{2})'
        
        last_found_date = fallback_date
        
        for line in lines:
            if not line.strip() or len(line.strip()) < 5:
                continue
            
            # Skip known header/footer lines
            line_lower = line.lower()
            if any(x in line_lower for x in ['page ', 'statement of account', 'head office', 'baniyas road', 'commercial registration', 'licensed by', 'electronically generated', 'www.emiratesnbd', 'tax registration', 'registered details', 'tel:', 'confirmation of', 'notice of disagreement', 'po box']):
                continue
            
            # Look for date
            parsed_date = None
            date_match = None
            for dp in date_patterns:
                date_match = re.search(dp, line)
                if date_match:
                    parsed_date = parse_date(date_match.group(1))
                    if parsed_date:
                        last_found_date = parsed_date
                        break
            
            amounts = re.findall(amount_pattern, line)
            if not amounts:
                continue
            
            # Skip balance-only lines (single amount with Cr at end and no transaction context)
            if len(amounts) == 1 and re.search(r'\d+\.\d{2}\s*cr\s*$', line_lower) and not any(x in line_lower for x in ['transfer', 'payment', 'salary', 'deposit', 'pos', 'atm']):
                continue
            
            entry_id += 1
            use_date = parsed_date or last_found_date or fallback_date
            
            first_amount = parse_amount(amounts[0])
            amount = first_amount
            
            # Emirates NBD debit keywords
            is_debit = any(x in line_lower for x in [
                'ddr', 'pos', 'atm', 'transfer', 'payment', 'loan', 'dds',
                'purchase', 'fee', 'charge', 'debit'
            ])
            is_credit = any(x in line_lower for x in [
                'salary', 'rma', 'incoming', 'refund', 'credit', 'deposit', 'cr'
            ])
            
            # Check for Cr suffix after last amount
            after_last = line[line.rfind(amounts[-1]) + len(amounts[-1]):].strip().lower()
            if after_last.startswith('cr'):
                is_credit = True
            elif after_last.startswith('dr'):
                is_debit = True
            
            # If multiple amounts: first is debit, second is credit, last is balance
            if len(amounts) >= 3:
                debit_amt = parse_amount(amounts[0])
                credit_amt = parse_amount(amounts[1])
                if debit_amt > 0 and credit_amt == 0:
                    amount = -abs(debit_amt)
                elif credit_amt > 0:
                    amount = abs(credit_amt)
                else:
                    amount = first_amount
            elif is_debit and not is_credit:
                amount = -abs(first_amount)
            elif is_credit:
                amount = abs(first_amount)
            
            # Extract description
            desc = line
            if date_match:
                desc = desc[desc.find(date_match.group(1)) + len(date_match.group(1)):]
            for a in amounts:
                desc = desc.replace(a, '', 1)
            desc = re.sub(r'[^\x20-\x7E]', ' ', desc)  # Remove non-ASCII (Arabic)
            description = re.sub(r'\s+', ' ', desc).strip()[:200]
            
            if amount != 0:
                entries.append({
                    "id": f"stmt_{entry_id}",
                    "date": use_date,
                    "description": description,
                    "reference": "",
                    "amount": amount,
                    "bank": "emirates_nbd"
                })
        
        return entries
    
    def _parse_adcb(self, text: str, fallback_date: str = None) -> List[Dict]:
        """Parser for ADCB statements"""
        # ADCB typically has: Date | Reference | Description | Debit | Credit | Balance
        return self._parse_generic(text, fallback_date)
    
    def _parse_fab(self, text: str, fallback_date: str = None) -> List[Dict]:
        """Parser for FAB (First Abu Dhabi Bank) statements"""
        # FAB format: Trans Date | Value Date | Description | Amount | Balance
        entries = []
        entry_id = 0
        
        lines = text.split('\n')
        date_pattern = r'(\d{2}[/\-]\d{2}[/\-]\d{4})'
        amount_pattern = r'([\d,]+\.\d{2})'
        
        for line in lines:
            date_match = re.search(date_pattern, line)
            if not date_match:
                continue
            
            amounts = re.findall(amount_pattern, line)
            if not amounts:
                continue
            
            entry_id += 1
            parsed_date = parse_date(date_match.group(1)) or fallback_date
            
            # FAB often shows DR/CR suffix
            amount = parse_amount(amounts[0])
            if 'dr' in line.lower() or amounts[0] in line and 'dr' in line[line.find(amounts[0]):line.find(amounts[0])+20].lower():
                amount = -abs(amount)
            
            # Extract description
            desc_start = line.find(date_match.group(1)) + len(date_match.group(1))
            desc_end = line.find(amounts[0]) if amounts else len(line)
            description = line[desc_start:desc_end].strip()
            description = re.sub(r'\s+', ' ', description)[:200]
            
            if description or amount != 0:
                entries.append({
                    "id": f"stmt_{entry_id}",
                    "date": parsed_date,
                    "description": description,
                    "reference": "",
                    "amount": amount,
                    "bank": "fab"
                })
        
        return entries
    
    def _parse_mashreq(self, text: str, fallback_date: str = None) -> List[Dict]:
        """Parser for Mashreq Bank statements"""
        # Mashreq format: Date | Description | Debit | Credit | Balance
        return self._parse_generic(text, fallback_date)
    
    def _parse_rakbank(self, text: str, fallback_date: str = None) -> List[Dict]:
        """Parser for RAK Bank statements"""
        return self._parse_generic(text, fallback_date)
    
    def _parse_dib(self, text: str, fallback_date: str = None) -> List[Dict]:
        """Parser for Dubai Islamic Bank statements"""
        # DIB may have Islamic banking terms
        entries = self._parse_generic(text, fallback_date)
        # Update bank name
        for entry in entries:
            entry["bank"] = "dib"
        return entries
    
    def _parse_cbd(self, text: str, fallback_date: str = None) -> List[Dict]:
        """Parser for Commercial Bank of Dubai statements"""
        return self._parse_generic(text, fallback_date)


def parse_bank_statement_pdf(content: bytes, filename: str, fallback_date: str) -> Tuple[List[Dict], str]:
    """
    Parse a bank statement PDF. Tries pdfplumber first, then OCR.
    """
    entries = []
    bank_name = "unknown"
    
    # --- Step 1: Try pdfplumber text extraction ---
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            all_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    all_text += text + "\n"
            
            if all_text.strip() and len(all_text.strip()) > 50:
                bank_name = detect_bank(all_text, filename)
                parser = BankStatementParser(bank_name)
                entries = parser.parse_ocr_text(all_text, fallback_date)
                if entries:
                    logger.info(f"Parsed {len(entries)} entries from {bank_name} (pdfplumber text)")
                    return entries, bank_name
            
            # Try table extraction
            for page in pdf.pages:
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        for row in table:
                            if row and any(row):
                                all_text += " ".join([str(c) for c in row if c]) + "\n"
            
            if all_text.strip() and len(all_text.strip()) > 50:
                bank_name = detect_bank(all_text, filename)
                parser = BankStatementParser(bank_name)
                entries = parser.parse_ocr_text(all_text, fallback_date)
                if entries:
                    logger.info(f"Parsed {len(entries)} entries from {bank_name} (pdfplumber tables)")
                    return entries, bank_name
    except Exception as e:
        logger.warning(f"pdfplumber extraction failed: {e}")
    
    # --- Step 2: OCR fallback with higher DPI ---
    try:
        from pdf2image import convert_from_bytes
        import pytesseract
    except ImportError:
        logger.error("OCR libraries not available")
        return [], bank_name
    
    try:
        images = convert_from_bytes(content, dpi=300)
    except Exception as e:
        logger.error(f"Failed to convert PDF to images: {e}")
        return [], bank_name
    
    all_text = ""
    for img in images:
        try:
            text = pytesseract.image_to_string(img, lang='eng')
            all_text += text + "\n"
        except Exception as e:
            logger.warning(f"OCR failed for page: {e}")
    
    if not all_text.strip():
        logger.warning("No text extracted from PDF")
        return [], "unknown"
    
    bank_name = detect_bank(all_text, filename)
    parser = BankStatementParser(bank_name)
    entries = parser.parse_ocr_text(all_text, fallback_date)
    
    logger.info(f"Parsed {len(entries)} entries from {bank_name} (OCR)")
    return entries, bank_name


def parse_bank_statement_csv(content: str, filename: str) -> Tuple[List[Dict], str]:
    """Parse a CSV bank statement"""
    import csv
    import io
    
    entries = []
    entry_id = 0
    
    # Detect bank from filename
    bank_name = detect_bank("", filename)
    
    reader = csv.DictReader(io.StringIO(content))
    headers = reader.fieldnames or []
    headers_lower = [h.lower() if h else "" for h in headers]
    
    # Find column indices
    date_cols = [i for i, h in enumerate(headers_lower) if any(x in h for x in ['date', 'trans', 'value'])]
    desc_cols = [i for i, h in enumerate(headers_lower) if any(x in h for x in ['description', 'narration', 'particulars', 'details'])]
    debit_cols = [i for i, h in enumerate(headers_lower) if any(x in h for x in ['debit', 'withdrawal', 'dr'])]
    credit_cols = [i for i, h in enumerate(headers_lower) if any(x in h for x in ['credit', 'deposit', 'cr'])]
    amount_cols = [i for i, h in enumerate(headers_lower) if 'amount' in h]
    ref_cols = [i for i, h in enumerate(headers_lower) if any(x in h for x in ['reference', 'ref', 'txn'])]
    
    for row in reader:
        values = list(row.values())
        entry_id += 1
        
        # Get date
        date_val = None
        for i in date_cols:
            if i < len(values) and values[i]:
                date_val = parse_date(str(values[i]))
                if date_val:
                    break
        
        # Get description
        description = ""
        for i in desc_cols:
            if i < len(values) and values[i]:
                description = str(values[i])[:200]
                break
        
        # Get amount
        amount = 0.0
        if debit_cols and credit_cols:
            debit = 0.0
            credit = 0.0
            for i in debit_cols:
                if i < len(values) and values[i]:
                    debit = parse_amount(str(values[i]))
                    break
            for i in credit_cols:
                if i < len(values) and values[i]:
                    credit = parse_amount(str(values[i]))
                    break
            amount = credit - debit
        elif amount_cols:
            for i in amount_cols:
                if i < len(values) and values[i]:
                    amount = parse_amount(str(values[i]))
                    break
        
        # Get reference
        reference = ""
        for i in ref_cols:
            if i < len(values) and values[i]:
                reference = str(values[i])[:50]
                break
        
        if date_val and (description or amount != 0):
            entries.append({
                "id": f"stmt_{entry_id}",
                "date": date_val,
                "description": description,
                "reference": reference,
                "amount": amount,
                "bank": bank_name
            })
    
    return entries, bank_name


def parse_bank_statement_excel(content: bytes, filename: str) -> Tuple[List[Dict], str]:
    """Parse an Excel bank statement"""
    from openpyxl import load_workbook
    import io
    
    entries = []
    entry_id = 0
    
    # Detect bank from filename
    bank_name = detect_bank("", filename)
    
    wb = load_workbook(filename=io.BytesIO(content), read_only=True)
    ws = wb.active
    
    # Get headers from first row
    headers = [str(cell.value).lower() if cell.value else "" for cell in ws[1]]
    
    # Find column indices
    date_col = next((i for i, h in enumerate(headers) if any(x in h for x in ['date', 'trans'])), 0)
    desc_col = next((i for i, h in enumerate(headers) if any(x in h for x in ['description', 'narration', 'particulars'])), 1)
    debit_col = next((i for i, h in enumerate(headers) if any(x in h for x in ['debit', 'withdrawal'])), -1)
    credit_col = next((i for i, h in enumerate(headers) if any(x in h for x in ['credit', 'deposit'])), -1)
    amount_col = next((i for i, h in enumerate(headers) if 'amount' in h), -1)
    ref_col = next((i for i, h in enumerate(headers) if any(x in h for x in ['reference', 'ref'])), -1)
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        
        entry_id += 1
        
        # Get date
        date_val = parse_date(str(row[date_col])) if date_col < len(row) and row[date_col] else None
        
        # Get description
        description = str(row[desc_col])[:200] if desc_col < len(row) and row[desc_col] else ""
        
        # Get amount
        amount = 0.0
        if debit_col >= 0 and credit_col >= 0:
            debit = parse_amount(str(row[debit_col])) if debit_col < len(row) and row[debit_col] else 0.0
            credit = parse_amount(str(row[credit_col])) if credit_col < len(row) and row[credit_col] else 0.0
            amount = credit - debit
        elif amount_col >= 0 and amount_col < len(row):
            amount = parse_amount(str(row[amount_col]))
        
        # Get reference
        reference = str(row[ref_col])[:50] if ref_col >= 0 and ref_col < len(row) and row[ref_col] else ""
        
        if date_val and (description or amount != 0):
            entries.append({
                "id": f"stmt_{entry_id}",
                "date": date_val,
                "description": description,
                "reference": reference,
                "amount": amount,
                "bank": bank_name
            })
    
    return entries, bank_name


# Supported banks info for UI
SUPPORTED_BANKS = [
    {"id": "emirates_nbd", "name": "Emirates NBD", "country": "UAE", "type": "bank"},
    {"id": "adcb", "name": "Abu Dhabi Commercial Bank (ADCB)", "country": "UAE", "type": "bank"},
    {"id": "fab", "name": "First Abu Dhabi Bank (FAB)", "country": "UAE", "type": "bank"},
    {"id": "mashreq", "name": "Mashreq Bank", "country": "UAE", "type": "bank"},
    {"id": "rakbank", "name": "RAK Bank", "country": "UAE", "type": "bank"},
    {"id": "dib", "name": "Dubai Islamic Bank (DIB)", "country": "UAE", "type": "bank"},
    {"id": "cbd", "name": "Commercial Bank of Dubai (CBD)", "country": "UAE", "type": "bank"},
    {"id": "cbi", "name": "Commercial Bank International (CBI)", "country": "UAE", "type": "bank"},
    {"id": "nbf", "name": "National Bank of Fujairah (NBF)", "country": "UAE", "type": "bank"},
    {"id": "ajman_bank", "name": "Ajman Bank", "country": "UAE", "type": "bank"},
    {"id": "generic", "name": "Other / Generic Bank", "country": "Any", "type": "bank"},
]

# Supported PSPs info for UI
SUPPORTED_PSPS = [
    {"id": "paytabs", "name": "PayTabs", "country": "UAE/MENA", "type": "psp"},
    {"id": "telr", "name": "Telr", "country": "UAE/MENA", "type": "psp"},
    {"id": "network_international", "name": "Network International (N-Genius)", "country": "UAE/MENA", "type": "psp"},
    {"id": "payfort", "name": "PayFort (Amazon Payment Services)", "country": "UAE/MENA", "type": "psp"},
    {"id": "checkout_com", "name": "Checkout.com", "country": "Global", "type": "psp"},
    {"id": "stripe", "name": "Stripe", "country": "Global", "type": "psp"},
    {"id": "paypal", "name": "PayPal", "country": "Global", "type": "psp"},
    {"id": "wise", "name": "Wise (TransferWise)", "country": "Global", "type": "psp"},
    {"id": "tap_payments", "name": "Tap Payments", "country": "UAE/MENA", "type": "psp"},
    {"id": "hyperpay", "name": "HyperPay", "country": "UAE/MENA", "type": "psp"},
    {"id": "payby", "name": "PayBy", "country": "UAE", "type": "psp"},
    {"id": "magnati", "name": "Magnati (FAB Payments)", "country": "UAE", "type": "psp"},
    {"id": "myfatoorah", "name": "MyFatoorah", "country": "UAE/MENA", "type": "psp"},
    {"id": "generic", "name": "Other / Generic PSP", "country": "Any", "type": "psp"},
]


class PSPStatementParser:
    """Parser for PSP (Payment Service Provider) statements"""
    
    def __init__(self, psp_name: str = "generic"):
        self.psp_name = psp_name
        self.entries = []
    
    def parse_ocr_text(self, text: str, fallback_date: str = None) -> List[Dict]:
        """Parse OCR text using PSP-specific rules"""
        # Get PSP-specific parser
        parser_method = getattr(self, f"_parse_{self.psp_name}", self._parse_generic)
        return parser_method(text, fallback_date)
    
    def _parse_generic(self, text: str, fallback_date: str = None) -> List[Dict]:
        """Generic PSP parser"""
        entries = []
        entry_id = 0
        
        lines = text.split('\n')
        
        # PSP statements often have different patterns
        # Common: Transaction ID, Date, Amount, Status, Fee
        date_pattern = r'(\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4})'
        amount_pattern = r'([\d,]+\.\d{2})'
        # Transaction ID patterns
        txn_pattern = r'([A-Z0-9]{8,})'
        
        for line in lines:
            if not line.strip():
                continue
            
            # Look for date
            date_match = re.search(date_pattern, line)
            if not date_match:
                continue
            
            # Find amounts
            amounts = re.findall(amount_pattern, line)
            if not amounts:
                continue
            
            entry_id += 1
            
            # Parse date
            parsed_date = parse_date(date_match.group(1)) or fallback_date
            
            # Get transaction ID if present
            txn_match = re.search(txn_pattern, line)
            reference = txn_match.group(1) if txn_match else ""
            
            # PSP amounts: usually gross amount, fee, net amount
            # First amount is typically the transaction amount
            amount = parse_amount(amounts[0])
            
            # Check for refund/chargeback indicators
            line_lower = line.lower()
            if any(x in line_lower for x in ['refund', 'chargeback', 'reversal', 'dispute']):
                amount = -abs(amount)
            
            # Extract description
            desc_start = line.find(date_match.group(1)) + len(date_match.group(1))
            desc_end = line.find(amounts[0]) if amounts else len(line)
            description = line[desc_start:desc_end].strip()
            description = re.sub(r'\s+', ' ', description)[:200]
            
            if description or amount != 0:
                entries.append({
                    "id": f"psp_{entry_id}",
                    "date": parsed_date,
                    "description": description,
                    "reference": reference,
                    "amount": amount,
                    "psp": self.psp_name
                })
        
        return entries
    
    def _parse_paytabs(self, text: str, fallback_date: str = None) -> List[Dict]:
        """Parser for PayTabs statements"""
        # PayTabs format: Transaction ID, Date, Amount, Currency, Status, Fee
        return self._parse_generic(text, fallback_date)
    
    def _parse_telr(self, text: str, fallback_date: str = None) -> List[Dict]:
        """Parser for Telr statements"""
        return self._parse_generic(text, fallback_date)
    
    def _parse_network_international(self, text: str, fallback_date: str = None) -> List[Dict]:
        """Parser for Network International (N-Genius) statements"""
        return self._parse_generic(text, fallback_date)
    
    def _parse_stripe(self, text: str, fallback_date: str = None) -> List[Dict]:
        """Parser for Stripe statements"""
        entries = []
        entry_id = 0
        
        lines = text.split('\n')
        
        # Stripe patterns
        date_pattern = r'(\w{3}\s+\d{1,2},?\s+\d{4}|\d{4}-\d{2}-\d{2})'
        amount_pattern = r'([\d,]+\.\d{2})'
        stripe_id_pattern = r'(ch_[a-zA-Z0-9]+|pi_[a-zA-Z0-9]+|py_[a-zA-Z0-9]+)'
        
        for line in lines:
            date_match = re.search(date_pattern, line)
            if not date_match:
                continue
            
            amounts = re.findall(amount_pattern, line)
            if not amounts:
                continue
            
            entry_id += 1
            
            parsed_date = parse_date(date_match.group(1)) or fallback_date
            
            stripe_id = re.search(stripe_id_pattern, line)
            reference = stripe_id.group(1) if stripe_id else ""
            
            amount = parse_amount(amounts[0])
            
            line_lower = line.lower()
            if any(x in line_lower for x in ['refund', 'dispute', 'chargeback']):
                amount = -abs(amount)
            
            desc_start = line.find(date_match.group(1)) + len(date_match.group(1))
            desc_end = line.find(amounts[0]) if amounts else len(line)
            description = line[desc_start:desc_end].strip()[:200]
            
            if description or amount != 0:
                entries.append({
                    "id": f"psp_{entry_id}",
                    "date": parsed_date,
                    "description": description,
                    "reference": reference,
                    "amount": amount,
                    "psp": "stripe"
                })
        
        return entries
    
    def _parse_paypal(self, text: str, fallback_date: str = None) -> List[Dict]:
        """Parser for PayPal statements"""
        entries = []
        entry_id = 0
        
        lines = text.split('\n')
        
        date_pattern = r'(\d{1,2}/\d{1,2}/\d{4}|\d{1,2}\s+\w{3}\s+\d{4})'
        amount_pattern = r'([\d,]+\.\d{2})'
        
        for line in lines:
            date_match = re.search(date_pattern, line)
            if not date_match:
                continue
            
            amounts = re.findall(amount_pattern, line)
            if not amounts:
                continue
            
            entry_id += 1
            
            parsed_date = parse_date(date_match.group(1)) or fallback_date
            amount = parse_amount(amounts[0])
            
            line_lower = line.lower()
            # PayPal indicators
            if any(x in line_lower for x in ['refund', 'reversal', 'chargeback', 'withdrawal']):
                amount = -abs(amount)
            elif any(x in line_lower for x in ['payment received', 'subscription']):
                amount = abs(amount)
            
            desc_start = line.find(date_match.group(1)) + len(date_match.group(1))
            desc_end = line.find(amounts[0]) if amounts else len(line)
            description = line[desc_start:desc_end].strip()[:200]
            
            if description or amount != 0:
                entries.append({
                    "id": f"psp_{entry_id}",
                    "date": parsed_date,
                    "description": description,
                    "reference": "",
                    "amount": amount,
                    "psp": "paypal"
                })
        
        return entries


def parse_psp_statement_pdf(content: bytes, filename: str, fallback_date: str) -> Tuple[List[Dict], str]:
    """
    Parse a PSP statement PDF using OCR and PSP-specific rules.
    
    Returns:
        Tuple of (entries list, detected PSP name)
    """
    try:
        from pdf2image import convert_from_bytes
        import pytesseract
    except ImportError:
        logger.error("OCR libraries not available")
        return [], "unknown"
    
    # Convert PDF to images
    try:
        images = convert_from_bytes(content, dpi=200)
    except Exception as e:
        logger.error(f"Failed to convert PDF to images: {e}")
        return [], "unknown"
    
    # OCR all pages
    all_text = ""
    for img in images:
        try:
            text = pytesseract.image_to_string(img)
            all_text += text + "\n"
        except Exception as e:
            logger.warning(f"OCR failed for page: {e}")
    
    if not all_text.strip():
        logger.warning("No text extracted from PDF")
        return [], "unknown"
    
    # Detect PSP
    psp_name = detect_psp(all_text, filename)
    
    # Parse using appropriate parser
    parser = PSPStatementParser(psp_name)
    entries = parser.parse_ocr_text(all_text, fallback_date)
    
    logger.info(f"Parsed {len(entries)} entries from {psp_name} statement")
    
    return entries, psp_name


def parse_psp_statement_csv(content: str, filename: str) -> Tuple[List[Dict], str]:
    """Parse a CSV PSP statement"""
    import csv
    import io
    
    entries = []
    entry_id = 0
    
    # Detect PSP from filename
    psp_name = detect_psp("", filename)
    
    reader = csv.DictReader(io.StringIO(content))
    headers = reader.fieldnames or []
    headers_lower = [h.lower() if h else "" for h in headers]
    
    # Find column indices - PSP specific columns
    date_cols = [i for i, h in enumerate(headers_lower) if any(x in h for x in ['date', 'created', 'timestamp', 'time'])]
    desc_cols = [i for i, h in enumerate(headers_lower) if any(x in h for x in ['description', 'type', 'name', 'customer'])]
    amount_cols = [i for i, h in enumerate(headers_lower) if any(x in h for x in ['amount', 'gross', 'total'])]
    fee_cols = [i for i, h in enumerate(headers_lower) if any(x in h for x in ['fee', 'charge', 'commission'])]
    net_cols = [i for i, h in enumerate(headers_lower) if any(x in h for x in ['net', 'payout'])]
    ref_cols = [i for i, h in enumerate(headers_lower) if any(x in h for x in ['reference', 'transaction', 'id', 'order'])]
    status_cols = [i for i, h in enumerate(headers_lower) if any(x in h for x in ['status', 'state'])]
    
    for row in reader:
        values = list(row.values())
        entry_id += 1
        
        # Get date
        date_val = None
        for i in date_cols:
            if i < len(values) and values[i]:
                date_val = parse_date(str(values[i]))
                if date_val:
                    break
        
        # Get description
        description = ""
        for i in desc_cols:
            if i < len(values) and values[i]:
                description = str(values[i])[:200]
                break
        
        # Get amount (prefer net amount for PSPs)
        amount = 0.0
        if net_cols:
            for i in net_cols:
                if i < len(values) and values[i]:
                    amount = parse_amount(str(values[i]))
                    break
        elif amount_cols:
            for i in amount_cols:
                if i < len(values) and values[i]:
                    amount = parse_amount(str(values[i]))
                    break
        
        # Get reference
        reference = ""
        for i in ref_cols:
            if i < len(values) and values[i]:
                reference = str(values[i])[:50]
                break
        
        # Check status for refunds
        status = ""
        for i in status_cols:
            if i < len(values) and values[i]:
                status = str(values[i]).lower()
                if any(x in status for x in ['refund', 'reversed', 'chargeback', 'dispute']):
                    amount = -abs(amount)
                break
        
        if date_val and (description or amount != 0):
            entries.append({
                "id": f"psp_{entry_id}",
                "date": date_val,
                "description": description,
                "reference": reference,
                "amount": amount,
                "psp": psp_name
            })
    
    return entries, psp_name

