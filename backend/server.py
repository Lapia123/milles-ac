from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, UploadFile, File, Form, BackgroundTasks, Body, Query
from fastapi.responses import HTMLResponse
from fastapi.security import HTTPBearer
from dotenv import load_dotenv
import asyncio
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
from enum import Enum
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import httpx
import base64
import aiosmtplib
import boto3
from botocore.config import Config as BotoConfig
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
from cache import (
    get_cached, set_cached, get_cache_key, CACHE_TTL,
    invalidate_vendor_cache, invalidate_ie_cache, invalidate_transaction_cache,
    invalidate_loan_cache, invalidate_treasury_cache, is_redis_available
)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection with connection pooling and timeouts
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(
    mongo_url,
    maxPoolSize=50,
    minPoolSize=10,
    maxIdleTimeMS=30000,
    serverSelectionTimeoutMS=5000,
    connectTimeoutMS=10000,
    socketTimeoutMS=30000,
    retryWrites=True,
    retryReads=True
)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 2

# Cloudflare R2 Configuration
R2_ACCOUNT_ID = os.environ.get('R2_ACCOUNT_ID')
R2_ACCESS_KEY_ID = os.environ.get('R2_ACCESS_KEY_ID')
R2_SECRET_ACCESS_KEY = os.environ.get('R2_SECRET_ACCESS_KEY')
R2_BUCKET_NAME = os.environ.get('R2_BUCKET_NAME')
R2_PUBLIC_URL = os.environ.get('R2_PUBLIC_URL', '').rstrip('/')

s3_client = boto3.client(
    's3',
    endpoint_url=f"https://{R2_ACCOUNT_ID}.r2.cloudflarestorage.com",
    aws_access_key_id=R2_ACCESS_KEY_ID,
    aws_secret_access_key=R2_SECRET_ACCESS_KEY,
    config=BotoConfig(signature_version='s3v4'),
    region_name='auto'
)

def upload_to_r2(file_content: bytes, filename: str, content_type: str = "application/octet-stream", folder: str = "uploads") -> str:
    """Upload a file to R2 and return the public URL."""
    ext = filename.rsplit('.', 1)[-1] if '.' in filename else ''
    key = f"{folder}/{uuid.uuid4().hex[:16]}_{filename}"
    s3_client.put_object(
        Bucket=R2_BUCKET_NAME,
        Key=key,
        Body=file_content,
        ContentType=content_type,
    )
    return f"{R2_PUBLIC_URL}/{key}"

# Create the main app
app = FastAPI(title="FX Broker Back-Office API")

# Create router with /api prefix
api_router = APIRouter(prefix="/api")

# Security
security = HTTPBearer(auto_error=False)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============== PAGINATION HELPER ==============
class PaginatedResponse(BaseModel):
    items: List
    total: int
    page: int
    page_size: int
    total_pages: int

async def paginate_query(collection, query: dict, page: int = 1, page_size: int = 50, 
                         sort_field: str = "created_at", sort_order: int = -1,
                         projection: dict = None) -> dict:
    """Helper function for paginated queries with caching"""
    if projection is None:
        projection = {"_id": 0}
    
    skip = (page - 1) * page_size
    
    # Get total count
    total = await collection.count_documents(query)
    total_pages = (total + page_size - 1) // page_size if total > 0 else 1
    
    # Get items
    cursor = collection.find(query, projection).sort(sort_field, sort_order).skip(skip).limit(page_size)
    items = await cursor.to_list(page_size)
    
    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages
    }

# ============== MODELS ==============

class UserRole:
    ADMIN = "admin"
    SUB_ADMIN = "sub_admin"
    ACCOUNTANT = "accountant"
    VENDOR = "vendor"

# ============== GRANULAR ACCESS CONTROL MODELS ==============

# All modules in the system
class Modules:
    DASHBOARD = "dashboard"
    CLIENTS = "clients"
    TRANSACTIONS = "transactions"
    TREASURY = "treasury"
    LP_MANAGEMENT = "lp_management"
    INCOME_EXPENSES = "income_expenses"
    LOANS = "loans"
    DEBTS = "debts"
    PSP = "psp"
    EXCHANGERS = "exchangers"
    RECONCILIATION = "reconciliation"
    AUDIT = "audit"
    LOGS = "logs"
    REPORTS = "reports"
    SETTINGS = "settings"
    USERS = "users"
    ROLES = "roles"
    MESSAGES = "messages"
    APPROVALS = "approvals"
    TRANSACTION_REQUESTS = "transaction_requests"

# Standard actions
class Actions:
    VIEW = "view"
    CREATE = "create"
    EDIT = "edit"
    DELETE = "delete"
    APPROVE = "approve"
    EXPORT = "export"

# All modules list
ALL_MODULES = [
    Modules.DASHBOARD, Modules.CLIENTS, Modules.TRANSACTIONS, Modules.TREASURY,
    Modules.LP_MANAGEMENT, Modules.INCOME_EXPENSES, Modules.LOANS, Modules.DEBTS,
    Modules.PSP, Modules.EXCHANGERS, Modules.RECONCILIATION, Modules.AUDIT,
    Modules.LOGS, Modules.REPORTS, Modules.SETTINGS, Modules.USERS, Modules.ROLES,
    Modules.MESSAGES,
    Modules.APPROVALS,
    Modules.TRANSACTION_REQUESTS
]

# All actions list
ALL_ACTIONS = [Actions.VIEW, Actions.CREATE, Actions.EDIT, Actions.DELETE, Actions.APPROVE, Actions.EXPORT]

# Module display names
MODULE_DISPLAY_NAMES = {
    Modules.DASHBOARD: "Dashboard",
    Modules.CLIENTS: "Clients",
    Modules.TRANSACTIONS: "Transactions",
    Modules.TREASURY: "Treasury",
    Modules.LP_MANAGEMENT: "LP Management",
    Modules.INCOME_EXPENSES: "Income & Expenses",
    Modules.LOANS: "Loans",
    Modules.DEBTS: "O/S Accounts",
    Modules.PSP: "PSP",
    Modules.EXCHANGERS: "Exchangers",
    Modules.RECONCILIATION: "Reconciliation",
    Modules.AUDIT: "Audit",
    Modules.LOGS: "Logs",
    Modules.REPORTS: "Reports",
    Modules.SETTINGS: "Settings",
    Modules.USERS: "Users",
    Modules.ROLES: "Roles & Permissions",
    Modules.MESSAGES: "Messages",
    Modules.APPROVALS: "Pending Approvals",
    Modules.TRANSACTION_REQUESTS: "Transaction Requests"
}

class RoleCreate(BaseModel):
    name: str
    display_name: str
    description: Optional[str] = None
    permissions: dict = {}  # {module: [actions]}
    is_system_role: bool = False
    hierarchy_level: int = 0  # Higher = more access

class RoleUpdate(BaseModel):
    display_name: Optional[str] = None
    description: Optional[str] = None
    permissions: Optional[dict] = None
    hierarchy_level: Optional[int] = None
    is_active: Optional[bool] = None

class UserPermissionOverride(BaseModel):
    user_id: str
    permissions: dict = {}  # Custom permission overrides for this user

class UserBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str
    email: str
    name: str
    role: str = UserRole.SUB_ADMIN
    picture: Optional[str] = None
    is_active: bool = True
    created_at: datetime

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: str = UserRole.SUB_ADMIN
    # Exchanger-specific fields (only used when role is 'vendor')
    deposit_commission: Optional[float] = 0.0
    withdrawal_commission: Optional[float] = 0.0

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    role_id: Optional[str] = None
    is_active: Optional[bool] = None

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: dict

class ClientStatus:
    PENDING = "pending"
    APPROVED = "approved"
    REJECTED = "rejected"
    SUSPENDED = "suspended"

class ClientBase(BaseModel):
    model_config = ConfigDict(extra="ignore")
    client_id: str
    first_name: str
    last_name: str
    email: str
    phone: Optional[str] = None
    country: Optional[str] = None
    mt5_number: Optional[str] = None
    crm_customer_id: Optional[str] = None
    kyc_status: str = ClientStatus.PENDING
    kyc_documents: List[str] = []
    notes: Optional[str] = None
    created_at: datetime
    updated_at: datetime

class ClientCreate(BaseModel):
    first_name: str
    last_name: str
    email: EmailStr
    phone: Optional[str] = None
    country: Optional[str] = None
    mt5_number: Optional[str] = None
    crm_customer_id: Optional[str] = None
    notes: Optional[str] = None
    tags: Optional[List[str]] = []

class ClientUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    country: Optional[str] = None
    mt5_number: Optional[str] = None
    crm_customer_id: Optional[str] = None
    kyc_status: Optional[str] = None
    notes: Optional[str] = None
    tags: Optional[List[str]] = None

# Treasury/Bank Account Models
class TreasuryAccountType:
    BANK = "bank"
    CRYPTO_WALLET = "crypto_wallet"
    PAYMENT_GATEWAY = "payment_gateway"
    USDT = "usdt"  # USDT Wallet

class TreasuryAccountStatus:
    ACTIVE = "active"
    INACTIVE = "inactive"

# LP (Liquidity Provider) Models
class LPAccountStatus:
    ACTIVE = "active"
    INACTIVE = "inactive"

class LPTransactionType:
    DEPOSIT = "deposit"
    WITHDRAWAL = "withdrawal"

class LPAccountCreate(BaseModel):
    lp_name: str
    account_number: Optional[str] = None
    bank_name: Optional[str] = None
    swift_code: Optional[str] = None
    currency: str = "USD"
    contact_person: Optional[str] = None
    contact_email: Optional[str] = None
    contact_phone: Optional[str] = None
    notes: Optional[str] = None

class LPAccountUpdate(BaseModel):
    lp_name: Optional[str] = None
    account_number: Optional[str] = None
    bank_name: Optional[str] = None
    swift_code: Optional[str] = None
    currency: Optional[str] = None
    contact_person: Optional[str] = None
    contact_email: Optional[str] = None
    contact_phone: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = None

class LPTransactionCreate(BaseModel):
    amount: float
    currency: str = "USD"
    treasury_account_id: Optional[str] = None  # Source/destination treasury
    reference: Optional[str] = None
    notes: Optional[str] = None

# Dealing P&L Models
class LPPnLEntry(BaseModel):
    lp_id: str
    lp_name: Optional[str] = None
    booked_pnl: float = 0
    floating_pnl: float = 0

class DealingPnLCreate(BaseModel):
    date: str  # YYYY-MM-DD format
    # MT5 Client Side
    mt5_booked_pnl: float = 0  # Client booked P&L (positive = client profit)
    mt5_floating_pnl: float = 0  # Today's running floating P&L
    # LP Hedging Side - Multiple LPs
    lp_entries: Optional[List[LPPnLEntry]] = []  # List of LP P&L entries
    # Legacy single LP fields (for backward compatibility)
    lp_booked_pnl: float = 0
    lp_floating_pnl: float = 0
    # Screenshots (base64 or URLs)
    mt5_screenshot: Optional[str] = None
    lp_screenshot: Optional[str] = None
    notes: Optional[str] = None

# PSP Models
class PSPStatus:
    ACTIVE = "active"
    INACTIVE = "inactive"

class PSPSettlementStatus:
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"

class CommissionPaidBy:
    CLIENT = "client"
    BROKER = "broker"

class PSPCreate(BaseModel):
    psp_name: str
    commission_rate: float  # percentage e.g., 2.5 for 2.5%
    reserve_fund_rate: float = 0  # percentage for reserve fund
    holding_days: int = 0  # days PSP holds funds before release
    settlement_days: int = 1  # T+1, T+2, etc.
    settlement_destination_id: str  # Treasury account ID
    min_settlement_amount: float = 0
    gateway_fee: float = 0  # Fixed fee per transaction
    refund_fee: float = 0  # Fee charged when processing refunds
    monthly_minimum_fee: float = 0  # Minimum monthly charge by PSP
    description: Optional[str] = None

class PSPUpdate(BaseModel):
    psp_name: Optional[str] = None
    commission_rate: Optional[float] = None
    reserve_fund_rate: Optional[float] = None
    holding_days: Optional[int] = None
    settlement_days: Optional[int] = None
    settlement_destination_id: Optional[str] = None
    min_settlement_amount: Optional[float] = None
    gateway_fee: Optional[float] = None
    refund_fee: Optional[float] = None
    monthly_minimum_fee: Optional[float] = None
    status: Optional[str] = None
    description: Optional[str] = None

# Vendor Models
class VendorStatus:
    ACTIVE = "active"
    INACTIVE = "inactive"

class VendorSettlementType:
    BANK = "bank"
    CASH = "cash"

class VendorSettlementStatus:
    PENDING = "pending"
    APPROVED = "approved"
    COMPLETED = "completed"
    REJECTED = "rejected"

class VendorCreate(BaseModel):
    vendor_name: str
    email: EmailStr
    password: str
    deposit_commission: float = 0  # percentage for deposits (bank)
    withdrawal_commission: float = 0  # percentage for withdrawals (bank)
    deposit_commission_cash: float = 0  # percentage for deposits (cash)
    withdrawal_commission_cash: float = 0  # percentage for withdrawals (cash)
    description: Optional[str] = None

class VendorUpdate(BaseModel):
    vendor_name: Optional[str] = None
    deposit_commission: Optional[float] = None
    withdrawal_commission: Optional[float] = None
    deposit_commission_cash: Optional[float] = None
    withdrawal_commission_cash: Optional[float] = None
    status: Optional[str] = None
    description: Optional[str] = None

# Income & Expense Models
class IncomeExpenseType:
    INCOME = "income"
    EXPENSE = "expense"

class IncomeCategory:
    COMMISSION = "commission"
    SERVICE_FEE = "service_fee"
    INTEREST = "interest"
    OTHER = "other"

class ExpenseCategory:
    BANK_FEE = "bank_fee"
    TRANSFER_CHARGE = "transfer_charge"
    VENDOR_PAYMENT = "vendor_payment"
    OPERATIONAL = "operational"
    MARKETING = "marketing"
    SOFTWARE = "software"
    OTHER = "other"

class IncomeExpenseCreate(BaseModel):
    entry_type: str  # income or expense
    category: Optional[str] = None
    custom_category: Optional[str] = None
    amount: float
    currency: str = "USD"
    base_currency: Optional[str] = None  # Payment currency (e.g., INR)
    base_amount: Optional[float] = None  # Amount in payment currency
    exchange_rate: Optional[float] = None  # Rate: 1 base_currency = ? USD
    treasury_account_id: Optional[str] = None  # Optional when vendor handles it
    vendor_id: Optional[str] = None  # If linked to an exchanger (money partner)
    vendor_supplier_id: Optional[str] = None  # If linked to a service supplier (rent, utilities)
    client_id: Optional[str] = None  # If linked to a client
    ie_category_id: Optional[str] = None  # Custom account category
    vendor_bank_account_name: Optional[str] = None
    vendor_bank_account_number: Optional[str] = None
    vendor_bank_ifsc: Optional[str] = None
    vendor_bank_branch: Optional[str] = None
    description: Optional[str] = None
    reference: Optional[str] = None
    date: Optional[str] = None  # ISO date string
    transaction_mode: Optional[str] = None  # "bank" or "cash"
    collecting_person_name: Optional[str] = None
    collecting_person_number: Optional[str] = None

class IncomeExpenseUpdate(BaseModel):
    category: Optional[str] = None
    custom_category: Optional[str] = None
    amount: Optional[float] = None
    description: Optional[str] = None
    reference: Optional[str] = None

# Loan Models
class LoanStatus:
    ACTIVE = "active"
    PARTIALLY_PAID = "partially_paid"
    FULLY_PAID = "fully_paid"
    OVERDUE = "overdue"
    PENDING_APPROVAL = "pending_approval"
    WRITTEN_OFF = "written_off"

class LoanType:
    SHORT_TERM = "short_term"  # < 1 year
    LONG_TERM = "long_term"    # > 1 year
    CREDIT_LINE = "credit_line"  # Revolving

class RepaymentMode:
    LUMP_SUM = "lump_sum"
    EMI = "emi"  # Monthly installments
    CUSTOM = "custom"  # Custom schedule

class LoanTransactionType:
    DISBURSEMENT = "disbursement"
    REPAYMENT = "repayment"
    INTEREST_PAYMENT = "interest_payment"
    PENALTY = "penalty"
    SWAP_OUT = "swap_out"  # Transferred to another borrower
    SWAP_IN = "swap_in"    # Received from another borrower
    WRITE_OFF = "write_off"
    REFINANCE = "refinance"

# ============== LOGGING SYSTEM ==============
class LogType(str, Enum):
    ACTIVITY = "activity"
    AUTH = "auth"
    AUDIT = "audit"
    ERROR = "error"

class LogAction(str, Enum):
    # Auth actions
    LOGIN = "login"
    LOGOUT = "logout"
    LOGIN_FAILED = "login_failed"
    PASSWORD_CHANGE = "password_change"
    # CRUD actions
    CREATE = "create"
    READ = "read"
    UPDATE = "update"
    DELETE = "delete"
    # Transaction actions
    APPROVE = "approve"
    REJECT = "reject"
    UPLOAD = "upload"
    EXPORT = "export"
    # System actions
    SYSTEM_ERROR = "system_error"

async def create_log(
    log_type: str,
    action: str,
    module: str,
    user_id: str = None,
    user_name: str = None,
    user_email: str = None,
    user_role: str = None,
    description: str = None,
    details: dict = None,
    ip_address: str = None,
    user_agent: str = None,
    reference_id: str = None,
    old_value: dict = None,
    new_value: dict = None,
    status: str = "success"
):
    """Create a log entry in the database"""
    try:
        log_entry = {
            "log_id": f"log_{uuid.uuid4().hex[:12]}",
            "log_type": log_type,
            "action": action,
            "module": module,
            "user_id": user_id,
            "user_name": user_name,
            "user_email": user_email,
            "user_role": user_role,
            "description": description,
            "details": details or {},
            "ip_address": ip_address,
            "user_agent": user_agent,
            "reference_id": reference_id,
            "old_value": old_value,
            "new_value": new_value,
            "status": status,
            "timestamp": datetime.now(timezone.utc).isoformat()
        }
        await db.system_logs.insert_one(log_entry)
    except Exception as e:
        logger.error(f"Failed to create log: {e}")

async def log_activity(request: Request, user: dict, action: str, module: str, description: str, reference_id: str = None, details: dict = None):
    """Helper to log activity with request info"""
    ip_address = request.client.host if request.client else None
    user_agent = request.headers.get("user-agent", "")
    await create_log(
        log_type="activity",
        action=action,
        module=module,
        user_id=user.get("user_id"),
        user_name=user.get("name"),
        user_email=user.get("email"),
        user_role=user.get("role"),
        description=description,
        details=details,
        ip_address=ip_address,
        user_agent=user_agent,
        reference_id=reference_id
    )

async def log_audit(request: Request, user: dict, action: str, module: str, reference_id: str, old_value: dict = None, new_value: dict = None, description: str = None):
    """Helper to log audit trail for financial changes"""
    ip_address = request.client.host if request.client else None
    user_agent = request.headers.get("user-agent", "")
    await create_log(
        log_type="audit",
        action=action,
        module=module,
        user_id=user.get("user_id"),
        user_name=user.get("name"),
        user_email=user.get("email"),
        user_role=user.get("role"),
        description=description,
        ip_address=ip_address,
        user_agent=user_agent,
        reference_id=reference_id,
        old_value=old_value,
        new_value=new_value
    )

class LoanCreate(BaseModel):
    vendor_id: Optional[str] = None  # Link to vendor (borrower company)
    borrower_name: str
    amount: float
    currency: str = "USD"
    interest_rate: float = 0  # Annual percentage (Simple interest)
    loan_type: str = LoanType.SHORT_TERM
    loan_date: str  # ISO date
    due_date: str  # ISO date
    repayment_mode: str = RepaymentMode.LUMP_SUM
    installment_amount: Optional[float] = None  # For EMI mode
    installment_frequency: Optional[str] = None  # monthly, weekly, etc.
    num_installments: Optional[int] = None  # Number of EMIs
    treasury_account_id: Optional[str] = None  # Source treasury account (optional if disburse_from_vendor_id is provided)
    disburse_from_vendor_id: Optional[str] = None  # Source vendor/exchanger account
    bank_details: Optional[str] = None  # Bank account details for Exchanger to see
    collateral: Optional[str] = None  # Security/collateral description
    notes: Optional[str] = None

class LoanUpdate(BaseModel):
    borrower_name: Optional[str] = None
    interest_rate: Optional[float] = None
    loan_type: Optional[str] = None
    due_date: Optional[str] = None
    repayment_mode: Optional[str] = None
    installment_amount: Optional[float] = None
    installment_frequency: Optional[str] = None
    num_installments: Optional[int] = None
    collateral: Optional[str] = None
    notes: Optional[str] = None

class LoanRepaymentCreate(BaseModel):
    amount: float
    currency: str = "USD"
    treasury_account_id: Optional[str] = None  # Where repayment goes (optional if credit_to_vendor_id provided)
    credit_to_vendor_id: Optional[str] = None  # Credit to vendor/exchanger instead of treasury
    payment_date: Optional[str] = None
    reference: Optional[str] = None
    notes: Optional[str] = None
    include_interest: bool = False  # If true, part of payment goes to interest
    exchange_rate: Optional[float] = None  # Custom exchange rate (payment currency -> loan currency)
    amount_in_loan_currency: Optional[float] = None  # Pre-calculated equivalent in loan currency

class LoanSwapRequest(BaseModel):
    target_vendor_id: Optional[str] = None  # New borrower vendor
    target_borrower_name: str  # New borrower name
    reason: Optional[str] = None
    adjust_terms: bool = False  # If true, allows term changes
    new_interest_rate: Optional[float] = None
    new_due_date: Optional[str] = None


# ============== VENDOR SUPPLIER MODELS (Service Suppliers - Rent, Utilities, etc.) ==============
class VendorSupplierStatus:
    ACTIVE = "active"
    INACTIVE = "inactive"

class VendorSupplierCreate(BaseModel):
    name: str
    contact_person: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    # Bank details
    bank_name: Optional[str] = None
    bank_account_name: Optional[str] = None
    bank_account_number: Optional[str] = None
    bank_ifsc: Optional[str] = None
    bank_branch: Optional[str] = None
    notes: Optional[str] = None

class VendorSupplierUpdate(BaseModel):
    name: Optional[str] = None
    contact_person: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    bank_name: Optional[str] = None
    bank_account_name: Optional[str] = None
    bank_account_number: Optional[str] = None
    bank_ifsc: Optional[str] = None
    bank_branch: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None

# ============== IE CATEGORY MODELS (Account Categories) ==============
class IECategoryType:
    INCOME = "income"
    EXPENSE = "expense"
    BOTH = "both"  # Can be used for both income and expense

class IECategoryCreate(BaseModel):
    name: str
    category_type: str = IECategoryType.BOTH  # income, expense, or both
    description: Optional[str] = None
    parent_category_id: Optional[str] = None  # For subcategories

class IECategoryUpdate(BaseModel):
    name: Optional[str] = None
    category_type: Optional[str] = None
    description: Optional[str] = None
    parent_category_id: Optional[str] = None
    is_active: Optional[bool] = None


# ============== LIVE FX RATE SERVICE ==============
FALLBACK_RATES_TO_USD = {
    "USD": 1.0, "EUR": 1.08, "GBP": 1.27, "AED": 0.27,
    "SAR": 0.27, "INR": 0.012, "JPY": 0.0067, "USDT": 1.0,
}

_fx_cache = {"rates": None, "fetched_at": None, "source": "fallback"}
FX_CACHE_TTL = timedelta(hours=1)

async def fetch_live_rates() -> dict:
    """Fetch live rates from open.er-api.com (USD base). Returns {currency: rate_to_usd}."""
    try:
        async with httpx.AsyncClient(timeout=10) as client_http:
            resp = await client_http.get("https://open.er-api.com/v6/latest/USD")
            resp.raise_for_status()
            data = resp.json()
            if data.get("result") != "success":
                raise ValueError("API returned non-success result")
            raw = data["rates"]  # e.g. {"USD":1, "AED":3.6725, ...}
            # Convert to "rate_to_usd" format: 1 unit of currency = X USD
            rates_to_usd = {}
            for code, val in raw.items():
                rates_to_usd[code] = round(1.0 / val, 8) if val else 0
            return rates_to_usd
    except Exception as e:
        logger.warning(f"Live FX fetch failed: {e}")
        return None

async def get_fx_rates() -> dict:
    """Return cached rates dict {currency: rate_to_usd}. Refreshes if stale."""
    now = datetime.now(timezone.utc)
    if (
        _fx_cache["rates"]
        and _fx_cache["fetched_at"]
        and (now - _fx_cache["fetched_at"]) < FX_CACHE_TTL
    ):
        return _fx_cache["rates"]
    live = await fetch_live_rates()
    if live:
        _fx_cache["rates"] = live
        _fx_cache["fetched_at"] = now
        _fx_cache["source"] = "live"
        return live
    if _fx_cache["rates"]:
        return _fx_cache["rates"]
    _fx_cache["rates"] = FALLBACK_RATES_TO_USD
    _fx_cache["source"] = "fallback"
    return FALLBACK_RATES_TO_USD

def convert_to_usd(amount: float, currency: str) -> float:
    """Sync wrapper – uses whatever is in cache (or fallback)."""
    rates = _fx_cache.get("rates") or FALLBACK_RATES_TO_USD
    rate = rates.get(currency.upper(), 1.0)
    return round(amount * rate, 2)

def convert_from_usd(amount: float, target_currency: str) -> float:
    """Convert USD amount to target currency."""
    rates = _fx_cache.get("rates") or FALLBACK_RATES_TO_USD
    rate = rates.get(target_currency.upper(), 1.0)
    if rate == 0:
        return amount
    return round(amount / rate, 2)

def convert_currency(amount: float, from_currency: str, to_currency: str) -> float:
    """Convert amount between any two currencies via USD as intermediate."""
    if from_currency.upper() == to_currency.upper():
        return amount
    usd_amount = convert_to_usd(amount, from_currency)
    return round(convert_from_usd(usd_amount, to_currency), 2)

class TreasuryAccountCreate(BaseModel):
    account_name: str
    account_type: str = TreasuryAccountType.BANK
    bank_name: Optional[str] = None
    account_number: Optional[str] = None
    routing_number: Optional[str] = None
    swift_code: Optional[str] = None
    currency: str = "USD"
    description: Optional[str] = None
    opening_balance: float = 0.0
    # USDT specific fields
    usdt_address: Optional[str] = None
    usdt_network: Optional[str] = None  # TRC20, ERC20, BEP20
    usdt_notes: Optional[str] = None

class TreasuryAccountUpdate(BaseModel):
    account_name: Optional[str] = None
    bank_name: Optional[str] = None
    account_number: Optional[str] = None
    routing_number: Optional[str] = None
    swift_code: Optional[str] = None
    currency: Optional[str] = None
    status: Optional[str] = None
    description: Optional[str] = None
    # USDT specific fields
    usdt_address: Optional[str] = None
    usdt_network: Optional[str] = None
    usdt_notes: Optional[str] = None

class TransactionType:
    DEPOSIT = "deposit"
    WITHDRAWAL = "withdrawal"
    TRANSFER = "transfer"
    COMMISSION = "commission"
    REBATE = "rebate"
    ADJUSTMENT = "adjustment"

class TransactionStatus:
    PENDING = "pending"
    APPROVED = "approved"
    COMPLETED = "completed"
    REJECTED = "rejected"
    CANCELLED = "cancelled"
    FAILED = "failed"

class TransactionCreate(BaseModel):
    client_id: str
    transaction_type: str
    amount: float
    currency: str = "USD"
    base_currency: str = "USD"
    base_amount: Optional[float] = None
    destination_type: str = "treasury"  # "treasury", "psp", "vendor", "bank", "usdt"
    destination_account_id: Optional[str] = None
    psp_id: Optional[str] = None
    vendor_id: Optional[str] = None
    commission_paid_by: Optional[str] = None  # "client" or "broker"
    description: Optional[str] = None
    reference: Optional[str] = None
    transaction_mode: Optional[str] = None  # "bank" or "cash"
    collecting_person_name: Optional[str] = None
    collecting_person_number: Optional[str] = None
    # Client bank details (for withdrawal to bank)
    client_bank_name: Optional[str] = None
    client_bank_account_name: Optional[str] = None
    client_bank_account_number: Optional[str] = None
    client_bank_swift_iban: Optional[str] = None
    client_bank_currency: Optional[str] = None
    # Client USDT details (for withdrawal to USDT)
    client_usdt_address: Optional[str] = None
    client_usdt_network: Optional[str] = None  # TRC20, ERC20, BEP20

class TransactionUpdate(BaseModel):
    status: Optional[str] = None
    description: Optional[str] = None
    rejection_reason: Optional[str] = None
    crm_reference: Optional[str] = None
    amount: Optional[float] = None
    base_amount: Optional[float] = None
    base_currency: Optional[str] = None
    exchange_rate: Optional[float] = None
    reference: Optional[str] = None
    transaction_date: Optional[str] = None
    client_tags: Optional[list] = None

# ============== HELPER FUNCTIONS ==============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode('utf-8'), hashed.encode('utf-8'))

def create_jwt_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> dict:
    # Check cookie first
    session_token = request.cookies.get("session_token")
    if session_token:
        session = await db.user_sessions.find_one({"session_token": session_token}, {"_id": 0})
        if session:
            expires_at = session.get("expires_at")
            if isinstance(expires_at, str):
                expires_at = datetime.fromisoformat(expires_at)
            if expires_at.tzinfo is None:
                expires_at = expires_at.replace(tzinfo=timezone.utc)
            if expires_at > datetime.now(timezone.utc):
                user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
                if user:
                    return user
    
    # Check Authorization header
    auth_header = request.headers.get("Authorization")
    if auth_header and auth_header.startswith("Bearer "):
        token = auth_header.split(" ")[1]
        # Check if it's a session token
        session = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
        if session:
            expires_at = session.get("expires_at")
            if isinstance(expires_at, str):
                expires_at = datetime.fromisoformat(expires_at)
            if expires_at.tzinfo is None:
                expires_at = expires_at.replace(tzinfo=timezone.utc)
            if expires_at > datetime.now(timezone.utc):
                user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
                if user:
                    return user
        
        # Try JWT token
        try:
            payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
            user = await db.users.find_one({"user_id": payload["user_id"]}, {"_id": 0})
            if user:
                return user
        except jwt.ExpiredSignatureError:
            pass
        except jwt.InvalidTokenError:
            pass
    
    raise HTTPException(status_code=401, detail="Not authenticated")

async def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

async def require_accountant_or_admin(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") not in [UserRole.ADMIN, UserRole.ACCOUNTANT]:
        raise HTTPException(status_code=403, detail="Accountant or Admin access required")
    return user

async def require_vendor(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != UserRole.VENDOR:
        raise HTTPException(status_code=403, detail="Vendor access required")
    return user

async def require_vendor_or_admin(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") not in [UserRole.ADMIN, UserRole.VENDOR]:
        raise HTTPException(status_code=403, detail="Vendor or Admin access required")
    return user

# ============== GRANULAR PERMISSION SYSTEM ==============

async def get_user_permissions(user_id: str) -> dict:
    """Get all permissions for a user (role permissions + user overrides)"""
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not user:
        return {}
    
    # Get role permissions
    role_name = user.get("role_id") or user.get("role", "")
    role = await db.roles.find_one({"role_id": role_name}, {"_id": 0})
    if not role:
        role = await db.roles.find_one({"name": role_name}, {"_id": 0})
    
    base_permissions = role.get("permissions", {}) if role else {}
    
    # Get user-specific overrides
    user_overrides = user.get("permission_overrides", {})
    
    # Merge: user overrides take precedence
    final_permissions = {**base_permissions}
    for module, actions in user_overrides.items():
        if module in final_permissions:
            final_permissions[module] = list(set(final_permissions[module] + actions))
        else:
            final_permissions[module] = actions
    
    return final_permissions

async def check_permission(user: dict, module: str, action: str) -> bool:
    """Check if user has permission for a specific action on a module"""
    # Super admin bypass - users with 'admin' role from old system
    if user.get("role") == UserRole.ADMIN and not user.get("role_id"):
        return True
    
    permissions = await get_user_permissions(user.get("user_id"))
    
    # Check if module exists and action is allowed
    if module in permissions:
        return action in permissions[module] or "*" in permissions[module]
    
    return False

def require_permission(module: str, action: str):
    """Decorator factory for permission checking"""
    async def permission_checker(user: dict = Depends(get_current_user)) -> dict:
        has_permission = await check_permission(user, module, action)
        if not has_permission:
            raise HTTPException(
                status_code=403, 
                detail=f"Permission denied: {action} on {module}"
            )
        return user
    return permission_checker

# Default role templates
DEFAULT_ROLES = [
    {
        "role_id": "admin",
        "name": "admin",
        "display_name": "Admin",
        "description": "Full system access with all permissions",
        "is_system_role": True,
        "hierarchy_level": 100,
        "permissions": {module: ALL_ACTIONS.copy() for module in ALL_MODULES}
    },
    {
        "role_id": "accountant",
        "name": "accountant",
        "display_name": "Accountant",
        "description": "Financial operations and approvals",
        "is_system_role": True,
        "hierarchy_level": 70,
        "permissions": {
            Modules.DASHBOARD: [Actions.VIEW],
            Modules.CLIENTS: [Actions.VIEW, Actions.CREATE, Actions.EDIT],
            Modules.TRANSACTIONS: ALL_ACTIONS.copy(),
            Modules.TREASURY: ALL_ACTIONS.copy(),
            Modules.LP_MANAGEMENT: ALL_ACTIONS.copy(),
            Modules.INCOME_EXPENSES: ALL_ACTIONS.copy(),
            Modules.LOANS: ALL_ACTIONS.copy(),
            Modules.DEBTS: [Actions.VIEW, Actions.CREATE, Actions.EDIT],
            Modules.PSP: ALL_ACTIONS.copy(),
            Modules.EXCHANGERS: ALL_ACTIONS.copy(),
            Modules.RECONCILIATION: [Actions.VIEW, Actions.CREATE],
            Modules.REPORTS: [Actions.VIEW, Actions.EXPORT],
            Modules.SETTINGS: [Actions.VIEW]
        }
    },
    {
        "role_id": "sub_admin",
        "name": "sub_admin",
        "display_name": "Sub Admin",
        "description": "Limited administrative access",
        "is_system_role": True,
        "hierarchy_level": 50,
        "permissions": {
            Modules.CLIENTS: [Actions.VIEW, Actions.CREATE, Actions.EDIT],
            Modules.TRANSACTIONS: [Actions.VIEW, Actions.CREATE],
            Modules.TREASURY: [Actions.VIEW],
            Modules.INCOME_EXPENSES: [Actions.VIEW],
            Modules.REPORTS: [Actions.VIEW, Actions.EXPORT]
        }
    },
    {
        "role_id": "exchanger",
        "name": "vendor",
        "display_name": "Exchanger",
        "description": "Exchanger/Vendor portal access",
        "is_system_role": True,
        "hierarchy_level": 20,
        "permissions": {
            Modules.DASHBOARD: [Actions.VIEW],
            Modules.TRANSACTIONS: [Actions.VIEW, Actions.APPROVE],
            Modules.INCOME_EXPENSES: [Actions.VIEW, Actions.APPROVE]
        }
    },
    {
        "role_id": "viewer",
        "name": "viewer",
        "display_name": "Viewer",
        "description": "Read-only access to selected modules",
        "is_system_role": True,
        "hierarchy_level": 10,
        "permissions": {
            Modules.DASHBOARD: [Actions.VIEW],
            Modules.CLIENTS: [Actions.VIEW],
            Modules.TRANSACTIONS: [Actions.VIEW],
            Modules.REPORTS: [Actions.VIEW]
        }
    }
]

async def initialize_default_roles():
    """Initialize default roles if they don't exist"""
    for role_data in DEFAULT_ROLES:
        existing = await db.roles.find_one({"role_id": role_data["role_id"]}, {"_id": 0})
        if not existing:
            now = datetime.now(timezone.utc)
            role_data["created_at"] = now.isoformat()
            role_data["updated_at"] = now.isoformat()
            role_data["is_active"] = True
            await db.roles.insert_one(role_data)
            logger.info(f"Created default role: {role_data['display_name']}")

# ============== AUTH ROUTES ==============

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    user_doc = {
        "user_id": user_id,
        "email": user_data.email,
        "password_hash": hash_password(user_data.password),
        "name": user_data.name,
        "role": user_data.role,
        "picture": None,
        "is_active": True,
        "created_at": now.isoformat()
    }
    
    await db.users.insert_one(user_doc)
    
    token = create_jwt_token(user_id, user_data.email, user_data.role)
    
    return TokenResponse(
        access_token=token,
        user={
            "user_id": user_id,
            "email": user_data.email,
            "name": user_data.name,
            "role": user_data.role
        }
    )

@api_router.post("/auth/login")
async def login(credentials: UserLogin, request: Request):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    ip_address = request.client.host if request.client else None
    user_agent = request.headers.get("user-agent", "")
    
    if not user:
        # Log failed login attempt
        await create_log(
            log_type="auth",
            action="login_failed",
            module="authentication",
            user_email=credentials.email,
            description=f"Failed login attempt: User not found",
            ip_address=ip_address,
            user_agent=user_agent,
            status="failed"
        )
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not verify_password(credentials.password, user.get("password_hash", "")):
        # Log failed login attempt
        await create_log(
            log_type="auth",
            action="login_failed",
            module="authentication",
            user_id=user.get("user_id"),
            user_email=credentials.email,
            user_name=user.get("name"),
            description=f"Failed login attempt: Invalid password",
            ip_address=ip_address,
            user_agent=user_agent,
            status="failed"
        )
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not user.get("is_active", True):
        raise HTTPException(status_code=403, detail="Account is disabled")
    
    # Check if 2FA is enabled
    security_settings = await db.app_settings.find_one({"setting_type": "security"}, {"_id": 0})
    twofa_enabled = security_settings.get("twofa_enabled", False) if security_settings else False
    
    if twofa_enabled:
        import random
        otp_code = str(random.randint(100000, 999999))
        now_otp = datetime.now(timezone.utc)
        await db.otp_codes.delete_many({"user_id": user["user_id"]})
        await db.otp_codes.insert_one({
            "user_id": user["user_id"], "email": user["email"],
            "code": otp_code, "attempts": 0,
            "created_at": now_otp.isoformat(),
            "expires_at": (now_otp + timedelta(minutes=5)).isoformat()
        })
        smtp_settings = await db.app_settings.find_one({"setting_type": "email"}, {"_id": 0})
        # Fallback to .env SMTP if DB settings not configured
        smtp_host = (smtp_settings or {}).get("smtp_host") or os.environ.get("SMTP_HOST", "smtp.gmail.com")
        smtp_port = (smtp_settings or {}).get("smtp_port") or int(os.environ.get("SMTP_PORT", "587"))
        smtp_email = (smtp_settings or {}).get("smtp_email") or os.environ.get("SMTP_USER", "")
        smtp_password = (smtp_settings or {}).get("smtp_password") or os.environ.get("SMTP_PASSWORD", "")
        smtp_from = (smtp_settings or {}).get("smtp_from_email") or os.environ.get("SMTP_FROM_EMAIL", smtp_email)
        
        if smtp_email and smtp_password:
            try:
                otp_html = f"""<div style="font-family:Arial,sans-serif;max-width:400px;margin:0 auto;padding:30px;background:#0B0C10;color:white;border-radius:8px;">
                    <h2 style="color:#66FCF1;text-align:center;margin:0 0 20px;">MILES CAPITALS</h2>
                    <p style="color:#C5C6C7;text-align:center;">Your login verification code:</p>
                    <div style="background:#1F2833;padding:20px;border-radius:8px;text-align:center;margin:20px 0;">
                        <span style="font-size:36px;font-weight:bold;letter-spacing:8px;color:#66FCF1;">{otp_code}</span>
                    </div>
                    <p style="color:#C5C6C7;text-align:center;font-size:12px;">This code expires in 5 minutes. Do not share it.</p></div>"""
                await send_email(
                    to_emails=[user["email"]], subject="Miles Capitals - Login Verification Code",
                    html_content=otp_html, smtp_host=smtp_host,
                    smtp_port=smtp_port, smtp_email=smtp_email,
                    smtp_password=smtp_password,
                    smtp_from_email=smtp_from
                )
                return {"access_token": "", "token_type": "bearer", "requires_2fa": True,
                    "message": f"Verification code sent to {user['email']}",
                    "user": {"user_id": user["user_id"], "email": user["email"], "name": user["name"], "role": user.get("role", "viewer")}}
            except Exception as e:
                logger.error(f"Failed to send OTP email: {e}")
    
    token = create_jwt_token(user["user_id"], user["email"], user["role"])
    
    # Log successful login
    await create_log(
        log_type="auth",
        action="login",
        module="authentication",
        user_id=user["user_id"],
        user_email=user["email"],
        user_name=user["name"],
        user_role=user["role"],
        description=f"User logged in successfully",
        ip_address=ip_address,
        user_agent=user_agent,
        status="success"
    )
    
    return TokenResponse(
        access_token=token,
        user={
            "user_id": user["user_id"],
            "email": user["email"],
            "name": user["name"],
            "role": user["role"]
        }
    )


@api_router.post("/auth/verify-otp")
async def verify_otp(request: Request, data: dict = Body(...)):
    """Verify 2FA OTP and issue JWT token"""
    email = data.get("email")
    otp_code = data.get("otp_code")
    if not email or not otp_code:
        raise HTTPException(status_code=400, detail="Email and OTP code are required")
    
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    otp_record = await db.otp_codes.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not otp_record:
        raise HTTPException(status_code=401, detail="No verification code found. Please login again.")
    
    expires_at = datetime.fromisoformat(otp_record["expires_at"].replace("Z", "+00:00"))
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if datetime.now(timezone.utc) > expires_at:
        await db.otp_codes.delete_many({"user_id": user["user_id"]})
        raise HTTPException(status_code=401, detail="Verification code expired. Please login again.")
    
    if otp_record.get("attempts", 0) >= 3:
        await db.otp_codes.delete_many({"user_id": user["user_id"]})
        raise HTTPException(status_code=401, detail="Too many attempts. Please login again.")
    
    if otp_record["code"] != otp_code:
        await db.otp_codes.update_one({"user_id": user["user_id"]}, {"$inc": {"attempts": 1}})
        remaining = 3 - otp_record.get("attempts", 0) - 1
        raise HTTPException(status_code=401, detail=f"Invalid code. {remaining} attempts remaining.")
    
    await db.otp_codes.delete_many({"user_id": user["user_id"]})
    token = create_jwt_token(user["user_id"], user["email"], user["role"])
    
    ip_address = request.client.host if request.client else None
    await create_log(log_type="auth", action="login", module="authentication",
        user_id=user["user_id"], user_email=user["email"], user_name=user["name"],
        user_role=user["role"], description="User logged in with 2FA verification",
        ip_address=ip_address, user_agent=request.headers.get("user-agent", ""), status="success")
    
    return {"access_token": token, "token_type": "bearer",
        "user": {"user_id": user["user_id"], "email": user["email"], "name": user["name"], "role": user.get("role", "viewer")}}


@api_router.post("/auth/change-password")
async def change_password(request: Request, data: dict = Body(...), user: dict = Depends(get_current_user)):
    """Change own password"""
    current_password = data.get("current_password")
    new_password = data.get("new_password")
    if not current_password or not new_password:
        raise HTTPException(status_code=400, detail="Current and new passwords are required")
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="New password must be at least 6 characters")
    
    user_doc = await db.users.find_one({"user_id": user["user_id"]})
    if not user_doc:
        raise HTTPException(status_code=404, detail="User not found")
    
    if not verify_password(current_password, user_doc.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Current password is incorrect")
    
    new_hash = hash_password(new_password)
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"password_hash": new_hash, "updated_at": datetime.now(timezone.utc).isoformat()}})
    
    await log_activity(request, user, "edit", "users", "Changed own password")
    return {"message": "Password changed successfully"}


@api_router.get("/auth/security-status")
async def get_user_security_status(user: dict = Depends(get_current_user)):
    """Get 2FA status for any logged-in user (no admin needed)"""
    settings = await db.app_settings.find_one({"setting_type": "security"}, {"_id": 0})
    return {
        "twofa_enabled": settings.get("twofa_enabled", False) if settings else False,
        "session_timeout_hours": settings.get("session_timeout_hours", 2) if settings else 2,
    }


@api_router.get("/auth/notification-preferences")
async def get_notification_preferences(user: dict = Depends(get_current_user)):
    """Get current user's notification preferences"""
    prefs = await db.user_preferences.find_one({"user_id": user["user_id"]}, {"_id": 0})
    return {
        "approval_notifications": prefs.get("approval_notifications", True) if prefs else True,
    }

@api_router.put("/auth/notification-preferences")
async def update_notification_preferences(data: dict = Body(...), user: dict = Depends(get_current_user)):
    """Update current user's notification preferences"""
    now = datetime.now(timezone.utc)
    updates = {"user_id": user["user_id"], "updated_at": now.isoformat()}
    if "approval_notifications" in data:
        updates["approval_notifications"] = bool(data["approval_notifications"])
    
    await db.user_preferences.update_one(
        {"user_id": user["user_id"]},
        {"$set": updates},
        upsert=True
    )
    return {"message": "Preferences updated"}



@api_router.post("/auth/forgot-password")
async def forgot_password(data: dict = Body(...)):
    """Send password reset OTP to user's email"""
    email = data.get("email")
    if not email:
        raise HTTPException(status_code=400, detail="Email is required")
    
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user:
        # Don't reveal if user exists
        return {"message": "If an account exists with this email, a reset code has been sent."}
    
    import random
    otp_code = str(random.randint(100000, 999999))
    now = datetime.now(timezone.utc)
    
    await db.password_resets.delete_many({"email": email})
    await db.password_resets.insert_one({
        "email": email, "user_id": user["user_id"],
        "code": otp_code, "attempts": 0,
        "created_at": now.isoformat(),
        "expires_at": (now + timedelta(minutes=10)).isoformat()
    })
    
    # Send reset email
    smtp_settings = await db.app_settings.find_one({"setting_type": "email"}, {"_id": 0})
    smtp_host = (smtp_settings or {}).get("smtp_host") or os.environ.get("SMTP_HOST", "smtp.gmail.com")
    smtp_port = (smtp_settings or {}).get("smtp_port") or int(os.environ.get("SMTP_PORT", "587"))
    smtp_email = (smtp_settings or {}).get("smtp_email") or os.environ.get("SMTP_USER", "")
    smtp_password = (smtp_settings or {}).get("smtp_password") or os.environ.get("SMTP_PASSWORD", "")
    smtp_from = (smtp_settings or {}).get("smtp_from_email") or os.environ.get("SMTP_FROM_EMAIL", smtp_email)
    
    if smtp_email and smtp_password:
        try:
            reset_html = f"""<div style="font-family:Arial,sans-serif;max-width:400px;margin:0 auto;padding:30px;background:#0B0C10;color:white;border-radius:8px;">
                <h2 style="color:#66FCF1;text-align:center;margin:0 0 20px;">MILES CAPITALS</h2>
                <p style="color:#C5C6C7;text-align:center;">Password Reset Code</p>
                <div style="background:#1F2833;padding:20px;border-radius:8px;text-align:center;margin:20px 0;">
                    <span style="font-size:36px;font-weight:bold;letter-spacing:8px;color:#66FCF1;">{otp_code}</span>
                </div>
                <p style="color:#C5C6C7;text-align:center;font-size:12px;">This code expires in 10 minutes. If you didn't request this, ignore this email.</p></div>"""
            await send_email(
                to_emails=[email], subject="Miles Capitals - Password Reset Code",
                html_content=reset_html, smtp_host=smtp_host, smtp_port=smtp_port,
                smtp_email=smtp_email, smtp_password=smtp_password, smtp_from_email=smtp_from
            )
        except Exception as e:
            logger.error(f"Failed to send reset email: {e}")
    
    return {"message": "If an account exists with this email, a reset code has been sent."}


@api_router.post("/auth/reset-password")
async def reset_password(data: dict = Body(...)):
    """Verify reset OTP and set new password"""
    email = data.get("email")
    otp_code = data.get("otp_code")
    new_password = data.get("new_password")
    
    if not email or not otp_code or not new_password:
        raise HTTPException(status_code=400, detail="Email, code, and new password are required")
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    
    reset_record = await db.password_resets.find_one({"email": email}, {"_id": 0})
    if not reset_record:
        raise HTTPException(status_code=401, detail="No reset code found. Please request again.")
    
    expires_at = datetime.fromisoformat(reset_record["expires_at"].replace("Z", "+00:00"))
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if datetime.now(timezone.utc) > expires_at:
        await db.password_resets.delete_many({"email": email})
        raise HTTPException(status_code=401, detail="Reset code expired. Please request again.")
    
    if reset_record.get("attempts", 0) >= 5:
        await db.password_resets.delete_many({"email": email})
        raise HTTPException(status_code=401, detail="Too many attempts. Please request again.")
    
    if reset_record["code"] != otp_code:
        await db.password_resets.update_one({"email": email}, {"$inc": {"attempts": 1}})
        remaining = 5 - reset_record.get("attempts", 0) - 1
        raise HTTPException(status_code=401, detail=f"Invalid code. {remaining} attempts remaining.")
    
    # Reset password
    new_hash = hash_password(new_password)
    await db.users.update_one({"user_id": reset_record["user_id"]}, {"$set": {"password_hash": new_hash, "updated_at": datetime.now(timezone.utc).isoformat()}})
    await db.password_resets.delete_many({"email": email})
    
    return {"message": "Password reset successfully. You can now login."}




@api_router.get("/settings/security")
async def get_security_settings(user: dict = Depends(require_permission(Modules.SETTINGS, Actions.VIEW))):
    """Get 2FA and session settings"""
    settings = await db.app_settings.find_one({"setting_type": "security"}, {"_id": 0})
    return {
        "twofa_enabled": settings.get("twofa_enabled", False) if settings else False,
        "session_timeout_hours": settings.get("session_timeout_hours", 2) if settings else 2,
    }

@api_router.put("/settings/security")
async def update_security_settings(request: Request, data: dict = Body(...), user: dict = Depends(require_permission(Modules.SETTINGS, Actions.EDIT))):
    """Update 2FA and session settings"""
    now = datetime.now(timezone.utc)
    updates = {"setting_type": "security", "updated_at": now.isoformat(), "updated_by": user["user_id"]}
    if "twofa_enabled" in data:
        updates["twofa_enabled"] = bool(data["twofa_enabled"])
    if "session_timeout_hours" in data:
        updates["session_timeout_hours"] = int(data["session_timeout_hours"])
    
    await db.app_settings.update_one({"setting_type": "security"}, {"$set": updates}, upsert=True)
    await log_activity(request, user, "edit", "settings", "Updated security settings")
    return {"message": "Security settings updated"}


@api_router.post("/auth/session")
async def process_session(request: Request, response: Response):
    """Process Google OAuth session_id"""
    body = await request.json()
    session_id = body.get("session_id")
    
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    
    async with httpx.AsyncClient() as http_client:
        try:
            resp = await http_client.get(
                "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
                headers={"X-Session-ID": session_id}
            )
            if resp.status_code != 200:
                raise HTTPException(status_code=401, detail="Invalid session")
            
            session_data = resp.json()
        except Exception as e:
            logger.error(f"Auth error: {e}")
            raise HTTPException(status_code=401, detail="Authentication failed")
    
    email = session_data.get("email")
    name = session_data.get("name")
    picture = session_data.get("picture")
    session_token = session_data.get("session_token")
    
    now = datetime.now(timezone.utc)
    
    user = await db.users.find_one({"email": email}, {"_id": 0})
    
    if user:
        user_id = user["user_id"]
        await db.users.update_one(
            {"email": email},
            {"$set": {"name": name, "picture": picture}}
        )
    else:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        user_doc = {
            "user_id": user_id,
            "email": email,
            "name": name,
            "picture": picture,
            "role": UserRole.SUB_ADMIN,
            "is_active": True,
            "created_at": now.isoformat()
        }
        await db.users.insert_one(user_doc)
    
    expires_at = now + timedelta(days=7)
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": expires_at.isoformat(),
        "created_at": now.isoformat()
    })
    
    response.set_cookie(
        key="session_token",
        value=session_token,
        httponly=True,
        secure=True,
        samesite="none",
        path="/",
        max_age=7 * 24 * 60 * 60
    )
    
    user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    return {
        "user_id": user["user_id"],
        "email": user["email"],
        "name": user["name"],
        "role": user.get("role", UserRole.SUB_ADMIN),
        "picture": user.get("picture")
    }

@api_router.get("/auth/me")
async def get_me(user: dict = Depends(get_current_user)):
    return {
        "user_id": user["user_id"],
        "email": user["email"],
        "name": user["name"],
        "role": user.get("role", UserRole.SUB_ADMIN),
        "picture": user.get("picture")
    }

@api_router.post("/auth/logout")
async def logout(request: Request, response: Response):
    session_token = request.cookies.get("session_token")
    if session_token:
        await db.user_sessions.delete_one({"session_token": session_token})
    response.delete_cookie("session_token", path="/")
    return {"message": "Logged out"}

# ============== USERS/ADMINS ROUTES ==============

@api_router.get("/users")
async def get_users(user: dict = Depends(require_permission(Modules.USERS, Actions.VIEW))):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return users

@api_router.post("/users")
async def create_user(request: Request, user_data: UserCreate, user: dict = Depends(require_permission(Modules.USERS, Actions.CREATE))):

    existing = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")
    
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    user_doc = {
        "user_id": user_id,
        "email": user_data.email,
        "password_hash": hash_password(user_data.password),
        "name": user_data.name,
        "role": user_data.role,
        "picture": None,
        "is_active": True,
        "created_at": now.isoformat()
    }
    
    # If role is vendor/exchanger, also create a vendor entity
    if user_data.role == UserRole.VENDOR:
        vendor_id = f"vendor_{uuid.uuid4().hex[:12]}"
        user_doc["vendor_id"] = vendor_id  # Link user to vendor
        
        vendor_doc = {
            "vendor_id": vendor_id,
            "user_id": user_id,
            "vendor_name": user_data.name,
            "email": user_data.email,
            "deposit_commission": user_data.deposit_commission or 0.0,
            "withdrawal_commission": user_data.withdrawal_commission or 0.0,
            "description": f"Exchanger created via User Management",
            "total_volume": 0.0,
            "total_commission": 0.0,
            "pending_settlement": 0.0,
            "status": VendorStatus.ACTIVE,
            "created_at": now.isoformat(),
            "updated_at": now.isoformat()
        }
        await db.vendors.insert_one(vendor_doc)
    
    await db.users.insert_one(user_doc)
    
    await log_activity(request, user, "create", "users", "Created user")

    return {"user_id": user_id, "email": user_data.email, "name": user_data.name, "role": user_data.role}

@api_router.put("/users/{user_id}")
async def update_user(request: Request, user_id: str, update_data: UserUpdate, user: dict = Depends(require_permission(Modules.USERS, Actions.EDIT))):

    updates = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No updates provided")
    
    result = await db.users.update_one({"user_id": user_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    updated_user = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})
    await log_activity(request, user, "edit", "users", "Updated user")

    return updated_user

@api_router.delete("/users/{user_id}")
async def delete_user(request: Request, user_id: str, user: dict = Depends(require_permission(Modules.USERS, Actions.DELETE))):

    if user["user_id"] == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    result = await db.users.delete_one({"user_id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    await log_activity(request, user, "delete", "users", "Deleted user")

    return {"message": "User deleted"}


# ============== CLIENT TAGS ROUTES ==============

class ClientTagCreate(BaseModel):
    name: str
    color: Optional[str] = "#3B82F6"

@api_router.get("/client-tags")
async def get_client_tags(user: dict = Depends(get_current_user)):
    """Get all predefined client tags"""
    tags = await db.client_tags.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    return tags

@api_router.post("/client-tags")
async def create_client_tag(tag: ClientTagCreate, user: dict = Depends(require_permission(Modules.CLIENTS, Actions.CREATE))):
    """Create a new client tag"""
    existing = await db.client_tags.find_one({"name": {"$regex": f"^{tag.name}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Tag already exists")
    
    tag_doc = {
        "tag_id": f"tag_{uuid.uuid4().hex[:8]}",
        "name": tag.name.strip(),
        "color": tag.color or "#3B82F6",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["user_id"]
    }
    await db.client_tags.insert_one(tag_doc)
    return {k: v for k, v in tag_doc.items() if k != "_id"}

@api_router.delete("/client-tags/{tag_id}")
async def delete_client_tag(tag_id: str, user: dict = Depends(require_permission(Modules.CLIENTS, Actions.DELETE))):
    """Delete a client tag"""
    result = await db.client_tags.delete_one({"tag_id": tag_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tag not found")
    return {"message": "Tag deleted"}


# ============== CLIENTS ROUTES ==============

@api_router.get("/clients")
async def get_clients(
    user: dict = Depends(require_permission(Modules.CLIENTS, Actions.VIEW)),
    status: Optional[str] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
):
    query = {}
    if status:
        query["kyc_status"] = status
    if search:
        query["$or"] = [
            {"first_name": {"$regex": search, "$options": "i"}},
            {"last_name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}}
        ]
    
    total = await db.clients.count_documents(query)
    total_pages = max(1, (total + page_size - 1) // page_size)
    skip = (page - 1) * page_size
    
    clients = await db.clients.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    
    # Get transaction summaries for fetched clients only
    client_ids = [c["client_id"] for c in clients]
    if client_ids:
        tx_pipeline = [
            {"$match": {"client_id": {"$in": client_ids}}},
            {"$group": {
                "_id": {
                    "client_id": "$client_id",
                    "type": "$transaction_type"
                },
                "total_amount": {"$sum": "$amount"},
                "count": {"$sum": 1}
            }}
        ]
        tx_summaries = await db.transactions.aggregate(tx_pipeline).to_list(10000)
    else:
        tx_summaries = []
    
    # Build lookup dict
    client_tx_map = {}
    for summary in tx_summaries:
        client_id = summary["_id"]["client_id"]
        tx_type = summary["_id"]["type"]
        if client_id not in client_tx_map:
            client_tx_map[client_id] = {"deposits": 0, "withdrawals": 0, "deposit_count": 0, "withdrawal_count": 0}
        if tx_type == "deposit":
            client_tx_map[client_id]["deposits"] = summary["total_amount"]
            client_tx_map[client_id]["deposit_count"] = summary["count"]
        elif tx_type == "withdrawal":
            client_tx_map[client_id]["withdrawals"] = summary["total_amount"]
            client_tx_map[client_id]["withdrawal_count"] = summary["count"]
    
    # Add summaries to clients
    for client in clients:
        tx_data = client_tx_map.get(client["client_id"], {})
        client["total_deposits"] = tx_data.get("deposits", 0)
        client["total_withdrawals"] = tx_data.get("withdrawals", 0)
        client["deposit_count"] = tx_data.get("deposit_count", 0)
        client["withdrawal_count"] = tx_data.get("withdrawal_count", 0)
        client["net_balance"] = client["total_deposits"] - client["total_withdrawals"]
        client["transaction_count"] = client["deposit_count"] + client["withdrawal_count"]
    
    return {"items": clients, "total": total, "page": page, "page_size": page_size, "total_pages": total_pages}

@api_router.get("/clients/{client_id}")
async def get_client(client_id: str, user: dict = Depends(require_permission(Modules.CLIENTS, Actions.VIEW))):
    client = await db.clients.find_one({"client_id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Get transaction summary
    pipeline = [
        {"$match": {"client_id": client_id}},
        {"$group": {
            "_id": "$transaction_type",
            "total_amount": {"$sum": "$amount"},
            "total_base_amount": {"$sum": {"$ifNull": ["$base_amount", "$amount"]}},
            "count": {"$sum": 1}
        }}
    ]
    
    tx_summary = await db.transactions.aggregate(pipeline).to_list(10)
    
    deposits = next((x for x in tx_summary if x["_id"] == "deposit"), {"total_amount": 0, "count": 0})
    withdrawals = next((x for x in tx_summary if x["_id"] == "withdrawal"), {"total_amount": 0, "count": 0})
    
    client["total_deposits"] = deposits.get("total_amount", 0)
    client["total_withdrawals"] = withdrawals.get("total_amount", 0)
    client["deposit_count"] = deposits.get("count", 0)
    client["withdrawal_count"] = withdrawals.get("count", 0)
    client["net_balance"] = deposits.get("total_amount", 0) - withdrawals.get("total_amount", 0)
    client["transaction_count"] = deposits.get("count", 0) + withdrawals.get("count", 0)
    
    # Get recent transactions
    recent_txs = await db.transactions.find(
        {"client_id": client_id},
        {"_id": 0}
    ).sort("created_at", -1).limit(10).to_list(10)
    client["recent_transactions"] = recent_txs
    
    return client

@api_router.post("/clients")
async def create_client(request: Request, client_data: ClientCreate, user: dict = Depends(require_permission(Modules.CLIENTS, Actions.CREATE))):

    existing = await db.clients.find_one({"email": client_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Client email already exists")
    
    client_id = f"client_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    client_doc = {
        "client_id": client_id,
        **client_data.model_dump(),
        "kyc_status": ClientStatus.PENDING,
        "kyc_documents": [],
        "created_at": now.isoformat(),
        "updated_at": now.isoformat()
    }
    
    await db.clients.insert_one(client_doc)
    
    await log_activity(request, user, "create", "clients", "Created client")

    return await db.clients.find_one({"client_id": client_id}, {"_id": 0})

@api_router.put("/clients/{client_id}")
async def update_client(request: Request, client_id: str, update_data: ClientUpdate, user: dict = Depends(require_permission(Modules.CLIENTS, Actions.EDIT))):

    updates = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No updates provided")
    
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.clients.update_one({"client_id": client_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    
    await log_activity(request, user, "edit", "clients", "Updated client")

    return await db.clients.find_one({"client_id": client_id}, {"_id": 0})

@api_router.delete("/clients/{client_id}")
async def delete_client(request: Request, client_id: str, user: dict = Depends(require_permission(Modules.CLIENTS, Actions.DELETE))):

    result = await db.clients.delete_one({"client_id": client_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Client not found")
    await log_activity(request, user, "delete", "clients", "Deleted client")

    return {"message": "Client deleted"}

@api_router.post("/clients/bulk-upload")
async def bulk_upload_clients(
    request: Request,
    file: UploadFile = File(...),
    user: dict = Depends(require_permission(Modules.CLIENTS, Actions.CREATE))
):
    """Bulk upload clients from Excel file"""
    import openpyxl
    import io
    
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty file")
    
    headers = [str(h).strip().lower() if h else '' for h in rows[0]]
    data_rows = rows[1:]
    
    # Map columns
    col_map = {}
    for i, h in enumerate(headers):
        if 'crm' in h and 'id' in h:
            col_map['crm_id'] = i
        elif h in ('name', 'full name', 'client name'):
            col_map['name'] = i
        elif 'first' in h and 'name' in h:
            col_map['first_name'] = i
        elif 'last' in h and 'name' in h:
            col_map['last_name'] = i
        elif 'email' in h:
            col_map['email'] = i
        elif 'phone' in h or 'mobile' in h:
            col_map['phone'] = i
        elif 'country' in h:
            col_map['country'] = i
    
    now = datetime.now(timezone.utc).isoformat()
    created = 0
    skipped = 0
    errors = []
    
    # Get existing emails for duplicate check
    existing_emails = set()
    async for doc in db.clients.find({}, {"email": 1, "_id": 0}):
        if doc.get("email"):
            existing_emails.add(doc["email"].lower().strip())
    
    bulk_docs = []
    for idx, row in enumerate(data_rows):
        try:
            # Parse name
            first_name = ''
            last_name = ''
            if 'name' in col_map:
                full_name = str(row[col_map['name']] or '').strip()
                parts = full_name.split(' ', 1)
                first_name = parts[0] if parts else ''
                last_name = parts[1] if len(parts) > 1 else ''
            if 'first_name' in col_map:
                first_name = str(row[col_map['first_name']] or '').strip()
            if 'last_name' in col_map:
                last_name = str(row[col_map['last_name']] or '').strip()
            
            email = str(row[col_map.get('email', -1)] or '').strip().lower() if 'email' in col_map else ''
            phone = str(row[col_map.get('phone', -1)] or '').strip() if 'phone' in col_map else ''
            country = str(row[col_map.get('country', -1)] or '').strip() if 'country' in col_map else ''
            crm_id = str(row[col_map.get('crm_id', -1)] or '').strip() if 'crm_id' in col_map else ''
            
            if not first_name or not email:
                skipped += 1
                continue
            
            if email in existing_emails:
                skipped += 1
                continue
            
            existing_emails.add(email)
            client_id = f"client_{uuid.uuid4().hex[:12]}"
            
            bulk_docs.append({
                "client_id": client_id,
                "first_name": first_name,
                "last_name": last_name,
                "email": email,
                "phone": phone,
                "country": country,
                "crm_customer_id": crm_id,
                "mt5_number": None,
                "notes": None,
                "kyc_status": "pending",
                "kyc_documents": [],
                "created_at": now,
                "updated_at": now,
            })
            created += 1
        except Exception as e:
            errors.append(f"Row {idx + 2}: {str(e)}")
            skipped += 1
    
    if bulk_docs:
        await db.clients.insert_many(bulk_docs)
    
    await log_activity(request, user, "create", "clients", f"Bulk uploaded {created} clients")
    
    return {
        "created": created,
        "skipped": skipped,
        "errors": errors[:10],
        "total_rows": len(data_rows),
    }

# ============== CLIENT BANK ACCOUNTS ROUTES ==============

@api_router.get("/clients/{client_id}/bank-accounts")
async def get_client_bank_accounts(client_id: str, user: dict = Depends(require_permission(Modules.CLIENTS, Actions.VIEW))):
    """Get all saved bank accounts for a client"""
    accounts = await db.client_bank_accounts.find({"client_id": client_id}, {"_id": 0}).to_list(100)
    return accounts

@api_router.post("/clients/{client_id}/bank-accounts")
async def create_client_bank_account(

    request: Request,

    client_id: str,
    bank_name: str = Form(...),
    account_name: str = Form(...),
    account_number: str = Form(...),
    swift_iban: str = Form(None),
    currency: str = Form("USD"),
    user: dict = Depends(require_permission(Modules.CLIENTS, Actions.CREATE))
):
    """Save a new bank account for a client"""
    # Check if client exists
    client = await db.clients.find_one({"client_id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Check for duplicate account
    existing = await db.client_bank_accounts.find_one({
        "client_id": client_id,
        "account_number": account_number,
        "bank_name": bank_name
    })
    if existing:
        return existing  # Return existing instead of creating duplicate
    
    bank_account_id = f"cba_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    bank_doc = {
        "bank_account_id": bank_account_id,
        "client_id": client_id,
        "bank_name": bank_name,
        "account_name": account_name,
        "account_number": account_number,
        "swift_iban": swift_iban,
        "currency": currency,
        "created_at": now.isoformat(),
        "created_by": user["user_id"]
    }
    
    await db.client_bank_accounts.insert_one(bank_doc)
    await log_activity(request, user, "create", "clients", "Added client bank account")

    return await db.client_bank_accounts.find_one({"bank_account_id": bank_account_id}, {"_id": 0})

@api_router.put("/clients/{client_id}/bank-accounts/{bank_account_id}")
async def update_client_bank_account(

    request: Request,

    client_id: str,
    bank_account_id: str,
    bank_name: str = Form(None),
    account_name: str = Form(None),
    account_number: str = Form(None),
    swift_iban: str = Form(None),
    currency: str = Form(None),
    user: dict = Depends(require_permission(Modules.CLIENTS, Actions.EDIT))
):
    """Update a client's bank account"""
    updates = {}
    if bank_name:
        updates["bank_name"] = bank_name
    if account_name:
        updates["account_name"] = account_name
    if account_number:
        updates["account_number"] = account_number
    if swift_iban is not None:
        updates["swift_iban"] = swift_iban
    if currency:
        updates["currency"] = currency
    
    if not updates:
        raise HTTPException(status_code=400, detail="No updates provided")
    
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.client_bank_accounts.update_one(
        {"bank_account_id": bank_account_id, "client_id": client_id},
        {"$set": updates}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Bank account not found")
    
    await log_activity(request, user, "edit", "clients", "Updated client bank account")

    return await db.client_bank_accounts.find_one({"bank_account_id": bank_account_id}, {"_id": 0})

@api_router.delete("/clients/{client_id}/bank-accounts/{bank_account_id}")
async def delete_client_bank_account(request: Request, client_id: str, bank_account_id: str, user: dict = Depends(require_permission(Modules.CLIENTS, Actions.DELETE))):

    """Delete a client's bank account"""
    result = await db.client_bank_accounts.delete_one({"bank_account_id": bank_account_id, "client_id": client_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bank account not found")
    await log_activity(request, user, "delete", "clients", "Deleted client bank account")

    return {"message": "Bank account deleted"}

# ============== TREASURY/BANK ACCOUNTS ROUTES ==============

@api_router.get("/treasury")
async def get_treasury_accounts(
    user: dict = Depends(require_permission(Modules.TREASURY, Actions.VIEW)),
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
):
    query = {}
    if search:
        query["$or"] = [
            {"account_name": {"$regex": search, "$options": "i"}},
            {"bank_name": {"$regex": search, "$options": "i"}},
            {"currency": {"$regex": search, "$options": "i"}},
        ]
    
    total = await db.treasury_accounts.count_documents(query)
    total_pages = max(1, (total + page_size - 1) // page_size)
    skip = (page - 1) * page_size
    
    accounts = await db.treasury_accounts.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    
    fx_settings = await db.app_settings.find_one({"setting_type": "manual_fx_rates"}, {"_id": 0})
    manual_rates = fx_settings.get("rates", {}) if fx_settings else {}
    
    for acc in accounts:
        currency = acc.get("currency", "USD")
        balance = acc.get("balance", 0)
        if currency == "USD":
            acc["balance_usd"] = balance
        elif currency in manual_rates and manual_rates[currency] > 0:
            acc["balance_usd"] = round(balance * manual_rates[currency], 2)
        else:
            acc["balance_usd"] = None
    
    return {"items": accounts, "total": total, "page": page, "page_size": page_size, "total_pages": total_pages}

@api_router.get("/treasury/{account_id}")
async def get_treasury_account(account_id: str, user: dict = Depends(require_permission(Modules.TREASURY, Actions.VIEW))):
    account = await db.treasury_accounts.find_one({"account_id": account_id}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=404, detail="Treasury account not found")
    return account

@api_router.post("/treasury")
async def create_treasury_account(account_data: TreasuryAccountCreate, request: Request, user: dict = Depends(require_permission(Modules.TREASURY, Actions.CREATE))):
    account_id = f"treasury_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    account_doc = {
        "account_id": account_id,
        **account_data.model_dump(exclude={"opening_balance"}),
        "balance": account_data.opening_balance,
        "opening_balance": account_data.opening_balance,
        "status": TreasuryAccountStatus.ACTIVE,
        "created_at": now.isoformat(),
        "updated_at": now.isoformat()
    }
    
    await db.treasury_accounts.insert_one(account_doc)
    
    # Log activity
    await log_activity(request, user, "create", "treasury", f"Created treasury account: {account_data.account_name}", reference_id=account_id)
    
    return await db.treasury_accounts.find_one({"account_id": account_id}, {"_id": 0})

@api_router.put("/treasury/{account_id}")
async def update_treasury_account(request: Request, account_id: str, update_data: TreasuryAccountUpdate, user: dict = Depends(require_permission(Modules.TREASURY, Actions.EDIT))):

    updates = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No updates provided")
    
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.treasury_accounts.update_one({"account_id": account_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Treasury account not found")
    
    await log_activity(request, user, "edit", "treasury", "Updated treasury account")

    return await db.treasury_accounts.find_one({"account_id": account_id}, {"_id": 0})

# Treasury Transaction History
@api_router.get("/treasury/{account_id}/history")
async def get_treasury_history(
    account_id: str, 
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    transaction_type: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    limit: int = 5000,
    user: dict = Depends(require_permission(Modules.TREASURY, Actions.VIEW))
):
    """Get transaction history for a treasury account"""
    account = await db.treasury_accounts.find_one({"account_id": account_id}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=404, detail="Treasury account not found")
    
    # Build query for treasury transactions
    query = {"account_id": account_id}
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    if transaction_type:
        query["transaction_type"] = transaction_type
    
    # Get treasury-specific transactions (these are the canonical records with proper currency conversion)
    treasury_txs = await db.treasury_transactions.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    
    # Adjust amounts for display - outflows should be negative
    outflow_types = ["debt_payment", "withdrawal", "transfer_out", "expense", "balance_adjustment_debit", "loan_disbursement"]
    for ttx in treasury_txs:
        if ttx.get("transaction_type") in outflow_types:
            ttx["amount"] = -abs(ttx.get("amount", 0))
    
    # Get transaction IDs that already have treasury transaction records
    existing_tx_ids = set()
    for ttx in treasury_txs:
        if ttx.get("transaction_id"):
            existing_tx_ids.add(ttx.get("transaction_id"))
    
    # Only get regular transactions that DON'T have treasury transaction records yet
    # (for backwards compatibility with old data that might not have treasury_transactions)
    tx_query = {
        "destination_account_id": account_id,
        "status": {"$in": ["approved", "completed"]}
    }
    if start_date:
        tx_query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in tx_query:
            tx_query["created_at"]["$lte"] = end_date
        else:
            tx_query["created_at"] = {"$lte": end_date}
    
    regular_txs = await db.transactions.find(tx_query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    
    # Convert regular transactions to history format ONLY if they don't already have a treasury_transaction
    account_currency = account.get("currency", "USD")
    for tx in regular_txs:
        if tx.get("transaction_id") not in existing_tx_ids:
            # For non-USD accounts, use base_amount if available and currency matches
            display_amount = tx.get("amount", 0)
            if account_currency != "USD":
                # If the transaction has base_amount in the same currency as the account, use it
                if tx.get("base_currency") == account_currency and tx.get("base_amount"):
                    display_amount = tx.get("base_amount")
                # Otherwise keep USD amount (no live FX conversion)
            
            treasury_txs.append({
                "treasury_transaction_id": tx.get("transaction_id"),
                "account_id": account_id,
                "transaction_type": tx.get("transaction_type"),
                "amount": display_amount if tx.get("transaction_type") == "deposit" else -display_amount,
                "currency": account_currency,
                "reference": f"{tx.get('transaction_type', '').capitalize()}: {tx.get('client_name', 'Unknown')} - {tx.get('reference', '')}",
                "client_id": tx.get("client_id"),
                "client_name": tx.get("client_name"),
                "created_at": tx.get("processed_at") or tx.get("created_at"),
                "created_by": tx.get("processed_by"),
                "created_by_name": tx.get("processed_by_name")
            })
    
    # Sort combined list by date (newest first)
    treasury_txs.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    
    # Calculate running balance (start from current balance, work backwards)
    current_balance = account.get("balance", 0)
    running = current_balance
    for tx in treasury_txs:
        tx["running_balance"] = round(running, 2)
        running -= (tx.get("amount", 0))
    
    # Paginate the combined result
    total = len(treasury_txs)
    total_pages = (total + page_size - 1) // page_size if total > 0 else 1
    skip = (page - 1) * page_size
    paginated = treasury_txs[skip:skip + page_size]
    
    return {
        "items": paginated,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages
    }

@api_router.delete("/treasury/{account_id}")
async def delete_treasury_account(request: Request, account_id: str, user: dict = Depends(require_permission(Modules.TREASURY, Actions.DELETE))):

    result = await db.treasury_accounts.delete_one({"account_id": account_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Treasury account not found")
    await log_activity(request, user, "delete", "treasury", "Deleted treasury account")

    return {"message": "Treasury account deleted"}

# Balance Fix / Opening Balance Adjustment
class BalanceFixRequest(BaseModel):
    actual_balance: float
    effective_date: Optional[str] = None
    reason: Optional[str] = None

@api_router.post("/treasury/{account_id}/balance-fix")
async def fix_treasury_balance(request: Request, account_id: str, fix: BalanceFixRequest, user: dict = Depends(require_admin)):
    """Fix treasury opening balance at effective date.
    Inserts an adjustment so the opening balance on that date matches the input.
    Then recalculates all running balances and the final account balance."""
    account = await db.treasury_accounts.find_one({"account_id": account_id}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=404, detail="Treasury account not found")
    
    # Effective date → adjustment placed at start-of-day (opening balance)
    raw_date = fix.effective_date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    day_str = raw_date[:10]  # YYYY-MM-DD
    day_start = f"{day_str}T00:00:00"
    
    outflow_types = ["debt_payment", "withdrawal", "transfer_out", "expense", "balance_adjustment_debit", "loan_disbursement"]
    
    # Get all transactions for this account sorted by date
    all_txs = await db.treasury_transactions.find(
        {"account_id": account_id},
        {"_id": 0, "treasury_transaction_id": 1, "created_at": 1, "amount": 1, "transaction_type": 1}
    ).sort("created_at", 1).to_list(None)
    
    # Remove any existing balance adjustments on the same date (re-fix scenario)
    existing_adj_ids = []
    for tx in all_txs:
        tx_date = tx.get("created_at", "")[:10]
        tx_type = tx.get("transaction_type", "")
        if tx_date == day_str and tx_type in ("balance_adjustment", "balance_adjustment_debit"):
            existing_adj_ids.append(tx["treasury_transaction_id"])
    
    if existing_adj_ids:
        await db.treasury_transactions.delete_many(
            {"treasury_transaction_id": {"$in": existing_adj_ids}}
        )
        # Refresh the transaction list after deletion
        all_txs = await db.treasury_transactions.find(
            {"account_id": account_id},
            {"_id": 0, "treasury_transaction_id": 1, "created_at": 1, "amount": 1, "transaction_type": 1}
        ).sort("created_at", 1).to_list(None)
    
    # Calculate running balance BEFORE the selected date's real transactions start.
    # Include any existing midnight adjustments (they are prior opening fixes).
    # "Real" transactions on this day start after 00:00:00.
    running_before_date = 0.0
    for tx in all_txs:
        tx_date = tx.get("created_at", "")
        # Include everything strictly before the day, PLUS any midnight entries (00:00:00)
        if tx_date > day_start:
            break
        amt = tx.get("amount", 0)
        if tx.get("transaction_type") in outflow_types:
            running_before_date -= abs(amt)
        else:
            running_before_date += abs(amt)
    
    running_before_date = round(running_before_date, 2)
    adjustment = round(fix.actual_balance - running_before_date, 2)
    
    if adjustment == 0:
        raise HTTPException(status_code=400, detail=f"Opening balance on {day_str} already matches {fix.actual_balance:,.2f}. No adjustment needed.")
    
    tx_id = f"ttx_{uuid.uuid4().hex[:12]}"
    # Place slightly after midnight so it sorts after any existing midnight entries
    adjustment_timestamp = f"{day_str}T00:00:00.500000"
    
    adjustment_entry = {
        "treasury_transaction_id": tx_id,
        "account_id": account_id,
        "transaction_type": "balance_adjustment",
        "amount": abs(adjustment),
        "currency": account.get("currency", "USD"),
        "reference": f"Opening Balance Fix: {fix.reason or 'Manual correction'}",
        "description": f"Opening balance on {day_str} fixed to {fix.actual_balance:,.2f}. Was {running_before_date:,.2f}. Reason: {fix.reason or 'Manual correction'}",
        "created_at": adjustment_timestamp,
        "created_by": user.get("email"),
        "created_by_name": user.get("name", user.get("email")),
        "adjustment_direction": "credit" if adjustment > 0 else "debit",
        "previous_balance": running_before_date,
        "new_balance": fix.actual_balance,
    }
    
    if adjustment < 0:
        adjustment_entry["transaction_type"] = "balance_adjustment_debit"
    
    await db.treasury_transactions.insert_one(adjustment_entry)
    adjustment_entry.pop("_id", None)
    
    # Recalculate ALL running balances from the beginning (including the new adjustment)
    all_txs_updated = await db.treasury_transactions.find(
        {"account_id": account_id},
        {"_id": 0, "treasury_transaction_id": 1, "created_at": 1, "amount": 1, "transaction_type": 1}
    ).sort("created_at", 1).to_list(None)
    
    running = 0.0
    for tx in all_txs_updated:
        amt = tx.get("amount", 0)
        if tx.get("transaction_type") in outflow_types:
            running -= abs(amt)
        else:
            running += abs(amt)
        await db.treasury_transactions.update_one(
            {"treasury_transaction_id": tx["treasury_transaction_id"]},
            {"$set": {"running_balance": round(running, 2)}}
        )
    
    # Set final account balance to the LAST running balance
    final_balance = round(running, 2)
    old_balance = account.get("balance", 0)
    await db.treasury_accounts.update_one(
        {"account_id": account_id},
        {"$set": {"balance": final_balance, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    await log_activity(request, user, "edit", "treasury", f"Opening balance fix on {day_str}: {account.get('account_name')} adjusted by {adjustment:,.2f} ({account.get('currency', 'USD')}). Account balance: {old_balance:,.2f} -> {final_balance:,.2f}")
    
    return {
        "message": f"Opening balance on {day_str} fixed to {fix.actual_balance:,.2f}. Account balance updated from {old_balance:,.2f} to {final_balance:,.2f}",
        "adjustment": adjustment,
        "treasury_transaction_id": tx_id,
        "effective_date": day_str,
        "opening_balance_on_date": fix.actual_balance,
        "new_account_balance": final_balance,
        "old_account_balance": old_balance
    }

# Inter-Treasury Transfer
class TreasuryTransferRequest(BaseModel):
    source_account_id: str
    destination_account_id: str
    amount: float
    exchange_rate: Optional[float] = 1.0
    notes: Optional[str] = None
    transfer_date: Optional[str] = None

@api_router.post("/treasury/transfer")
async def inter_treasury_transfer(request: Request, transfer: TreasuryTransferRequest, user: dict = Depends(require_permission(Modules.TREASURY, Actions.CREATE))):

    """Transfer funds between treasury accounts"""
    if transfer.source_account_id == transfer.destination_account_id:
        raise HTTPException(status_code=400, detail="Source and destination accounts must be different")
    
    if transfer.amount <= 0:
        raise HTTPException(status_code=400, detail="Transfer amount must be positive")
    
    # Get source account
    source = await db.treasury_accounts.find_one({"account_id": transfer.source_account_id}, {"_id": 0})
    if not source:
        raise HTTPException(status_code=404, detail="Source account not found")
    
    # Get destination account
    destination = await db.treasury_accounts.find_one({"account_id": transfer.destination_account_id}, {"_id": 0})
    if not destination:
        raise HTTPException(status_code=404, detail="Destination account not found")
    
    # Check sufficient balance
    if source.get("balance", 0) < transfer.amount:
        raise HTTPException(status_code=400, detail="Insufficient balance in source account")
    
    now = datetime.now(timezone.utc)
    transfer_id = f"trf_{uuid.uuid4().hex[:12]}"
    
    # Use provided transfer_date or current time
    if transfer.transfer_date:
        record_date = f"{transfer.transfer_date}T00:00:00" if 'T' not in transfer.transfer_date else transfer.transfer_date
    else:
        record_date = now.isoformat()
    
    # Calculate destination amount based on exchange rate
    destination_amount = round(transfer.amount * (transfer.exchange_rate or 1.0), 2)
    
    # Deduct from source
    await db.treasury_accounts.update_one(
        {"account_id": transfer.source_account_id},
        {"$inc": {"balance": -transfer.amount}, "$set": {"updated_at": now.isoformat()}}
    )
    
    # Add to destination
    await db.treasury_accounts.update_one(
        {"account_id": transfer.destination_account_id},
        {"$inc": {"balance": destination_amount}, "$set": {"updated_at": now.isoformat()}}
    )
    
    # Record source transaction (transfer out)
    source_tx_id = f"ttx_{uuid.uuid4().hex[:12]}"
    source_tx_doc = {
        "treasury_transaction_id": source_tx_id,
        "account_id": transfer.source_account_id,
        "transaction_type": "transfer_out",
        "amount": -transfer.amount,
        "currency": source.get("currency", "USD"),
        "reference": f"Transfer to {destination.get('account_name')}",
        "transfer_id": transfer_id,
        "related_account_id": transfer.destination_account_id,
        "related_account_name": destination.get("account_name"),
        "exchange_rate": transfer.exchange_rate,
        "destination_amount": destination_amount,
        "destination_currency": destination.get("currency", "USD"),
        "notes": transfer.notes,
        "created_at": record_date,
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    await db.treasury_transactions.insert_one(source_tx_doc)
    
    # Record destination transaction (transfer in)
    dest_tx_id = f"ttx_{uuid.uuid4().hex[:12]}"
    dest_tx_doc = {
        "treasury_transaction_id": dest_tx_id,
        "account_id": transfer.destination_account_id,
        "transaction_type": "transfer_in",
        "amount": destination_amount,
        "currency": destination.get("currency", "USD"),
        "reference": f"Transfer from {source.get('account_name')}",
        "transfer_id": transfer_id,
        "related_account_id": transfer.source_account_id,
        "related_account_name": source.get("account_name"),
        "exchange_rate": transfer.exchange_rate,
        "source_amount": transfer.amount,
        "source_currency": source.get("currency", "USD"),
        "notes": transfer.notes,
        "created_at": record_date,
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    await db.treasury_transactions.insert_one(dest_tx_doc)
    
    # Return transfer details
    await log_activity(request, user, "create", "treasury", "Inter-treasury transfer")

    return {
        "transfer_id": transfer_id,
        "source_account": source.get("account_name"),
        "source_currency": source.get("currency", "USD"),
        "source_amount": transfer.amount,
        "destination_account": destination.get("account_name"),
        "destination_currency": destination.get("currency", "USD"),
        "destination_amount": destination_amount,
        "exchange_rate": transfer.exchange_rate,
        "notes": transfer.notes,
        "created_at": now.isoformat(),
        "created_by_name": user["name"]
    }

# ============== LP (LIQUIDITY PROVIDER) ROUTES ==============

@api_router.get("/lp")
async def get_lp_accounts(
    user: dict = Depends(require_permission(Modules.LP_MANAGEMENT, Actions.VIEW)),
    search: Optional[str] = None,
    status: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
):
    """Get LP accounts with pagination"""
    query = {}
    if status:
        query["status"] = status
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"provider": {"$regex": search, "$options": "i"}},
        ]
    total = await db.lp_accounts.count_documents(query)
    total_pages = max(1, (total + page_size - 1) // page_size)
    skip = (page - 1) * page_size
    accounts = await db.lp_accounts.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    return {"items": accounts, "total": total, "page": page, "page_size": page_size, "total_pages": total_pages}

@api_router.get("/lp/dashboard")
async def get_lp_dashboard(user: dict = Depends(require_permission(Modules.LP_MANAGEMENT, Actions.VIEW))):
    """Get LP dashboard with summary statistics"""
    accounts = await db.lp_accounts.find({}, {"_id": 0}).to_list(1000)
    
    total_balance = sum(a.get("balance", 0) for a in accounts)
    total_deposits = sum(a.get("total_deposits", 0) for a in accounts)
    total_withdrawals = sum(a.get("total_withdrawals", 0) for a in accounts)
    active_count = sum(1 for a in accounts if a.get("status") == LPAccountStatus.ACTIVE)
    
    # Get recent transactions
    recent_txs = await db.lp_transactions.find({}, {"_id": 0}).sort("created_at", -1).limit(10).to_list(10)
    
    return {
        "total_balance": total_balance,
        "total_deposits": total_deposits,
        "total_withdrawals": total_withdrawals,
        "active_lp_count": active_count,
        "total_lp_count": len(accounts),
        "accounts": accounts,
        "recent_transactions": recent_txs
    }

@api_router.get("/lp/{lp_id}")
async def get_lp_account(lp_id: str, user: dict = Depends(require_permission(Modules.LP_MANAGEMENT, Actions.VIEW))):
    """Get a specific LP account"""
    account = await db.lp_accounts.find_one({"lp_id": lp_id}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=404, detail="LP account not found")
    return account

@api_router.post("/lp")
async def create_lp_account(request: Request, lp_data: LPAccountCreate, user: dict = Depends(require_permission(Modules.LP_MANAGEMENT, Actions.CREATE))):

    """Create a new LP account"""
    lp_id = f"lp_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    account_doc = {
        "lp_id": lp_id,
        "lp_name": lp_data.lp_name,
        "account_number": lp_data.account_number,
        "bank_name": lp_data.bank_name,
        "swift_code": lp_data.swift_code,
        "currency": lp_data.currency,
        "contact_person": lp_data.contact_person,
        "contact_email": lp_data.contact_email,
        "contact_phone": lp_data.contact_phone,
        "notes": lp_data.notes,
        "balance": 0,
        "total_deposits": 0,
        "total_withdrawals": 0,
        "transaction_count": 0,
        "status": LPAccountStatus.ACTIVE,
        "created_at": now.isoformat(),
        "updated_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    
    await db.lp_accounts.insert_one(account_doc)
    await log_activity(request, user, "create", "lp_management", "Created LP account")

    return await db.lp_accounts.find_one({"lp_id": lp_id}, {"_id": 0})

@api_router.put("/lp/{lp_id}")
async def update_lp_account(request: Request, lp_id: str, lp_data: LPAccountUpdate, user: dict = Depends(require_permission(Modules.LP_MANAGEMENT, Actions.EDIT))):

    """Update an LP account"""
    account = await db.lp_accounts.find_one({"lp_id": lp_id}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=404, detail="LP account not found")
    
    updates = {k: v for k, v in lp_data.model_dump().items() if v is not None}
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.lp_accounts.update_one({"lp_id": lp_id}, {"$set": updates})
    await log_activity(request, user, "edit", "lp_management", "Updated LP account")

    return await db.lp_accounts.find_one({"lp_id": lp_id}, {"_id": 0})

@api_router.get("/lp/{lp_id}/transactions")
async def get_lp_transactions(lp_id: str, user: dict = Depends(require_permission(Modules.LP_MANAGEMENT, Actions.VIEW)),
                              page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=200)):
    """Get transactions for a specific LP account"""
    account = await db.lp_accounts.find_one({"lp_id": lp_id}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=404, detail="LP account not found")
    
    return await paginate_query(db.lp_transactions, {"lp_id": lp_id}, page, page_size)

@api_router.post("/lp/{lp_id}/deposit")
async def create_lp_deposit(request: Request, lp_id: str, tx_data: LPTransactionCreate, user: dict = Depends(require_permission(Modules.LP_MANAGEMENT, Actions.CREATE))):

    """Create a deposit to LP (sending funds TO the LP)"""
    account = await db.lp_accounts.find_one({"lp_id": lp_id}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=404, detail="LP account not found")
    
    if tx_data.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be positive")
    
    now = datetime.now(timezone.utc)
    tx_id = f"lptx_{uuid.uuid4().hex[:12]}"
    
    # If treasury account specified, deduct from it
    treasury_name = None
    treasury_deduct_amount = tx_data.amount
    if tx_data.treasury_account_id:
        treasury = await db.treasury_accounts.find_one({"account_id": tx_data.treasury_account_id}, {"_id": 0})
        if not treasury:
            raise HTTPException(status_code=404, detail="Treasury account not found")
        
        # Convert amount if currencies differ
        treasury_currency = treasury.get("currency", "USD")
        if treasury_currency.upper() != tx_data.currency.upper():
            treasury_deduct_amount = convert_currency(tx_data.amount, tx_data.currency, treasury_currency)
        
        if treasury.get("balance", 0) < treasury_deduct_amount:
            raise HTTPException(status_code=400, detail=f"Insufficient treasury balance. Required: {treasury_deduct_amount:,.2f} {treasury_currency}")
        
        treasury_name = treasury.get("account_name")
        
        # Deduct from treasury (in treasury's currency)
        await db.treasury_accounts.update_one(
            {"account_id": tx_data.treasury_account_id},
            {"$inc": {"balance": -treasury_deduct_amount}, "$set": {"updated_at": now.isoformat()}}
        )
        
        # Record treasury transaction
        conversion_note = f" (Converted from {tx_data.amount:,.2f} {tx_data.currency})" if treasury_currency.upper() != tx_data.currency.upper() else ""
        await db.treasury_transactions.insert_one({
            "treasury_transaction_id": f"ttx_{uuid.uuid4().hex[:12]}",
            "account_id": tx_data.treasury_account_id,
            "account_name": treasury_name,
            "transaction_type": "lp_deposit",
            "amount": -treasury_deduct_amount,
            "currency": treasury_currency,
            "original_amount": tx_data.amount,
            "original_currency": tx_data.currency,
            "reference": f"Deposit to LP: {account['lp_name']}{conversion_note}",
            "lp_transaction_id": tx_id,
            "created_at": now.isoformat(),
            "created_by": user["user_id"],
            "created_by_name": user["name"]
        })
    
    # Create LP transaction
    tx_doc = {
        "lp_transaction_id": tx_id,
        "lp_id": lp_id,
        "lp_name": account["lp_name"],
        "transaction_type": LPTransactionType.DEPOSIT,
        "amount": tx_data.amount,
        "currency": tx_data.currency,
        "treasury_account_id": tx_data.treasury_account_id,
        "treasury_name": treasury_name,
        "reference": tx_data.reference or f"DEP-{tx_id[-8:].upper()}",
        "notes": tx_data.notes,
        "balance_before": account.get("balance", 0),
        "balance_after": account.get("balance", 0) + tx_data.amount,
        "created_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    
    await db.lp_transactions.insert_one(tx_doc)
    
    # Update LP account balance
    await db.lp_accounts.update_one(
        {"lp_id": lp_id},
        {
            "$inc": {"balance": tx_data.amount, "total_deposits": tx_data.amount, "transaction_count": 1},
            "$set": {"updated_at": now.isoformat()}
        }
    )
    
    tx_doc.pop("_id", None)
    await log_activity(request, user, "create", "lp_management", "LP deposit recorded")

    return tx_doc

@api_router.post("/lp/{lp_id}/withdraw")
async def create_lp_withdrawal(request: Request, lp_id: str, tx_data: LPTransactionCreate, user: dict = Depends(require_permission(Modules.LP_MANAGEMENT, Actions.CREATE))):

    """Create a withdrawal from LP (receiving funds FROM the LP)"""
    account = await db.lp_accounts.find_one({"lp_id": lp_id}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=404, detail="LP account not found")
    
    if tx_data.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be positive")
    
    if account.get("balance", 0) < tx_data.amount:
        raise HTTPException(status_code=400, detail="Insufficient LP balance")
    
    now = datetime.now(timezone.utc)
    tx_id = f"lptx_{uuid.uuid4().hex[:12]}"
    
    # If treasury account specified, add to it
    treasury_name = None
    treasury_add_amount = tx_data.amount
    if tx_data.treasury_account_id:
        treasury = await db.treasury_accounts.find_one({"account_id": tx_data.treasury_account_id}, {"_id": 0})
        if not treasury:
            raise HTTPException(status_code=404, detail="Treasury account not found")
        
        treasury_name = treasury.get("account_name")
        
        # Convert amount if currencies differ
        treasury_currency = treasury.get("currency", "USD")
        if treasury_currency.upper() != tx_data.currency.upper():
            treasury_add_amount = convert_currency(tx_data.amount, tx_data.currency, treasury_currency)
        
        # Add to treasury (in treasury's currency)
        await db.treasury_accounts.update_one(
            {"account_id": tx_data.treasury_account_id},
            {"$inc": {"balance": treasury_add_amount}, "$set": {"updated_at": now.isoformat()}}
        )
        
        # Record treasury transaction
        conversion_note = f" (Converted from {tx_data.amount:,.2f} {tx_data.currency})" if treasury_currency.upper() != tx_data.currency.upper() else ""
        await db.treasury_transactions.insert_one({
            "treasury_transaction_id": f"ttx_{uuid.uuid4().hex[:12]}",
            "account_id": tx_data.treasury_account_id,
            "account_name": treasury_name,
            "transaction_type": "lp_withdrawal",
            "amount": treasury_add_amount,
            "currency": treasury_currency,
            "original_amount": tx_data.amount,
            "original_currency": tx_data.currency,
            "reference": f"Withdrawal from LP: {account['lp_name']}{conversion_note}",
            "lp_transaction_id": tx_id,
            "created_at": now.isoformat(),
            "created_by": user["user_id"],
            "created_by_name": user["name"]
        })
    
    # Create LP transaction
    tx_doc = {
        "lp_transaction_id": tx_id,
        "lp_id": lp_id,
        "lp_name": account["lp_name"],
        "transaction_type": LPTransactionType.WITHDRAWAL,
        "amount": tx_data.amount,
        "currency": tx_data.currency,
        "treasury_account_id": tx_data.treasury_account_id,
        "treasury_name": treasury_name,
        "reference": tx_data.reference or f"WTH-{tx_id[-8:].upper()}",
        "notes": tx_data.notes,
        "balance_before": account.get("balance", 0),
        "balance_after": account.get("balance", 0) - tx_data.amount,
        "created_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    
    await db.lp_transactions.insert_one(tx_doc)
    
    # Update LP account balance
    await db.lp_accounts.update_one(
        {"lp_id": lp_id},
        {
            "$inc": {"balance": -tx_data.amount, "total_withdrawals": tx_data.amount, "transaction_count": 1},
            "$set": {"updated_at": now.isoformat()}
        }
    )
    
    tx_doc.pop("_id", None)
    await log_activity(request, user, "create", "lp_management", "LP withdrawal recorded")

    return tx_doc

@api_router.get("/lp/export/csv")
async def export_lp_csv(user: dict = Depends(require_permission(Modules.LP_MANAGEMENT, Actions.EXPORT))):
    """Export LP accounts and transactions to CSV"""
    import csv
    import io
    
    accounts = await db.lp_accounts.find({}, {"_id": 0}).to_list(1000)
    transactions = await db.lp_transactions.find({}, {"_id": 0}).sort("created_at", -1).to_list(5000)
    
    output = io.StringIO()
    
    # Write accounts
    output.write("=== LP ACCOUNTS ===\n")
    if accounts:
        writer = csv.DictWriter(output, fieldnames=["lp_id", "lp_name", "currency", "balance", "total_deposits", "total_withdrawals", "status"])
        writer.writeheader()
        for acc in accounts:
            writer.writerow({k: acc.get(k, "") for k in writer.fieldnames})
    
    output.write("\n=== LP TRANSACTIONS ===\n")
    if transactions:
        writer = csv.DictWriter(output, fieldnames=["lp_transaction_id", "lp_name", "transaction_type", "amount", "currency", "reference", "created_at", "created_by_name"])
        writer.writeheader()
        for tx in transactions:
            writer.writerow({k: tx.get(k, "") for k in writer.fieldnames})
    
    from fastapi.responses import StreamingResponse
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=lp_export_{datetime.now().strftime('%Y%m%d')}.csv"}
    )


# ============== DEALING P&L ROUTES ==============

@api_router.get("/dealing-pnl")
async def get_dealing_pnl_records(
    limit: int = 30,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(require_permission(Modules.LP_MANAGEMENT, Actions.VIEW))
):
    """Get all dealing P&L records with calculated values"""
    query = {}
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    records = await db.dealing_pnl.find(query, {"_id": 0}).sort("date", -1).limit(limit).to_list(limit)
    
    # Calculate derived values for each record
    result = []
    for i, record in enumerate(records):
        # Get previous day's floating values
        prev_mt5_floating = 0
        prev_lp_entries = {}
        
        # Find previous record by date
        prev_record = await db.dealing_pnl.find_one(
            {"date": {"$lt": record["date"]}},
            {"_id": 0},
            sort=[("date", -1)]
        )
        if prev_record:
            prev_mt5_floating = prev_record.get("mt5_floating_pnl", 0)
            # Build map of previous LP floating values
            for lp_entry in prev_record.get("lp_entries", []):
                prev_lp_entries[lp_entry.get("lp_id")] = lp_entry.get("floating_pnl", 0)
        
        # Calculate MT5 dealing P&L
        mt5_booked = record.get("mt5_booked_pnl", 0)
        mt5_floating = record.get("mt5_floating_pnl", 0)
        mt5_floating_change = mt5_floating - prev_mt5_floating
        broker_mt5_pnl = -mt5_booked - mt5_floating_change
        
        # Calculate LP dealing P&L for each LP
        lp_entries_with_calc = []
        total_lp_booked = 0
        total_lp_floating = 0
        total_lp_floating_change = 0
        total_broker_lp_pnl = 0
        
        for lp_entry in record.get("lp_entries", []):
            lp_id = lp_entry.get("lp_id")
            lp_booked = lp_entry.get("booked_pnl", 0)
            lp_floating = lp_entry.get("floating_pnl", 0)
            prev_floating = prev_lp_entries.get(lp_id, 0)
            floating_change = lp_floating - prev_floating
            lp_pnl = lp_booked + floating_change
            
            lp_entries_with_calc.append({
                **lp_entry,
                "prev_floating": prev_floating,
                "floating_change": floating_change,
                "lp_pnl": lp_pnl,
            })
            
            total_lp_booked += lp_booked
            total_lp_floating += lp_floating
            total_lp_floating_change += floating_change
            total_broker_lp_pnl += lp_pnl
        
        # Total Dealing P&L
        total_dealing_pnl = broker_mt5_pnl + total_broker_lp_pnl
        
        result.append({
            **record,
            "lp_entries": lp_entries_with_calc,
            "prev_mt5_floating": prev_mt5_floating,
            "mt5_floating_change": mt5_floating_change,
            "broker_mt5_pnl": broker_mt5_pnl,
            "total_lp_booked": total_lp_booked,
            "total_lp_floating": total_lp_floating,
            "total_lp_floating_change": total_lp_floating_change,
            "broker_lp_pnl": total_broker_lp_pnl,
            "total_dealing_pnl": total_dealing_pnl,
        })
    
    return result


@api_router.get("/dealing-pnl/summary")
async def get_dealing_pnl_summary(
    days: int = 30,
    user: dict = Depends(require_permission(Modules.LP_MANAGEMENT, Actions.VIEW))
):
    """Get summary statistics for dealing P&L"""
    from datetime import timedelta
    
    end_date = datetime.now(timezone.utc).date()
    start_date = end_date - timedelta(days=days)
    
    records = await db.dealing_pnl.find(
        {"date": {"$gte": start_date.isoformat(), "$lte": end_date.isoformat()}},
        {"_id": 0}
    ).sort("date", -1).to_list(1000)
    
    if not records:
        return {
            "total_dealing_pnl": 0,
            "total_mt5_booked": 0,
            "total_lp_booked": 0,
            "record_count": 0,
            "profitable_days": 0,
            "loss_days": 0,
            "best_day": None,
            "worst_day": None,
        }
    
    # Calculate totals
    total_dealing_pnl = 0
    total_mt5_booked = 0
    total_lp_booked = 0
    profitable_days = 0
    loss_days = 0
    best_day = {"date": None, "pnl": float('-inf')}
    worst_day = {"date": None, "pnl": float('inf')}
    
    prev_record = None
    for record in reversed(records):  # Process in chronological order
        mt5_booked = record.get("mt5_booked_pnl", 0)
        mt5_floating = record.get("mt5_floating_pnl", 0)
        lp_booked = record.get("lp_booked_pnl", 0)
        lp_floating = record.get("lp_floating_pnl", 0)
        
        prev_mt5_floating = prev_record.get("mt5_floating_pnl", 0) if prev_record else 0
        prev_lp_floating = prev_record.get("lp_floating_pnl", 0) if prev_record else 0
        
        mt5_floating_change = mt5_floating - prev_mt5_floating
        lp_floating_change = lp_floating - prev_lp_floating
        
        broker_mt5_pnl = -mt5_booked - mt5_floating_change
        broker_lp_pnl = lp_booked + lp_floating_change
        day_pnl = broker_mt5_pnl + broker_lp_pnl
        
        total_dealing_pnl += day_pnl
        total_mt5_booked += mt5_booked
        total_lp_booked += lp_booked
        
        if day_pnl > 0:
            profitable_days += 1
        elif day_pnl < 0:
            loss_days += 1
        
        if day_pnl > best_day["pnl"]:
            best_day = {"date": record["date"], "pnl": day_pnl}
        if day_pnl < worst_day["pnl"]:
            worst_day = {"date": record["date"], "pnl": day_pnl}
        
        prev_record = record
    
    return {
        "total_dealing_pnl": round(total_dealing_pnl, 2),
        "total_mt5_booked": round(total_mt5_booked, 2),
        "total_lp_booked": round(total_lp_booked, 2),
        "record_count": len(records),
        "profitable_days": profitable_days,
        "loss_days": loss_days,
        "best_day": best_day if best_day["date"] else None,
        "worst_day": worst_day if worst_day["date"] else None,
    }


@api_router.post("/dealing-pnl")
async def create_dealing_pnl(request: Request, data: DealingPnLCreate, user: dict = Depends(require_permission(Modules.LP_MANAGEMENT, Actions.CREATE))):

    """Create or update a dealing P&L record for a specific date"""
    now = datetime.now(timezone.utc)
    
    # Check if record already exists for this date
    existing = await db.dealing_pnl.find_one({"date": data.date}, {"_id": 0})
    
    # Process LP entries - fetch LP names if not provided
    lp_entries = []
    for entry in (data.lp_entries or []):
        lp_entry = {
            "lp_id": entry.lp_id,
            "lp_name": entry.lp_name,
            "booked_pnl": entry.booked_pnl,
            "floating_pnl": entry.floating_pnl,
        }
        # Fetch LP name if not provided
        if not lp_entry["lp_name"]:
            lp = await db.lp_accounts.find_one({"lp_id": entry.lp_id}, {"_id": 0, "lp_name": 1})
            if lp:
                lp_entry["lp_name"] = lp.get("lp_name")
        lp_entries.append(lp_entry)
    
    record = {
        "date": data.date,
        "mt5_booked_pnl": data.mt5_booked_pnl,
        "mt5_floating_pnl": data.mt5_floating_pnl,
        "lp_entries": lp_entries,
        "mt5_screenshot": data.mt5_screenshot,
        "lp_screenshot": data.lp_screenshot,
        "notes": data.notes,
        "updated_at": now.isoformat(),
        "updated_by": user["user_id"],
        "updated_by_name": user["name"],
    }
    
    if existing:
        await db.dealing_pnl.update_one(
            {"date": data.date},
            {"$set": record}
        )
        return {"message": "Dealing P&L record updated", "date": data.date}
    else:
        record["record_id"] = f"dpnl_{uuid.uuid4().hex[:12]}"
        record["created_at"] = now.isoformat()
        record["created_by"] = user["user_id"]
        record["created_by_name"] = user["name"]
        await db.dealing_pnl.insert_one(record)
        await log_activity(request, user, "create", "lp_management", "Created Dealing P&L entry")

        return {"message": "Dealing P&L record created", "date": data.date, "record_id": record["record_id"]}


@api_router.get("/dealing-pnl/{date}")
async def get_dealing_pnl_by_date(date: str, user: dict = Depends(require_permission(Modules.LP_MANAGEMENT, Actions.VIEW))):
    """Get dealing P&L record for a specific date"""
    record = await db.dealing_pnl.find_one({"date": date}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="No record found for this date")
    return record


@api_router.delete("/dealing-pnl/{date}")
async def delete_dealing_pnl(request: Request, date: str, user: dict = Depends(require_permission(Modules.LP_MANAGEMENT, Actions.DELETE))):

    """Delete a dealing P&L record"""
    result = await db.dealing_pnl.delete_one({"date": date})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Record not found")
    await log_activity(request, user, "delete", "lp_management", "Deleted Dealing P&L entry")

    return {"message": "Dealing P&L record deleted", "date": date}


@api_router.post("/dealing-pnl/{date}/send-email")
async def send_dealing_pnl_email(date: str, user: dict = Depends(require_permission(Modules.LP_MANAGEMENT, Actions.EXPORT))):
    """Send Dealing P&L email notification for a specific date"""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    
    # Get the record
    record = await db.dealing_pnl.find_one({"date": date}, {"_id": 0})
    if not record:
        raise HTTPException(status_code=404, detail="No record found for this date")
    
    # Get email settings
    settings = await db.app_settings.find_one({"setting_type": "email"}, {"_id": 0})
    if not settings:
        raise HTTPException(status_code=400, detail="Email settings not configured")
    
    director_emails = settings.get("director_emails", [])
    if not director_emails:
        raise HTTPException(status_code=400, detail="No director emails configured")
    
    # Get previous day's record for calculations
    prev_record = await db.dealing_pnl.find_one(
        {"date": {"$lt": date}},
        {"_id": 0},
        sort=[("date", -1)]
    )
    
    prev_mt5_floating = prev_record.get("mt5_floating_pnl", 0) if prev_record else 0
    prev_lp_entries = {}
    if prev_record:
        for lp in prev_record.get("lp_entries", []):
            prev_lp_entries[lp.get("lp_id")] = lp.get("floating_pnl", 0)
    
    # Calculate MT5 P&L
    mt5_booked = record.get("mt5_booked_pnl", 0)
    mt5_floating = record.get("mt5_floating_pnl", 0)
    mt5_floating_change = mt5_floating - prev_mt5_floating
    broker_mt5_pnl = -mt5_booked - mt5_floating_change
    
    # Calculate LP P&L
    total_lp_booked = 0
    total_broker_lp_pnl = 0
    lp_rows = ""
    
    for lp in record.get("lp_entries", []):
        lp_id = lp.get("lp_id")
        lp_name = lp.get("lp_name", lp_id)
        lp_booked = lp.get("booked_pnl", 0)
        lp_floating = lp.get("floating_pnl", 0)
        prev_floating = prev_lp_entries.get(lp_id, 0)
        floating_change = lp_floating - prev_floating
        lp_pnl = lp_booked + floating_change
        
        total_lp_booked += lp_booked
        total_broker_lp_pnl += lp_pnl
        
        pnl_color = "#4ade80" if lp_pnl >= 0 else "#f87171"
        lp_rows += f"<tr><td style='padding:8px;border-bottom:1px solid #333;color:white;'>{lp_name}</td><td style='padding:8px;border-bottom:1px solid #333;color:white;text-align:right;'>${lp_booked:,.0f}</td><td style='padding:8px;border-bottom:1px solid #333;color:white;text-align:right;'>${lp_floating:,.0f}</td><td style='padding:8px;border-bottom:1px solid #333;color:{pnl_color};text-align:right;font-weight:bold;'>${lp_pnl:+,.0f}</td></tr>"
    
    total_dealing_pnl = broker_mt5_pnl + total_broker_lp_pnl
    total_color = "#4ade80" if total_dealing_pnl >= 0 else "#f87171"
    mt5_color = "#4ade80" if broker_mt5_pnl >= 0 else "#f87171"
    lp_color = "#4ade80" if total_broker_lp_pnl >= 0 else "#f87171"
    
    # Generate email HTML
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
    </head>
    <body style="margin:0;padding:0;background-color:#f5f5f5;font-family:'Segoe UI',Arial,sans-serif;">
        <div style="max-width:600px;margin:0 auto;background-color:#0B0C10;color:white;">
            <div style="background:linear-gradient(135deg,#1F2833 0%,#0B0C10 100%);padding:30px;text-align:center;border-bottom:3px solid #66FCF1;">
                <h1 style="color:#66FCF1;margin:0;font-size:24px;letter-spacing:2px;">MILES CAPITALS</h1>
                <p style="color:#C5C6C7;margin:10px 0 0;font-size:14px;">Dealing P&L Report - {date}</p>
            </div>
            
            <div style="padding:30px;">
                <div style="background-color:#1F2833;border-radius:8px;padding:20px;margin-bottom:20px;">
                    <h2 style="color:#66FCF1;font-size:16px;margin:0 0 20px;text-transform:uppercase;letter-spacing:1px;border-bottom:1px solid #66FCF1;padding-bottom:10px;">📈 Daily Dealing P&L</h2>
                    
                    <div style="text-align:center;background-color:#0B0C10;border-radius:6px;padding:25px;margin-bottom:20px;">
                        <div style="color:#C5C6C7;font-size:11px;text-transform:uppercase;letter-spacing:1px;">TOTAL DEALING P&L</div>
                        <div style="color:{total_color};font-size:42px;font-weight:bold;margin-top:5px;">${total_dealing_pnl:+,.0f}</div>
                        <div style="color:#C5C6C7;font-size:12px;margin-top:5px;">USD</div>
                    </div>
                    
                    <table style="width:100%;border-collapse:collapse;margin-bottom:20px;">
                        <tr>
                            <td style="width:50%;padding:15px;background-color:#0B0C10;border-radius:6px;">
                                <div style="color:#C5C6C7;font-size:11px;text-transform:uppercase;">MT5 Broker P&L</div>
                                <div style="color:{mt5_color};font-size:24px;font-weight:bold;margin-top:5px;">${broker_mt5_pnl:+,.0f}</div>
                            </td>
                            <td style="width:10px;"></td>
                            <td style="width:50%;padding:15px;background-color:#0B0C10;border-radius:6px;">
                                <div style="color:#C5C6C7;font-size:11px;text-transform:uppercase;">LP Hedging P&L</div>
                                <div style="color:{lp_color};font-size:24px;font-weight:bold;margin-top:5px;">${total_broker_lp_pnl:+,.0f}</div>
                            </td>
                        </tr>
                    </table>
                    
                    <div style="background-color:#0B0C10;border-radius:6px;padding:15px;margin-bottom:15px;">
                        <h4 style="color:#66FCF1;font-size:12px;margin:0 0 10px;text-transform:uppercase;">MT5 Details</h4>
                        <table style="width:100%;border-collapse:collapse;">
                            <tr>
                                <td style="color:#C5C6C7;font-size:12px;padding:5px 0;">Client Booked P&L:</td>
                                <td style="color:white;font-size:14px;text-align:right;">${mt5_booked:+,.0f}</td>
                            </tr>
                            <tr>
                                <td style="color:#C5C6C7;font-size:12px;padding:5px 0;">Running Floating:</td>
                                <td style="color:white;font-size:14px;text-align:right;">${mt5_floating:,.0f}</td>
                            </tr>
                            <tr>
                                <td style="color:#C5C6C7;font-size:12px;padding:5px 0;">Floating Change:</td>
                                <td style="color:white;font-size:14px;text-align:right;">${mt5_floating_change:+,.0f}</td>
                            </tr>
                        </table>
                    </div>
                    
                    {f'''<div style="background-color:#0B0C10;border-radius:6px;padding:15px;">
                        <h4 style="color:#66FCF1;font-size:12px;margin:0 0 10px;text-transform:uppercase;">LP Breakdown</h4>
                        <table style="width:100%;border-collapse:collapse;">
                            <tr>
                                <th style="padding:8px;border-bottom:1px solid #333;color:#66FCF1;font-size:11px;text-align:left;">LP</th>
                                <th style="padding:8px;border-bottom:1px solid #333;color:#66FCF1;font-size:11px;text-align:right;">Booked</th>
                                <th style="padding:8px;border-bottom:1px solid #333;color:#66FCF1;font-size:11px;text-align:right;">Floating</th>
                                <th style="padding:8px;border-bottom:1px solid #333;color:#66FCF1;font-size:11px;text-align:right;">P&L</th>
                            </tr>
                            {lp_rows}
                        </table>
                    </div>''' if lp_rows else ''}
                </div>
            </div>
            
            <div style="background-color:#1F2833;padding:20px;text-align:center;border-top:1px solid #333;">
                <p style="color:#C5C6C7;font-size:12px;margin:0;">This is an automated Dealing P&L report from Miles Capitals</p>
                <p style="color:#C5C6C7;font-size:12px;margin:5px 0 0;">Submitted by: {user.get('name', 'Unknown')}</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Send email
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"Dealing P&L Report - {date} - ${total_dealing_pnl:+,.0f} USD"
        msg['From'] = settings.get("smtp_user", "")
        msg['To'] = ", ".join(director_emails)
        
        msg.attach(MIMEText(html, 'html'))
        
        with smtplib.SMTP(settings.get("smtp_host", "smtp.gmail.com"), settings.get("smtp_port", 587)) as server:
            server.starttls()
            server.login(settings.get("smtp_user", ""), settings.get("smtp_password", ""))
            server.sendmail(msg['From'], director_emails, msg.as_string())
        
        # Log the email
        await db.logs.insert_one({
            "log_id": f"log_{uuid.uuid4().hex[:12]}",
            "type": "dealing_pnl_email",
            "action": "send_dealing_pnl_email",
            "date": date,
            "total_dealing_pnl": total_dealing_pnl,
            "recipients": director_emails,
            "sent_by": user["user_id"],
            "sent_by_name": user["name"],
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "status": "success"
        })
        
        return {
            "message": "Dealing P&L email sent successfully",
            "date": date,
            "total_dealing_pnl": total_dealing_pnl,
            "recipients": director_emails
        }
        
    except Exception as e:
        logger.error(f"Failed to send dealing P&L email: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to send email: {str(e)}")


# ============== PSP ROUTES ==============

@api_router.get("/psp")
async def get_psps(user: dict = Depends(require_permission(Modules.PSP, Actions.VIEW))):
    psps = await db.psps.find({}, {"_id": 0}).to_list(1000)
    
    # Batch fetch treasury accounts to avoid N+1 queries
    treasury_ids = list(set(psp.get("settlement_destination_id") for psp in psps if psp.get("settlement_destination_id")))
    treasury_map = {}
    if treasury_ids:
        treasuries = await db.treasury_accounts.find({"account_id": {"$in": treasury_ids}}, {"_id": 0}).to_list(len(treasury_ids))
        treasury_map = {t["account_id"]: t for t in treasuries}
    
    for psp in psps:
        if psp.get("settlement_destination_id"):
            dest = treasury_map.get(psp["settlement_destination_id"])
            psp["settlement_destination_name"] = dest["account_name"] if dest else "Unknown"
            psp["settlement_destination_bank"] = dest.get("bank_name") if dest else None
    return psps

@api_router.get("/psp/{psp_id}")
async def get_psp(psp_id: str, user: dict = Depends(require_permission(Modules.PSP, Actions.VIEW))):
    psp = await db.psps.find_one({"psp_id": psp_id}, {"_id": 0})
    if not psp:
        raise HTTPException(status_code=404, detail="PSP not found")
    return psp

@api_router.post("/psp")
async def create_psp(request: Request, psp_data: PSPCreate, user: dict = Depends(require_permission(Modules.PSP, Actions.CREATE))):

    # Verify settlement destination exists
    dest = await db.treasury_accounts.find_one({"account_id": psp_data.settlement_destination_id}, {"_id": 0})
    if not dest:
        raise HTTPException(status_code=404, detail="Settlement destination account not found")
    
    psp_id = f"psp_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    psp_doc = {
        "psp_id": psp_id,
        **psp_data.model_dump(),
        "total_volume": 0.0,
        "total_commission": 0.0,
        "pending_settlement": 0.0,
        "status": PSPStatus.ACTIVE,
        "created_at": now.isoformat(),
        "updated_at": now.isoformat()
    }
    
    await db.psps.insert_one(psp_doc)
    await log_activity(request, user, "create", "psp", "Created PSP")

    return await db.psps.find_one({"psp_id": psp_id}, {"_id": 0})

@api_router.put("/psp/{psp_id}")
async def update_psp(request: Request, psp_id: str, update_data: PSPUpdate, user: dict = Depends(require_permission(Modules.PSP, Actions.EDIT))):

    updates = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No updates provided")
    
    # Verify settlement destination if being updated
    if updates.get("settlement_destination_id"):
        dest = await db.treasury_accounts.find_one({"account_id": updates["settlement_destination_id"]}, {"_id": 0})
        if not dest:
            raise HTTPException(status_code=404, detail="Settlement destination account not found")
    
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.psps.update_one({"psp_id": psp_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="PSP not found")
    
    await log_activity(request, user, "edit", "psp", "Updated PSP")

    return await db.psps.find_one({"psp_id": psp_id}, {"_id": 0})

@api_router.delete("/psp/{psp_id}")
async def delete_psp(request: Request, psp_id: str, user: dict = Depends(require_permission(Modules.PSP, Actions.DELETE))):

    result = await db.psps.delete_one({"psp_id": psp_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="PSP not found")
    await log_activity(request, user, "delete", "psp", "Deleted PSP")

    return {"message": "PSP deleted"}

# PSP Settlements
@api_router.get("/psp/{psp_id}/settlements")
async def get_psp_settlements(
    psp_id: str,
    page: int = 1,
    page_size: int = 20,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(require_permission(Modules.PSP, Actions.VIEW))
):
    query = {"psp_id": psp_id}
    if date_from:
        query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"
    
    total = await db.psp_settlements.count_documents(query)
    skip = (page - 1) * page_size
    settlements = await db.psp_settlements.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    # Enrich each settlement with payment currency info from its transactions
    for stl in settlements:
        tx_ids = stl.get("transaction_ids", [])
        if tx_ids:
            txs = await db.transactions.find(
                {"transaction_id": {"$in": tx_ids}},
                {"_id": 0, "base_currency": 1, "exchange_rate": 1, "base_amount": 1}
            ).to_list(len(tx_ids))
            currencies = set(tx.get("base_currency") for tx in txs if tx.get("base_currency") and tx.get("base_currency") != "USD")
            if currencies:
                stl["payment_currency"] = ", ".join(sorted(currencies))
                rates = [tx.get("exchange_rate") for tx in txs if tx.get("exchange_rate")]
                stl["avg_exchange_rate"] = round(sum(rates) / len(rates), 4) if rates else None
                total_base = sum(tx.get("base_amount", 0) for tx in txs if tx.get("base_amount"))
                stl["base_gross_amount"] = round(total_base, 2) if total_base else None
            else:
                stl["payment_currency"] = "USD"
        else:
            stl["payment_currency"] = stl.get("treasury_currency", "USD")
    return {"items": settlements, "total": total, "page": page, "page_size": page_size, "total_pages": max(1, -(-total // page_size))}

# Get transactions included in a specific settlement
@api_router.get("/psp/{psp_id}/settlement/{settlement_id}/transactions")
async def get_settlement_transactions(psp_id: str, settlement_id: str, user: dict = Depends(require_permission(Modules.PSP, Actions.VIEW))):
    settlement = await db.psp_settlements.find_one({"settlement_id": settlement_id, "psp_id": psp_id}, {"_id": 0})
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement not found")
    tx_ids = settlement.get("transaction_ids", [])
    if not tx_ids:
        return []
    txs = await db.transactions.find(
        {"transaction_id": {"$in": tx_ids}},
        {"_id": 0, "transaction_id": 1, "reference": 1, "client_name": 1, "amount": 1, "base_amount": 1, "base_currency": 1, "exchange_rate": 1, "created_at": 1, "transaction_date": 1, "transaction_type": 1, "psp_commission_amount": 1, "psp_reserve_fund_amount": 1, "psp_chargeback_amount": 1, "psp_extra_charges": 1, "psp_extra_commission": 1, "psp_withdrawal_extra_commission": 1, "psp_gateway_fee": 1, "psp_net_amount": 1}
    ).to_list(len(tx_ids))
    return txs

@api_router.get("/psp-settlements")
async def get_all_settlements(
    user: dict = Depends(require_permission(Modules.PSP, Actions.VIEW)),
    status: Optional[str] = None,
    limit: int = 100
):
    query = {}
    if status:
        query["status"] = status
    settlements = await db.psp_settlements.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return settlements

@api_router.post("/psp/{psp_id}/settle")
async def create_settlement(request: Request, psp_id: str, user: dict = Depends(require_permission(Modules.PSP, Actions.CREATE))):

    """Create a settlement for pending PSP transactions"""
    psp = await db.psps.find_one({"psp_id": psp_id}, {"_id": 0})
    if not psp:
        raise HTTPException(status_code=404, detail="PSP not found")
    
    # Get pending transactions for this PSP
    pending_txs = await db.transactions.find({
        "psp_id": psp_id,
        "status": {"$in": [TransactionStatus.APPROVED, TransactionStatus.COMPLETED]},
        "settled": {"$ne": True}
    }, {"_id": 0}).to_list(1000)
    
    if not pending_txs:
        raise HTTPException(status_code=400, detail="No pending transactions to settle")
    
    # Calculate totals
    gross_amount = sum(tx["amount"] for tx in pending_txs)
    
    # Commission
    commission_rate = psp.get("commission_rate", 0) / 100
    commission_amount = round(gross_amount * commission_rate, 2)
    
    # Reserve Fund (from PSP default rate)
    reserve_fund_rate = psp.get("reserve_fund_rate", psp.get("chargeback_rate", 0)) / 100
    reserve_fund_amount = round(gross_amount * reserve_fund_rate, 2)
    
    # Extra charges (sum from individual transactions)
    total_extra_charges = sum(tx.get("psp_extra_charges", 0) for tx in pending_txs)
    
    # Gateway fees (per transaction)
    gateway_fee = psp.get("gateway_fee", 0)
    total_gateway_fees = round(gateway_fee * len(pending_txs), 2)
    
    # Individual transaction reserve fund amounts (override PSP rate if set per transaction)
    total_tx_reserve = sum(tx.get("psp_reserve_fund_amount", tx.get("psp_chargeback_amount", 0)) for tx in pending_txs)
    # Use transaction-level amounts if any, otherwise use PSP rate
    final_reserve_fund = total_tx_reserve if total_tx_reserve > 0 else reserve_fund_amount
    
    # Net amount = Gross - Commission - Reserve Fund - Extra Charges - Gateway Fees
    net_amount = gross_amount - commission_amount - final_reserve_fund - total_extra_charges - total_gateway_fees
    
    settlement_id = f"stl_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    settlement_date = now + timedelta(days=psp.get("settlement_days", 1))
    
    settlement_doc = {
        "settlement_id": settlement_id,
        "psp_id": psp_id,
        "psp_name": psp["psp_name"],
        "gross_amount": gross_amount,
        "commission_rate": psp.get("commission_rate", 0),
        "commission_amount": commission_amount,
        "reserve_fund_rate": psp.get("reserve_fund_rate", psp.get("chargeback_rate", 0)),
        "reserve_fund_amount": final_reserve_fund,
        "chargeback_rate": psp.get("reserve_fund_rate", psp.get("chargeback_rate", 0)),
        "chargeback_amount": final_reserve_fund,
        "extra_charges": total_extra_charges,
        "gateway_fees": total_gateway_fees,
        "total_deductions": commission_amount + final_reserve_fund + total_extra_charges + total_gateway_fees,
        "net_amount": net_amount,
        "holding_days": psp.get("holding_days", 0),
        "transaction_count": len(pending_txs),
        "transaction_ids": [tx["transaction_id"] for tx in pending_txs],
        "settlement_destination_id": psp["settlement_destination_id"],
        "status": PSPSettlementStatus.PENDING,
        "expected_settlement_date": settlement_date.isoformat(),
        "created_at": now.isoformat(),
        "settled_at": None,
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    
    await db.psp_settlements.insert_one(settlement_doc)
    
    # Mark transactions as pending settlement
    await db.transactions.update_many(
        {"transaction_id": {"$in": [tx["transaction_id"] for tx in pending_txs]}},
        {"$set": {"settlement_id": settlement_id, "settlement_status": "pending"}}
    )
    
    # Update PSP stats
    await db.psps.update_one(
        {"psp_id": psp_id},
        {"$inc": {"pending_settlement": net_amount}}
    )
    
    await log_activity(request, user, "create", "psp", "Created PSP settlement")

    return await db.psp_settlements.find_one({"settlement_id": settlement_id}, {"_id": 0})

@api_router.post("/psp-settlements/{settlement_id}/complete")
async def complete_settlement(request: Request, settlement_id: str, user: dict = Depends(require_permission(Modules.PSP, Actions.APPROVE))):

    """Mark a settlement as completed and transfer funds to treasury"""
    settlement = await db.psp_settlements.find_one({"settlement_id": settlement_id}, {"_id": 0})
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement not found")
    
    if settlement["status"] == PSPSettlementStatus.COMPLETED:
        raise HTTPException(status_code=400, detail="Settlement already completed")
    
    now = datetime.now(timezone.utc)
    
    # Get destination treasury account for currency conversion
    dest = await db.treasury_accounts.find_one({"account_id": settlement["settlement_destination_id"]}, {"_id": 0})
    dest_currency = dest.get("currency", "USD") if dest else "USD"
    
    # Settlement amounts are in USD (transaction currency); convert to treasury currency
    settlement_net = settlement["net_amount"]
    treasury_amount = convert_currency(settlement_net, "USD", dest_currency)
    
    # Update treasury balance (in treasury account's currency)
    await db.treasury_accounts.update_one(
        {"account_id": settlement["settlement_destination_id"]},
        {"$inc": {"balance": treasury_amount}, "$set": {"updated_at": now.isoformat()}}
    )
    
    # Update settlement status
    await db.psp_settlements.update_one(
        {"settlement_id": settlement_id},
        {"$set": {
            "status": PSPSettlementStatus.COMPLETED,
            "settled_at": now.isoformat()
        }}
    )
    
    # Mark transactions as settled
    await db.transactions.update_many(
        {"settlement_id": settlement_id},
        {"$set": {"settled": True, "settlement_status": "completed"}}
    )
    
    # Update PSP stats
    psp = await db.psps.find_one({"psp_id": settlement["psp_id"]}, {"_id": 0})
    if psp:
        await db.psps.update_one(
            {"psp_id": settlement["psp_id"]},
            {
                "$inc": {
                    "total_volume": settlement["gross_amount"],
                    "total_commission": settlement["commission_amount"],
                    "pending_settlement": -settlement["net_amount"]
                }
            }
        )
    
    await log_activity(request, user, "approve", "psp", "Completed PSP settlement")

    return await db.psp_settlements.find_one({"settlement_id": settlement_id}, {"_id": 0})

# Get pending PSP transactions (not yet settled)
@api_router.get("/psp/{psp_id}/pending-transactions")
async def get_psp_pending_transactions(
    psp_id: str,
    page: int = 1,
    page_size: int = 20,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(require_permission(Modules.PSP, Actions.VIEW))
):
    """Get approved DEPOSIT transactions for a PSP that haven't been settled"""
    query = {
        "psp_id": psp_id,
        "destination_type": "psp",
        "transaction_type": "deposit",
        "status": TransactionStatus.APPROVED,
        "settled": {"$ne": True}
    }
    if date_from:
        query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"
    
    total = await db.transactions.count_documents(query)
    skip = (page - 1) * page_size
    transactions = await db.transactions.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    return {"items": transactions, "total": total, "page": page, "page_size": page_size, "total_pages": max(1, -(-total // page_size))}

@api_router.get("/psp/{psp_id}/withdrawal-transactions")
async def get_psp_withdrawal_transactions(
    psp_id: str,
    page: int = 1,
    page_size: int = 20,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(require_permission(Modules.PSP, Actions.VIEW))
):
    """Get withdrawal transactions for a PSP (unsettled only)"""
    query = {
        "psp_id": psp_id,
        "destination_type": "psp",
        "transaction_type": "withdrawal",
        "status": {"$in": [TransactionStatus.APPROVED, TransactionStatus.PENDING]},
        "settled": {"$ne": True}
    }
    if date_from:
        query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"
    
    total = await db.transactions.count_documents(query)
    skip = (page - 1) * page_size
    transactions = await db.transactions.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    return {"items": transactions, "total": total, "page": page, "page_size": page_size, "total_pages": max(1, -(-total // page_size))}

@api_router.post("/psp/{psp_id}/deposit-extra-commission")
async def update_psp_deposit_extra_commission(
    request: Request,
    psp_id: str,
    user: dict = Depends(require_permission(Modules.PSP, Actions.EDIT))
):
    """Add extra commission to a PSP deposit transaction"""
    body = await request.json()
    transaction_id = body.get("transaction_id")
    extra_commission = float(body.get("extra_commission", 0))
    extra_commission_note = body.get("note", "")
    
    if not transaction_id:
        raise HTTPException(status_code=400, detail="Transaction ID required")
    
    tx = await db.transactions.find_one({"transaction_id": transaction_id, "psp_id": psp_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    old_extra = tx.get("psp_extra_commission", 0) or 0
    
    # Recalculate net amount: gross - commission - extra charges - reserve fund - extra commission
    gross = tx.get("amount", 0)
    commission = tx.get("psp_commission_amount", 0)
    extra_charges = tx.get("psp_extra_charges", 0) or 0
    reserve_fund = tx.get("psp_reserve_fund_amount", tx.get("psp_chargeback_amount", 0)) or 0
    new_net = round(gross - commission - extra_charges - reserve_fund - extra_commission, 2)
    
    await db.transactions.update_one(
        {"transaction_id": transaction_id},
        {"$set": {
            "psp_extra_commission": round(extra_commission, 2),
            "psp_extra_commission_note": extra_commission_note,
            "psp_net_amount": new_net,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update PSP pending_settlement: adjust by the difference in extra commission
    diff = extra_commission - old_extra
    if diff != 0:
        await db.psps.update_one(
            {"psp_id": psp_id},
            {"$inc": {"pending_settlement": -diff}}
        )
    
    await log_activity(request, user, "edit", "psp", f"Added extra commission ${extra_commission} to PSP deposit {transaction_id}")
    return {"message": "Extra commission updated", "new_net_amount": new_net}

@api_router.post("/psp/{psp_id}/withdrawal-extra-commission")
async def update_psp_withdrawal_extra_commission(
    request: Request,
    psp_id: str,
    user: dict = Depends(require_permission(Modules.PSP, Actions.EDIT))
):
    """Add extra commission to a PSP withdrawal transaction"""
    body = await request.json()
    transaction_id = body.get("transaction_id")
    extra_commission = float(body.get("extra_commission", 0))
    extra_commission_note = body.get("note", "")
    
    if not transaction_id:
        raise HTTPException(status_code=400, detail="Transaction ID required")
    
    tx = await db.transactions.find_one({"transaction_id": transaction_id, "psp_id": psp_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    await db.transactions.update_one(
        {"transaction_id": transaction_id},
        {"$set": {
            "psp_withdrawal_extra_commission": round(extra_commission, 2),
            "psp_withdrawal_extra_commission_note": extra_commission_note,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await log_activity(request, user, "edit", "psp", f"Added extra commission ${extra_commission} to PSP withdrawal {transaction_id}")
    return {"message": "Extra commission updated"}

# Get PSP dashboard summary
@api_router.get("/psp-summary")
async def get_psp_summary(
    page: int = 1,
    page_size: int = 20,
    user: dict = Depends(require_permission(Modules.PSP, Actions.VIEW))
):
    """Get summary of all PSPs with pending settlements"""
    total = await db.psps.count_documents({"status": PSPStatus.ACTIVE})
    skip = (page - 1) * page_size
    psps = await db.psps.find({"status": PSPStatus.ACTIVE}, {"_id": 0}).skip(skip).limit(page_size).to_list(page_size)
    now = datetime.now(timezone.utc)
    
    result = []
    for psp in psps:
        # Get pending transactions count and amount
        pending_txs = await db.transactions.find({
            "psp_id": psp["psp_id"],
            "destination_type": "psp",
            "transaction_type": "deposit",
            "status": TransactionStatus.APPROVED,
            "settled": {"$ne": True}
        }, {"_id": 0}).to_list(1000)
        
        # Get unsettled approved withdrawals
        withdrawal_txs = await db.transactions.find({
            "psp_id": psp["psp_id"],
            "destination_type": "psp",
            "transaction_type": "withdrawal",
            "status": TransactionStatus.APPROVED,
            "settled": {"$ne": True}
        }, {"_id": 0}).to_list(1000)
        
        pending_count = len(pending_txs)
        pending_amount_gross = sum(tx.get("psp_net_amount", tx.get("amount", 0)) for tx in pending_txs)
        
        # Check for overdue settlements
        overdue_count = 0
        for tx in pending_txs:
            exp_date = tx.get("psp_expected_settlement_date")
            if exp_date:
                exp_dt = datetime.fromisoformat(exp_date.replace('Z', '+00:00'))
                if exp_dt.tzinfo is None:
                    exp_dt = exp_dt.replace(tzinfo=timezone.utc)
                if exp_dt < now:
                    overdue_count += 1
        
        # Calculate reserve fund held from pending transactions
        reserve_fund_rate = psp.get("reserve_fund_rate", psp.get("chargeback_rate", 0)) / 100
        reserve_from_pending = 0
        for tx in pending_txs:
            rf = tx.get("psp_reserve_fund_amount", tx.get("psp_chargeback_amount", 0))
            if rf > 0:
                reserve_from_pending += rf
            else:
                reserve_from_pending += round(tx.get("amount", 0) * reserve_fund_rate, 2)

        # Pending Amount = Deposit Net - Reserve Fund - Withdrawals - Withdrawal Extra Commission
        withdrawal_total = sum(tx.get("amount", 0) for tx in withdrawal_txs)
        withdrawal_extra_comm = sum(tx.get("psp_withdrawal_extra_commission", 0) or 0 for tx in withdrawal_txs)
        pending_amount = max(round(pending_amount_gross - reserve_from_pending - withdrawal_total - withdrawal_extra_comm, 2), 0)

        # Total reserve held includes pending + settled unreleased
        total_reserve_held = reserve_from_pending

        # Also count released/unreleased reserve funds from settled transactions
        settled_with_reserve = await db.transactions.find({
            "psp_id": psp["psp_id"],
            "destination_type": "psp",
            "settled": True,
            "$or": [
                {"psp_reserve_fund_amount": {"$gt": 0}},
                {"psp_chargeback_amount": {"$gt": 0}}
            ]
        }, {"_id": 0, "psp_reserve_fund_amount": 1, "psp_chargeback_amount": 1, "reserve_fund_released": 1}).to_list(10000)
        
        held_from_settled = sum(
            tx.get("psp_reserve_fund_amount", tx.get("psp_chargeback_amount", 0))
            for tx in settled_with_reserve if not tx.get("reserve_fund_released")
        )
        total_reserve_held += held_from_settled

        # Get settlement destination
        dest = await db.treasury_accounts.find_one({"account_id": psp.get("settlement_destination_id")}, {"_id": 0})
        
        result.append({
            **psp,
            "pending_transactions_count": pending_count,
            "pending_amount": pending_amount,
            "overdue_count": overdue_count,
            "total_reserve_fund_held": round(total_reserve_held, 2),
            "settlement_destination_name": dest["account_name"] if dest else "Unknown",
            "settlement_destination_bank": dest.get("bank_name") if dest else None
        })
    
    return {"items": result, "total": total, "page": page, "page_size": page_size, "total_pages": max(1, -(-total // page_size))}

# Model for PSP transaction charges
class PSPTransactionCharges(BaseModel):
    reserve_fund_amount: float = 0  # Reserve fund amount for this transaction
    extra_charges: float = 0  # Extra charges for this transaction
    charges_description: Optional[str] = None  # Description of charges

@api_router.put("/psp/transactions/{transaction_id}/charges")
async def update_psp_transaction_charges(

    request: Request,

    transaction_id: str,
    charges: PSPTransactionCharges,
    user: dict = Depends(require_permission(Modules.PSP, Actions.EDIT))
):
    """Record reserve fund and extra charges on a PSP transaction"""
    tx = await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    if tx.get("destination_type") != "psp":
        raise HTTPException(status_code=400, detail="Transaction is not a PSP transaction")
    
    if tx.get("settled"):
        raise HTTPException(status_code=400, detail="Cannot update charges on settled transaction")
    
    now = datetime.now(timezone.utc)
    
    # Calculate old net for pending_settlement adjustment
    old_extra_charges = tx.get("psp_extra_charges", 0) or 0
    old_reserve_fund = tx.get("psp_reserve_fund_amount", tx.get("psp_chargeback_amount", 0)) or 0
    extra_commission = tx.get("psp_extra_commission", 0) or 0
    
    # Calculate new net amount (include extra_commission if it exists)
    gross_amount = tx.get("amount", 0)
    commission = tx.get("psp_commission_amount", 0)
    new_net = round(gross_amount - commission - charges.reserve_fund_amount - charges.extra_charges - extra_commission, 2)
    
    updates = {
        "psp_reserve_fund_amount": charges.reserve_fund_amount,
        "psp_chargeback_amount": charges.reserve_fund_amount,
        "psp_extra_charges": charges.extra_charges,
        "psp_charges_description": charges.charges_description,
        "psp_total_deductions": commission + charges.reserve_fund_amount + charges.extra_charges + extra_commission,
        "psp_net_amount": new_net,
        "charges_updated_at": now.isoformat(),
        "charges_updated_by": user["user_id"],
        "charges_updated_by_name": user["name"]
    }
    
    await db.transactions.update_one(
        {"transaction_id": transaction_id},
        {"$set": updates}
    )
    
    # Adjust PSP pending_settlement by the difference in deductions
    old_deductions = old_extra_charges + old_reserve_fund
    new_deductions = charges.extra_charges + charges.reserve_fund_amount
    deduction_diff = new_deductions - old_deductions
    if deduction_diff != 0 and tx.get("psp_id"):
        await db.psps.update_one(
            {"psp_id": tx["psp_id"]},
            {"$inc": {"pending_settlement": round(-deduction_diff, 2)}}
        )
    
    await log_activity(request, user, "edit", "psp", "Updated PSP transaction charges")

    return await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})

# Mark single PSP transaction as awaiting settlement (holding period)
@api_router.post("/psp/transactions/{transaction_id}/mark-awaiting")
async def mark_psp_transaction_awaiting(
    transaction_id: str, 
    user: dict = Depends(require_permission(Modules.PSP, Actions.EDIT))
):
    """Mark a PSP transaction as awaiting settlement (in holding period)"""
    tx = await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    if tx.get("destination_type") != "psp":
        raise HTTPException(status_code=400, detail="Transaction is not a PSP transaction")
    
    if tx.get("settled"):
        raise HTTPException(status_code=400, detail="Transaction already settled")
    
    now = datetime.now(timezone.utc)
    
    # Get PSP info for holding days
    psp = await db.psps.find_one({"psp_id": tx.get("psp_id")}, {"_id": 0})
    holding_days = psp.get("holding_days", 0) if psp else 0
    release_date = (now + timedelta(days=holding_days)).isoformat() if holding_days > 0 else now.isoformat()
    
    # Mark transaction as awaiting settlement
    await db.transactions.update_one(
        {"transaction_id": transaction_id},
        {"$set": {
            "settlement_status": "awaiting",
            "psp_holding_release_date": release_date,
            "awaiting_marked_at": now.isoformat(),
            "awaiting_marked_by": user["user_id"],
            "awaiting_marked_by_name": user["name"]
        }}
    )
    
    return await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})


# Record actual payment received from PSP (completes settlement)
@api_router.post("/psp/transactions/{transaction_id}/record-payment")
async def record_psp_payment(
    transaction_id: str, 
    destination_account_id: Optional[str] = None,
    actual_amount_received: Optional[float] = None,
    user: dict = Depends(require_permission(Modules.PSP, Actions.EDIT))
):
    """Record actual payment received from PSP and transfer to treasury"""
    tx = await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    if tx.get("destination_type") != "psp":
        raise HTTPException(status_code=400, detail="Transaction is not a PSP transaction")
    
    if tx.get("settled"):
        raise HTTPException(status_code=400, detail="Transaction already settled")
    
    now = datetime.now(timezone.utc)
    
    # Get PSP info for destination
    psp = await db.psps.find_one({"psp_id": tx.get("psp_id")}, {"_id": 0})
    dest_account_id = destination_account_id or (psp.get("settlement_destination_id") if psp else None)
    
    if not dest_account_id:
        raise HTTPException(status_code=400, detail="No destination account specified")
    
    # Get destination account
    dest = await db.treasury_accounts.find_one({"account_id": dest_account_id}, {"_id": 0})
    if not dest:
        raise HTTPException(status_code=404, detail="Destination treasury account not found")
    
    # Amount to settle - use actual amount if provided, otherwise calculated net
    expected_amount = tx.get("psp_net_amount", tx.get("amount", 0))
    settle_amount = actual_amount_received if actual_amount_received is not None else expected_amount
    variance = settle_amount - expected_amount if actual_amount_received is not None else 0
    
    # Currency conversion: convert from transaction currency to treasury account currency
    tx_currency = tx.get("currency", "USD")
    dest_currency = dest.get("currency", "USD")
    treasury_amount = convert_currency(settle_amount, tx_currency, dest_currency)
    
    # Update treasury balance (in treasury account's currency)
    await db.treasury_accounts.update_one(
        {"account_id": dest_account_id},
        {"$inc": {"balance": treasury_amount}, "$set": {"updated_at": now.isoformat()}}
    )
    
    # Add treasury transaction record
    treasury_tx_id = f"ttx_{uuid.uuid4().hex[:12]}"
    conversion_note = f" (Converted: {tx_currency} {settle_amount:,.2f} -> {dest_currency} {treasury_amount:,.2f})" if tx_currency != dest_currency else ""
    treasury_tx = {
        "treasury_transaction_id": treasury_tx_id,
        "account_id": dest_account_id,
        "account_name": dest["account_name"],
        "transaction_type": "psp_settlement",
        "amount": treasury_amount,
        "currency": dest_currency,
        "original_amount": settle_amount,
        "original_currency": tx_currency,
        "reference": f"PSP Settlement - {tx.get('reference', transaction_id)}",
        "description": f"Settlement from {tx.get('psp_name', 'PSP')} - Expected: {tx_currency} {expected_amount:,.2f}, Received: {tx_currency} {settle_amount:,.2f}{conversion_note}",
        "related_transaction_id": transaction_id,
        "psp_id": tx.get("psp_id"),
        "psp_name": tx.get("psp_name"),
        "created_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    await db.treasury_transactions.insert_one(treasury_tx)
    
    # Create PSP settlement record for Settlement History
    settlement_id = f"stl_{uuid.uuid4().hex[:12]}"
    gross_amount = tx.get("amount", 0)
    commission_amount = tx.get("psp_commission_amount", 0)
    reserve_fund_amount = tx.get("psp_reserve_fund_amount", tx.get("psp_chargeback_amount", 0))
    extra_charges = tx.get("psp_extra_charges", 0)
    
    settlement_doc = {
        "settlement_id": settlement_id,
        "psp_id": tx.get("psp_id"),
        "psp_name": tx.get("psp_name", psp.get("psp_name") if psp else "Unknown PSP"),
        "gross_amount": gross_amount,
        "commission_rate": psp.get("commission_rate", 0) if psp else 0,
        "commission_amount": commission_amount,
        "reserve_fund_rate": psp.get("reserve_fund_rate", psp.get("chargeback_rate", 0)) if psp else 0,
        "reserve_fund_amount": reserve_fund_amount,
        "chargeback_rate": psp.get("reserve_fund_rate", psp.get("chargeback_rate", 0)) if psp else 0,
        "chargeback_amount": reserve_fund_amount,
        "extra_charges": extra_charges,
        "gateway_fees": tx.get("psp_gateway_fee", 0),
        "total_deductions": commission_amount + reserve_fund_amount + extra_charges,
        "net_amount": settle_amount,
        "actual_amount_received": settle_amount,
        "variance": variance,
        "holding_days": psp.get("holding_days", 0) if psp else 0,
        "transaction_count": 1,
        "transaction_ids": [transaction_id],
        "settlement_destination_id": dest_account_id,
        "status": PSPSettlementStatus.COMPLETED,
        "expected_settlement_date": now.isoformat(),
        "created_at": now.isoformat(),
        "settled_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"],
        "reference": tx.get("reference", transaction_id)
    }
    await db.psp_settlements.insert_one(settlement_doc)
    
    # Mark transaction as settled
    await db.transactions.update_one(
        {"transaction_id": transaction_id},
        {"$set": {
            "settled": True,
            "settlement_status": "completed",
            "settled_at": now.isoformat(),
            "settled_by": user["user_id"],
            "settled_by_name": user["name"],
            "settlement_destination_id": dest_account_id,
            "settlement_destination_name": dest["account_name"],
            "psp_actual_amount_received": settle_amount,
            "psp_settlement_variance": variance,
            "settlement_id": settlement_id
        }}
    )
    
    # Update PSP stats
    if psp:
        commission_amount = tx.get("psp_commission_amount", 0)
        await db.psps.update_one(
            {"psp_id": tx.get("psp_id")},
            {
                "$inc": {
                    "total_volume": tx.get("amount", 0),
                    "total_commission": commission_amount,
                    "pending_settlement": -expected_amount
                }
            }
        )
    
    return await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})


# Legacy endpoint - Mark single PSP transaction as settled (immediate)
@api_router.post("/psp/transactions/{transaction_id}/settle")
async def settle_psp_transaction(

    request: Request,

    transaction_id: str, 
    destination_account_id: Optional[str] = None,
    user: dict = Depends(require_permission(Modules.PSP, Actions.APPROVE))
):
    """Mark a single PSP transaction as settled and transfer to treasury (immediate settlement)"""
    tx = await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    if tx.get("destination_type") != "psp":
        raise HTTPException(status_code=400, detail="Transaction is not a PSP transaction")
    
    if tx.get("settled"):
        raise HTTPException(status_code=400, detail="Transaction already settled")
    
    if tx.get("status") not in [TransactionStatus.APPROVED, TransactionStatus.COMPLETED, TransactionStatus.PENDING]:
        raise HTTPException(status_code=400, detail="Transaction cannot be settled in current status")
    
    now = datetime.now(timezone.utc)
    
    # Get PSP info for destination
    psp = await db.psps.find_one({"psp_id": tx.get("psp_id")}, {"_id": 0})
    dest_account_id = destination_account_id or (psp.get("settlement_destination_id") if psp else None)
    
    if not dest_account_id:
        raise HTTPException(status_code=400, detail="No destination account specified")
    
    # Get destination account
    dest = await db.treasury_accounts.find_one({"account_id": dest_account_id}, {"_id": 0})
    if not dest:
        raise HTTPException(status_code=404, detail="Destination treasury account not found")
    
    # Amount to settle (net amount after commission)
    settle_amount = tx.get("psp_net_amount", tx.get("amount", 0))
    
    # Currency conversion: convert from transaction currency to treasury account currency
    tx_currency = tx.get("currency", "USD")
    dest_currency = dest.get("currency", "USD")
    treasury_amount = convert_currency(settle_amount, tx_currency, dest_currency)
    
    # Update treasury balance (in treasury account's currency)
    await db.treasury_accounts.update_one(
        {"account_id": dest_account_id},
        {"$inc": {"balance": treasury_amount}, "$set": {"updated_at": now.isoformat()}}
    )
    
    # Add treasury transaction record
    treasury_tx_id = f"ttx_{uuid.uuid4().hex[:12]}"
    conversion_note = f" (Converted: {tx_currency} {settle_amount:,.2f} -> {dest_currency} {treasury_amount:,.2f})" if tx_currency != dest_currency else ""
    treasury_tx = {
        "treasury_transaction_id": treasury_tx_id,
        "account_id": dest_account_id,
        "account_name": dest["account_name"],
        "transaction_type": "psp_settlement",
        "amount": treasury_amount,
        "currency": dest_currency,
        "original_amount": settle_amount,
        "original_currency": tx_currency,
        "reference": f"PSP Settlement - {tx.get('reference', transaction_id)}",
        "description": f"Settlement from {tx.get('psp_name', 'PSP')}{conversion_note}",
        "related_transaction_id": transaction_id,
        "psp_id": tx.get("psp_id"),
        "psp_name": tx.get("psp_name"),
        "created_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    await db.treasury_transactions.insert_one(treasury_tx)
    
    # Mark transaction as settled
    await db.transactions.update_one(
        {"transaction_id": transaction_id},
        {"$set": {
            "settled": True,
            "settlement_status": "completed",
            "settled_at": now.isoformat(),
            "settled_by": user["user_id"],
            "settled_by_name": user["name"],
            "settlement_destination_id": dest_account_id,
            "settlement_destination_name": dest["account_name"]
        }}
    )
    
    # Update PSP stats
    if psp:
        commission_amount = tx.get("psp_commission_amount", 0)
        await db.psps.update_one(
            {"psp_id": tx.get("psp_id")},
            {
                "$inc": {
                    "total_volume": tx.get("amount", 0),
                    "total_commission": commission_amount,
                    "pending_settlement": -settle_amount
                }
            }
        )
    
    await log_activity(request, user, "approve", "psp", "Settled PSP transaction")

    return await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})


# Compound/Batch Settlement: Settle selected PSP transactions as one lump sum
class BatchSettleRequest(BaseModel):
    transaction_ids: List[str]
    destination_account_id: Optional[str] = None
    settlement_date: Optional[str] = None

@api_router.post("/psp/{psp_id}/settle-batch")
async def batch_settle_psp_transactions(
    request: Request,
    psp_id: str,
    body: BatchSettleRequest,
    user: dict = Depends(require_permission(Modules.PSP, Actions.APPROVE))
):
    """Create a compound settlement from selected PSP transactions (lump sum to treasury)"""
    psp = await db.psps.find_one({"psp_id": psp_id}, {"_id": 0})
    if not psp:
        raise HTTPException(status_code=404, detail="PSP not found")

    if not body.transaction_ids:
        raise HTTPException(status_code=400, detail="No transactions selected")

    # Fetch all selected transactions
    selected_txs = await db.transactions.find({
        "transaction_id": {"$in": body.transaction_ids},
        "psp_id": psp_id,
        "destination_type": "psp",
    }, {"_id": 0}).to_list(len(body.transaction_ids))

    if len(selected_txs) != len(body.transaction_ids):
        raise HTTPException(status_code=400, detail="Some transactions not found or do not belong to this PSP")

    # Validate none are already settled
    already_settled = [tx["transaction_id"] for tx in selected_txs if tx.get("settled")]
    if already_settled:
        raise HTTPException(status_code=400, detail=f"Transactions already settled: {', '.join(already_settled)}")

    # Calculate compound totals
    gross_amount = sum(tx.get("amount", 0) for tx in selected_txs)
    total_commission = sum(tx.get("psp_commission_amount", 0) for tx in selected_txs)
    total_reserve = sum(tx.get("psp_reserve_fund_amount", tx.get("psp_chargeback_amount", 0)) for tx in selected_txs)
    total_extra = sum(tx.get("psp_extra_charges", 0) for tx in selected_txs)
    total_gateway = sum(tx.get("psp_gateway_fee", 0) for tx in selected_txs)
    total_deductions = total_commission + total_reserve + total_extra + total_gateway
    net_amount = gross_amount - total_deductions

    # Determine destination treasury account
    dest_account_id = body.destination_account_id or psp.get("settlement_destination_id")
    if not dest_account_id:
        raise HTTPException(status_code=400, detail="No destination account specified")

    dest = await db.treasury_accounts.find_one({"account_id": dest_account_id}, {"_id": 0})
    if not dest:
        raise HTTPException(status_code=404, detail="Destination treasury account not found")

    dest_currency = dest.get("currency", "USD")
    
    # Use actual transaction exchange rates when treasury currency matches payment currency
    # This ensures the exact amount that PSP sends to bank is credited to treasury
    base_currencies = set(tx.get("base_currency") for tx in selected_txs if tx.get("base_currency") and tx.get("base_currency") != "USD")
    if len(base_currencies) == 1 and dest_currency in base_currencies:
        # Treasury currency matches payment currency — calculate net directly from base amounts
        base_gross = sum(tx.get("base_amount", 0) for tx in selected_txs if tx.get("base_amount"))
        rates = [tx.get("exchange_rate") for tx in selected_txs if tx.get("exchange_rate")]
        avg_rate = sum(rates) / len(rates) if rates else 1
        base_deductions = total_deductions / avg_rate if avg_rate else 0
        treasury_amount = base_gross - base_deductions
    else:
        treasury_amount = convert_currency(net_amount, "USD", dest_currency)

    now = datetime.now(timezone.utc)
    settlement_id = f"stl_{uuid.uuid4().hex[:12]}"
    
    # Use provided settlement_date or current time for treasury/settlement records
    if body.settlement_date:
        settle_date = f"{body.settlement_date}T00:00:00" if 'T' not in body.settlement_date else body.settlement_date
    else:
        settle_date = now.isoformat()

    # Create ONE compound settlement record
    settlement_doc = {
        "settlement_id": settlement_id,
        "psp_id": psp_id,
        "psp_name": psp["psp_name"],
        "settlement_type": "compound",
        "gross_amount": round(gross_amount, 2),
        "commission_amount": round(total_commission, 2),
        "reserve_fund_amount": round(total_reserve, 2),
        "extra_charges": round(total_extra, 2),
        "gateway_fees": round(total_gateway, 2),
        "total_deductions": round(total_deductions, 2),
        "net_amount": round(net_amount, 2),
        "treasury_amount": round(treasury_amount, 2),
        "treasury_currency": dest_currency,
        "transaction_count": len(selected_txs),
        "transaction_ids": [tx["transaction_id"] for tx in selected_txs],
        "settlement_destination_id": dest_account_id,
        "settlement_destination_name": dest["account_name"],
        "status": PSPSettlementStatus.PENDING,
        "expected_settlement_date": settle_date,
        "settlement_date": body.settlement_date or now.strftime("%Y-%m-%d"),
        "created_at": settle_date,
        "settled_at": None,
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    await db.psp_settlements.insert_one(settlement_doc)

    # Mark selected transactions as pending settlement (treasury credit deferred to approval)
    await db.transactions.update_many(
        {"transaction_id": {"$in": body.transaction_ids}},
        {"$set": {
            "settlement_id": settlement_id,
            "settlement_status": "pending",
        }}
    )

    await log_activity(request, user, "create", "psp", f"Compound settlement created (pending approval): {len(selected_txs)} transactions, net ${net_amount:,.2f}")

    result = await db.psp_settlements.find_one({"settlement_id": settlement_id}, {"_id": 0})
    return result



# Net Settlement: Settle ALL pending deposits + withdrawals as one net amount
class NetSettleRequest(BaseModel):
    destination_account_id: Optional[str] = None
    settlement_date: Optional[str] = None

@api_router.post("/psp/{psp_id}/net-settle")
async def net_settle_psp(
    request: Request,
    psp_id: str,
    body: NetSettleRequest,
    user: dict = Depends(require_permission(Modules.PSP, Actions.APPROVE))
):
    """Create a net settlement from ALL pending deposits and withdrawals"""
    psp = await db.psps.find_one({"psp_id": psp_id}, {"_id": 0})
    if not psp:
        raise HTTPException(status_code=404, detail="PSP not found")

    # Fetch all unsettled deposits
    deposit_txs = await db.transactions.find({
        "psp_id": psp_id,
        "destination_type": "psp",
        "transaction_type": "deposit",
        "status": TransactionStatus.APPROVED,
        "settled": {"$ne": True}
    }, {"_id": 0}).to_list(10000)

    # Fetch all unsettled approved withdrawals
    withdrawal_txs = await db.transactions.find({
        "psp_id": psp_id,
        "destination_type": "psp",
        "transaction_type": "withdrawal",
        "status": TransactionStatus.APPROVED,
        "settled": {"$ne": True}
    }, {"_id": 0}).to_list(10000)

    if not deposit_txs and not withdrawal_txs:
        raise HTTPException(status_code=400, detail="No pending transactions to settle")

    # Calculate deposit totals
    dep_gross = sum(tx.get("amount", 0) for tx in deposit_txs)
    dep_commission = sum(tx.get("psp_commission_amount", 0) or 0 for tx in deposit_txs)
    dep_reserve = sum((tx.get("psp_reserve_fund_amount", 0) or tx.get("psp_chargeback_amount", 0) or 0) for tx in deposit_txs)
    dep_extra_charges = sum(tx.get("psp_extra_charges", 0) or 0 for tx in deposit_txs)
    dep_extra_comm = sum(tx.get("psp_extra_commission", 0) or 0 for tx in deposit_txs)
    dep_gateway = sum(tx.get("psp_gateway_fee", 0) or 0 for tx in deposit_txs)
    dep_total_deductions = dep_commission + dep_reserve + dep_extra_charges + dep_extra_comm + dep_gateway
    dep_net = dep_gross - dep_total_deductions

    # Calculate withdrawal totals
    wdr_gross = sum(tx.get("amount", 0) for tx in withdrawal_txs)
    wdr_extra_comm = sum(tx.get("psp_withdrawal_extra_commission", 0) or 0 for tx in withdrawal_txs)
    wdr_total = wdr_gross + wdr_extra_comm

    # Net amount = deposit net - withdrawal total
    net_amount = round(dep_net - wdr_total, 2)

    if net_amount <= 0:
        raise HTTPException(status_code=400, detail=f"Net settlement amount is ${net_amount:,.2f}. Cannot settle zero or negative amounts.")

    # Determine destination treasury account
    dest_account_id = body.destination_account_id or psp.get("settlement_destination_id")
    if not dest_account_id:
        raise HTTPException(status_code=400, detail="No destination account specified")

    dest = await db.treasury_accounts.find_one({"account_id": dest_account_id}, {"_id": 0})
    if not dest:
        raise HTTPException(status_code=404, detail="Destination treasury account not found")

    dest_currency = dest.get("currency", "USD")
    treasury_amount = convert_currency(net_amount, "USD", dest_currency)

    now = datetime.now(timezone.utc)
    settlement_id = f"stl_{uuid.uuid4().hex[:12]}"

    if body.settlement_date:
        settle_date = f"{body.settlement_date}T00:00:00" if 'T' not in body.settlement_date else body.settlement_date
    else:
        settle_date = now.isoformat()

    all_tx_ids = [tx["transaction_id"] for tx in deposit_txs] + [tx["transaction_id"] for tx in withdrawal_txs]

    settlement_doc = {
        "settlement_id": settlement_id,
        "psp_id": psp_id,
        "psp_name": psp["psp_name"],
        "settlement_type": "net_settlement",
        "deposit_count": len(deposit_txs),
        "withdrawal_count": len(withdrawal_txs),
        "deposit_gross": round(dep_gross, 2),
        "deposit_deductions": round(dep_total_deductions, 2),
        "deposit_net": round(dep_net, 2),
        "withdrawal_gross": round(wdr_gross, 2),
        "withdrawal_extra_commission": round(wdr_extra_comm, 2),
        "withdrawal_total": round(wdr_total, 2),
        "gross_amount": round(dep_gross, 2),
        "commission_amount": round(dep_commission, 2),
        "reserve_fund_amount": round(dep_reserve, 2),
        "extra_charges": round(dep_extra_charges + dep_extra_comm, 2),
        "gateway_fees": round(dep_gateway, 2),
        "total_deductions": round(dep_total_deductions + wdr_total, 2),
        "net_amount": round(net_amount, 2),
        "treasury_amount": round(treasury_amount, 2),
        "treasury_currency": dest_currency,
        "transaction_count": len(all_tx_ids),
        "transaction_ids": all_tx_ids,
        "settlement_destination_id": dest_account_id,
        "settlement_destination_name": dest["account_name"],
        "status": PSPSettlementStatus.PENDING,
        "expected_settlement_date": settle_date,
        "settlement_date": body.settlement_date or now.strftime("%Y-%m-%d"),
        "created_at": settle_date,
        "settled_at": None,
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    await db.psp_settlements.insert_one(settlement_doc)

    # Mark ALL transactions as pending settlement (treasury credit deferred to approval)
    await db.transactions.update_many(
        {"transaction_id": {"$in": all_tx_ids}},
        {"$set": {
            "settlement_id": settlement_id,
            "settlement_status": "pending",
        }}
    )

    await log_activity(request, user, "create", "psp", f"Net settlement created (pending approval): {len(deposit_txs)} deposits + {len(withdrawal_txs)} withdrawals, net ${net_amount:,.2f}")

    result = await db.psp_settlements.find_one({"settlement_id": settlement_id}, {"_id": 0})
    return result



# Migration endpoint: Backfill existing settled PSP transactions into psp_settlements
@api_router.post("/psp/backfill-settlements")
async def backfill_psp_settlements(user: dict = Depends(require_permission(Modules.PSP, Actions.CREATE))):
    """Backfill historical settled PSP transactions into psp_settlements collection."""
    # Find all settled PSP transactions that don't have a settlement_id
    settled_txs = await db.transactions.find({
        "destination_type": "psp",
        "settled": True,
        "$or": [
            {"settlement_id": {"$exists": False}},
            {"settlement_id": None}
        ]
    }, {"_id": 0}).to_list(10000)
    
    if not settled_txs:
        return {"message": "No transactions to backfill", "count": 0}
    
    backfilled = 0
    now = datetime.now(timezone.utc)
    
    for tx in settled_txs:
        # Get PSP info
        psp = await db.psps.find_one({"psp_id": tx.get("psp_id")}, {"_id": 0})
        
        # Create settlement record
        settlement_id = f"stl_{uuid.uuid4().hex[:12]}"
        gross_amount = tx.get("amount", 0)
        commission_amount = tx.get("psp_commission_amount", 0)
        reserve_fund_amount = tx.get("psp_reserve_fund_amount", tx.get("psp_chargeback_amount", 0))
        extra_charges = tx.get("psp_extra_charges", 0)
        settle_amount = tx.get("psp_actual_amount_received", tx.get("psp_net_amount", gross_amount - commission_amount - reserve_fund_amount - extra_charges))
        
        settlement_doc = {
            "settlement_id": settlement_id,
            "psp_id": tx.get("psp_id"),
            "psp_name": tx.get("psp_name", psp.get("psp_name") if psp else "Unknown PSP"),
            "gross_amount": gross_amount,
            "commission_rate": psp.get("commission_rate", 0) if psp else 0,
            "commission_amount": commission_amount,
            "reserve_fund_rate": psp.get("reserve_fund_rate", psp.get("chargeback_rate", 0)) if psp else 0,
            "reserve_fund_amount": reserve_fund_amount,
            "chargeback_rate": psp.get("reserve_fund_rate", psp.get("chargeback_rate", 0)) if psp else 0,
            "chargeback_amount": reserve_fund_amount,
            "extra_charges": extra_charges,
            "gateway_fees": tx.get("psp_gateway_fee", 0),
            "total_deductions": commission_amount + reserve_fund_amount + extra_charges,
            "net_amount": settle_amount,
            "actual_amount_received": settle_amount,
            "variance": tx.get("psp_settlement_variance", 0),
            "holding_days": psp.get("holding_days", 0) if psp else 0,
            "transaction_count": 1,
            "transaction_ids": [tx.get("transaction_id")],
            "settlement_destination_id": tx.get("settlement_destination_id", psp.get("settlement_destination_id") if psp else None),
            "status": PSPSettlementStatus.COMPLETED,
            "expected_settlement_date": tx.get("psp_expected_settlement_date"),
            "created_at": tx.get("settled_at", now.isoformat()),
            "settled_at": tx.get("settled_at", now.isoformat()),
            "created_by": tx.get("settled_by", user["user_id"]),
            "created_by_name": tx.get("settled_by_name", user["name"]),
            "reference": tx.get("reference", tx.get("transaction_id"))
        }
        
        await db.psp_settlements.insert_one(settlement_doc)
        
        # Update the transaction with settlement_id
        await db.transactions.update_one(
            {"transaction_id": tx.get("transaction_id")},
            {"$set": {"settlement_id": settlement_id}}
        )
        
        backfilled += 1
    
    return {"message": f"Successfully backfilled {backfilled} settlements", "count": backfilled}


# Migration endpoint: Fix treasury transaction dates to match linked income/expense entries
@api_router.post("/treasury/sync-ie-dates")
async def sync_treasury_ie_dates(user: dict = Depends(require_permission(Modules.TREASURY, Actions.EDIT))):
    """Fix treasury transaction dates to match their linked income/expense entry dates."""
    # Find all treasury transactions with income_expense_id
    txs = await db.treasury_transactions.find({"income_expense_id": {"$exists": True}}, {"_id": 0}).to_list(10000)
    
    if not txs:
        return {"message": "No transactions to fix", "count": 0}
    
    fixed = 0
    for tx in txs:
        ie_id = tx.get("income_expense_id")
        if not ie_id:
            continue
            
        # Get the linked income/expense entry
        entry = await db.income_expenses.find_one({"entry_id": ie_id}, {"_id": 0})
        if not entry or not entry.get("date"):
            continue
        
        entry_date = entry.get("date")
        # The date format should be YYYY-MM-DD, convert to full ISO format
        new_created_at = f"{entry_date}T12:00:00+00:00"
        
        # Update treasury transaction created_at to match entry date
        await db.treasury_transactions.update_one(
            {"treasury_transaction_id": tx.get("treasury_transaction_id")},
            {"$set": {"created_at": new_created_at}}
        )
        fixed += 1
    
    return {"message": f"Successfully fixed {fixed} treasury transaction dates", "count": fixed}


# Migration endpoint: Assign role_ids to existing users
@api_router.post("/users/assign-role-ids")
async def assign_role_ids(user: dict = Depends(require_permission(Modules.USERS, Actions.EDIT))):
    """Assign role_ids to existing users based on their role field."""
    # Role mapping from old role names to new role_ids
    role_mapping = {
        "admin": "admin",
        "super_admin": "super_admin",
        "accountant": "accountant", 
        "sub_admin": "sub_admin",
        "vendor": "exchanger",  # vendors/exchangers get exchanger role
        "exchanger": "exchanger",
        "viewer": "viewer"
    }
    
    # Find users without role_id
    users = await db.users.find({"role_id": {"$exists": False}}, {"_id": 0}).to_list(1000)
    users.extend(await db.users.find({"role_id": None}, {"_id": 0}).to_list(1000))
    
    # Remove duplicates
    seen_ids = set()
    unique_users = []
    for u in users:
        if u["user_id"] not in seen_ids:
            seen_ids.add(u["user_id"])
            unique_users.append(u)
    
    updated = 0
    for u in unique_users:
        old_role = u.get("role", "viewer")
        new_role_id = role_mapping.get(old_role, "viewer")
        
        await db.users.update_one(
            {"user_id": u["user_id"]},
            {"$set": {"role_id": new_role_id}}
        )
        updated += 1
        logger.info(f"Assigned role_id '{new_role_id}' to user {u.get('email')}")
    
    return {"message": f"Assigned role_ids to {updated} users", "count": updated}


# ============== RESERVE FUND MANAGEMENT ==============

@api_router.get("/psps/{psp_id}/reserve-funds")
async def get_psp_reserve_funds(psp_id: str, user: dict = Depends(require_permission(Modules.PSP, Actions.VIEW))):
    """Get reserve fund ledger for a PSP — all transactions with reserve fund amounts."""
    psp = await db.psps.find_one({"psp_id": psp_id}, {"_id": 0})
    if not psp:
        raise HTTPException(status_code=404, detail="PSP not found")
    
    holding_days = psp.get("holding_days", 0)
    now = datetime.now(timezone.utc)
    
    # Get all PSP transactions that have reserve fund amounts
    txs = await db.transactions.find(
        {
            "psp_id": psp_id,
            "destination_type": "psp",
            "status": {"$in": [TransactionStatus.PENDING, TransactionStatus.APPROVED, TransactionStatus.COMPLETED]},
            "$or": [
                {"psp_reserve_fund_amount": {"$gt": 0}},
                {"psp_chargeback_amount": {"$gt": 0}},
                {"settled": True}
            ]
        },
        {"_id": 0}
    ).sort("created_at", -1).to_list(10000)
    
    reserve_fund_rate = psp.get("reserve_fund_rate", psp.get("chargeback_rate", 0)) / 100
    
    ledger = []
    total_held = 0
    total_released = 0
    due_this_week = 0
    
    for tx in txs:
        rf_amount = tx.get("psp_reserve_fund_amount", tx.get("psp_chargeback_amount", 0))
        if rf_amount <= 0:
            rf_amount = round(tx.get("amount", 0) * reserve_fund_rate, 2)
        if rf_amount <= 0:
            continue
        
        # Calculate dates
        created_str = tx.get("created_at", "")
        try:
            created_dt = datetime.fromisoformat(created_str.replace('Z', '+00:00'))
            if created_dt.tzinfo is None:
                created_dt = created_dt.replace(tzinfo=timezone.utc)
        except:
            created_dt = now
        
        release_date = created_dt + timedelta(days=holding_days) if holding_days > 0 else created_dt
        days_remaining = max(0, (release_date - now).days)
        is_released = tx.get("reserve_fund_released", False)
        
        # Determine status
        if is_released:
            status = "released"
            total_released += rf_amount
        elif release_date <= now:
            status = "due"
            total_held += rf_amount
            if (release_date - now).days >= -7:
                due_this_week += rf_amount
        else:
            status = "held"
            total_held += rf_amount
            if days_remaining <= 7:
                due_this_week += rf_amount
        
        ledger.append({
            "transaction_id": tx["transaction_id"],
            "reference": tx.get("reference", tx["transaction_id"]),
            "client_name": tx.get("client_name", ""),
            "amount": tx.get("amount", 0),
            "currency": tx.get("currency", "USD"),
            "reserve_fund_amount": rf_amount,
            "hold_date": created_dt.isoformat(),
            "release_date": release_date.isoformat(),
            "days_remaining": days_remaining,
            "status": status,
            "released_at": tx.get("reserve_fund_released_at"),
            "released_by_name": tx.get("reserve_fund_released_by_name"),
        })
    
    return {
        "ledger": ledger,
        "summary": {
            "total_held": round(total_held, 2),
            "total_released": round(total_released, 2),
            "due_this_week": round(due_this_week, 2),
            "holding_days": holding_days,
            "reserve_fund_rate": psp.get("reserve_fund_rate", psp.get("chargeback_rate", 0)),
            "total_entries": len(ledger),
        }
    }

@api_router.post("/psps/reserve-funds/{transaction_id}/release")
async def release_reserve_fund(request: Request, transaction_id: str, user: dict = Depends(require_permission(Modules.PSP, Actions.APPROVE))):

    """Mark a reserve fund as released back."""
    tx = await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    if tx.get("reserve_fund_released"):
        raise HTTPException(status_code=400, detail="Reserve fund already released")
    
    now = datetime.now(timezone.utc)
    rf_amount = tx.get("psp_reserve_fund_amount", tx.get("psp_chargeback_amount", 0))
    
    # Fall back to calculating from PSP rate if amount not stored (legacy transactions)
    if not rf_amount or rf_amount <= 0:
        psp = await db.psps.find_one({"psp_id": tx.get("psp_id")}, {"_id": 0})
        if psp:
            rf_rate = psp.get("reserve_fund_rate", psp.get("chargeback_rate", 0)) / 100
            rf_amount = round(tx.get("amount", 0) * rf_rate, 2)
    
    await db.transactions.update_one(
        {"transaction_id": transaction_id},
        {"$set": {
            "reserve_fund_released": True,
            "reserve_fund_released_at": now.isoformat(),
            "reserve_fund_released_by": user["user_id"],
            "reserve_fund_released_by_name": user["name"],
        }}
    )
    
    # Credit the reserve fund back to treasury
    psp = await db.psps.find_one({"psp_id": tx.get("psp_id")}, {"_id": 0})
    if psp and psp.get("settlement_destination_id") and rf_amount > 0:
        dest = await db.treasury_accounts.find_one({"account_id": psp["settlement_destination_id"]}, {"_id": 0})
        tx_currency = tx.get("currency", "USD")
        dest_currency = dest.get("currency", "USD") if dest else "USD"
        treasury_amount = convert_currency(rf_amount, tx_currency, dest_currency)
        
        await db.treasury_accounts.update_one(
            {"account_id": psp["settlement_destination_id"]},
            {"$inc": {"balance": treasury_amount}, "$set": {"updated_at": now.isoformat()}}
        )
        conversion_note = f" (Converted: {tx_currency} {rf_amount:,.2f} -> {dest_currency} {treasury_amount:,.2f})" if tx_currency != dest_currency else ""
        # Create treasury transaction record
        await db.treasury_transactions.insert_one({
            "treasury_transaction_id": f"ttx_{uuid.uuid4().hex[:12]}",
            "account_id": psp["settlement_destination_id"],
            "account_name": dest["account_name"] if dest else None,
            "transaction_type": "reserve_fund_release",
            "amount": treasury_amount,
            "currency": dest_currency,
            "original_amount": rf_amount,
            "original_currency": tx_currency,
            "description": f"Reserve fund release - {tx.get('reference', transaction_id)} from {psp.get('psp_name', '')}{conversion_note}",
            "reference": f"RF-{transaction_id}",
            "related_transaction_id": transaction_id,
            "psp_id": tx.get("psp_id"),
            "psp_name": psp.get("psp_name"),
            "created_at": now.isoformat(),
            "created_by": user["user_id"],
            "created_by_name": user["name"],
        })
    
    await log_activity(request, user, "approve", "psp", "Released reserve fund")

    return {"message": "Reserve fund released", "amount": rf_amount, "transaction_id": transaction_id}

@api_router.post("/psps/reserve-funds/bulk-release")
async def bulk_release_reserve_funds(request: Request, user: dict = Depends(require_permission(Modules.PSP, Actions.APPROVE))):
    """Bulk release multiple reserve funds."""
    data = await request.json()
    tx_ids = data.get("transaction_ids", [])
    if not tx_ids:
        raise HTTPException(status_code=400, detail="No transaction IDs provided")
    
    now = datetime.now(timezone.utc)
    released_count = 0
    total_released = 0
    
    for tx_id in tx_ids:
        tx = await db.transactions.find_one({"transaction_id": tx_id, "reserve_fund_released": {"$ne": True}}, {"_id": 0})
        if not tx:
            continue
        rf_amount = tx.get("psp_reserve_fund_amount", tx.get("psp_chargeback_amount", 0))
        # Fall back to calculating from PSP rate if amount not stored (legacy transactions)
        if not rf_amount or rf_amount <= 0:
            psp_for_rate = await db.psps.find_one({"psp_id": tx.get("psp_id")}, {"_id": 0})
            if psp_for_rate:
                rf_rate = psp_for_rate.get("reserve_fund_rate", psp_for_rate.get("chargeback_rate", 0)) / 100
                rf_amount = round(tx.get("amount", 0) * rf_rate, 2)
        if rf_amount <= 0:
            continue
        
        await db.transactions.update_one(
            {"transaction_id": tx_id},
            {"$set": {
                "reserve_fund_released": True,
                "reserve_fund_released_at": now.isoformat(),
                "reserve_fund_released_by": user["user_id"],
                "reserve_fund_released_by_name": user["name"],
            }}
        )
        
        psp = await db.psps.find_one({"psp_id": tx.get("psp_id")}, {"_id": 0})
        if psp and psp.get("settlement_destination_id"):
            dest = await db.treasury_accounts.find_one({"account_id": psp["settlement_destination_id"]}, {"_id": 0})
            tx_currency = tx.get("currency", "USD")
            dest_currency = dest.get("currency", "USD") if dest else "USD"
            treasury_amount = convert_currency(rf_amount, tx_currency, dest_currency)
            
            await db.treasury_accounts.update_one(
                {"account_id": psp["settlement_destination_id"]},
                {"$inc": {"balance": treasury_amount}, "$set": {"updated_at": now.isoformat()}}
            )
            conversion_note = f" (Converted: {tx_currency} {rf_amount:,.2f} -> {dest_currency} {treasury_amount:,.2f})" if tx_currency != dest_currency else ""
            await db.treasury_transactions.insert_one({
                "treasury_transaction_id": f"ttx_{uuid.uuid4().hex[:12]}",
                "account_id": psp["settlement_destination_id"],
                "account_name": dest["account_name"] if dest else None,
                "transaction_type": "reserve_fund_release",
                "amount": treasury_amount,
                "currency": dest_currency,
                "original_amount": rf_amount,
                "original_currency": tx_currency,
                "description": f"Reserve fund release - {tx.get('reference', tx_id)} from {psp.get('psp_name', '')}{conversion_note}",
                "reference": f"RF-{tx_id}",
                "related_transaction_id": tx_id,
                "psp_id": tx.get("psp_id"),
                "psp_name": psp.get("psp_name"),
                "created_at": now.isoformat(),
                "created_by": user["user_id"],
                "created_by_name": user["name"],
            })
        
        released_count += 1
        total_released += rf_amount
    
    return {"message": f"Released {released_count} reserve funds", "total_released": round(total_released, 2), "count": released_count}

@api_router.get("/psps/reserve-funds/global-summary")
async def get_global_reserve_fund_summary(user: dict = Depends(require_permission(Modules.PSP, Actions.VIEW))):
    """Get global reserve fund summary across all PSPs."""
    psps = await db.psps.find({"status": PSPStatus.ACTIVE}, {"_id": 0, "psp_id": 1, "psp_name": 1, "reserve_fund_rate": 1, "chargeback_rate": 1, "holding_days": 1}).to_list(1000)
    now = datetime.now(timezone.utc)
    
    total_held = 0
    total_released = 0
    due_for_release = 0
    
    for psp in psps:
        rf_rate = psp.get("reserve_fund_rate", psp.get("chargeback_rate", 0)) / 100
        holding = psp.get("holding_days", 0)
        
        txs = await db.transactions.find(
            {"psp_id": psp["psp_id"], "destination_type": "psp",
             "$or": [{"psp_reserve_fund_amount": {"$gt": 0}}, {"psp_chargeback_amount": {"$gt": 0}}]},
            {"_id": 0, "psp_reserve_fund_amount": 1, "psp_chargeback_amount": 1, "reserve_fund_released": 1, "created_at": 1, "amount": 1}
        ).to_list(10000)
        
        for tx in txs:
            rf = tx.get("psp_reserve_fund_amount", tx.get("psp_chargeback_amount", 0))
            if rf <= 0:
                rf = round(tx.get("amount", 0) * rf_rate, 2)
            if rf <= 0:
                continue
            
            if tx.get("reserve_fund_released"):
                total_released += rf
            else:
                total_held += rf
                try:
                    created = datetime.fromisoformat(tx.get("created_at", "").replace('Z', '+00:00'))
                    if created.tzinfo is None:
                        created = created.replace(tzinfo=timezone.utc)
                    release_dt = created + timedelta(days=holding)
                    if release_dt <= now:
                        due_for_release += rf
                except:
                    pass
    
    return {
        "total_held": round(total_held, 2),
        "total_released": round(total_released, 2),
        "due_for_release": round(due_for_release, 2),
    }

# ============== VENDOR ROUTES ==============

@api_router.get("/vendors")
async def get_vendors(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    search: Optional[str] = None,
    user: dict = Depends(require_permission(Modules.EXCHANGERS, Actions.VIEW))
):
    # Check cache first
    cache_key = get_cache_key("vendors:list", page=page, page_size=page_size, search=search)
    cached = get_cached(cache_key)
    if cached:
        return cached
    
    # Build query
    query = {}
    if search:
        query["vendor_name"] = {"$regex": search, "$options": "i"}
    
    # Get paginated vendors
    skip = (page - 1) * page_size
    total = await db.vendors.count_documents(query)
    vendors = await db.vendors.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    
    # Batch fetch treasury accounts to avoid N+1 queries
    treasury_ids = list(set(v.get("settlement_destination_id") for v in vendors if v.get("settlement_destination_id")))
    treasury_map = {}
    if treasury_ids:
        treasuries = await db.treasury_accounts.find({"account_id": {"$in": treasury_ids}}, {"_id": 0}).to_list(len(treasury_ids))
        treasury_map = {t["account_id"]: t for t in treasuries}
    
    # Batch fetch all pending transactions for all vendors
    vendor_ids = [v["vendor_id"] for v in vendors]
    pending_txs_all = await db.transactions.find({
        "vendor_id": {"$in": vendor_ids},
        "destination_type": "vendor",
        "status": {"$in": [TransactionStatus.APPROVED, TransactionStatus.COMPLETED]},
        "settled": {"$ne": True}
    }, {"_id": 0}).to_list(10000)
    
    # Batch fetch all completed income/expense entries for all vendors
    # IMPORTANT: Exclude converted_to_loan entries - they're tracked under Loans, not settlement
    ie_entries_all = await db.income_expenses.find({
        "vendor_id": {"$in": vendor_ids},
        "status": "completed",
        "converted_to_loan": {"$ne": True},  # Exclude converted entries to prevent double-counting
        "settled": {"$ne": True}
    }, {"_id": 0}).to_list(10000)
    
    # Batch fetch completed loan transactions involving vendors
    loan_txs_all = await db.loan_transactions.find({
        "$or": [
            {"source_vendor_id": {"$in": vendor_ids}},
            {"credit_to_vendor_id": {"$in": vendor_ids}}
        ],
        "status": "completed",
        "settled": {"$ne": True}
    }, {"_id": 0}).to_list(10000)
    
    # Group transactions and IE entries by vendor_id
    from collections import defaultdict
    pending_by_vendor = defaultdict(list)
    for tx in pending_txs_all:
        pending_by_vendor[tx["vendor_id"]].append(tx)
    
    ie_by_vendor = defaultdict(list)
    for ie in ie_entries_all:
        ie_by_vendor[ie["vendor_id"]].append(ie)
    
    # Group loan transactions by vendor
    loan_tx_by_vendor = defaultdict(list)
    for ltx in loan_txs_all:
        if ltx.get("source_vendor_id"):
            loan_tx_by_vendor[ltx["source_vendor_id"]].append({"type": "out", "tx": ltx})
        if ltx.get("credit_to_vendor_id"):
            loan_tx_by_vendor[ltx["credit_to_vendor_id"]].append({"type": "in", "tx": ltx})
    
    # Populate vendor data
    for vendor in vendors:
        if vendor.get("settlement_destination_id"):
            dest = treasury_map.get(vendor["settlement_destination_id"])
            vendor["settlement_destination_name"] = dest["account_name"] if dest else "Unknown"
            vendor["settlement_destination_bank"] = dest.get("bank_name") if dest else None
        
        pending_txs = pending_by_vendor.get(vendor["vendor_id"], [])
        vendor["pending_transactions_count"] = len(pending_txs)
        
        # Calculate net pending amount by currency
        currency_breakdown = {}
        
        def ensure_currency(currency):
            if currency not in currency_breakdown:
                currency_breakdown[currency] = {
                    "deposits_base": 0, "withdrawals_base": 0,
                    "deposits_usd": 0, "withdrawals_usd": 0,
                    "commission_base": 0, "commission_usd": 0
                }
        
        for tx in pending_txs:
            currency = tx.get("base_currency") or tx.get("currency", "USD")
            ensure_currency(currency)
            
            base_amount = tx.get("base_amount") or tx.get("amount", 0)
            usd_amount = tx.get("amount", 0)
            commission_base = tx.get("vendor_commission_base_amount", 0)
            commission_usd = tx.get("vendor_commission_amount", 0)
            
            if tx.get("transaction_type") == "deposit":
                currency_breakdown[currency]["deposits_base"] += base_amount
                currency_breakdown[currency]["deposits_usd"] += usd_amount
            else:
                currency_breakdown[currency]["withdrawals_base"] += base_amount
                currency_breakdown[currency]["withdrawals_usd"] += usd_amount
            
            currency_breakdown[currency]["commission_base"] += commission_base
            currency_breakdown[currency]["commission_usd"] += commission_usd
        
        # Include income/expense entries: income = Money In, expense = Money Out
        # Use base_currency/base_amount (payment currency) when available
        for ie in ie_by_vendor.get(vendor["vendor_id"], []):
            currency = ie.get("base_currency") or ie.get("currency", "USD")
            ensure_currency(currency)
            
            base_amount = ie.get("base_amount") or ie.get("amount", 0)
            usd_amount = ie.get("amount_usd") or ie.get("amount", 0)
            commission_base = ie.get("vendor_commission_base_amount", 0)
            commission_usd = ie.get("vendor_commission_amount", 0)
            
            if ie.get("entry_type") == "income":
                currency_breakdown[currency]["deposits_base"] += base_amount
                currency_breakdown[currency]["deposits_usd"] += usd_amount
            else:
                currency_breakdown[currency]["withdrawals_base"] += base_amount
                currency_breakdown[currency]["withdrawals_usd"] += usd_amount
            
            currency_breakdown[currency]["commission_base"] += commission_base
            currency_breakdown[currency]["commission_usd"] += commission_usd
        
        # Include loan transactions: repayments TO vendor = Money In, disbursements FROM vendor = Money Out
        for loan_entry in loan_tx_by_vendor.get(vendor["vendor_id"], []):
            ltx = loan_entry["tx"]
            currency = ltx.get("currency", "USD")
            ensure_currency(currency)
            
            amount = ltx.get("amount", 0)
            commission_base = ltx.get("vendor_commission_base_amount", 0)
            commission_amount = ltx.get("vendor_commission_amount", 0)
            
            if loan_entry["type"] == "in":  # Repayment TO vendor
                currency_breakdown[currency]["deposits_base"] += amount
                currency_breakdown[currency]["deposits_usd"] += amount
            else:  # Disbursement FROM vendor
                currency_breakdown[currency]["withdrawals_base"] += amount
                currency_breakdown[currency]["withdrawals_usd"] += amount
            
            # Add loan commission
            currency_breakdown[currency]["commission_base"] += commission_base
            currency_breakdown[currency]["commission_usd"] += commission_amount
        
        # Build settlement by currency for list view
        settlement_by_currency = []
        total_net_usd = 0
        
        # Get custom settled amounts for this vendor
        custom_settled_map = {}
        custom_stls = await db.vendor_settlements.aggregate([
            {"$match": {"vendor_id": vendor["vendor_id"], "settlement_mode": "custom", "status": VendorSettlementStatus.APPROVED}},
            {"$group": {"_id": "$source_currency", "total": {"$sum": "$gross_amount"}}}
        ]).to_list(50)
        for cs in custom_stls:
            custom_settled_map[cs["_id"]] = cs["total"]
        
        for currency, data in currency_breakdown.items():
            cs_amount = custom_settled_map.get(currency, 0)
            net_base = (data["deposits_base"] - data["withdrawals_base"]) - data["commission_base"] - cs_amount
            net_usd = (data["deposits_usd"] - data["withdrawals_usd"]) - data["commission_usd"] - cs_amount
            total_net_usd += net_usd
            settlement_by_currency.append({
                "currency": currency,
                "amount": net_base,
                "usd_equivalent": net_usd,
                "commission_base": data["commission_base"],
                "custom_settled": cs_amount,
            })
        
        vendor["settlement_by_currency"] = settlement_by_currency
        vendor["pending_amount"] = total_net_usd
    
    # Build response with pagination
    response = {
        "items": vendors,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size if total > 0 else 1
    }
    
    # Cache the response
    set_cached(cache_key, response, CACHE_TTL['vendors_list'])
    
    return response

@api_router.get("/vendors/{vendor_id}")
async def get_vendor(vendor_id: str, user: dict = Depends(require_permission(Modules.EXCHANGERS, Actions.VIEW))):
    vendor = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Calculate settlement balance by currency (unsettled approved/completed transactions)
    # Settlement = (Money In - Money Out - Commission)
    # Money In = deposits + income + loan repayments TO vendor
    # Money Out = withdrawals + expense + loan disbursements FROM vendor
    settlement_pipeline = [
        {"$match": {
            "vendor_id": vendor_id,
            "status": {"$in": ["approved", "completed"]},
            "settled": {"$ne": True}
        }},
        {"$group": {
            "_id": {"$ifNull": ["$base_currency", "$currency"]},
            "deposit_amount": {
                "$sum": {
                    "$cond": [
                        {"$eq": ["$transaction_type", "deposit"]},
                        {"$ifNull": ["$base_amount", "$amount"]},
                        0
                    ]
                }
            },
            "withdrawal_amount": {
                "$sum": {
                    "$cond": [
                        {"$eq": ["$transaction_type", "withdrawal"]},
                        {"$ifNull": ["$base_amount", "$amount"]},
                        0
                    ]
                }
            },
            "deposit_usd": {
                "$sum": {
                    "$cond": [
                        {"$eq": ["$transaction_type", "deposit"]},
                        "$amount",
                        0
                    ]
                }
            },
            "withdrawal_usd": {
                "$sum": {
                    "$cond": [
                        {"$eq": ["$transaction_type", "withdrawal"]},
                        "$amount",
                        0
                    ]
                }
            },
            "deposit_count": {
                "$sum": {"$cond": [{"$eq": ["$transaction_type", "deposit"]}, 1, 0]}
            },
            "withdrawal_count": {
                "$sum": {"$cond": [{"$eq": ["$transaction_type", "withdrawal"]}, 1, 0]}
            },
            "total_commission_usd": {"$sum": {"$ifNull": ["$vendor_commission_amount", 0]}},
            "total_commission_base": {"$sum": {"$ifNull": ["$vendor_commission_base_amount", 0]}}
        }}
    ]
    
    settlement_by_currency = await db.transactions.aggregate(settlement_pipeline).to_list(100)
    
    # Fetch loan transactions involving this vendor (disbursements FROM vendor and repayments TO vendor)
    # Disbursements from vendor = Money OUT (like withdrawals)
    # Repayments to vendor = Money IN (like deposits)
    loan_tx_pipeline = [
        {"$match": {
            "$or": [
                {"source_vendor_id": vendor_id},  # Disbursements from this vendor
                {"credit_to_vendor_id": vendor_id}  # Repayments to this vendor
            ],
            "status": "completed",
            "settled": {"$ne": True}
        }},
        {"$group": {
            "_id": "$currency",
            # Loan repayments TO vendor = Money IN
            "loan_in_amount": {
                "$sum": {
                    "$cond": [
                        {"$eq": ["$credit_to_vendor_id", vendor_id]},
                        "$amount",
                        0
                    ]
                }
            },
            # Loan disbursements FROM vendor = Money OUT
            "loan_out_amount": {
                "$sum": {
                    "$cond": [
                        {"$eq": ["$source_vendor_id", vendor_id]},
                        "$amount",
                        0
                    ]
                }
            },
            "loan_in_count": {
                "$sum": {"$cond": [{"$eq": ["$credit_to_vendor_id", vendor_id]}, 1, 0]}
            },
            "loan_out_count": {
                "$sum": {"$cond": [{"$eq": ["$source_vendor_id", vendor_id]}, 1, 0]}
            },
            # Loan commission (in transaction currency)
            "loan_commission_amount": {"$sum": {"$ifNull": ["$vendor_commission_amount", 0]}},
            "loan_commission_base": {"$sum": {"$ifNull": ["$vendor_commission_base_amount", 0]}}
        }}
    ]
    loan_tx_by_currency = await db.loan_transactions.aggregate(loan_tx_pipeline).to_list(100)
    loan_tx_map = {item["_id"] or "USD": item for item in loan_tx_by_currency}
    
    # Also fetch completed income/expense entries for this vendor
    # IMPORTANT: Exclude converted_to_loan entries - they're tracked under Loans, not as settlement
    # Group by base_currency (payment currency) - this is what the exchanger actually handles
    ie_pipeline = [
        {"$match": {
            "vendor_id": vendor_id,
            "status": "completed",
            "converted_to_loan": {"$ne": True},  # Exclude converted entries to prevent double-counting
            "settled": {"$ne": True}
        }},
        {"$group": {
            "_id": {"$ifNull": ["$base_currency", "$currency"]},  # Group by payment currency
            "income_base": {
                "$sum": {"$cond": [{"$eq": ["$entry_type", "income"]}, {"$ifNull": ["$base_amount", "$amount"]}, 0]}
            },
            "expense_base": {
                "$sum": {"$cond": [{"$eq": ["$entry_type", "expense"]}, {"$ifNull": ["$base_amount", "$amount"]}, 0]}
            },
            "income_usd": {
                "$sum": {"$cond": [{"$eq": ["$entry_type", "income"]}, {"$ifNull": ["$amount_usd", "$amount"]}, 0]}
            },
            "expense_usd": {
                "$sum": {"$cond": [{"$eq": ["$entry_type", "expense"]}, {"$ifNull": ["$amount_usd", "$amount"]}, 0]}
            },
            "income_count": {
                "$sum": {"$cond": [{"$eq": ["$entry_type", "income"]}, 1, 0]}
            },
            "expense_count": {
                "$sum": {"$cond": [{"$eq": ["$entry_type", "expense"]}, 1, 0]}
            },
            "ie_commission_usd": {"$sum": {"$ifNull": ["$vendor_commission_amount", 0]}},
            "ie_commission_base": {"$sum": {"$ifNull": ["$vendor_commission_base_amount", 0]}}
        }}
    ]
    ie_by_currency = await db.income_expenses.aggregate(ie_pipeline).to_list(100)
    ie_map = {item["_id"] or "USD": item for item in ie_by_currency}
    
    # Merge transactions and IE data into settlement_by_currency
    currency_data = {}
    for item in settlement_by_currency:
        curr = item["_id"] or "USD"
        currency_data[curr] = {
            "tx_deposit": item["deposit_amount"],
            "tx_withdrawal": item["withdrawal_amount"],
            "tx_deposit_usd": item["deposit_usd"],
            "tx_withdrawal_usd": item["withdrawal_usd"],
            "tx_deposit_count": item["deposit_count"],
            "tx_withdrawal_count": item["withdrawal_count"],
            "tx_commission_usd": item["total_commission_usd"],
            "tx_commission_base": item["total_commission_base"],
            "ie_in": 0, "ie_out": 0, "ie_in_usd": 0, "ie_out_usd": 0,
            "ie_in_count": 0, "ie_out_count": 0,
            "ie_commission_usd": 0, "ie_commission_base": 0,
            "loan_in": 0, "loan_out": 0, "loan_in_usd": 0, "loan_out_usd": 0,
            "loan_in_count": 0, "loan_out_count": 0,
            "loan_commission_usd": 0, "loan_commission_base": 0,
        }
    
    for curr, ie_item in ie_map.items():
        if curr not in currency_data:
            currency_data[curr] = {
                "tx_deposit": 0, "tx_withdrawal": 0,
                "tx_deposit_usd": 0, "tx_withdrawal_usd": 0,
                "tx_deposit_count": 0, "tx_withdrawal_count": 0,
                "tx_commission_usd": 0, "tx_commission_base": 0,
                "ie_in": 0, "ie_out": 0, "ie_in_usd": 0, "ie_out_usd": 0,
                "ie_in_count": 0, "ie_out_count": 0,
                "ie_commission_usd": 0, "ie_commission_base": 0,
                "loan_in": 0, "loan_out": 0, "loan_in_usd": 0, "loan_out_usd": 0,
                "loan_in_count": 0, "loan_out_count": 0,
                "loan_commission_usd": 0, "loan_commission_base": 0,
            }
        currency_data[curr]["ie_in"] += ie_item["income_base"]
        currency_data[curr]["ie_in_usd"] += ie_item["income_usd"]
        currency_data[curr]["ie_in_count"] += ie_item["income_count"]
        currency_data[curr]["ie_out"] += ie_item["expense_base"]
        currency_data[curr]["ie_out_usd"] += ie_item["expense_usd"]
        currency_data[curr]["ie_out_count"] += ie_item["expense_count"]
        currency_data[curr]["ie_commission_usd"] += ie_item["ie_commission_usd"]
        currency_data[curr]["ie_commission_base"] += ie_item["ie_commission_base"]
    
    for curr, loan_item in loan_tx_map.items():
        if curr not in currency_data:
            currency_data[curr] = {
                "tx_deposit": 0, "tx_withdrawal": 0,
                "tx_deposit_usd": 0, "tx_withdrawal_usd": 0,
                "tx_deposit_count": 0, "tx_withdrawal_count": 0,
                "tx_commission_usd": 0, "tx_commission_base": 0,
                "ie_in": 0, "ie_out": 0, "ie_in_usd": 0, "ie_out_usd": 0,
                "ie_in_count": 0, "ie_out_count": 0,
                "ie_commission_usd": 0, "ie_commission_base": 0,
                "loan_in": 0, "loan_out": 0, "loan_in_usd": 0, "loan_out_usd": 0,
                "loan_in_count": 0, "loan_out_count": 0,
                "loan_commission_usd": 0, "loan_commission_base": 0,
            }
        currency_data[curr]["loan_in"] += loan_item["loan_in_amount"]
        currency_data[curr]["loan_in_usd"] += loan_item["loan_in_amount"]
        currency_data[curr]["loan_in_count"] += loan_item["loan_in_count"]
        currency_data[curr]["loan_out"] += loan_item["loan_out_amount"]
        currency_data[curr]["loan_out_usd"] += loan_item["loan_out_amount"]
        currency_data[curr]["loan_out_count"] += loan_item["loan_out_count"]
        currency_data[curr]["loan_commission_usd"] += loan_item.get("loan_commission_amount", 0)
        currency_data[curr]["loan_commission_base"] += loan_item.get("loan_commission_base", 0)
    
    # Get approved custom settlements for this vendor (partial payments not linked to specific transactions)
    custom_settlements = await db.vendor_settlements.aggregate([
        {"$match": {
            "vendor_id": vendor_id,
            "settlement_mode": "custom",
            "status": VendorSettlementStatus.APPROVED
        }},
        {"$group": {
            "_id": "$source_currency",
            "total_settled": {"$sum": "$gross_amount"}
        }}
    ]).to_list(100)
    
    # Inject custom settled amounts into currency_data
    for cs in custom_settlements:
        curr = cs["_id"]
        if curr not in currency_data:
            currency_data[curr] = {
                "tx_deposit": 0, "tx_withdrawal": 0, "tx_deposit_usd": 0, "tx_withdrawal_usd": 0,
                "tx_deposit_count": 0, "tx_withdrawal_count": 0,
                "tx_commission_usd": 0, "tx_commission_base": 0,
                "ie_in": 0, "ie_out": 0, "ie_in_usd": 0, "ie_out_usd": 0,
                "ie_in_count": 0, "ie_out_count": 0,
                "ie_commission_usd": 0, "ie_commission_base": 0,
                "loan_in": 0, "loan_out": 0, "loan_in_usd": 0, "loan_out_usd": 0,
                "loan_in_count": 0, "loan_out_count": 0,
                "loan_commission_usd": 0, "loan_commission_base": 0,
            }
        currency_data[curr]["custom_settled"] = cs["total_settled"]
        currency_data[curr]["custom_settled_usd"] = cs["total_settled"]  # same if source_currency matches
    
    vendor["settlement_by_currency"] = [
        {
            "currency": curr,
            # Total calculations
            "total_in": d["tx_deposit"] + d["ie_in"] + d["loan_in"],
            "total_out": d["tx_withdrawal"] + d["ie_out"] + d["loan_out"],
            "total_commission_base": d["tx_commission_base"] + d["ie_commission_base"] + d["loan_commission_base"],
            "total_commission_usd": d["tx_commission_usd"] + d["ie_commission_usd"] + d["loan_commission_usd"],
            "custom_settled": d.get("custom_settled", 0),
            "amount": (d["tx_deposit"] + d["ie_in"] + d["loan_in"]) - (d["tx_withdrawal"] + d["ie_out"] + d["loan_out"]) - (d["tx_commission_base"] + d["ie_commission_base"] + d["loan_commission_base"]) - d.get("custom_settled", 0),
            "usd_equivalent": (d["tx_deposit_usd"] + d["ie_in_usd"] + d["loan_in_usd"]) - (d["tx_withdrawal_usd"] + d["ie_out_usd"] + d["loan_out_usd"]) - (d["tx_commission_usd"] + d["ie_commission_usd"] + d["loan_commission_usd"]) - d.get("custom_settled_usd", 0),
            # Breakdown
            "deposit_amount": d["tx_deposit"], "withdrawal_amount": d["tx_withdrawal"],
            "ie_in": d["ie_in"], "ie_out": d["ie_out"],
            "loan_in": d["loan_in"], "loan_out": d["loan_out"],
            "tx_commission_base": d["tx_commission_base"],
            "ie_commission_base": d["ie_commission_base"],
            "loan_commission_base": d["loan_commission_base"],
            "commission_earned_usd": d["tx_commission_usd"] + d["ie_commission_usd"] + d["loan_commission_usd"],
            "commission_earned_base": d["tx_commission_base"] + d["ie_commission_base"] + d["loan_commission_base"],
            "deposit_count": d["tx_deposit_count"],
            "withdrawal_count": d["tx_withdrawal_count"],
            "transaction_count": d["tx_deposit_count"] + d["tx_withdrawal_count"] + d["ie_in_count"] + d["ie_out_count"] + d["loan_in_count"] + d["loan_out_count"]
        }
        for curr, d in currency_data.items()
    ]
    
    return vendor

@api_router.post("/vendors")
async def create_vendor(vendor_data: VendorCreate, request: Request, user: dict = Depends(require_permission(Modules.EXCHANGERS, Actions.CREATE))):
    # Check if email already exists
    existing = await db.users.find_one({"email": vendor_data.email}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    vendor_id = f"vendor_{uuid.uuid4().hex[:12]}"
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    # Create user account for vendor
    user_doc = {
        "user_id": user_id,
        "email": vendor_data.email,
        "password_hash": hash_password(vendor_data.password),
        "name": vendor_data.vendor_name,
        "role": UserRole.VENDOR,
        "vendor_id": vendor_id,  # Link to vendor
        "picture": None,
        "is_active": True,
        "created_at": now.isoformat()
    }
    await db.users.insert_one(user_doc)
    
    # Create vendor record
    vendor_doc = {
        "vendor_id": vendor_id,
        "user_id": user_id,
        "vendor_name": vendor_data.vendor_name,
        "email": vendor_data.email,
        "deposit_commission": vendor_data.deposit_commission,
        "withdrawal_commission": vendor_data.withdrawal_commission,
        "description": vendor_data.description,
        "total_volume": 0.0,
        "total_commission": 0.0,
        "pending_settlement": 0.0,
        "status": VendorStatus.ACTIVE,
        "created_at": now.isoformat(),
        "updated_at": now.isoformat()
    }
    
    await db.vendors.insert_one(vendor_doc)
    
    # Log activity
    await log_activity(request, user, "create", "exchangers", f"Created exchanger: {vendor_data.vendor_name}", reference_id=vendor_id)
    
    return await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})

@api_router.put("/vendors/{vendor_id}")
async def update_vendor(request: Request, vendor_id: str, update_data: VendorUpdate, user: dict = Depends(require_permission(Modules.EXCHANGERS, Actions.EDIT))):

    updates = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No updates provided")
    
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    result = await db.vendors.update_one({"vendor_id": vendor_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Update vendor name in user record if changed
    if updates.get("vendor_name"):
        vendor = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
        if vendor:
            await db.users.update_one({"user_id": vendor["user_id"]}, {"$set": {"name": updates["vendor_name"]}})
    
    await log_activity(request, user, "edit", "exchangers", "Updated exchanger")

    return await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})

@api_router.delete("/vendors/{vendor_id}")
async def delete_vendor(request: Request, vendor_id: str, user: dict = Depends(require_permission(Modules.EXCHANGERS, Actions.DELETE))):

    vendor = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Delete vendor user account
    await db.users.delete_one({"user_id": vendor["user_id"]})
    
    # Delete vendor record
    await db.vendors.delete_one({"vendor_id": vendor_id})
    
    await log_activity(request, user, "delete", "exchangers", "Deleted exchanger")

    return {"message": "Vendor deleted"}

# Get vendor's assigned transactions (for vendor portal)
@api_router.get("/vendors/{vendor_id}/transactions")
async def get_vendor_transactions(
    vendor_id: str, 
    status: Optional[str] = None,
    user: dict = Depends(require_permission(Modules.EXCHANGERS, Actions.VIEW))
):
    # Vendors can only see their own transactions
    if user.get("role") == UserRole.VENDOR:
        user_vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
        if not user_vendor or user_vendor["vendor_id"] != vendor_id:
            raise HTTPException(status_code=403, detail="Access denied")
    
    query = {"vendor_id": vendor_id, "destination_type": "vendor"}
    if status:
        query["status"] = status
    
    transactions = await db.transactions.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return transactions

# Get current vendor info (for logged in vendor)
@api_router.get("/vendor/me")
async def get_my_vendor_info(user: dict = Depends(require_vendor)):
    vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Get pending transactions
    pending_txs = await db.transactions.find({
        "vendor_id": vendor["vendor_id"],
        "destination_type": "vendor",
        "status": TransactionStatus.PENDING
    }, {"_id": 0}).to_list(1000)
    
    vendor["pending_transactions"] = pending_txs
    vendor["pending_count"] = len(pending_txs)
    
    # Calculate settlement balance by currency (unsettled approved/completed transactions)
    # Settlement = (Money In - Money Out - Commission)
    # Money In = deposits + income + loan repayments TO vendor
    # Money Out = withdrawals + expense + loan disbursements FROM vendor
    settlement_pipeline = [
        {"$match": {
            "vendor_id": vendor["vendor_id"],
            "status": {"$in": ["approved", "completed"]},
            "settled": {"$ne": True}
        }},
        {"$group": {
            "_id": {"$ifNull": ["$base_currency", "$currency"]},
            "deposit_amount": {
                "$sum": {
                    "$cond": [
                        {"$eq": ["$transaction_type", "deposit"]},
                        {"$ifNull": ["$base_amount", "$amount"]},
                        0
                    ]
                }
            },
            "withdrawal_amount": {
                "$sum": {
                    "$cond": [
                        {"$eq": ["$transaction_type", "withdrawal"]},
                        {"$ifNull": ["$base_amount", "$amount"]},
                        0
                    ]
                }
            },
            "deposit_usd": {
                "$sum": {
                    "$cond": [
                        {"$eq": ["$transaction_type", "deposit"]},
                        "$amount",
                        0
                    ]
                }
            },
            "withdrawal_usd": {
                "$sum": {
                    "$cond": [
                        {"$eq": ["$transaction_type", "withdrawal"]},
                        "$amount",
                        0
                    ]
                }
            },
            "total_commission_usd": {"$sum": {"$ifNull": ["$vendor_commission_amount", 0]}},
            "total_commission_base": {"$sum": {"$ifNull": ["$vendor_commission_base_amount", 0]}},
            "deposit_count": {
                "$sum": {"$cond": [{"$eq": ["$transaction_type", "deposit"]}, 1, 0]}
            },
            "withdrawal_count": {
                "$sum": {"$cond": [{"$eq": ["$transaction_type", "withdrawal"]}, 1, 0]}
            }
        }}
    ]
    
    settlement_by_currency = await db.transactions.aggregate(settlement_pipeline).to_list(100)
    
    # Fetch loan transactions involving this vendor
    loan_tx_pipeline = [
        {"$match": {
            "$or": [
                {"source_vendor_id": vendor["vendor_id"]},
                {"credit_to_vendor_id": vendor["vendor_id"]}
            ],
            "status": "completed",
            "settled": {"$ne": True}
        }},
        {"$group": {
            "_id": "$currency",
            "loan_in_amount": {
                "$sum": {
                    "$cond": [
                        {"$eq": ["$credit_to_vendor_id", vendor["vendor_id"]]},
                        "$amount",
                        0
                    ]
                }
            },
            "loan_out_amount": {
                "$sum": {
                    "$cond": [
                        {"$eq": ["$source_vendor_id", vendor["vendor_id"]]},
                        "$amount",
                        0
                    ]
                }
            },
            "loan_in_count": {
                "$sum": {"$cond": [{"$eq": ["$credit_to_vendor_id", vendor["vendor_id"]]}, 1, 0]}
            },
            "loan_out_count": {
                "$sum": {"$cond": [{"$eq": ["$source_vendor_id", vendor["vendor_id"]]}, 1, 0]}
            },
            # Loan commission (in transaction currency)
            "loan_commission_amount": {"$sum": {"$ifNull": ["$vendor_commission_amount", 0]}},
            "loan_commission_base": {"$sum": {"$ifNull": ["$vendor_commission_base_amount", 0]}}
        }}
    ]
    loan_tx_by_currency = await db.loan_transactions.aggregate(loan_tx_pipeline).to_list(100)
    loan_tx_map = {item["_id"] or "USD": item for item in loan_tx_by_currency}
    
    # Also fetch approved/completed income/expense entries for this vendor
    # IMPORTANT: Exclude "converted_to_loan" status - when an expense is converted to loan,
    # it's a reclassification, NOT a cash settlement event. The loan module tracks it instead.
    # Group by base_currency (payment currency) - this is what the exchanger actually handles
    ie_pipeline = [
        {"$match": {
            "vendor_id": vendor["vendor_id"],
            "status": {"$in": ["approved", "completed"]},
            "converted_to_loan": {"$ne": True},  # Exclude converted entries to prevent double-counting
            "settled": {"$ne": True}
        }},
        {"$group": {
            "_id": {"$ifNull": ["$base_currency", "$currency"]},  # Group by payment currency
            "income_base": {
                "$sum": {"$cond": [{"$eq": ["$entry_type", "income"]}, {"$ifNull": ["$base_amount", "$amount"]}, 0]}
            },
            "expense_base": {
                "$sum": {"$cond": [{"$eq": ["$entry_type", "expense"]}, {"$ifNull": ["$base_amount", "$amount"]}, 0]}
            },
            "income_usd": {
                "$sum": {"$cond": [{"$eq": ["$entry_type", "income"]}, {"$ifNull": ["$amount_usd", "$amount"]}, 0]}
            },
            "expense_usd": {
                "$sum": {"$cond": [{"$eq": ["$entry_type", "expense"]}, {"$ifNull": ["$amount_usd", "$amount"]}, 0]}
            },
            "income_count": {
                "$sum": {"$cond": [{"$eq": ["$entry_type", "income"]}, 1, 0]}
            },
            "expense_count": {
                "$sum": {"$cond": [{"$eq": ["$entry_type", "expense"]}, 1, 0]}
            },
            "ie_commission_usd": {"$sum": {"$ifNull": ["$vendor_commission_amount", 0]}},
            "ie_commission_base": {"$sum": {"$ifNull": ["$vendor_commission_base_amount", 0]}}
        }}
    ]
    ie_by_currency = await db.income_expenses.aggregate(ie_pipeline).to_list(100)
    ie_map = {item["_id"] or "USD": item for item in ie_by_currency}
    
    # Merge transactions and IE data
    currency_data = {}
    for item in settlement_by_currency:
        curr = item["_id"] or "USD"
        currency_data[curr] = {
            "deposit_amount": item["deposit_amount"],
            "withdrawal_amount": item["withdrawal_amount"],
            "deposit_usd": item["deposit_usd"],
            "withdrawal_usd": item["withdrawal_usd"],
            "deposit_count": item["deposit_count"],
            "withdrawal_count": item["withdrawal_count"],
            "commission_usd": item["total_commission_usd"],
            "commission_base": item["total_commission_base"],
        }
    
    for curr, ie_item in ie_map.items():
        if curr not in currency_data:
            currency_data[curr] = {
                "deposit_amount": 0, "withdrawal_amount": 0,
                "deposit_usd": 0, "withdrawal_usd": 0,
                "deposit_count": 0, "withdrawal_count": 0,
                "commission_usd": 0, "commission_base": 0,
            }
        currency_data[curr]["deposit_amount"] += ie_item["income_base"]
        currency_data[curr]["deposit_usd"] += ie_item["income_usd"]
        currency_data[curr]["deposit_count"] += ie_item["income_count"]
        currency_data[curr]["withdrawal_amount"] += ie_item["expense_base"]
        currency_data[curr]["withdrawal_usd"] += ie_item["expense_usd"]
        currency_data[curr]["withdrawal_count"] += ie_item["expense_count"]
        currency_data[curr]["commission_usd"] += ie_item["ie_commission_usd"]
        currency_data[curr]["commission_base"] += ie_item["ie_commission_base"]
    
    # Add loan transactions to settlement
    for curr, loan_item in loan_tx_map.items():
        if curr not in currency_data:
            currency_data[curr] = {
                "deposit_amount": 0, "withdrawal_amount": 0,
                "deposit_usd": 0, "withdrawal_usd": 0,
                "deposit_count": 0, "withdrawal_count": 0,
                "commission_usd": 0, "commission_base": 0,
            }
        currency_data[curr]["deposit_amount"] += loan_item["loan_in_amount"]
        currency_data[curr]["deposit_usd"] += loan_item["loan_in_amount"]
        currency_data[curr]["deposit_count"] += loan_item["loan_in_count"]
        currency_data[curr]["withdrawal_amount"] += loan_item["loan_out_amount"]
        currency_data[curr]["withdrawal_usd"] += loan_item["loan_out_amount"]
        currency_data[curr]["withdrawal_count"] += loan_item["loan_out_count"]
        # Add loan commission to total commission
        currency_data[curr]["commission_usd"] += loan_item.get("loan_commission_amount", 0)
        currency_data[curr]["commission_base"] += loan_item.get("loan_commission_base", 0)
    
    # Get approved custom settlements for this vendor
    custom_settled_portal = await db.vendor_settlements.aggregate([
        {"$match": {"vendor_id": vendor["vendor_id"], "settlement_mode": "custom", "status": VendorSettlementStatus.APPROVED}},
        {"$group": {"_id": "$source_currency", "total": {"$sum": "$gross_amount"}}}
    ]).to_list(50)
    custom_map_portal = {cs["_id"]: cs["total"] for cs in custom_settled_portal}
    
    vendor["settlement_by_currency"] = [
        {
            "currency": curr,
            "amount": (d["deposit_amount"] - d["withdrawal_amount"]) - d["commission_base"] - custom_map_portal.get(curr, 0),
            "usd_equivalent": (d["deposit_usd"] - d["withdrawal_usd"]) - d["commission_usd"] - custom_map_portal.get(curr, 0),
            "custom_settled": custom_map_portal.get(curr, 0),
            "deposit_amount": d["deposit_amount"],
            "withdrawal_amount": d["withdrawal_amount"],
            "commission_earned_usd": d["commission_usd"],
            "commission_earned_base": d["commission_base"],
            "deposit_count": d["deposit_count"],
            "withdrawal_count": d["withdrawal_count"],
            "transaction_count": d["deposit_count"] + d["withdrawal_count"]
        }
        for curr, d in currency_data.items()
    ]
    
    return vendor


@api_router.get("/vendor/settlements")
async def get_vendor_my_settlements(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    user: dict = Depends(require_vendor)
):
    """Get settlements for the logged-in vendor"""
    vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    return await paginate_query(db.vendor_settlements, {"vendor_id": vendor["vendor_id"]}, page, page_size)



# Vendor: Get all assigned transactions with filters
@api_router.get("/vendor/transactions")
async def get_vendor_all_transactions(
    status: Optional[str] = None,
    transaction_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    user: dict = Depends(require_vendor),
):
    vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")

    query = {"vendor_id": vendor["vendor_id"], "destination_type": "vendor"}
    if status and status != "all":
        query["status"] = status
    if transaction_type and transaction_type != "all":
        query["transaction_type"] = transaction_type
    if date_from:
        query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"

    return await paginate_query(db.transactions, query, page, page_size)


# Vendor: Export transactions to Excel
@api_router.get("/vendor/transactions/export/excel")
async def vendor_export_excel(
    status: Optional[str] = None,
    transaction_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(require_vendor),
):
    from fastapi.responses import StreamingResponse
    import openpyxl
    from io import BytesIO

    vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")

    query = {"vendor_id": vendor["vendor_id"], "destination_type": "vendor"}
    if status and status != "all":
        query["status"] = status
    if transaction_type and transaction_type != "all":
        query["transaction_type"] = transaction_type
    if date_from:
        query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"

    txs = await db.transactions.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["Reference", "Type", "Client", "Amount", "Currency", "Commission (USD)", "Mode", "Status", "Date"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="0B3D91", end_color="0B3D91", fill_type="solid")

    for i, tx in enumerate(txs, 2):
        ws.cell(row=i, column=1, value=tx.get("reference", ""))
        ws.cell(row=i, column=2, value=tx.get("transaction_type", ""))
        ws.cell(row=i, column=3, value=tx.get("client_name", ""))
        ws.cell(row=i, column=4, value=tx.get("base_amount") or tx.get("amount", 0))
        ws.cell(row=i, column=5, value=tx.get("base_currency") or tx.get("currency", "USD"))
        ws.cell(row=i, column=6, value=tx.get("vendor_commission_amount", 0) or 0)
        ws.cell(row=i, column=7, value=tx.get("transaction_mode", "bank"))
        ws.cell(row=i, column=8, value=tx.get("status", ""))
        ws.cell(row=i, column=9, value=tx.get("created_at", "")[:10] if tx.get("created_at") else "")

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=transactions_{vendor['vendor_name'].replace(' ', '_')}.xlsx"}
    )


# Vendor: Export transactions to PDF
@api_router.get("/vendor/transactions/export/pdf")
async def vendor_export_pdf(
    status: Optional[str] = None,
    transaction_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: dict = Depends(require_vendor),
):
    from fastapi.responses import StreamingResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from io import BytesIO

    vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")

    query = {"vendor_id": vendor["vendor_id"], "destination_type": "vendor"}
    if status and status != "all":
        query["status"] = status
    if transaction_type and transaction_type != "all":
        query["transaction_type"] = transaction_type
    if date_from:
        query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"

    txs = await db.transactions.find(query, {"_id": 0}).sort("created_at", -1).to_list(5000)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, spaceAfter=6)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=9, textColor=colors.grey, spaceAfter=14)

    elements.append(Paragraph(f"Transaction Report — {vendor['vendor_name']}", title_style))
    filter_parts = []
    if status and status != "all":
        filter_parts.append(f"Status: {status}")
    if transaction_type and transaction_type != "all":
        filter_parts.append(f"Type: {transaction_type}")
    if date_from:
        filter_parts.append(f"From: {date_from}")
    if date_to:
        filter_parts.append(f"To: {date_to}")
    elements.append(Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}  |  {' | '.join(filter_parts) if filter_parts else 'All transactions'}", subtitle_style))

    data = [["Reference", "Type", "Client", "Amount", "Currency", "Commission", "Mode", "Status", "Date"]]
    for tx in txs:
        data.append([
            tx.get("reference", ""),
            tx.get("transaction_type", "").capitalize(),
            tx.get("client_name", ""),
            f"{tx.get('base_amount') or tx.get('amount', 0):,.2f}",
            tx.get("base_currency") or tx.get("currency", "USD"),
            f"${tx.get('vendor_commission_amount', 0) or 0:,.2f}",
            tx.get("transaction_mode", "bank").capitalize(),
            tx.get("status", "").capitalize(),
            tx.get("created_at", "")[:10] if tx.get("created_at") else "",
        ])

    col_widths = [1.1*inch, 0.75*inch, 1.3*inch, 0.9*inch, 0.7*inch, 0.85*inch, 0.65*inch, 0.75*inch, 0.85*inch]
    table = RLTable(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0B3D91')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (3, 0), (5, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    # Summary
    total_deposits = sum(1 for t in txs if t.get("transaction_type") == "deposit")
    total_withdrawals = sum(1 for t in txs if t.get("transaction_type") == "withdrawal")
    total_commission = sum(t.get("vendor_commission_amount", 0) or 0 for t in txs)
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"Total: {len(txs)} transactions  |  Deposits: {total_deposits}  |  Withdrawals: {total_withdrawals}  |  Total Commission: ${total_commission:,.2f}", subtitle_style))

    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=transactions_{vendor['vendor_name'].replace(' ', '_')}.pdf"}
    )


# Vendor approve transaction
@api_router.post("/vendor/transactions/{transaction_id}/approve")
async def vendor_approve_transaction(request: Request, transaction_id: str, user: dict = Depends(require_permission(Modules.TRANSACTIONS, Actions.APPROVE))):

    vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    tx = await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    if tx.get("vendor_id") != vendor["vendor_id"]:
        raise HTTPException(status_code=403, detail="Transaction does not belong to this vendor")
    
    if tx["status"] != TransactionStatus.PENDING:
        raise HTTPException(status_code=400, detail="Transaction is not pending")
    
    # Only withdrawals require proof upload before approving
    # Deposits can be approved without proof
    if tx["transaction_type"] == TransactionType.WITHDRAWAL and not tx.get("vendor_proof_image"):
        raise HTTPException(status_code=400, detail="Please upload proof screenshot before approving withdrawal")
    
    now = datetime.now(timezone.utc)
    
    # Calculate commission based on transaction type and mode
    tx_mode = tx.get("transaction_mode", "bank")
    commission_rate = 0.0
    if tx["transaction_type"] == TransactionType.DEPOSIT:
        if tx_mode == "cash":
            commission_rate = vendor.get("deposit_commission_cash", vendor.get("deposit_commission", 0)) / 100
        else:
            commission_rate = vendor.get("deposit_commission", 0) / 100
    elif tx["transaction_type"] == TransactionType.WITHDRAWAL:
        if tx_mode == "cash":
            commission_rate = vendor.get("withdrawal_commission_cash", vendor.get("withdrawal_commission", 0)) / 100
        else:
            commission_rate = vendor.get("withdrawal_commission", 0) / 100
    
    # Calculate commission in BASE/PAYMENT currency (original currency)
    base_amount = tx.get("base_amount") or tx["amount"]
    base_currency = tx.get("base_currency") or tx.get("currency", "USD")
    commission_amount_base = round(base_amount * commission_rate, 2)
    # Calculate USD commission separately
    commission_amount_usd = round(tx["amount"] * commission_rate, 2)
    
    updates = {
        "status": TransactionStatus.APPROVED,
        "processed_by": user["user_id"],
        "processed_by_name": user["name"],
        "processed_at": now.isoformat(),
        "vendor_commission_rate": commission_rate * 100,  # Store as percentage
        "vendor_commission_amount": commission_amount_usd,  # USD commission amount
        "vendor_commission_base_amount": commission_amount_base,  # Base currency commission amount
        "vendor_commission_base_currency": base_currency  # Base currency code
    }
    
    await db.transactions.update_one({"transaction_id": transaction_id}, {"$set": updates})
    
    # Update vendor's total commission and volume
    await db.vendors.update_one(
        {"vendor_id": vendor["vendor_id"]},
        {
            "$inc": {
                "total_commission": commission_amount_usd,
                "total_volume": tx["amount"]
            },
            "$set": {"updated_at": now.isoformat()}
        }
    )
    
    await log_activity(request, user, "approve", "transactions", "Vendor approved transaction")

    return await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})

# Vendor reject transaction
@api_router.post("/vendor/transactions/{transaction_id}/reject")
async def vendor_reject_transaction(

    request: Request,

    transaction_id: str, 
    reason: str = "",
    user: dict = Depends(require_permission(Modules.TRANSACTIONS, Actions.APPROVE))
):
    vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    tx = await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    if tx.get("vendor_id") != vendor["vendor_id"]:
        raise HTTPException(status_code=403, detail="Transaction does not belong to this vendor")
    
    if tx["status"] != TransactionStatus.PENDING:
        raise HTTPException(status_code=400, detail="Transaction is not pending")
    
    now = datetime.now(timezone.utc)
    
    updates = {
        "status": TransactionStatus.REJECTED,
        "rejection_reason": reason,
        "processed_by": user["user_id"],
        "processed_by_name": user["name"],
        "processed_at": now.isoformat()
    }
    
    await db.transactions.update_one({"transaction_id": transaction_id}, {"$set": updates})
    
    await log_activity(request, user, "reject", "transactions", "Vendor rejected transaction")

    return await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})

# Vendor complete withdrawal with screenshot upload
@api_router.post("/vendor/transactions/{transaction_id}/upload-proof")
async def vendor_upload_proof(

    request: Request,

    transaction_id: str,
    proof_image: UploadFile = File(...),
    user: dict = Depends(require_vendor)
):
    """Vendor uploads proof of payment for withdrawal transactions"""
    vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    tx = await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    if tx.get("vendor_id") != vendor["vendor_id"]:
        raise HTTPException(status_code=403, detail="Transaction does not belong to this vendor")
    
    content = await proof_image.read()
    proof_image_url = upload_to_r2(content, proof_image.filename or "proof.png", proof_image.content_type or "image/png", "proofs")
    
    now = datetime.now(timezone.utc)
    await db.transactions.update_one(
        {"transaction_id": transaction_id},
        {"$set": {
            "vendor_proof_image": proof_image_url,
            "vendor_proof_uploaded_at": now.isoformat(),
            "vendor_proof_uploaded_by": user["user_id"],
            "vendor_proof_uploaded_by_name": user["name"]
        }}
    )
    
    await log_activity(request, user, "edit", "transactions", "Vendor uploaded proof")

    return {"message": "Proof uploaded successfully", "transaction_id": transaction_id}

# Vendor complete withdrawal with screenshot upload
@api_router.post("/vendor/transactions/{transaction_id}/complete")
async def vendor_complete_withdrawal(

    request: Request,

    transaction_id: str,
    proof_image: UploadFile = File(...),
    user: dict = Depends(require_vendor)
):
    vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    tx = await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    if tx.get("vendor_id") != vendor["vendor_id"]:
        raise HTTPException(status_code=403, detail="Transaction does not belong to this vendor")
    
    if tx.get("transaction_type") != TransactionType.WITHDRAWAL:
        raise HTTPException(status_code=400, detail="Only withdrawals can be completed with proof")
    
    if tx["status"] not in [TransactionStatus.PENDING, TransactionStatus.APPROVED]:
        raise HTTPException(status_code=400, detail="Transaction cannot be completed in current status")
    
    now = datetime.now(timezone.utc)
    
    # Handle proof image upload
    content = await proof_image.read()
    proof_image_url = upload_to_r2(content, proof_image.filename or "proof.png", proof_image.content_type or "image/png", "proofs")
    
    updates = {
        "status": TransactionStatus.COMPLETED,
        "vendor_proof_image": proof_image_url,
        "processed_by": user["user_id"],
        "processed_by_name": user["name"],
        "processed_at": now.isoformat()
    }
    
    await db.transactions.update_one({"transaction_id": transaction_id}, {"$set": updates})
    
    await log_activity(request, user, "edit", "transactions", "Vendor completed withdrawal")

    return await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})

# Vendor Loan Transactions - Get pending loan transactions for exchanger
@api_router.get("/vendor/loan-transactions")
async def get_vendor_loan_transactions(request: Request,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    user: dict = Depends(require_vendor)
):
    """Get loan transactions assigned to this vendor (disbursements from or repayments to)"""
    vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    query = {
        "$or": [
            {"source_vendor_id": vendor["vendor_id"]},
            {"credit_to_vendor_id": vendor["vendor_id"]}
        ]
    }
    skip = (page - 1) * page_size
    total = await db.loan_transactions.count_documents(query)
    total_pages = (total + page_size - 1) // page_size if total > 0 else 1
    
    transactions = await db.loan_transactions.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    
    # Enrich with loan details
    loan_ids = list(set([tx.get("loan_id") for tx in transactions if tx.get("loan_id")]))
    loan_map = {}
    if loan_ids:
        loans = await db.loans.find({"loan_id": {"$in": loan_ids}}, {"_id": 0}).to_list(500)
        loan_map = {loan["loan_id"]: loan for loan in loans}
    
    for tx in transactions:
        loan = loan_map.get(tx.get("loan_id"))
        if loan:
            tx["bank_details"] = loan.get("bank_details")
            tx["borrower_name"] = loan.get("borrower_name")
    
    return {"items": transactions, "total": total, "page": page, "page_size": page_size, "total_pages": total_pages}

# Vendor approve loan transaction with proof upload
@api_router.post("/vendor/loan-transactions/{transaction_id}/approve")
async def vendor_approve_loan_transaction(
    request: Request,
    transaction_id: str,
    proof_image: UploadFile = File(...),
    user: dict = Depends(require_vendor)
):
    """Vendor approves loan transaction with proof screenshot"""
    vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    tx = await db.loan_transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Loan transaction not found")
    
    # Verify this vendor is related to the transaction
    if tx.get("source_vendor_id") != vendor["vendor_id"] and tx.get("credit_to_vendor_id") != vendor["vendor_id"]:
        raise HTTPException(status_code=403, detail="Transaction does not belong to this vendor")
    
    if tx.get("status") != "pending_vendor":
        raise HTTPException(status_code=400, detail="Transaction is not pending vendor approval")
    
    now = datetime.now(timezone.utc)
    
    # Handle proof image upload
    content = await proof_image.read()
    proof_image_url = upload_to_r2(content, proof_image.filename or "proof.png", proof_image.content_type or "image/png", "proofs")
    
    updates = {
        "status": "completed",
        "vendor_proof_image": proof_image_url,
        "approved_by": user["user_id"],
        "approved_by_name": user["name"],
        "approved_at": now.isoformat()
    }
    
    await db.loan_transactions.update_one({"transaction_id": transaction_id}, {"$set": updates})
    
    await log_activity(request, user, "approve", "loans", "Vendor approved loan transaction")

    return await db.loan_transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})

# Vendor reject loan transaction
@api_router.post("/vendor/loan-transactions/{transaction_id}/reject")
async def vendor_reject_loan_transaction(
    request: Request,
    transaction_id: str,
    reason: str = "",
    user: dict = Depends(require_vendor)
):
    """Vendor rejects loan transaction"""
    vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    tx = await db.loan_transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Loan transaction not found")
    
    # Verify this vendor is related to the transaction
    if tx.get("source_vendor_id") != vendor["vendor_id"] and tx.get("credit_to_vendor_id") != vendor["vendor_id"]:
        raise HTTPException(status_code=403, detail="Transaction does not belong to this vendor")
    
    if tx.get("status") != "pending_vendor":
        raise HTTPException(status_code=400, detail="Transaction is not pending vendor approval")
    
    now = datetime.now(timezone.utc)
    
    updates = {
        "status": "rejected",
        "rejection_reason": reason,
        "rejected_by": user["user_id"],
        "rejected_by_name": user["name"],
        "rejected_at": now.isoformat()
    }
    
    await db.loan_transactions.update_one({"transaction_id": transaction_id}, {"$set": updates})
    
    await log_activity(request, user, "reject", "loans", "Vendor rejected loan transaction")

    return await db.loan_transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})

# Vendor settlements
@api_router.get("/vendors/{vendor_id}/settlements")
async def get_vendor_settlements(vendor_id: str, user: dict = Depends(require_permission(Modules.EXCHANGERS, Actions.VIEW))):
    settlements = await db.vendor_settlements.find({"vendor_id": vendor_id}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return settlements

@api_router.get("/settlements/{settlement_id}/statement")
async def get_settlement_statement(settlement_id: str, user: dict = Depends(require_permission(Modules.EXCHANGERS, Actions.VIEW))):
    """Get full settlement statement with underlying transactions, I&E entries, and loan transactions."""
    settlement = await db.vendor_settlements.find_one({"settlement_id": settlement_id}, {"_id": 0})
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement not found")
    
    # Fetch regular transactions
    tx_ids = settlement.get("transaction_ids", [])
    transactions = []
    if tx_ids:
        transactions = await db.transactions.find(
            {"transaction_id": {"$in": tx_ids}},
            {"_id": 0, "transaction_id": 1, "transaction_type": 1, "amount": 1, "currency": 1,
             "base_amount": 1, "base_currency": 1, "client_name": 1, "reference": 1,
             "created_at": 1, "status": 1}
        ).to_list(1000)
    
    # Fetch I&E entries
    ie_ids = settlement.get("ie_entry_ids", [])
    ie_entries = []
    if ie_ids:
        ie_entries = await db.income_expenses.find(
            {"entry_id": {"$in": ie_ids}},
            {"_id": 0, "entry_id": 1, "entry_type": 1, "amount": 1, "currency": 1,
             "base_amount": 1, "base_currency": 1, "category": 1, "description": 1,
             "vendor_commission_base_amount": 1, "created_at": 1, "date": 1}
        ).to_list(1000)
    
    # Fetch Loan transactions
    loan_ids = settlement.get("loan_tx_ids", [])
    loan_entries = []
    if loan_ids:
        loan_txs = await db.loan_transactions.find(
            {"transaction_id": {"$in": loan_ids}},
            {"_id": 0, "transaction_id": 1, "transaction_type": 1, "amount": 1, "currency": 1,
             "borrower_name": 1, "vendor_commission_base_amount": 1, "created_at": 1,
             "source_vendor_id": 1, "credit_to_vendor_id": 1}
        ).to_list(1000)
        loan_entries = loan_txs
    
    vendor = await db.vendors.find_one({"vendor_id": settlement.get("vendor_id")}, {"_id": 0, "vendor_name": 1, "contact_person": 1, "email": 1, "phone": 1})
    return {
        "settlement": settlement,
        "transactions": transactions,
        "ie_entries": ie_entries,
        "loan_entries": loan_entries,
        "vendor": vendor or {},
    }

# Admin settle vendor balance
class VendorSettlementRequest(BaseModel):
    settlement_type: str  # "bank" or "cash"
    destination_account_id: Optional[str] = None  # Treasury account (optional for direct transfers)
    commission_amount: float = 0  # Manual commission entry
    charges_amount: float = 0  # Additional charges/fees
    charges_description: Optional[str] = None
    # Multi-currency support
    source_currency: str = "USD"  # Currency of transactions
    destination_currency: str = "USD"  # Currency of destination treasury
    exchange_rate: float = 1.0  # Conversion rate
    settlement_amount_in_dest_currency: Optional[float] = None  # Final amount in destination currency
    # Custom/partial amount settlement
    settlement_mode: str = "full"  # "full" = settle all, "custom" = partial amount
    custom_amount: Optional[float] = None  # Amount for custom settlement
    custom_currency: Optional[str] = None  # Currency of custom amount
    is_direct_transfer: bool = False  # True = just record (no treasury link)
    notes: Optional[str] = None

@api_router.post("/vendors/{vendor_id}/settle")
async def settle_vendor_balance(

    request: Request,

    vendor_id: str, 
    settlement_request: VendorSettlementRequest,
    user: dict = Depends(require_permission(Modules.EXCHANGERS, Actions.CREATE))
):
    vendor = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    is_custom = settlement_request.settlement_mode == "custom"
    is_direct = settlement_request.is_direct_transfer
    
    # Validate: non-direct transfers require a destination account
    if not is_direct and not settlement_request.destination_account_id:
        raise HTTPException(status_code=400, detail="Please select a settlement destination account")
    
    # Validate: custom mode requires amount and currency
    if is_custom:
        if not settlement_request.custom_amount or settlement_request.custom_amount <= 0:
            raise HTTPException(status_code=400, detail="Please enter a valid custom settlement amount")
        if not settlement_request.custom_currency:
            raise HTTPException(status_code=400, detail="Please select a currency for the custom settlement")
    
    pending_txs = []
    pending_ie = []
    pending_loans = []
    ie_entry_ids = []
    loan_tx_ids = []
    gross_amount = 0
    source_currency = settlement_request.source_currency
    
    if is_custom:
        # Custom/partial: use the provided amount, no transactions marked
        gross_amount = settlement_request.custom_amount
        source_currency = settlement_request.custom_currency
    else:
        # Full: get all approved transactions for this vendor that haven't been settled
        pending_txs = await db.transactions.find({
            "vendor_id": vendor_id,
            "destination_type": "vendor",
            "status": {"$in": [TransactionStatus.APPROVED, TransactionStatus.COMPLETED]},
            "settled": {"$ne": True}
        }, {"_id": 0}).to_list(1000)
        
        # Also get approved IE entries for this vendor that haven't been settled
        pending_ie = await db.income_expenses.find({
            "vendor_id": vendor_id,
            "status": {"$in": ["completed", "approved"]},
            "settled": {"$ne": True},
            "converted_to_loan": {"$ne": True},
        }, {"_id": 0}).to_list(1000)
        
        # Also get completed loan transactions for this vendor that haven't been settled
        pending_loans = await db.loan_transactions.find({
            "$or": [
                {"source_vendor_id": vendor_id},
                {"credit_to_vendor_id": vendor_id}
            ],
            "status": "completed",
            "settled": {"$ne": True}
        }, {"_id": 0}).to_list(1000)
        
        if not pending_txs and not pending_ie and not pending_loans:
            raise HTTPException(status_code=400, detail="No pending transactions to settle")
        
        # Calculate NET amounts
        for tx in pending_txs:
            if tx.get("base_currency") == source_currency and tx.get("base_amount"):
                tx_amount = tx["base_amount"]
            elif tx.get("currency") == source_currency:
                tx_amount = tx["amount"]
            else:
                tx_amount = tx["amount"]
            if tx.get("transaction_type") == "deposit":
                gross_amount += tx_amount
            elif tx.get("transaction_type") == "withdrawal":
                gross_amount -= tx_amount
            else:
                gross_amount += tx_amount
        
        for ie in pending_ie:
            if ie.get("base_currency") == source_currency and ie.get("base_amount"):
                ie_amount = ie["base_amount"]
            elif ie.get("currency") == source_currency:
                ie_amount = ie["amount"]
            else:
                ie_amount = ie.get("amount_usd", ie["amount"])
            if ie["entry_type"] == "income":
                gross_amount += ie_amount
            else:
                gross_amount -= ie_amount
            ie_entry_ids.append(ie["entry_id"])
        
        for ltx in pending_loans:
            ltx_amount = ltx.get("amount", 0)
            if ltx.get("credit_to_vendor_id") == vendor_id:
                gross_amount += ltx_amount
            elif ltx.get("source_vendor_id") == vendor_id:
                gross_amount -= ltx_amount
            loan_tx_ids.append(ltx["transaction_id"])
    
    # Manual commission and charges
    commission_amount = settlement_request.commission_amount
    charges_amount = settlement_request.charges_amount
    
    # Net amount before currency conversion
    net_amount_source = gross_amount - commission_amount - charges_amount
    
    # Calculate settlement amount in destination currency
    if settlement_request.settlement_amount_in_dest_currency is not None:
        # Admin entered final amount directly
        settlement_amount = settlement_request.settlement_amount_in_dest_currency
    else:
        # Use exchange rate
        settlement_amount = round(net_amount_source * settlement_request.exchange_rate, 2)
    
    settlement_id = f"vstl_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    dest = None
    dest_account_id = settlement_request.destination_account_id
    if dest_account_id and not is_direct:
        dest = await db.treasury_accounts.find_one({"account_id": dest_account_id}, {"_id": 0})
        if not dest:
            raise HTTPException(status_code=404, detail="Settlement destination account not found")
    
    settlement_doc = {
        "settlement_id": settlement_id,
        "vendor_id": vendor_id,
        "vendor_name": vendor["vendor_name"],
        "settlement_type": settlement_request.settlement_type,
        "settlement_mode": settlement_request.settlement_mode,
        "is_direct_transfer": is_direct,
        "gross_amount": gross_amount,
        "source_currency": source_currency,
        "commission_amount": commission_amount,
        "charges_amount": charges_amount,
        "charges_description": settlement_request.charges_description,
        "net_amount_source": net_amount_source,
        "exchange_rate": settlement_request.exchange_rate,
        "destination_currency": settlement_request.destination_currency if dest else source_currency,
        "settlement_amount": settlement_amount,
        "transaction_count": len(pending_txs) + len(pending_ie) + len(pending_loans),
        "transaction_ids": [tx["transaction_id"] for tx in pending_txs],
        "ie_entry_ids": ie_entry_ids,
        "loan_tx_ids": loan_tx_ids,
        "settlement_destination_id": dest_account_id if not is_direct else None,
        "settlement_destination_name": dest["account_name"] if dest else ("Direct Transfer" if is_direct else None),
        "notes": settlement_request.notes,
        "status": VendorSettlementStatus.PENDING,
        "created_at": now.isoformat(),
        "settled_at": None,
        "approved_at": None,
        "approved_by": None,
        "approved_by_name": None,
        "rejection_reason": None,
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    
    await db.vendor_settlements.insert_one(settlement_doc)
    
    # Only mark specific transactions as pending settlement for full settlements
    if not is_custom:
        await db.transactions.update_many(
            {"transaction_id": {"$in": [tx["transaction_id"] for tx in pending_txs]}},
            {"$set": {"settlement_id": settlement_id, "settlement_status": "pending_approval"}}
        )
        
        if ie_entry_ids:
            await db.income_expenses.update_many(
                {"entry_id": {"$in": ie_entry_ids}},
                {"$set": {"settlement_id": settlement_id, "settlement_status": "pending_approval"}}
            )
        
        if loan_tx_ids:
            await db.loan_transactions.update_many(
                {"transaction_id": {"$in": loan_tx_ids}},
                {"$set": {"settlement_id": settlement_id, "settlement_status": "pending_approval"}}
            )
    
    mode_label = "Custom partial" if is_custom else "Full"
    dest_label = "Direct Transfer" if is_direct else (dest["account_name"] if dest else "N/A")
    await log_activity(request, user, "create", "exchangers", f"{mode_label} settlement for {vendor['vendor_name']}: {settlement_amount:,.2f} {settlement_request.destination_currency if dest else source_currency} → {dest_label}")

    # Send settlement approval notification email (fire and forget)
    import asyncio
    asyncio.create_task(send_approval_notification("settlement", {
        "settlement_id": settlement_id,
        "vendor_name": vendor["vendor_name"],
        "gross_amount": gross_amount,
        "currency": source_currency,
        "net_amount": settlement_amount,
        "dest_currency": settlement_request.destination_currency if dest else source_currency,
        "tx_count": len(pending_txs) + len(pending_ie) + len(pending_loans),
        "created_by": user["name"],
        "settlement_mode": settlement_request.settlement_mode,
        "is_direct_transfer": is_direct,
    }))

    # Notify the exchanger about settlement
    asyncio.create_task(send_exchanger_notification("settlement", vendor_id, {
        "settlement_id": settlement_id,
        "gross_display": f"{gross_amount:,.2f} {source_currency}",
        "net_display": f"{settlement_amount:,.2f} {(settlement_request.destination_currency if dest else source_currency)}",
        "tx_count": len(pending_txs) + len(pending_ie) + len(pending_loans),
    }))

    return await db.vendor_settlements.find_one({"settlement_id": settlement_id}, {"_id": 0})

# Get all pending settlements (for approval page)
@api_router.get("/settlements/pending")
async def get_pending_settlements(user: dict = Depends(require_permission(Modules.EXCHANGERS, Actions.VIEW))):
    """Get all pending vendor settlements awaiting approval"""
    settlements = await db.vendor_settlements.find(
        {"status": VendorSettlementStatus.PENDING}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    return settlements

# Approve vendor settlement
@api_router.post("/settlements/{settlement_id}/approve")
async def approve_settlement(request: Request, settlement_id: str, user: dict = Depends(require_permission(Modules.EXCHANGERS, Actions.APPROVE))):

    """Approve a pending vendor settlement"""
    settlement = await db.vendor_settlements.find_one({"settlement_id": settlement_id}, {"_id": 0})
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement not found")
    
    if settlement["status"] != VendorSettlementStatus.PENDING:
        raise HTTPException(status_code=400, detail="Settlement is not pending")
    
    now = datetime.now(timezone.utc)
    
    # Update settlement status
    await db.vendor_settlements.update_one(
        {"settlement_id": settlement_id},
        {"$set": {
            "status": VendorSettlementStatus.APPROVED,
            "approved_at": now.isoformat(),
            "approved_by": user["user_id"],
            "approved_by_name": user["name"],
            "settled_at": now.isoformat()
        }}
    )
    
    # Mark transactions as fully settled
    await db.transactions.update_many(
        {"settlement_id": settlement_id},
        {"$set": {"settled": True, "settlement_status": "completed"}}
    )
    
    # Mark IE entries as fully settled
    ie_entry_ids = settlement.get("ie_entry_ids", [])
    if ie_entry_ids:
        await db.income_expenses.update_many(
            {"entry_id": {"$in": ie_entry_ids}},
            {"$set": {"settled": True, "settlement_status": "completed"}}
        )
    
    # Mark Loan transactions as fully settled
    loan_tx_ids = settlement.get("loan_tx_ids", [])
    if loan_tx_ids:
        await db.loan_transactions.update_many(
            {"transaction_id": {"$in": loan_tx_ids}},
            {"$set": {"settled": True, "settlement_status": "completed"}}
        )
    
    # Update treasury balance with settlement amount
    await db.treasury_accounts.update_one(
        {"account_id": settlement["settlement_destination_id"]},
        {"$inc": {"balance": settlement["settlement_amount"]}, "$set": {"updated_at": now.isoformat()}}
    )
    
    # Update vendor stats
    await db.vendors.update_one(
        {"vendor_id": settlement["vendor_id"]},
        {
            "$inc": {
                "total_volume": settlement["gross_amount"],
                "total_commission": settlement["commission_amount"]
            }
        }
    )
    
    # Record treasury transaction for history
    treasury_tx_id = f"ttx_{uuid.uuid4().hex[:12]}"
    treasury_tx_doc = {
        "treasury_transaction_id": treasury_tx_id,
        "account_id": settlement["settlement_destination_id"],
        "transaction_type": "settlement_in",
        "amount": settlement["settlement_amount"],
        "currency": settlement["destination_currency"],
        "reference": f"Vendor Settlement: {settlement['vendor_name']}",
        "settlement_id": settlement_id,
        "vendor_id": settlement["vendor_id"],
        "created_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    await db.treasury_transactions.insert_one(treasury_tx_doc)
    
    await log_activity(request, user, "approve", "exchangers", "Approved settlement")

    return await db.vendor_settlements.find_one({"settlement_id": settlement_id}, {"_id": 0})

# Reject vendor settlement
@api_router.post("/settlements/{settlement_id}/reject")
async def reject_settlement(request: Request, settlement_id: str, reason: str = "", user: dict = Depends(require_permission(Modules.EXCHANGERS, Actions.APPROVE))):

    """Reject a pending vendor settlement"""
    settlement = await db.vendor_settlements.find_one({"settlement_id": settlement_id}, {"_id": 0})
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement not found")
    
    if settlement["status"] != VendorSettlementStatus.PENDING:
        raise HTTPException(status_code=400, detail="Settlement is not pending")
    
    now = datetime.now(timezone.utc)
    
    # Update settlement status
    await db.vendor_settlements.update_one(
        {"settlement_id": settlement_id},
        {"$set": {
            "status": VendorSettlementStatus.REJECTED,
            "rejection_reason": reason,
            "approved_at": now.isoformat(),
            "approved_by": user["user_id"],
            "approved_by_name": user["name"]
        }}
    )
    
    # Reset transaction settlement status so they can be settled again
    await db.transactions.update_many(
        {"settlement_id": settlement_id},
        {"$set": {"settlement_id": None, "settlement_status": None}}
    )
    
    # Reset IE entries
    ie_entry_ids = settlement.get("ie_entry_ids", [])
    if ie_entry_ids:
        await db.income_expenses.update_many(
            {"entry_id": {"$in": ie_entry_ids}},
            {"$set": {"settlement_id": None, "settlement_status": None}}
        )
    
    # Reset Loan transactions
    loan_tx_ids = settlement.get("loan_tx_ids", [])
    if loan_tx_ids:
        await db.loan_transactions.update_many(
            {"transaction_id": {"$in": loan_tx_ids}},
            {"$set": {"settlement_id": None, "settlement_status": None}}
        )
    
    await log_activity(request, user, "reject", "exchangers", "Rejected settlement")

    return await db.vendor_settlements.find_one({"settlement_id": settlement_id}, {"_id": 0})

# ============== TRANSACTIONS ROUTES ==============

@api_router.get("/transactions")
async def get_transactions(
    user: dict = Depends(require_permission(Modules.TRANSACTIONS, Actions.VIEW)),
    client_id: Optional[str] = None,
    transaction_type: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    client_tag: Optional[str] = None,
    page: int = 1,
    page_size: int = 25,
    limit: int = 100
):
    """Get transactions with pagination and filtering"""
    # Use page_size if provided, otherwise fall back to limit for backwards compatibility
    actual_limit = min(page_size, limit, 100)  # Cap at 100 for performance
    skip = (page - 1) * actual_limit
    
    # Build query
    query = {}
    if client_id:
        query["client_id"] = client_id
    if transaction_type:
        query["transaction_type"] = transaction_type
    if status:
        query["status"] = status
    if date_from:
        query["created_at"] = {"$gte": date_from}
    if date_to:
        if "created_at" in query:
            query["created_at"]["$lte"] = date_to + "T23:59:59"
        else:
            query["created_at"] = {"$lte": date_to + "T23:59:59"}
    if search:
        query["$or"] = [
            {"reference": {"$regex": search, "$options": "i"}},
            {"client_name": {"$regex": search, "$options": "i"}},
            {"transaction_id": {"$regex": search, "$options": "i"}}
        ]
    if client_tag:
        query["client_tags"] = client_tag
    
    # Try cache first
    cache_key = get_cache_key("transactions:list", page=page, page_size=actual_limit, 
                              client_id=client_id, transaction_type=transaction_type, status=status,
                              search=search, date_from=date_from, date_to=date_to, client_tag=client_tag)
    cached = get_cached(cache_key)
    if cached:
        return cached
    
    # Get total count for pagination
    total = await db.transactions.count_documents(query)
    total_pages = (total + actual_limit - 1) // actual_limit
    
    transactions = await db.transactions.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(actual_limit).to_list(actual_limit)
    
    # Enrich transactions with client email and destination names
    client_ids = list(set(tx.get("client_id") for tx in transactions if tx.get("client_id")))
    clients_map = {}
    if client_ids:
        clients = await db.clients.find({"client_id": {"$in": client_ids}}, {"_id": 0, "client_id": 1, "email": 1}).to_list(len(client_ids))
        clients_map = {c["client_id"]: c.get("email", "") for c in clients}
    
    # Enrich with destination names for transactions missing them
    psp_ids = list(set(tx.get("psp_id") for tx in transactions if tx.get("psp_id") and not tx.get("psp_name")))
    treasury_ids = list(set(tx.get("destination_account_id") for tx in transactions if tx.get("destination_account_id") and not tx.get("destination_account_name") and tx.get("destination_type") in ["treasury", "usdt"]))
    vendor_ids = list(set(tx.get("vendor_id") for tx in transactions if tx.get("vendor_id") and not tx.get("vendor_name")))
    
    psps_map = {}
    if psp_ids:
        psps_list = await db.psps.find({"psp_id": {"$in": psp_ids}}, {"_id": 0, "psp_id": 1, "psp_name": 1}).to_list(len(psp_ids))
        psps_map = {p["psp_id"]: p["psp_name"] for p in psps_list}
    treasury_map = {}
    if treasury_ids:
        treasury_list = await db.treasury_accounts.find({"account_id": {"$in": treasury_ids}}, {"_id": 0, "account_id": 1, "account_name": 1}).to_list(len(treasury_ids))
        treasury_map = {t["account_id"]: t["account_name"] for t in treasury_list}
    vendors_map = {}
    if vendor_ids:
        vendors_list = await db.vendors.find({"vendor_id": {"$in": vendor_ids}}, {"_id": 0, "vendor_id": 1, "vendor_name": 1}).to_list(len(vendor_ids))
        vendors_map = {v["vendor_id"]: v["vendor_name"] for v in vendors_list}
    
    for tx in transactions:
        tx["client_email"] = clients_map.get(tx.get("client_id"), "")
        if tx.get("psp_id") and not tx.get("psp_name"):
            tx["psp_name"] = psps_map.get(tx["psp_id"])
        if tx.get("destination_account_id") and not tx.get("destination_account_name") and tx.get("destination_type") in ["treasury", "usdt"]:
            tx["destination_account_name"] = treasury_map.get(tx["destination_account_id"])
        if tx.get("vendor_id") and not tx.get("vendor_name"):
            tx["vendor_name"] = vendors_map.get(tx["vendor_id"])
    
    result = {
        "items": transactions,
        "total": total,
        "page": page,
        "page_size": actual_limit,
        "total_pages": total_pages
    }
    
    # Cache the result
    set_cached(cache_key, result, CACHE_TTL.get('transactions', 30))
    
    return result

@api_router.get("/transactions/pending")
async def get_pending_transactions(user: dict = Depends(require_permission(Modules.TRANSACTIONS, Actions.VIEW))):
    """Get all pending transactions for accountant approval"""
    transactions = await db.transactions.find(
        {"status": TransactionStatus.PENDING}, 
        {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    return transactions

# ============================================================================
# UNIFIED PENDING APPROVALS ENDPOINTS
# ============================================================================

@api_router.get("/pending-approvals/all")
async def get_all_pending_approvals(user: dict = Depends(require_permission(Modules.TRANSACTIONS, Actions.VIEW))):
    """Get all pending items across all modules for the Accountant Dashboard"""
    # Pending Income/Expenses
    pending_ie = await db.income_expenses.find(
        {"status": "pending"},
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    
    # Pending Loans (pending_approval)
    pending_loans = await db.loans.find(
        {"status": "pending_approval"},
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    
    # Pending Loan Repayments
    pending_repayments = await db.loan_repayments.find(
        {"status": "pending_approval"},
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    # Enrich with loan details
    for rep in pending_repayments:
        loan = await db.loans.find_one({"loan_id": rep["loan_id"]}, {"_id": 0, "borrower_name": 1, "amount": 1, "currency": 1})
        if loan:
            rep["borrower_name"] = loan.get("borrower_name", "Unknown")
    
    # Pending PSP Settlements
    pending_psp_settlements = await db.psp_settlements.find(
        {"status": PSPSettlementStatus.PENDING},
        {"_id": 0}
    ).sort("created_at", -1).to_list(500)
    # Enrich with destination info
    for stl in pending_psp_settlements:
        dest = await db.treasury_accounts.find_one(
            {"account_id": stl.get("settlement_destination_id")},
            {"_id": 0, "account_name": 1, "bank_name": 1, "currency": 1}
        )
        if dest:
            stl["settlement_destination_name"] = dest.get("account_name", "Unknown")
            stl["settlement_destination_bank"] = dest.get("bank_name")
            stl["settlement_destination_currency"] = dest.get("currency", "USD")
    
    return {
        "income_expenses": pending_ie,
        "loans": pending_loans,
        "loan_repayments": pending_repayments,
        "psp_settlements": pending_psp_settlements,
        "counts": {
            "income_expenses": len(pending_ie),
            "loans": len(pending_loans),
            "loan_repayments": len(pending_repayments),
            "psp_settlements": len(pending_psp_settlements),
        }
    }


# ---- Income/Expense Approve/Reject ----

@api_router.post("/income-expenses/{entry_id}/approve")
async def approve_income_expense(request: Request, entry_id: str, user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.APPROVE))):
    """Approve a pending income/expense entry and execute treasury operations"""
    entry = await db.income_expenses.find_one({"entry_id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    if entry.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Entry is not pending approval")
    
    now = datetime.now(timezone.utc)
    
    # Get treasury account
    treasury = None
    treasury_account_id = entry.get("treasury_account_id")
    if treasury_account_id:
        treasury = await db.treasury_accounts.find_one({"account_id": treasury_account_id}, {"_id": 0})
    
    # Execute treasury operations
    if treasury:
        if entry["entry_type"] == "income":
            await db.treasury_accounts.update_one(
                {"account_id": treasury_account_id},
                {"$inc": {"balance": entry["amount"]}, "$set": {"updated_at": now.isoformat()}}
            )
            ttx_amount = entry["amount"]
        else:
            await db.treasury_accounts.update_one(
                {"account_id": treasury_account_id},
                {"$inc": {"balance": -entry["amount"]}, "$set": {"updated_at": now.isoformat()}}
            )
            ttx_amount = -entry["amount"]
        
        # Record treasury transaction
        tx_id = f"ttx_{uuid.uuid4().hex[:12]}"
        await db.treasury_transactions.insert_one({
            "treasury_transaction_id": tx_id,
            "account_id": treasury_account_id,
            "transaction_type": entry["entry_type"],
            "amount": ttx_amount,
            "currency": entry.get("currency", "USD"),
            "reference": f"{entry['entry_type'].capitalize()}: {entry.get('description') or entry.get('category', 'N/A')}",
            "income_expense_id": entry_id,
            "created_at": now.isoformat(),
            "created_by": user["user_id"],
            "created_by_name": user["name"]
        })
    
    # Update status to approved
    await db.income_expenses.update_one(
        {"entry_id": entry_id},
        {"$set": {
            "status": "approved",
            "approved_at": now.isoformat(),
            "approved_by": user["user_id"],
            "approved_by_name": user["name"]
        }}
    )
    
    await log_activity(request, user, "approve", "income_expenses", f"Approved {entry['entry_type']}: {entry.get('description', 'N/A')}", reference_id=entry_id)
    return {"message": "Income/Expense entry approved successfully"}


@api_router.post("/income-expenses/{entry_id}/reject")
async def reject_income_expense(request: Request, entry_id: str, reason: str = "", user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.APPROVE))):
    """Reject a pending income/expense entry"""
    entry = await db.income_expenses.find_one({"entry_id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    if entry.get("status") != "pending":
        raise HTTPException(status_code=400, detail="Entry is not pending approval")
    
    now = datetime.now(timezone.utc)
    await db.income_expenses.update_one(
        {"entry_id": entry_id},
        {"$set": {
            "status": "rejected",
            "rejected_at": now.isoformat(),
            "rejected_by": user["user_id"],
            "rejected_by_name": user["name"],
            "rejection_reason": reason
        }}
    )
    
    await log_activity(request, user, "reject", "income_expenses", f"Rejected {entry['entry_type']}: {entry.get('description', 'N/A')}", reference_id=entry_id)
    return {"message": "Income/Expense entry rejected"}


# ---- Loan Disbursement Approve/Reject ----

@api_router.post("/loans/{loan_id}/approve-disbursement")
async def approve_loan_disbursement(request: Request, loan_id: str, user: dict = Depends(require_permission(Modules.LOANS, Actions.APPROVE))):
    """Approve a pending loan disbursement and execute treasury deduction"""
    loan = await db.loans.find_one({"loan_id": loan_id}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    if loan.get("status") != "pending_approval":
        raise HTTPException(status_code=400, detail="Loan is not pending approval")
    
    now = datetime.now(timezone.utc)
    
    # Execute treasury deduction if from treasury
    if loan.get("source_treasury_id"):
        treasury = await db.treasury_accounts.find_one({"account_id": loan["source_treasury_id"]}, {"_id": 0})
        if not treasury:
            raise HTTPException(status_code=404, detail="Source treasury account not found")
        
        if treasury.get("balance", 0) < loan["amount"]:
            raise HTTPException(status_code=400, detail="Insufficient balance in treasury account")
        
        treasury_currency = treasury.get("currency", "USD")
        treasury_deduct_amount = loan["amount"]
        if treasury_currency.upper() != loan["currency"].upper():
            treasury_deduct_amount = convert_currency(loan["amount"], loan["currency"], treasury_currency)
        
        await db.treasury_accounts.update_one(
            {"account_id": loan["source_treasury_id"]},
            {"$inc": {"balance": -treasury_deduct_amount}, "$set": {"updated_at": now.isoformat()}}
        )
        
        # Record treasury transaction
        conversion_note = f" (Converted from {loan['amount']:,.2f} {loan['currency']})" if treasury_currency.upper() != loan["currency"].upper() else ""
        tx_id = f"ttx_{uuid.uuid4().hex[:12]}"
        await db.treasury_transactions.insert_one({
            "treasury_transaction_id": tx_id,
            "account_id": loan["source_treasury_id"],
            "transaction_type": "loan_disbursement",
            "amount": -treasury_deduct_amount,
            "currency": treasury_currency,
            "original_amount": loan["amount"],
            "original_currency": loan["currency"],
            "reference": f"Loan to {loan['borrower_name']}{conversion_note}",
            "loan_id": loan_id,
            "created_at": now.isoformat(),
            "created_by": user["user_id"],
            "created_by_name": user["name"]
        })
    
    # Update loan status to active
    await db.loans.update_one(
        {"loan_id": loan_id},
        {"$set": {
            "status": LoanStatus.ACTIVE,
            "approved_at": now.isoformat(),
            "approved_by": user["user_id"],
            "approved_by_name": user["name"]
        }}
    )
    
    # Update loan transaction status
    await db.loan_transactions.update_one(
        {"loan_id": loan_id, "transaction_type": LoanTransactionType.DISBURSEMENT, "status": "pending_approval"},
        {"$set": {"status": "completed"}}
    )
    
    await log_activity(request, user, "approve", "loans", f"Approved loan disbursement to {loan['borrower_name']}: {loan['amount']} {loan['currency']}", reference_id=loan_id)
    return {"message": "Loan disbursement approved successfully"}


@api_router.post("/loans/{loan_id}/reject-disbursement")
async def reject_loan_disbursement(request: Request, loan_id: str, reason: str = "", user: dict = Depends(require_permission(Modules.LOANS, Actions.APPROVE))):
    """Reject a pending loan disbursement"""
    loan = await db.loans.find_one({"loan_id": loan_id}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    if loan.get("status") != "pending_approval":
        raise HTTPException(status_code=400, detail="Loan is not pending approval")
    
    now = datetime.now(timezone.utc)
    await db.loans.update_one(
        {"loan_id": loan_id},
        {"$set": {
            "status": "rejected",
            "rejected_at": now.isoformat(),
            "rejected_by": user["user_id"],
            "rejected_by_name": user["name"],
            "rejection_reason": reason
        }}
    )
    
    # Update loan transaction status
    await db.loan_transactions.update_one(
        {"loan_id": loan_id, "transaction_type": LoanTransactionType.DISBURSEMENT, "status": "pending_approval"},
        {"$set": {"status": "rejected"}}
    )
    
    await log_activity(request, user, "reject", "loans", f"Rejected loan disbursement to {loan['borrower_name']}", reference_id=loan_id)
    return {"message": "Loan disbursement rejected"}


# ---- Loan Repayment Approve/Reject ----

@api_router.post("/loan-repayments/{repayment_id}/approve")
async def approve_loan_repayment(request: Request, repayment_id: str, user: dict = Depends(require_permission(Modules.LOANS, Actions.APPROVE))):
    """Approve a pending loan repayment - execute treasury credit and update loan totals"""
    repayment = await db.loan_repayments.find_one({"repayment_id": repayment_id}, {"_id": 0})
    if not repayment:
        raise HTTPException(status_code=404, detail="Repayment not found")
    if repayment.get("status") != "pending_approval":
        raise HTTPException(status_code=400, detail="Repayment is not pending approval")
    
    loan = await db.loans.find_one({"loan_id": repayment["loan_id"]}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Associated loan not found")
    
    now = datetime.now(timezone.utc)
    repayment_amount_in_loan_currency = repayment.get("amount_in_loan_currency", repayment["amount"])
    
    # Update loan totals
    new_total_repaid = loan.get("total_repaid", 0) + repayment_amount_in_loan_currency
    outstanding = loan["amount"] + loan.get("total_interest", 0) - new_total_repaid
    
    new_status = loan["status"]
    if outstanding <= 0:
        new_status = LoanStatus.FULLY_PAID
    elif new_total_repaid > 0:
        new_status = LoanStatus.PARTIALLY_PAID
    
    await db.loans.update_one(
        {"loan_id": repayment["loan_id"]},
        {
            "$set": {
                "total_repaid": new_total_repaid,
                "outstanding_balance": max(0, outstanding),
                "status": new_status,
                "updated_at": now.isoformat()
            },
            "$inc": {"repayment_count": 1}
        }
    )
    
    # Credit treasury if from treasury
    if repayment.get("treasury_account_id"):
        treasury = await db.treasury_accounts.find_one({"account_id": repayment["treasury_account_id"]}, {"_id": 0})
        if treasury:
            treasury_currency = treasury.get("currency", "USD")
            treasury_credit_amount = repayment["amount"]
            if treasury_currency.upper() != repayment["currency"].upper():
                treasury_credit_amount = convert_currency(repayment["amount"], repayment["currency"], treasury_currency)
            
            await db.treasury_accounts.update_one(
                {"account_id": repayment["treasury_account_id"]},
                {"$inc": {"balance": treasury_credit_amount}, "$set": {"updated_at": now.isoformat()}}
            )
            
            conversion_note = f" (Converted from {repayment['amount']:,.2f} {repayment['currency']})" if treasury_currency.upper() != repayment["currency"].upper() else ""
            tx_id = f"ttx_{uuid.uuid4().hex[:12]}"
            await db.treasury_transactions.insert_one({
                "treasury_transaction_id": tx_id,
                "account_id": repayment["treasury_account_id"],
                "transaction_type": "loan_repayment",
                "amount": treasury_credit_amount,
                "currency": treasury_currency,
                "original_amount": repayment["amount"],
                "original_currency": repayment["currency"],
                "reference": f"Loan repayment from {loan['borrower_name']}{conversion_note}",
                "loan_id": repayment["loan_id"],
                "repayment_id": repayment_id,
                "created_at": now.isoformat(),
                "created_by": user["user_id"],
                "created_by_name": user["name"]
            })
    
    # Update repayment status
    await db.loan_repayments.update_one(
        {"repayment_id": repayment_id},
        {"$set": {
            "status": "approved",
            "approved_at": now.isoformat(),
            "approved_by": user["user_id"],
            "approved_by_name": user["name"]
        }}
    )
    
    # Update loan transaction status
    await db.loan_transactions.update_one(
        {"loan_id": repayment["loan_id"], "transaction_type": LoanTransactionType.REPAYMENT, "status": "pending_approval"},
        {"$set": {"status": "completed"}}
    )
    
    await log_activity(request, user, "approve", "loans", f"Approved loan repayment from {loan['borrower_name']}: {repayment['amount']} {repayment['currency']}", reference_id=repayment_id)
    return {"message": "Loan repayment approved successfully"}


@api_router.post("/loan-repayments/{repayment_id}/reject")
async def reject_loan_repayment(request: Request, repayment_id: str, reason: str = "", user: dict = Depends(require_permission(Modules.LOANS, Actions.APPROVE))):
    """Reject a pending loan repayment"""
    repayment = await db.loan_repayments.find_one({"repayment_id": repayment_id}, {"_id": 0})
    if not repayment:
        raise HTTPException(status_code=404, detail="Repayment not found")
    if repayment.get("status") != "pending_approval":
        raise HTTPException(status_code=400, detail="Repayment is not pending approval")
    
    now = datetime.now(timezone.utc)
    await db.loan_repayments.update_one(
        {"repayment_id": repayment_id},
        {"$set": {
            "status": "rejected",
            "rejected_at": now.isoformat(),
            "rejected_by": user["user_id"],
            "rejected_by_name": user["name"],
            "rejection_reason": reason
        }}
    )
    
    # Update loan transaction status
    await db.loan_transactions.update_one(
        {"loan_id": repayment["loan_id"], "transaction_type": LoanTransactionType.REPAYMENT, "status": "pending_approval"},
        {"$set": {"status": "rejected"}}
    )
    
    await log_activity(request, user, "reject", "loans", f"Rejected loan repayment", reference_id=repayment_id)
    return {"message": "Loan repayment rejected"}


# ---- PSP Settlement Approve (uses existing complete logic) / Reject ----

@api_router.post("/psp-settlements/{settlement_id}/approve")
async def approve_psp_settlement(request: Request, settlement_id: str, user: dict = Depends(require_permission(Modules.PSP, Actions.APPROVE))):
    """Approve a pending PSP settlement and execute treasury credit"""
    settlement = await db.psp_settlements.find_one({"settlement_id": settlement_id}, {"_id": 0})
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement not found")
    if settlement["status"] != PSPSettlementStatus.PENDING:
        raise HTTPException(status_code=400, detail="Settlement is not pending approval")
    
    now = datetime.now(timezone.utc)
    
    # Get destination treasury account
    dest = await db.treasury_accounts.find_one({"account_id": settlement["settlement_destination_id"]}, {"_id": 0})
    dest_currency = dest.get("currency", "USD") if dest else "USD"
    
    # Calculate treasury amount
    treasury_amount = settlement.get("treasury_amount")
    if not treasury_amount:
        treasury_amount = convert_currency(settlement["net_amount"], "USD", dest_currency)
    
    # Credit treasury
    await db.treasury_accounts.update_one(
        {"account_id": settlement["settlement_destination_id"]},
        {"$inc": {"balance": treasury_amount}, "$set": {"updated_at": now.isoformat()}}
    )
    
    # Record treasury transaction
    treasury_tx_id = f"ttx_{uuid.uuid4().hex[:12]}"
    stl_type = settlement.get("settlement_type", "standard")
    conversion_note = f" (Converted: USD {settlement['net_amount']:,.2f} -> {dest_currency} {treasury_amount:,.2f})" if dest_currency != "USD" else ""
    await db.treasury_transactions.insert_one({
        "treasury_transaction_id": treasury_tx_id,
        "account_id": settlement["settlement_destination_id"],
        "account_name": dest["account_name"] if dest else "Unknown",
        "transaction_type": "psp_settlement",
        "amount": treasury_amount,
        "currency": dest_currency,
        "original_amount": settlement["net_amount"],
        "original_currency": "USD",
        "reference": f"PSP {stl_type.capitalize()} Settlement - {settlement_id}",
        "description": f"{stl_type.capitalize()} settlement ({settlement.get('transaction_count', 0)} txns) from {settlement.get('psp_name', 'Unknown')}{conversion_note}",
        "related_settlement_id": settlement_id,
        "psp_id": settlement.get("psp_id"),
        "psp_name": settlement.get("psp_name"),
        "created_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    })
    
    # Mark settlement as completed
    await db.psp_settlements.update_one(
        {"settlement_id": settlement_id},
        {"$set": {
            "status": PSPSettlementStatus.COMPLETED,
            "settled_at": now.isoformat(),
            "approved_by": user["user_id"],
            "approved_by_name": user["name"]
        }}
    )
    
    # Mark all transactions as settled
    tx_ids = settlement.get("transaction_ids", [])
    if tx_ids:
        await db.transactions.update_many(
            {"transaction_id": {"$in": tx_ids}},
            {"$set": {
                "settled": True,
                "settlement_id": settlement_id,
                "settlement_status": "completed",
                "settled_at": now.isoformat(),
                "settled_by": user["user_id"],
                "settled_by_name": user["name"],
                "settlement_destination_id": settlement["settlement_destination_id"],
                "settlement_destination_name": dest["account_name"] if dest else "Unknown"
            }}
        )
    
    # Update PSP stats
    psp = await db.psps.find_one({"psp_id": settlement["psp_id"]}, {"_id": 0})
    if psp:
        current_pending = psp.get("pending_settlement", 0) or 0
        new_pending = max(current_pending - settlement["net_amount"], 0)
        await db.psps.update_one(
            {"psp_id": settlement["psp_id"]},
            {
                "$set": {"pending_settlement": round(new_pending, 2)},
                "$inc": {
                    "total_volume": settlement.get("gross_amount", 0),
                    "total_commission": settlement.get("commission_amount", 0)
                }
            }
        )
    
    await log_activity(request, user, "approve", "psp", f"Approved PSP settlement: {settlement.get('psp_name')} - ${settlement['net_amount']:,.2f}", reference_id=settlement_id)
    return {"message": "PSP settlement approved successfully"}


@api_router.post("/psp-settlements/{settlement_id}/reject")
async def reject_psp_settlement(request: Request, settlement_id: str, reason: str = "", user: dict = Depends(require_permission(Modules.PSP, Actions.APPROVE))):
    """Reject a pending PSP settlement"""
    settlement = await db.psp_settlements.find_one({"settlement_id": settlement_id}, {"_id": 0})
    if not settlement:
        raise HTTPException(status_code=404, detail="Settlement not found")
    if settlement["status"] != PSPSettlementStatus.PENDING:
        raise HTTPException(status_code=400, detail="Settlement is not pending approval")
    
    now = datetime.now(timezone.utc)
    
    # Revert transactions back to unsettled state
    tx_ids = settlement.get("transaction_ids", [])
    if tx_ids:
        await db.transactions.update_many(
            {"transaction_id": {"$in": tx_ids}},
            {"$unset": {"settlement_id": "", "settlement_status": ""}}
        )
    
    # Mark settlement as rejected
    await db.psp_settlements.update_one(
        {"settlement_id": settlement_id},
        {"$set": {
            "status": "rejected",
            "rejected_at": now.isoformat(),
            "rejected_by": user["user_id"],
            "rejected_by_name": user["name"],
            "rejection_reason": reason
        }}
    )
    
    await log_activity(request, user, "reject", "psp", f"Rejected PSP settlement: {settlement.get('psp_name')}", reference_id=settlement_id)
    return {"message": "PSP settlement rejected"}



@api_router.get("/transactions/form-data")
async def get_transaction_form_data(user: dict = Depends(require_permission(Modules.TRANSACTIONS, Actions.VIEW))):
    """Return all dropdown data needed for the Create Transaction form.
    Only requires Transactions permission, so any role that can view/create transactions
    can also see PSPs, treasury accounts, vendors, and clients in the form dropdowns."""
    clients_cursor = db.clients.find({}, {"_id": 0, "client_id": 1, "first_name": 1, "last_name": 1, "email": 1}).sort("created_at", -1).limit(200)
    treasury_cursor = db.treasury_accounts.find({"status": "active"}, {"_id": 0, "account_id": 1, "account_name": 1, "bank_name": 1, "currency": 1, "account_type": 1, "balance": 1}).sort("account_name", 1)
    psp_cursor = db.psps.find({}, {"_id": 0, "status": 1, "psp_id": 1, "psp_name": 1, "commission_rate": 1, "settlement_days": 1}).sort("psp_name", 1)
    vendors_cursor = db.vendors.find({"status": "active"}, {"_id": 0, "vendor_id": 1, "vendor_name": 1, "deposit_commission": 1, "withdrawal_commission": 1, "status": 1}).sort("vendor_name", 1)

    clients, treasury, psps, vendors = await asyncio.gather(
        clients_cursor.to_list(200),
        treasury_cursor.to_list(100),
        psp_cursor.to_list(100),
        vendors_cursor.to_list(100),
    )
    return {
        "clients": clients,
        "treasury_accounts": treasury,
        "psps": psps,
        "vendors": vendors,
    }

@api_router.get("/transactions/bulk-template")
async def download_bulk_template(
    format: str = "csv",
    user: dict = Depends(require_permission(Modules.TRANSACTIONS, Actions.CREATE))
):
    """Download a template for bulk transaction upload"""
    import io
    headers_list = [
        "Client Email", "Type", "Payment Currency", "Amount", 
        "Exchange Rate", "Destination Type", "Destination", 
        "Transaction Date", "Reference", "CRM Reference", "Description"
    ]
    sample_rows = [
        ["client@example.com", "deposit", "AED", "3700", "0.2723", "treasury", "test usd", "2026-03-01", "", "CRM001", "Client deposit"],
        ["client@example.com", "withdrawal", "USD", "500", "", "bank", "", "2026-03-01", "", "", "Client withdrawal"],
    ]
    
    if format == "xlsx":
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        ws.append(headers_list)
        for row in sample_rows:
            ws.append(row)
        ns = wb.create_sheet("Notes")
        ns.append(["Column", "Required", "Notes"])
        ns.append(["Client Email", "Yes", "Must match an existing client email"])
        ns.append(["Type", "Yes", "deposit or withdrawal"])
        ns.append(["Payment Currency", "No", "USD if empty. e.g. AED, EUR, GBP, INR"])
        ns.append(["Amount", "Yes", "Amount in Payment Currency"])
        ns.append(["Exchange Rate", "No", "Required if currency != USD. Rate: 1 PaymentCurrency = ? USD"])
        ns.append(["Destination Type", "Yes", "treasury, psp, vendor, bank, usdt"])
        ns.append(["Destination", "No", "Name of treasury account, PSP, or exchanger. Required for treasury/psp/vendor"])
        ns.append(["Transaction Date", "No", "YYYY-MM-DD format. Today if empty"])
        ns.append(["Reference", "No", "Auto-generated if empty"])
        ns.append(["CRM Reference", "No", "Optional CRM reference"])
        ns.append(["Description", "No", "Optional description"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from starlette.responses import StreamingResponse
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": "attachment; filename=bulk_transactions_template.xlsx"})
    else:
        buf = io.StringIO()
        import csv as csv_mod
        writer = csv_mod.writer(buf)
        writer.writerow(headers_list)
        for row in sample_rows:
            writer.writerow(row)
        buf.seek(0)
        from starlette.responses import StreamingResponse
        return StreamingResponse(io.BytesIO(buf.getvalue().encode()), media_type="text/csv",
                                 headers={"Content-Disposition": "attachment; filename=bulk_transactions_template.csv"})

@api_router.get("/transactions/{transaction_id}")
async def get_transaction(transaction_id: str, user: dict = Depends(require_permission(Modules.TRANSACTIONS, Actions.VIEW))):
    transaction = await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return transaction


@api_router.post("/transactions/bulk-validate")
async def bulk_validate_transactions(
    request: Request,
    file: UploadFile = File(...),
    user: dict = Depends(require_permission(Modules.TRANSACTIONS, Actions.CREATE))
):
    """Parse and validate a bulk transaction upload file. Returns validation results without creating transactions."""
    import pandas as pd
    import io
    
    content = await file.read()
    filename = file.filename.lower()
    
    try:
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            df = pd.read_excel(io.BytesIO(content), sheet_name=0)
        elif filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(content))
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported file format '{filename.split('.')[-1]}'. Please save as CSV or Excel (.xlsx). If using Apple Numbers, go to File > Export To > CSV or Excel.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}. Make sure the file is a valid CSV or Excel (.xlsx) file.")
    
    # Normalize column names
    col_map = {}
    for col in df.columns:
        cl = str(col).strip().lower().replace(' ', '_')
        if 'client' in cl and ('email' in cl or 'id' in cl): col_map[col] = 'client_email'
        elif cl in ['type', 'transaction_type']: col_map[col] = 'type'
        elif 'payment' in cl and 'currency' in cl or cl == 'currency': col_map[col] = 'payment_currency'
        elif cl in ['amount', 'amt']: col_map[col] = 'amount'
        elif 'exchange' in cl and 'rate' in cl: col_map[col] = 'exchange_rate'
        elif 'destination' in cl and 'type' in cl: col_map[col] = 'destination_type'
        elif cl in ['destination', 'dest']: col_map[col] = 'destination'
        elif 'transaction' in cl and 'date' in cl or cl == 'date': col_map[col] = 'transaction_date'
        elif cl in ['reference', 'ref']: col_map[col] = 'reference'
        elif 'crm' in cl: col_map[col] = 'crm_reference'
        elif cl in ['description', 'desc', 'notes']: col_map[col] = 'description'
    
    df = df.rename(columns=col_map)
    
    # Build lookup caches
    all_clients = await db.clients.find({}, {"_id": 0, "client_id": 1, "email": 1, "full_name": 1}).to_list(10000)
    email_to_client = {c["email"].lower(): c for c in all_clients if c.get("email")}
    
    all_treasury = await db.treasury_accounts.find({}, {"_id": 0, "account_id": 1, "account_name": 1, "currency": 1}).to_list(1000)
    treasury_name_map = {a["account_name"].lower(): a for a in all_treasury}
    
    all_psps = await db.psps.find({}, {"_id": 0, "psp_id": 1, "psp_name": 1}).to_list(1000)
    psp_name_map = {p["psp_name"].lower(): p for p in all_psps}
    
    all_vendors = await db.vendors.find({}, {"_id": 0, "vendor_id": 1, "vendor_name": 1}).to_list(1000)
    vendor_name_map = {v["vendor_name"].lower(): v for v in all_vendors}
    
    rows = []
    has_errors = False
    
    for idx, row in df.iterrows():
        errors = []
        row_data = {}
        
        # Client validation
        client_email = str(row.get('client_email', '')).strip().lower()
        if not client_email or client_email == 'nan':
            errors.append("Client Email is required")
        elif client_email not in email_to_client:
            errors.append(f"Client '{client_email}' not found")
        else:
            c = email_to_client[client_email]
            row_data['client_id'] = c['client_id']
            row_data['client_name'] = c.get('full_name', client_email)
            row_data['client_email'] = client_email
        
        # Type validation
        tx_type = str(row.get('type', '')).strip().lower()
        if tx_type not in ['deposit', 'withdrawal']:
            errors.append("Type must be 'deposit' or 'withdrawal'")
        row_data['type'] = tx_type
        
        # Amount validation
        try:
            amount_raw = float(row.get('amount', 0))
            if amount_raw <= 0:
                errors.append("Amount must be positive")
            row_data['amount_raw'] = amount_raw
        except (ValueError, TypeError):
            errors.append("Amount must be a valid number")
            row_data['amount_raw'] = 0
        
        # Currency & exchange rate
        currency = str(row.get('payment_currency', 'USD')).strip().upper()
        if not currency or currency == 'NAN':
            currency = 'USD'
        row_data['payment_currency'] = currency
        
        exchange_rate = None
        if currency != 'USD':
            try:
                er = row.get('exchange_rate')
                if er is not None and str(er).strip() and str(er).strip().lower() != 'nan':
                    exchange_rate = float(er)
                    if exchange_rate <= 0:
                        errors.append("Exchange Rate must be positive")
                else:
                    errors.append(f"Exchange Rate required for currency {currency}")
            except (ValueError, TypeError):
                errors.append("Exchange Rate must be a valid number")
        row_data['exchange_rate'] = exchange_rate
        
        # Calculate USD amount
        if currency == 'USD':
            row_data['usd_amount'] = row_data['amount_raw']
            row_data['base_amount'] = None
        else:
            row_data['base_amount'] = row_data['amount_raw']
            row_data['usd_amount'] = round(row_data['amount_raw'] * (exchange_rate or 0), 2)
        
        # Destination type
        dest_type = str(row.get('destination_type', '')).strip().lower()
        if not dest_type or dest_type == 'nan':
            errors.append("Destination Type is required")
        elif dest_type not in ['treasury', 'psp', 'vendor', 'bank', 'usdt']:
            errors.append(f"Invalid Destination Type '{dest_type}'. Must be: treasury, psp, vendor, bank, usdt")
        row_data['destination_type'] = dest_type
        
        # Destination name resolution
        dest_name = str(row.get('destination', '')).strip()
        if dest_name.lower() == 'nan':
            dest_name = ''
        row_data['destination_name'] = dest_name
        row_data['destination_id'] = None
        row_data['psp_id'] = None
        row_data['vendor_id'] = None
        
        if dest_type in ['treasury', 'usdt']:
            if dest_name:
                matched = treasury_name_map.get(dest_name.lower())
                if matched:
                    row_data['destination_id'] = matched['account_id']
                    row_data['destination_resolved'] = matched['account_name']
                else:
                    errors.append(f"Treasury account '{dest_name}' not found")
            else:
                errors.append("Destination name required for treasury/usdt type")
        elif dest_type == 'psp':
            if dest_name:
                matched = psp_name_map.get(dest_name.lower())
                if matched:
                    row_data['psp_id'] = matched['psp_id']
                    row_data['destination_resolved'] = matched['psp_name']
                else:
                    errors.append(f"PSP '{dest_name}' not found")
            else:
                errors.append("Destination name required for PSP type")
        elif dest_type == 'vendor':
            if dest_name:
                matched = vendor_name_map.get(dest_name.lower())
                if matched:
                    row_data['vendor_id'] = matched['vendor_id']
                    row_data['destination_resolved'] = matched['vendor_name']
                else:
                    errors.append(f"Exchanger '{dest_name}' not found")
            else:
                errors.append("Destination name required for vendor/exchanger type")
        
        # Transaction date
        tx_date = str(row.get('transaction_date', '')).strip()
        if not tx_date or tx_date.lower() == 'nan' or tx_date.lower() == 'nat':
            tx_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        else:
            try:
                parsed = datetime.strptime(tx_date[:10], "%Y-%m-%d")
                tx_date = parsed.strftime("%Y-%m-%d")
            except ValueError:
                errors.append(f"Invalid date format '{tx_date}'. Use YYYY-MM-DD")
                tx_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        row_data['transaction_date'] = tx_date
        
        # Reference
        ref = str(row.get('reference', '')).strip()
        if ref.lower() == 'nan':
            ref = ''
        row_data['reference'] = ref
        
        # CRM Reference
        crm_ref = str(row.get('crm_reference', '')).strip()
        if crm_ref.lower() == 'nan':
            crm_ref = ''
        row_data['crm_reference'] = crm_ref
        
        # Description
        desc = str(row.get('description', '')).strip()
        if desc.lower() == 'nan':
            desc = ''
        row_data['description'] = desc
        
        if errors:
            has_errors = True
        
        rows.append({
            "row_number": idx + 2,  # +2 for 1-indexed + header row
            "errors": errors,
            "valid": len(errors) == 0,
            "data": row_data
        })
    
    return {
        "total_rows": len(rows),
        "valid_rows": sum(1 for r in rows if r["valid"]),
        "error_rows": sum(1 for r in rows if not r["valid"]),
        "has_errors": has_errors,
        "can_proceed": not has_errors and len(rows) > 0,
        "rows": rows
    }


@api_router.post("/transactions/bulk-create")
async def bulk_create_transactions(
    request: Request,
    body: dict = Body(...),
    user: dict = Depends(require_permission(Modules.TRANSACTIONS, Actions.CREATE))
):
    """Create transactions from validated bulk data"""
    rows = body.get("rows", [])
    if not rows:
        raise HTTPException(status_code=400, detail="No rows to process")
    
    now = datetime.now(timezone.utc)
    created = []
    
    # Lookup caches for enrichment
    psp_cache = {}
    vendor_cache = {}
    treasury_cache = {}
    
    for row_item in rows:
        data = row_item.get("data", {})
        if not row_item.get("valid"):
            continue
        
        tx_id = f"tx_{uuid.uuid4().hex[:12]}"
        ref = data.get("reference") or f"REF{uuid.uuid4().hex[:8].upper()}"
        tx_type = data["type"]
        currency = data.get("payment_currency", "USD")
        usd_amount = data.get("usd_amount", 0)
        base_amount = data.get("base_amount")
        exchange_rate = data.get("exchange_rate")
        dest_type = data.get("destination_type", "bank")
        
        # Enrich PSP info
        psp_name = None
        psp_commission_rate = None
        psp_commission_amount = None
        psp_reserve_fund_amount = 0
        psp_net_amount = usd_amount
        if data.get("psp_id"):
            if data["psp_id"] not in psp_cache:
                psp_cache[data["psp_id"]] = await db.psps.find_one({"psp_id": data["psp_id"]}, {"_id": 0})
            psp_info = psp_cache.get(data["psp_id"])
            if psp_info:
                psp_name = psp_info["psp_name"]
                psp_commission_rate = psp_info.get("commission_rate", 0)
                comm = psp_commission_rate / 100
                psp_commission_amount = round(usd_amount * comm, 2)
                rf_rate = psp_info.get("reserve_fund_rate", psp_info.get("chargeback_rate", 0)) / 100
                if tx_type == "deposit" and rf_rate > 0:
                    psp_reserve_fund_amount = round(usd_amount * rf_rate, 2)
                psp_net_amount = round(usd_amount - psp_commission_amount - psp_reserve_fund_amount, 2)
        
        # Enrich vendor info
        vendor_name = None
        v_comm_rate = 0
        v_comm_amt = 0
        v_comm_base_amt = 0
        v_comm_base_currency = None
        if data.get("vendor_id"):
            if data["vendor_id"] not in vendor_cache:
                vendor_cache[data["vendor_id"]] = await db.vendors.find_one({"vendor_id": data["vendor_id"]}, {"_id": 0})
            vendor_info = vendor_cache.get(data["vendor_id"])
            if vendor_info:
                vendor_name = vendor_info["vendor_name"]
                v_comm_rate = vendor_info.get("deposit_commission" if tx_type == "deposit" else "withdrawal_commission", 0)
                if v_comm_rate > 0:
                    # USD commission
                    v_comm_amt = round(usd_amount * v_comm_rate / 100, 2)
                    # Base/payment currency commission
                    v_base = base_amount if (currency and currency != "USD" and base_amount) else usd_amount
                    v_comm_base_amt = round(v_base * v_comm_rate / 100, 2)
                    v_comm_base_currency = currency if (currency and currency != "USD") else "USD"
        
        # Enrich treasury info
        dest_account_name = None
        if data.get("destination_id"):
            if data["destination_id"] not in treasury_cache:
                treasury_cache[data["destination_id"]] = await db.treasury_accounts.find_one({"account_id": data["destination_id"]}, {"_id": 0})
            treasury_info = treasury_cache.get(data["destination_id"])
            if treasury_info:
                dest_account_name = treasury_info["account_name"]
        
        tx_doc = {
            "transaction_id": tx_id,
            "client_id": data["client_id"],
            "client_name": data.get("client_name", ""),
            "transaction_type": tx_type,
            "amount": usd_amount,
            "currency": "USD",
            "base_currency": currency if currency != "USD" else None,
            "base_amount": base_amount,
            "exchange_rate": exchange_rate,
            "destination_type": dest_type,
            "destination_account_id": data.get("destination_id"),
            "destination_account_name": dest_account_name,
            "vendor_id": data.get("vendor_id"),
            "vendor_name": vendor_name,
            "psp_id": data.get("psp_id"),
            "psp_name": psp_name,
            "psp_commission_rate": psp_commission_rate,
            "psp_commission_amount": psp_commission_amount,
            "psp_reserve_fund_amount": psp_reserve_fund_amount if psp_reserve_fund_amount > 0 else None,
            "psp_chargeback_amount": psp_reserve_fund_amount if psp_reserve_fund_amount > 0 else None,
            "psp_net_amount": psp_net_amount if data.get("psp_id") else None,
            "psp_total_deductions": round((psp_commission_amount or 0) + psp_reserve_fund_amount, 2) if data.get("psp_id") else None,
            "vendor_commission_rate": v_comm_rate if v_comm_rate > 0 else None,
            "vendor_commission_amount": v_comm_amt if v_comm_amt > 0 else None,
            "vendor_commission_base_amount": v_comm_base_amt if v_comm_base_amt > 0 else None,
            "vendor_commission_base_currency": v_comm_base_currency,
            "vendor_deposit_commission": vendor_info.get("deposit_commission") if (data.get("vendor_id") and vendor_cache.get(data["vendor_id"]) and tx_type == "deposit") else None,
            "vendor_withdrawal_commission": vendor_info.get("withdrawal_commission") if (data.get("vendor_id") and vendor_cache.get(data["vendor_id"]) and tx_type == "withdrawal") else None,
            "transaction_mode": "bank",
            "status": TransactionStatus.PENDING,
            "description": data.get("description", ""),
            "reference": ref,
            "crm_reference": data.get("crm_reference") or None,
            "transaction_date": data.get("transaction_date") or now.strftime("%Y-%m-%d"),
            "proof_image": None,
            "created_by": user["user_id"],
            "created_by_name": user["name"],
            "processed_by": None,
            "processed_by_name": None,
            "rejection_reason": None,
            "settled": False,
            "settlement_id": None,
            "settlement_status": None,
            "request_id": None,
            "bulk_upload": True,
            "created_at": now.isoformat(),
            "processed_at": None
        }
        await db.transactions.insert_one(tx_doc)
        created.append({"transaction_id": tx_id, "reference": ref, "client_name": data.get("client_name"), "amount": usd_amount})
    
    await log_activity(request, user, "create", "transactions", f"Bulk created {len(created)} transactions")
    
    return {"created": len(created), "transactions": created}


@api_router.post("/transactions")
async def create_transaction(

    request: Request,

    client_id: str = Form(...),
    transaction_type: str = Form(...),
    amount: float = Form(...),
    currency: str = Form("USD"),
    base_currency: str = Form("USD"),
    base_amount: Optional[float] = Form(None),
    exchange_rate: Optional[float] = Form(None),
    destination_type: str = Form("treasury"),
    destination_account_id: Optional[str] = Form(None),
    psp_id: Optional[str] = Form(None),
    vendor_id: Optional[str] = Form(None),
    commission_paid_by: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    reference: Optional[str] = Form(None),
    # Client bank details (for withdrawal to bank)
    client_bank_name: Optional[str] = Form(None),
    client_bank_account_name: Optional[str] = Form(None),
    client_bank_account_number: Optional[str] = Form(None),
    client_bank_swift_iban: Optional[str] = Form(None),
    client_bank_currency: Optional[str] = Form(None),
    save_bank_to_client: Optional[str] = Form(None),  # 'true' to save bank to client profile
    # Client USDT details (for withdrawal to USDT)
    client_usdt_address: Optional[str] = Form(None),
    client_usdt_network: Optional[str] = Form(None),
    transaction_mode: Optional[str] = Form("bank"),
    collecting_person_name: Optional[str] = Form(None),
    collecting_person_number: Optional[str] = Form(None),
    crm_reference: Optional[str] = Form(None),
    transaction_date: Optional[str] = Form(None),
    client_tags: Optional[str] = Form(None),
    proof_image: Optional[UploadFile] = File(None),
    user: dict = Depends(require_permission(Modules.TRANSACTIONS, Actions.CREATE))
):
    try:
        return await _create_transaction_impl(
            request, client_id, transaction_type, amount, currency, base_currency,
            base_amount, exchange_rate, destination_type, destination_account_id,
            psp_id, vendor_id, commission_paid_by, description, reference,
            client_bank_name, client_bank_account_name, client_bank_account_number,
            client_bank_swift_iban, client_bank_currency, save_bank_to_client,
            client_usdt_address, client_usdt_network, transaction_mode,
            collecting_person_name, collecting_person_number, crm_reference,
            transaction_date, client_tags, proof_image, user
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Transaction creation failed: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Transaction creation failed: {str(e)}")

async def _create_transaction_impl(
    request, client_id, transaction_type, amount, currency, base_currency,
    base_amount, exchange_rate, destination_type, destination_account_id,
    psp_id, vendor_id, commission_paid_by, description, reference,
    client_bank_name, client_bank_account_name, client_bank_account_number,
    client_bank_swift_iban, client_bank_currency, save_bank_to_client,
    client_usdt_address, client_usdt_network, transaction_mode,
    collecting_person_name, collecting_person_number, crm_reference,
    transaction_date, client_tags_str, proof_image, user
):
    now = datetime.now(timezone.utc)
    
    # ===== INPUT VALIDATION =====
    if not client_id or not client_id.strip():
        raise HTTPException(status_code=400, detail="Client is required")
    if not transaction_type or transaction_type not in ["deposit", "withdrawal"]:
        raise HTTPException(status_code=400, detail="Invalid transaction type")
    if not amount or amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than 0")
    if destination_type == "psp" and not psp_id:
        raise HTTPException(status_code=400, detail="PSP selection is required for PSP destination")
    if destination_type == "vendor" and not vendor_id:
        raise HTTPException(status_code=400, detail="Exchanger selection is required for vendor destination")
    if destination_type in ["treasury", "usdt"] and not destination_account_id:
        raise HTTPException(status_code=400, detail="Account selection is required for treasury/USDT destination")

    # ===== DUPLICATE DETECTION =====
    # Check 1: If reference is provided, ensure it's unique
    if reference:
        existing_by_ref = await db.transactions.find_one({"reference": reference}, {"_id": 0})
        if existing_by_ref:
            raise HTTPException(
                status_code=400, 
                detail=f"Duplicate transaction: Reference '{reference}' already exists (Transaction ID: {existing_by_ref['transaction_id']})"
            )
    
    # Check CRM reference uniqueness
    if crm_reference and crm_reference.strip():
        existing_crm = await db.transactions.find_one({"crm_reference": crm_reference.strip()}, {"_id": 0})
        if existing_crm:
            raise HTTPException(
                status_code=400,
                detail=f"CRM Reference '{crm_reference}' already exists (Transaction ID: {existing_crm['transaction_id']})"
            )
    
    # Check 2: Same client, type, amount within 5 minutes (prevents accidental double-submit)
    five_minutes_ago = (now - timedelta(minutes=5)).isoformat()
    duplicate_query = {
        "client_id": client_id,
        "transaction_type": transaction_type,
        "amount": amount,
        "created_at": {"$gte": five_minutes_ago}
    }
    # Add destination filters for more precise matching
    if destination_type == "psp" and psp_id:
        duplicate_query["psp_id"] = psp_id
    elif destination_type == "vendor" and vendor_id:
        duplicate_query["vendor_id"] = vendor_id
    elif destination_type == "treasury" and destination_account_id:
        duplicate_query["destination_account_id"] = destination_account_id
    
    recent_duplicate = await db.transactions.find_one(duplicate_query, {"_id": 0})
    if recent_duplicate:
        raise HTTPException(
            status_code=400, 
            detail=f"Possible duplicate: Similar transaction created {recent_duplicate.get('created_at', 'recently')} (Transaction ID: {recent_duplicate['transaction_id']}). Wait 5 minutes or use a unique reference."
        )
    # ===== END DUPLICATE DETECTION =====
    
    # Verify client exists
    client = await db.clients.find_one({"client_id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Verify destination based on type
    destination_account = None
    psp_info = None
    vendor_info = None
    
    if destination_type == "treasury" and destination_account_id:
        destination_account = await db.treasury_accounts.find_one({"account_id": destination_account_id}, {"_id": 0})
        if not destination_account:
            raise HTTPException(status_code=404, detail="Destination account not found")
    elif destination_type == "usdt" and destination_account_id:
        # For USDT deposits, select the USDT treasury account
        destination_account = await db.treasury_accounts.find_one({"account_id": destination_account_id}, {"_id": 0})
        if not destination_account:
            raise HTTPException(status_code=404, detail="USDT treasury account not found")
    elif destination_type == "psp" and psp_id:
        psp_info = await db.psps.find_one({"psp_id": psp_id}, {"_id": 0})
        if not psp_info:
            raise HTTPException(status_code=404, detail="PSP not found")
    elif destination_type == "vendor" and vendor_id:
        vendor_info = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
        if not vendor_info:
            raise HTTPException(status_code=404, detail="Vendor not found")
    
    # Save bank account to client profile if requested
    if destination_type in ["bank", "vendor"] and save_bank_to_client == "true" and client_bank_name and client_bank_account_number:
        existing_bank = await db.client_bank_accounts.find_one({
            "client_id": client_id,
            "account_number": client_bank_account_number,
            "bank_name": client_bank_name
        })
        if not existing_bank:
            bank_account_id = f"cba_{uuid.uuid4().hex[:12]}"
            bank_doc = {
                "bank_account_id": bank_account_id,
                "client_id": client_id,
                "bank_name": client_bank_name,
                "account_name": client_bank_account_name,
                "account_number": client_bank_account_number,
                "swift_iban": client_bank_swift_iban,
                "currency": client_bank_currency or "USD",
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": user["user_id"]
            }
            await db.client_bank_accounts.insert_one(bank_doc)
    
    tx_id = f"tx_{uuid.uuid4().hex[:12]}"
    # now is already defined at the top for duplicate detection
    
    # Handle proof image upload
    proof_image_url = None
    if proof_image:
        content = await proof_image.read()
        proof_image_url = upload_to_r2(content, proof_image.filename or "proof.png", proof_image.content_type or "image/png", "proofs")
    
    # Calculate USD amount if base currency is different
    usd_amount = amount
    actual_exchange_rate = exchange_rate
    if base_currency and base_currency != "USD" and base_amount:
        # Use user-provided exchange rate if available, otherwise fall back to convert_to_usd
        if exchange_rate and exchange_rate > 0:
            usd_amount = round(base_amount * exchange_rate, 2)
        else:
            usd_amount = convert_to_usd(base_amount, base_currency)
            # Calculate the implicit exchange rate for storage
            if base_amount > 0:
                actual_exchange_rate = round(usd_amount / base_amount, 6)
    
    # Calculate PSP commission if applicable
    commission_amount = 0.0
    net_amount = usd_amount
    expected_settlement_date = None
    
    if destination_type == "psp" and psp_info:
        commission_rate = psp_info.get("commission_rate", 0) / 100
        commission_amount = round(usd_amount * commission_rate, 2)
        
        # If commission paid by client, net amount is reduced
        if commission_paid_by == CommissionPaidBy.CLIENT:
            net_amount = usd_amount - commission_amount
        # If commission paid by broker, net amount stays same (broker absorbs)
        
        # Calculate expected settlement date
        settlement_days = psp_info.get("settlement_days", 1)
        expected_settlement_date = (now + timedelta(days=settlement_days)).isoformat()
        
        # Calculate reserve fund amount per-transaction
        reserve_fund_rate_pct = psp_info.get("reserve_fund_rate", psp_info.get("chargeback_rate", 0))
        psp_reserve_fund_amount = round(usd_amount * reserve_fund_rate_pct / 100, 2)
        
        # Calculate holding release date (when funds will be released from PSP holding)
        holding_days = psp_info.get("holding_days", 0)
        holding_release_date = (now + timedelta(days=holding_days)).isoformat() if holding_days > 0 else None
    
    # ===== BROKER COMMISSION CALCULATION =====
    broker_commission_rate = 0.0
    broker_commission_amount = 0.0
    broker_commission_base = 0.0
    commission_settings = await db.app_settings.find_one({"setting_type": "commission"}, {"_id": 0})
    if commission_settings and commission_settings.get("commission_enabled"):
        if transaction_type == TransactionType.DEPOSIT:
            broker_commission_rate = commission_settings.get("deposit_commission_rate", 0)
        elif transaction_type == TransactionType.WITHDRAWAL:
            broker_commission_rate = commission_settings.get("withdrawal_commission_rate", 0)
        if broker_commission_rate > 0:
            broker_commission_amount = round(usd_amount * broker_commission_rate / 100, 2)
            b_base = base_amount if (base_currency and base_currency != "USD" and base_amount) else usd_amount
            broker_commission_base = round(b_base * broker_commission_rate / 100, 2)

    # Calculate vendor commission at creation time
    vendor_commission_rate = 0.0
    vendor_commission_amount = 0.0
    vendor_commission_base_amount = 0.0
    if destination_type == "vendor" and vendor_info:
        tx_mode = transaction_mode or "bank"
        if transaction_type == TransactionType.DEPOSIT:
            if tx_mode == "cash":
                vendor_commission_rate = vendor_info.get("deposit_commission_cash", vendor_info.get("deposit_commission", 0))
            else:
                vendor_commission_rate = vendor_info.get("deposit_commission", 0)
        elif transaction_type == TransactionType.WITHDRAWAL:
            if tx_mode == "cash":
                vendor_commission_rate = vendor_info.get("withdrawal_commission_cash", vendor_info.get("withdrawal_commission", 0))
            else:
                vendor_commission_rate = vendor_info.get("withdrawal_commission", 0)
        if vendor_commission_rate > 0:
            # Calculate commission in BASE currency (base_amount), not USD
            v_base = base_amount if (base_currency and base_currency != "USD" and base_amount) else usd_amount
            vendor_commission_base_amount = round(v_base * vendor_commission_rate / 100, 2)
            # USD commission (separate)
            vendor_commission_amount = round(usd_amount * vendor_commission_rate / 100, 2)

    # Parse client_tags: either from form (comma-separated string) or from client defaults
    tx_client_tags = []
    if client_tags_str:
        tx_client_tags = [t.strip() for t in client_tags_str.split(",") if t.strip()]
    elif client.get("tags"):
        tx_client_tags = client["tags"]
    
    tx_doc = {
        "transaction_id": tx_id,
        "client_id": client_id,
        "client_name": f"{client['first_name']} {client['last_name']}",
        "client_tags": tx_client_tags,
        "transaction_type": transaction_type,
        "amount": usd_amount,
        "currency": "USD",
        "base_currency": base_currency or "USD",
        "base_amount": base_amount if base_currency != "USD" else None,
        "exchange_rate": actual_exchange_rate if (base_currency and base_currency != "USD") else None,
        "destination_type": destination_type,
        "destination_account_id": destination_account_id if destination_type in ["treasury", "usdt"] else None,
        "destination_account_name": destination_account["account_name"] if destination_account else None,
        "destination_bank_name": destination_account["bank_name"] if destination_account else None,
        # Client bank details (for withdrawal to bank or vendor)
        "client_bank_name": client_bank_name if destination_type in ["bank", "vendor"] else None,
        "client_bank_account_name": client_bank_account_name if destination_type in ["bank", "vendor"] else None,
        "client_bank_account_number": client_bank_account_number if destination_type in ["bank", "vendor"] else None,
        "client_bank_swift_iban": client_bank_swift_iban if destination_type in ["bank", "vendor"] else None,
        "client_bank_currency": client_bank_currency if destination_type in ["bank", "vendor"] else None,
        # Client USDT details (for withdrawal to USDT)
        "client_usdt_address": client_usdt_address if destination_type == "usdt" else None,
        "client_usdt_network": client_usdt_network if destination_type == "usdt" else None,
        "psp_id": psp_id if destination_type == "psp" else None,
        "psp_name": psp_info["psp_name"] if psp_info else None,
        "psp_commission_rate": psp_info["commission_rate"] if psp_info else None,
        "psp_commission_amount": commission_amount if psp_info else None,
        "psp_commission_paid_by": commission_paid_by if psp_info else None,
        "psp_net_amount": net_amount if psp_info else None,
        "psp_expected_settlement_date": expected_settlement_date,
        "psp_holding_days": psp_info.get("holding_days", 0) if psp_info else None,
        "psp_holding_release_date": holding_release_date if psp_info else None,
        "psp_reserve_fund_rate": psp_info.get("reserve_fund_rate", psp_info.get("chargeback_rate", 0)) if psp_info else None,
        "psp_chargeback_rate": psp_info.get("reserve_fund_rate", psp_info.get("chargeback_rate", 0)) if psp_info else None,
        "psp_reserve_fund_amount": psp_reserve_fund_amount if psp_info else None,
        "psp_chargeback_amount": psp_reserve_fund_amount if psp_info else None,
        "vendor_id": vendor_id if destination_type == "vendor" else None,
        "vendor_name": vendor_info["vendor_name"] if vendor_info else None,
        "vendor_deposit_commission": vendor_info["deposit_commission"] if vendor_info and transaction_type == TransactionType.DEPOSIT else None,
        "vendor_withdrawal_commission": vendor_info["withdrawal_commission"] if vendor_info and transaction_type == TransactionType.WITHDRAWAL else None,
        "vendor_commission_rate": vendor_commission_rate if vendor_commission_rate > 0 else None,
        "vendor_commission_amount": vendor_commission_amount if vendor_commission_amount > 0 else None,
        "vendor_commission_base_amount": vendor_commission_base_amount if vendor_commission_base_amount > 0 else None,
        "vendor_commission_base_currency": base_currency if vendor_commission_base_amount > 0 else None,
        "vendor_proof_image": None,
        "accountant_proof_image": None,
        "transaction_mode": transaction_mode or "bank",
        "collecting_person_name": collecting_person_name if transaction_mode == "cash" else None,
        "collecting_person_number": collecting_person_number if transaction_mode == "cash" else None,
        "status": TransactionStatus.PENDING,
        "description": description,
        "reference": reference or f"REF{uuid.uuid4().hex[:8].upper()}",
        "crm_reference": crm_reference.strip() if crm_reference else None,
        "transaction_date": transaction_date or now.strftime("%Y-%m-%d"),
        "proof_image": proof_image_url,
        "created_by": user["user_id"],
        "created_by_name": user["name"],
        "processed_by": None,
        "processed_by_name": None,
        "rejection_reason": None,
        "settled": False,
        "settlement_id": None,
        "settlement_status": None,
        "broker_commission_rate": broker_commission_rate,
        "broker_commission_amount": broker_commission_amount,
        "broker_commission_base_amount": broker_commission_base,
        "broker_commission_base_currency": base_currency or "USD",
        "created_at": now.isoformat(),
        "processed_at": None
    }
    
    await db.transactions.insert_one(tx_doc)
    
    # Invalidate transaction cache
    invalidate_transaction_cache()
    
    # Update PSP pending balance if this is a PSP transaction
    if destination_type == "psp" and psp_info:
        # Deposits add to pending_settlement (PSP owes us), withdrawals subtract
        psp_delta = net_amount if transaction_type == "deposit" else -net_amount
        await db.psps.update_one(
            {"psp_id": psp_id},
            {"$inc": {"pending_settlement": psp_delta}}
        )
    
    # Update vendor pending balance if this is a vendor transaction
    if destination_type == "vendor" and vendor_info:
        await db.vendors.update_one(
            {"vendor_id": vendor_id},
            {"$inc": {"pending_settlement": usd_amount}}
        )
    
    result = await db.transactions.find_one({"transaction_id": tx_id}, {"_id": 0})
    await log_activity(request, user, "create", "transactions", "Created transaction")

    # Send notifications (fire and forget)
    import asyncio
    if destination_type == "vendor" and vendor_id:
        # Vendor transaction → only notify the exchanger, NOT approvers
        amt_display = f"{base_amount:,.2f} {base_currency}" if base_currency and base_currency != "USD" and base_amount else f"${usd_amount:,.2f} USD"
        asyncio.create_task(send_exchanger_notification("transaction", vendor_id, {
            "reference": result.get("reference", tx_id),
            "type": transaction_type,
            "client": result.get("client_name", "Unknown"),
            "amount_display": amt_display,
        }))
    else:
        # Non-vendor transaction → notify approvers
        asyncio.create_task(send_approval_notification("transaction", {
            "reference": result.get("reference", tx_id),
            "type": transaction_type,
            "client": result.get("client_name", "Unknown"),
            "amount": usd_amount,
            "base_amount": base_amount,
            "base_currency": base_currency,
            "destination": result.get("destination_account_name") or result.get("psp_name") or destination_type,
            "created_by": user["name"]
        }))

    return result

@api_router.put("/transactions/{transaction_id}")
async def update_transaction(request: Request, transaction_id: str, update_data: TransactionUpdate, user: dict = Depends(require_permission(Modules.TRANSACTIONS, Actions.EDIT))):

    updates = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No updates provided")
    
    tx = await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    now = datetime.now(timezone.utc)
    
    # Editable fields only allowed on pending transactions
    editable_fields = {"crm_reference", "amount", "base_amount", "base_currency", "exchange_rate", "reference", "transaction_date"}
    has_editable_fields = any(k in editable_fields for k in updates)
    if has_editable_fields:
        if tx["status"] != TransactionStatus.PENDING:
            raise HTTPException(status_code=400, detail="These fields can only be edited on pending transactions")
    
    # client_tags can be updated on any status
    if "client_tags" in updates:
        updates["client_tags"] = updates["client_tags"] if updates["client_tags"] else []
    
    # Validate CRM reference uniqueness if being updated
    if "crm_reference" in updates and updates["crm_reference"]:
        existing_crm = await db.transactions.find_one(
            {"crm_reference": updates["crm_reference"].strip(), "transaction_id": {"$ne": transaction_id}}, {"_id": 0}
        )
        if existing_crm:
            raise HTTPException(status_code=400, detail=f"CRM Reference '{updates['crm_reference']}' already exists")
        updates["crm_reference"] = updates["crm_reference"].strip()
    
    # Validate reference uniqueness if being updated
    if "reference" in updates and updates["reference"]:
        existing_ref = await db.transactions.find_one(
            {"reference": updates["reference"].strip(), "transaction_id": {"$ne": transaction_id}}, {"_id": 0}
        )
        if existing_ref:
            raise HTTPException(status_code=400, detail=f"Reference '{updates['reference']}' already exists")
        updates["reference"] = updates["reference"].strip()
    
    # Validate amounts
    if "amount" in updates:
        updates["amount"] = float(updates["amount"])
        if updates["amount"] <= 0:
            raise HTTPException(status_code=400, detail="Amount must be positive")
    if "base_amount" in updates and updates["base_amount"] is not None:
        updates["base_amount"] = float(updates["base_amount"])
        if updates["base_amount"] <= 0:
            raise HTTPException(status_code=400, detail="Base amount must be positive")
    if "exchange_rate" in updates and updates["exchange_rate"] is not None:
        updates["exchange_rate"] = float(updates["exchange_rate"])
        if updates["exchange_rate"] <= 0:
            raise HTTPException(status_code=400, detail="Exchange rate must be positive")
    
    # Auto-calculate USD amount if base_amount and exchange_rate are provided
    new_base = updates.get("base_amount", tx.get("base_amount"))
    new_rate = updates.get("exchange_rate", tx.get("exchange_rate"))
    new_base_currency = updates.get("base_currency", tx.get("base_currency", "USD"))
    if new_base_currency and new_base_currency != "USD" and new_base and new_rate:
        updates["amount"] = round(float(new_base) * float(new_rate), 2)
    
    # If approving/completing or rejecting transaction
    if updates.get("status") in [TransactionStatus.APPROVED, TransactionStatus.COMPLETED, TransactionStatus.REJECTED]:
        updates["processed_by"] = user["user_id"]
        updates["processed_by_name"] = user["name"]
        updates["processed_at"] = now.isoformat()
        
        # Update treasury balance if completed
        if updates.get("status") == TransactionStatus.COMPLETED and tx.get("destination_account_id"):
            balance_change = tx["amount"] if tx["transaction_type"] == TransactionType.DEPOSIT else -tx["amount"]
            await db.treasury_accounts.update_one(
                {"account_id": tx["destination_account_id"]},
                {"$inc": {"balance": balance_change}, "$set": {"updated_at": now.isoformat()}}
            )
    
    updates["updated_at"] = now.isoformat()
    await db.transactions.update_one({"transaction_id": transaction_id}, {"$set": updates})
    
    await log_activity(request, user, "edit", "transactions", "Updated transaction")

    return await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})

@api_router.put("/transactions/{transaction_id}/assign")
async def assign_transaction_destination(
    request: Request,
    transaction_id: str,
    data: dict = Body(...),
    user: dict = Depends(require_permission(Modules.TRANSACTIONS, Actions.EDIT))
):
    """Assign/change destination for a pending transaction (e.g., assign exchanger)"""
    tx = await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    if tx["status"] != TransactionStatus.PENDING:
        raise HTTPException(status_code=400, detail="Only pending transactions can be edited")
    
    updates = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    dest_type = data.get("destination_type")
    vendor_id = data.get("vendor_id")
    dest_account_id = data.get("destination_account_id")
    
    if dest_type:
        updates["destination_type"] = dest_type
    
    if dest_type == "vendor" and vendor_id:
        vendor = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
        if not vendor:
            raise HTTPException(status_code=404, detail="Exchanger not found")
        updates["vendor_id"] = vendor_id
        updates["vendor_name"] = vendor.get("vendor_name")
        updates["destination_type"] = "vendor"
        
        # Calculate commission
        base_currency = tx.get("base_currency", "USD")
        base_amount = tx.get("base_amount") or tx.get("amount", 0)
        usd_amount = tx.get("amount", 0)
        comm_rate = vendor.get("withdrawal_commission", 0) if tx["transaction_type"] == "withdrawal" else vendor.get("deposit_commission", 0)
        if comm_rate > 0:
            updates["vendor_commission_rate"] = comm_rate
            updates["vendor_commission_base_amount"] = round(base_amount * comm_rate / 100, 2)
            updates["vendor_commission_amount"] = round(usd_amount * comm_rate / 100, 2)
            updates["vendor_commission_base_currency"] = base_currency
        
        # Notify exchanger
        import asyncio
        amt_display = f"{base_amount:,.2f} {base_currency}" if base_currency != "USD" else f"${usd_amount:,.2f} USD"
        asyncio.create_task(send_exchanger_notification("transaction", vendor_id, {
            "reference": tx.get("reference", transaction_id),
            "type": tx["transaction_type"],
            "client": tx.get("client_name", "Unknown"),
            "amount_display": amt_display,
        }))
    
    if dest_account_id:
        updates["destination_account_id"] = dest_account_id
        dest_acc = await db.treasury_accounts.find_one({"account_id": dest_account_id}, {"_id": 0})
        if dest_acc:
            updates["destination_account_name"] = dest_acc.get("account_name")
            updates["destination_bank_name"] = dest_acc.get("bank_name")
    
    # Allow editing description and reference
    if data.get("description") is not None:
        updates["description"] = data["description"]
    if data.get("crm_reference") is not None:
        updates["crm_reference"] = data["crm_reference"]
    
    await db.transactions.update_one({"transaction_id": transaction_id}, {"$set": updates})
    await log_activity(request, user, "edit", "transactions", f"Assigned transaction to {updates.get('vendor_name', dest_type)}")
    
    return await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})



@api_router.post("/transactions/{transaction_id}/approve")
async def approve_transaction(

    request: Request,

    transaction_id: str, 
    source_account_id: Optional[str] = None,
    bank_receipt_date: Optional[str] = None,
    require_proof: bool = True,
    user: dict = Depends(require_permission(Modules.TRANSACTIONS, Actions.APPROVE))
):
    """Approve a pending transaction"""
    tx = await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    if tx["status"] != TransactionStatus.PENDING:
        raise HTTPException(status_code=400, detail="Transaction is not pending")
    
    now = datetime.now(timezone.utc)
    
    # Use bank_receipt_date for treasury records if provided, otherwise use current time
    # Normalize to ISO datetime format for consistent querying
    if bank_receipt_date:
        treasury_date = f"{bank_receipt_date}T00:00:00" if 'T' not in bank_receipt_date else bank_receipt_date
    else:
        treasury_date = now.isoformat()
    
    updates = {
        "status": TransactionStatus.APPROVED,
        "processed_by": user["user_id"],
        "processed_by_name": user["name"],
        "processed_at": now.isoformat()
    }
    
    if bank_receipt_date:
        updates["bank_receipt_date"] = bank_receipt_date
    
    # For deposits, require proof of payment screenshot
    if tx["transaction_type"] == TransactionType.DEPOSIT:
        if require_proof and not tx.get("accountant_proof_image"):
            raise HTTPException(status_code=400, detail="Proof of payment screenshot is required for deposit approvals")
    
    # For withdrawals with bank/usdt destination, require source account
    if tx["transaction_type"] == TransactionType.WITHDRAWAL:
        if tx.get("destination_type") in ["bank", "usdt"]:
            if not source_account_id:
                raise HTTPException(status_code=400, detail="Source account is required for withdrawal approvals")
            
            # Check if source is a PSP (prefixed with psp_)
            is_psp_source = source_account_id.startswith("psp_")
            
            if is_psp_source:
                # Source is a PSP account
                psp_id = source_account_id  # Already has psp_ prefix
                psp_account = await db.psps.find_one({"psp_id": psp_id}, {"_id": 0})
                if not psp_account:
                    raise HTTPException(status_code=404, detail="PSP account not found")
                
                psp_currency = psp_account.get("currency", "USD")
                tx_currency = tx.get("currency", "USD")
                withdrawal_amount = tx["amount"]
                
                # Convert currencies if needed
                if tx.get("base_currency") == psp_currency and tx.get("base_amount"):
                    withdrawal_amount = tx["base_amount"]
                elif tx_currency == "USD" and psp_currency != "USD":
                    withdrawal_amount = convert_from_usd(tx["amount"], psp_currency)
                elif tx_currency != "USD" and psp_currency == "USD":
                    withdrawal_amount = convert_to_usd(tx["amount"], tx_currency)
                elif tx_currency != psp_currency:
                    usd_amount = convert_to_usd(tx["amount"], tx_currency)
                    withdrawal_amount = convert_from_usd(usd_amount, psp_currency)
                
                withdrawal_amount = round(withdrawal_amount, 2)
                
                psp_balance = psp_account.get("current_balance", 0)
                if psp_balance < withdrawal_amount:
                    raise HTTPException(status_code=400, detail=f"Insufficient PSP balance. Required: {withdrawal_amount:,.2f} {psp_currency}, Available: {psp_balance:,.2f} {psp_currency}")
                
                # Deduct from PSP balance
                await db.psps.update_one(
                    {"psp_id": psp_id},
                    {"$inc": {"current_balance": -withdrawal_amount}, "$set": {"updated_at": now.isoformat()}}
                )
                
                updates["source_account_id"] = psp_id
                updates["source_account_name"] = psp_account.get("psp_name")
                updates["source_type"] = "psp"
                updates["withdrawal_amount_in_source_currency"] = withdrawal_amount
                updates["source_currency"] = psp_currency
                
                # Record PSP transaction entry
                treasury_tx_id = f"ttx_{uuid.uuid4().hex[:12]}"
                treasury_tx = {
                    "treasury_transaction_id": treasury_tx_id,
                    "account_id": psp_id,
                    "account_type": "psp",
                    "transaction_type": "withdrawal",
                    "amount": -withdrawal_amount,
                    "currency": psp_currency,
                    "transaction_id": transaction_id,
                    "reference": tx.get("reference", ""),
                    "client_name": tx.get("client_name", ""),
                    "notes": f"Withdrawal from PSP {psp_account.get('psp_name')} for {tx.get('client_name', 'Unknown')}",
                    "created_at": treasury_date,
                    "created_by": user["user_id"],
                    "created_by_name": user["name"]
                }
                await db.treasury_transactions.insert_one(treasury_tx)
            else:
                # Source is a Treasury account (existing logic)
                source_account = await db.treasury_accounts.find_one({"account_id": source_account_id}, {"_id": 0})
                if not source_account:
                    raise HTTPException(status_code=404, detail="Source account not found")
            
                # Calculate withdrawal amount in source account's currency
                source_currency = source_account.get("currency", "USD")
                tx_currency = tx.get("currency", "USD")
                withdrawal_amount = tx["amount"]
                
                # PRIORITY: Use manual base_amount if source currency matches base_currency
                if tx.get("base_currency") == source_currency and tx.get("base_amount"):
                    withdrawal_amount = tx["base_amount"]
                elif tx_currency == "USD" and source_currency != "USD":
                    if tx.get("exchange_rate") and tx.get("base_amount") and tx.get("base_currency") == source_currency:
                        withdrawal_amount = tx["base_amount"]
                    else:
                        withdrawal_amount = convert_from_usd(tx["amount"], source_currency)
                elif tx_currency != "USD" and source_currency == "USD":
                    withdrawal_amount = convert_to_usd(tx["amount"], tx_currency)
                elif tx_currency != source_currency:
                    usd_amount = convert_to_usd(tx["amount"], tx_currency)
                    withdrawal_amount = convert_from_usd(usd_amount, source_currency)
                
                if source_account.get("balance", 0) < withdrawal_amount:
                    raise HTTPException(status_code=400, detail=f"Insufficient balance in source account. Required: {withdrawal_amount:,.2f} {source_currency}, Available: {source_account.get('balance', 0):,.2f} {source_currency}")
                
                # Deduct from source account
                await db.treasury_accounts.update_one(
                    {"account_id": source_account_id},
                    {"$inc": {"balance": -withdrawal_amount}, "$set": {"updated_at": now.isoformat()}}
                )
                
                updates["source_account_id"] = source_account_id
                updates["source_account_name"] = source_account.get("account_name")
                updates["source_type"] = "treasury"
                updates["withdrawal_amount_in_source_currency"] = withdrawal_amount
                updates["source_currency"] = source_currency
                
                # Record treasury transaction
                treasury_tx_id = f"ttx_{uuid.uuid4().hex[:12]}"
                original_amt = tx.get("base_amount") if tx.get("base_amount") else tx["amount"]
                original_curr = tx.get("base_currency") if tx.get("base_currency") else tx_currency
                manual_rate = tx.get("exchange_rate") if tx.get("exchange_rate") else (withdrawal_amount / tx["amount"] if tx["amount"] > 0 else 1)
            
                treasury_tx_doc = {
                    "treasury_transaction_id": treasury_tx_id,
                    "account_id": source_account_id,
                    "transaction_type": "withdrawal",
                    "amount": -withdrawal_amount,
                    "currency": source_currency,
                    "original_amount": original_amt,
                    "original_currency": original_curr,
                    "exchange_rate": manual_rate,
                    "reference": f"Withdrawal: {tx.get('client_name', 'Client')} - {tx.get('reference', '')}",
                    "transaction_id": transaction_id,
                    "client_id": tx.get("client_id"),
                    "created_at": treasury_date,
                    "created_by": user["user_id"],
                    "created_by_name": user["name"]
                }
                await db.treasury_transactions.insert_one(treasury_tx_doc)
        
        # For withdrawals with treasury destination, deduct from that treasury account
        elif tx.get("destination_type") == "treasury" and tx.get("destination_account_id"):
            dest_account = await db.treasury_accounts.find_one({"account_id": tx["destination_account_id"]}, {"_id": 0})
            if dest_account:
                dest_currency = dest_account.get("currency", "USD")
                tx_currency = tx.get("currency", "USD")
                withdrawal_amount = tx["amount"]
                
                if tx.get("base_currency") == dest_currency and tx.get("base_amount"):
                    withdrawal_amount = tx["base_amount"]
                elif tx_currency == "USD" and dest_currency != "USD":
                    if tx.get("exchange_rate") and tx.get("base_amount") and tx.get("base_currency") == dest_currency:
                        withdrawal_amount = tx["base_amount"]
                    else:
                        withdrawal_amount = convert_from_usd(tx["amount"], dest_currency)
                elif tx_currency != "USD" and dest_currency == "USD":
                    withdrawal_amount = convert_to_usd(tx["amount"], tx_currency)
                elif tx_currency != dest_currency:
                    usd_amount = convert_to_usd(tx["amount"], tx_currency)
                    withdrawal_amount = convert_from_usd(usd_amount, dest_currency)
                
                if dest_account.get("balance", 0) < withdrawal_amount:
                    raise HTTPException(status_code=400, detail=f"Insufficient balance in treasury account. Required: {withdrawal_amount:,.2f} {dest_currency}, Available: {dest_account.get('balance', 0):,.2f} {dest_currency}")
                
                # Deduct from treasury account
                await db.treasury_accounts.update_one(
                    {"account_id": tx["destination_account_id"]},
                    {"$inc": {"balance": -withdrawal_amount}, "$set": {"updated_at": now.isoformat()}}
                )
                
                updates["source_account_id"] = tx["destination_account_id"]
                updates["source_account_name"] = dest_account.get("account_name")
                updates["withdrawal_amount_in_source_currency"] = withdrawal_amount
                updates["source_currency"] = dest_currency
                
                # Record treasury transaction
                treasury_tx_id = f"ttx_{uuid.uuid4().hex[:12]}"
                original_amt = tx.get("base_amount") if tx.get("base_amount") else tx["amount"]
                original_curr = tx.get("base_currency") if tx.get("base_currency") else tx_currency
                manual_rate = tx.get("exchange_rate") if tx.get("exchange_rate") else (withdrawal_amount / tx["amount"] if tx["amount"] > 0 else 1)
                
                treasury_tx_doc = {
                    "treasury_transaction_id": treasury_tx_id,
                    "account_id": tx["destination_account_id"],
                    "transaction_type": "withdrawal",
                    "amount": -withdrawal_amount,
                    "currency": dest_currency,
                    "original_amount": original_amt,
                    "original_currency": original_curr,
                    "exchange_rate": manual_rate,
                    "reference": f"Withdrawal: {tx.get('client_name', 'Client')} - {tx.get('reference', '')}",
                    "transaction_id": transaction_id,
                    "client_id": tx.get("client_id"),
                    "created_at": treasury_date,
                    "created_by": user["user_id"],
                    "created_by_name": user["name"]
                }
                await db.treasury_transactions.insert_one(treasury_tx_doc)
    
    # Handle PSP-destination withdrawals - deduct from PSP balance
    if tx.get("destination_type") == "psp" and tx.get("psp_id") and tx["transaction_type"] == TransactionType.WITHDRAWAL:
        psp = await db.psps.find_one({"psp_id": tx["psp_id"]}, {"_id": 0})
        if psp:
            withdrawal_amount = round(tx["amount"], 2)
            
            # Deduct from PSP current_balance
            await db.psps.update_one(
                {"psp_id": tx["psp_id"]},
                {
                    "$inc": {"current_balance": -withdrawal_amount},
                    "$set": {"updated_at": now.isoformat()}
                }
            )
            
            updates["source_account_id"] = tx["psp_id"]
            updates["source_account_name"] = psp.get("psp_name")
            updates["source_type"] = "psp"
            updates["withdrawal_amount_in_source_currency"] = withdrawal_amount
            updates["source_currency"] = psp.get("currency", "USD")
            
            # Record PSP transaction
            treasury_tx_id = f"ttx_{uuid.uuid4().hex[:12]}"
            psp_tx_doc = {
                "treasury_transaction_id": treasury_tx_id,
                "account_id": tx["psp_id"],
                "account_type": "psp",
                "transaction_type": "withdrawal",
                "amount": -withdrawal_amount,
                "currency": psp.get("currency", "USD"),
                "reference": f"Withdrawal via PSP: {tx.get('client_name', 'Client')} - {tx.get('reference', '')}",
                "transaction_id": transaction_id,
                "client_id": tx.get("client_id"),
                "created_at": treasury_date,
                "created_by": user["user_id"],
                "created_by_name": user["name"]
            }
            await db.treasury_transactions.insert_one(psp_tx_doc)

    # Update treasury balance for deposits going to treasury
    if tx.get("destination_account_id") and tx["transaction_type"] == TransactionType.DEPOSIT:
        # Get the destination account to check its currency
        dest_account = await db.treasury_accounts.find_one({"account_id": tx["destination_account_id"]}, {"_id": 0})
        
        if dest_account:
            dest_currency = dest_account.get("currency", "USD")
            tx_currency = tx.get("currency", "USD")
            deposit_amount = tx["amount"]
            
            # Check if transaction has base_amount in the same currency as destination
            if tx.get("base_currency") == dest_currency and tx.get("base_amount"):
                # Use the actual base_amount (e.g., 100,000 AED deposited to AED account)
                deposit_amount = tx["base_amount"]
            elif tx_currency == "USD" and dest_currency != "USD":
                # Convert from USD to destination currency
                deposit_amount = convert_from_usd(tx["amount"], dest_currency)
            elif tx_currency != "USD" and dest_currency == "USD":
                # Convert from source currency to USD
                deposit_amount = convert_to_usd(tx["amount"], tx_currency)
            elif tx_currency != dest_currency:
                # Convert via USD as intermediate
                usd_amount = convert_to_usd(tx["amount"], tx_currency)
                deposit_amount = convert_from_usd(usd_amount, dest_currency)
            
            await db.treasury_accounts.update_one(
                {"account_id": tx["destination_account_id"]},
                {"$inc": {"balance": deposit_amount}, "$set": {"updated_at": now.isoformat()}}
            )
            
            # Record treasury transaction for deposit
            treasury_tx_id = f"ttx_{uuid.uuid4().hex[:12]}"
            # Determine the original amount and currency for reference
            original_amt = tx.get("base_amount") if tx.get("base_amount") else tx["amount"]
            original_curr = tx.get("base_currency") if tx.get("base_currency") else tx_currency
            
            treasury_tx_doc = {
                "treasury_transaction_id": treasury_tx_id,
                "account_id": tx["destination_account_id"],
                "transaction_type": "deposit",
                "amount": deposit_amount,
                "currency": dest_currency,
                "original_amount": original_amt,
                "original_currency": original_curr,
                "exchange_rate": tx.get("exchange_rate") or (deposit_amount / tx["amount"] if tx["amount"] > 0 else 1),
                "reference": f"Deposit: {tx.get('client_name', 'Client')} - {tx.get('reference', '')}",
                "transaction_id": transaction_id,
                "client_id": tx.get("client_id"),
                "created_at": treasury_date,
                "created_by": user["user_id"],
                "created_by_name": user["name"]
            }
            await db.treasury_transactions.insert_one(treasury_tx_doc)
    
    await db.transactions.update_one({"transaction_id": transaction_id}, {"$set": updates})
    
    # Invalidate transaction cache
    invalidate_transaction_cache()
    invalidate_treasury_cache()
    
    await log_activity(request, user, "approve", "transactions", "Approved transaction")

    return await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})

@api_router.post("/transactions/{transaction_id}/upload-proof")
async def upload_transaction_proof(

    request: Request,

    transaction_id: str,
    proof_image: UploadFile = File(...),
    user: dict = Depends(require_permission(Modules.TRANSACTIONS, Actions.EDIT))
):
    """Upload proof of payment for deposit and withdrawal transactions"""
    tx = await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # Allow uploading proof for pending deposits and withdrawals
    if tx["transaction_type"] not in [TransactionType.WITHDRAWAL, TransactionType.DEPOSIT]:
        raise HTTPException(status_code=400, detail="Proof upload is only for deposit and withdrawal transactions")
    
    content = await proof_image.read()
    proof_image_url = upload_to_r2(content, proof_image.filename or "proof.png", proof_image.content_type or "image/png", "proofs")
    
    now = datetime.now(timezone.utc)
    await db.transactions.update_one(
        {"transaction_id": transaction_id},
        {"$set": {
            "accountant_proof_image": proof_image_url,
            "proof_uploaded_at": now.isoformat(),
            "proof_uploaded_by": user["user_id"],
            "proof_uploaded_by_name": user["name"]
        }}
    )
    
    await log_activity(request, user, "edit", "transactions", "Uploaded transaction proof")

    return {"message": "Proof uploaded successfully", "transaction_id": transaction_id}

@api_router.post("/transactions/{transaction_id}/reject")
async def reject_transaction(request: Request, transaction_id: str, reason: str = "", user: dict = Depends(require_permission(Modules.TRANSACTIONS, Actions.APPROVE))):

    """Reject a pending transaction"""
    tx = await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    if tx["status"] != TransactionStatus.PENDING:
        raise HTTPException(status_code=400, detail="Transaction is not pending")
    
    now = datetime.now(timezone.utc)
    
    updates = {
        "status": TransactionStatus.REJECTED,
        "rejection_reason": reason,
        "processed_by": user["user_id"],
        "processed_by_name": user["name"],
        "processed_at": now.isoformat()
    }
    
    await db.transactions.update_one({"transaction_id": transaction_id}, {"$set": updates})
    
    # Invalidate transaction cache
    invalidate_transaction_cache()
    
    await log_activity(request, user, "reject", "transactions", "Rejected transaction")

    return await db.transactions.find_one({"transaction_id": transaction_id}, {"_id": 0})


# ============== TRANSACTION REQUESTS ==============

@api_router.get("/transaction-requests")
async def get_transaction_requests(
    status: Optional[str] = None,
    transaction_type: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    user: dict = Depends(require_permission(Modules.TRANSACTION_REQUESTS, Actions.VIEW))
):
    query = {}
    # Handle approved/rejected as transaction_status filters
    tx_status_filter = None
    if status and status != "all":
        if status in ("approved", "rejected"):
            query["status"] = "processed"
            tx_status_filter = status
        else:
            query["status"] = status
    if transaction_type and transaction_type != "all":
        query["transaction_type"] = transaction_type
    if search:
        query["$or"] = [
            {"client_name": {"$regex": search, "$options": "i"}},
            {"reference": {"$regex": search, "$options": "i"}},
            {"crm_reference": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}},
        ]
    if date_from or date_to:
        date_q = {}
        if date_from:
            date_q["$gte"] = date_from + "T00:00:00"
        if date_to:
            date_q["$lte"] = date_to + "T23:59:59"
        query["created_at"] = date_q
    result = await paginate_query(db.transaction_requests, query, page, page_size)
    
    # Enrich processed requests with their transaction's approval status
    tx_ids = [r["transaction_id"] for r in result.get("items", []) if r.get("transaction_id")]
    if tx_ids:
        txs = await db.transactions.find(
            {"transaction_id": {"$in": tx_ids}},
            {"_id": 0, "transaction_id": 1, "status": 1}
        ).to_list(len(tx_ids))
        tx_status_map = {tx["transaction_id"]: tx["status"] for tx in txs}
        for req in result.get("items", []):
            if req.get("transaction_id"):
                req["transaction_status"] = tx_status_map.get(req["transaction_id"])
    
    # Filter by transaction status if requested
    if tx_status_filter:
        result["items"] = [r for r in result.get("items", []) if r.get("transaction_status") == tx_status_filter]
        result["total"] = len(result["items"])
    
    return result

@api_router.get("/transaction-requests/pending-count")
async def get_pending_request_count(user: dict = Depends(get_current_user)):
    count = await db.transaction_requests.count_documents({"status": "pending"})
    return {"count": count}

@api_router.get("/transaction-requests/{request_id}")
async def get_transaction_request(request_id: str, user: dict = Depends(require_permission(Modules.TRANSACTION_REQUESTS, Actions.VIEW))):
    req = await db.transaction_requests.find_one({"request_id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    return req

@api_router.post("/transaction-requests")
async def create_transaction_request(
    request: Request,
    client_id: str = Form(...),
    transaction_type: str = Form(...),
    amount: float = Form(...),
    currency: str = Form("USD"),
    base_currency: str = Form("USD"),
    base_amount: Optional[float] = Form(None),
    exchange_rate: Optional[float] = Form(None),
    destination_type: str = Form("bank"),
    destination_account_id: Optional[str] = Form(None),
    psp_id: Optional[str] = Form(None),
    vendor_id: Optional[str] = Form(None),
    reference: Optional[str] = Form(None),
    crm_reference: Optional[str] = Form(None),
    description: Optional[str] = Form(None),
    client_bank_name: Optional[str] = Form(None),
    client_bank_account_name: Optional[str] = Form(None),
    client_bank_account_number: Optional[str] = Form(None),
    client_bank_swift_iban: Optional[str] = Form(None),
    client_bank_currency: Optional[str] = Form(None),
    client_usdt_address: Optional[str] = Form(None),
    client_usdt_network: Optional[str] = Form(None),
    transaction_date: Optional[str] = Form(None),
    client_tags: Optional[str] = Form(None),
    proof_image: Optional[UploadFile] = File(None),
    user: dict = Depends(require_permission(Modules.TRANSACTION_REQUESTS, Actions.CREATE))
):
    now = datetime.now(timezone.utc)
    
    # CRM reference uniqueness
    if crm_reference and crm_reference.strip():
        existing = await db.transaction_requests.find_one({"crm_reference": crm_reference.strip()}, {"_id": 0})
        if existing:
            raise HTTPException(status_code=400, detail=f"CRM Reference '{crm_reference}' already exists")
    
    # Get client info
    client = await db.clients.find_one({"client_id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Parse client_tags
    tx_client_tags = []
    if client_tags:
        tx_client_tags = [t.strip() for t in client_tags.split(",") if t.strip()]
    elif client.get("tags"):
        tx_client_tags = client["tags"]
    
    # Handle proof image
    proof_url = None
    if proof_image and proof_image.filename:
        content = await proof_image.read()
        proof_url = upload_to_r2(content, proof_image.filename, proof_image.content_type or "image/png", "proofs")
    
    request_id = f"txreq_{uuid.uuid4().hex[:12]}"
    
    doc = {
        "request_id": request_id,
        "transaction_type": transaction_type,
        "client_id": client_id,
        "client_name": f"{client.get('first_name', '')} {client.get('last_name', '')}".strip(),
        "client_tags": tx_client_tags,
        "amount": amount,
        "currency": currency,
        "base_currency": base_currency,
        "base_amount": base_amount,
        "exchange_rate": exchange_rate,
        "destination_type": destination_type,
        "destination_account_id": destination_account_id,
        "psp_id": psp_id,
        "vendor_id": vendor_id,
        "reference": reference,
        "crm_reference": crm_reference.strip() if crm_reference else None,
        "transaction_date": transaction_date or now.strftime("%Y-%m-%d"),
        "description": description,
        "client_bank_name": client_bank_name,
        "client_bank_account_name": client_bank_account_name,
        "client_bank_account_number": client_bank_account_number,
        "client_bank_swift_iban": client_bank_swift_iban,
        "client_bank_currency": client_bank_currency,
        "client_usdt_address": client_usdt_address,
        "client_usdt_network": client_usdt_network,
        "proof_image": proof_url,
        "status": "pending",
        "created_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"],
        "processed_at": None,
        "processed_by": None,
        "processed_by_name": None,
        "transaction_id": None,
    }
    
    await db.transaction_requests.insert_one(doc)
    await log_activity(request, user, "create", "transaction_requests", f"Created {transaction_type} request")
    
    # Auto-process deposits immediately
    if transaction_type == "deposit":
        tx_id = f"tx_{uuid.uuid4().hex[:12]}"
        vendor_info = None
        psp_info = None
        destination_account = None
        
        if vendor_id:
            vendor_info = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0})
        if psp_id:
            psp_info = await db.psps.find_one({"psp_id": psp_id}, {"_id": 0})
        if destination_account_id and destination_type in ["treasury", "usdt"]:
            destination_account = await db.treasury_accounts.find_one({"account_id": destination_account_id}, {"_id": 0})
        
        # Calculate vendor commission
        v_comm_rate = 0.0
        v_comm_amt = 0.0
        v_comm_base = 0.0
        if destination_type == "vendor" and vendor_info:
            v_comm_rate = vendor_info.get("deposit_commission", 0)
            if v_comm_rate > 0:
                v_base = base_amount if (base_currency and base_currency != "USD" and base_amount) else amount
                v_comm_base = round(v_base * v_comm_rate / 100, 2)
                v_comm_amt = round(amount * v_comm_rate / 100, 2)
        
        # Calculate PSP commission if applicable
        psp_commission_amount = 0.0
        psp_net_amount = amount
        psp_expected_settlement_date = None
        psp_reserve_fund_amount = 0.0
        psp_holding_release_date = None
        if destination_type == "psp" and psp_info:
            comm_rate = psp_info.get("commission_rate", 0) / 100
            psp_commission_amount = round(amount * comm_rate, 2)
            psp_net_amount = amount - psp_commission_amount
            settlement_days = psp_info.get("settlement_days", 1)
            psp_expected_settlement_date = (now + timedelta(days=settlement_days)).isoformat()
            reserve_fund_rate_pct = psp_info.get("reserve_fund_rate", psp_info.get("chargeback_rate", 0))
            psp_reserve_fund_amount = round(amount * reserve_fund_rate_pct / 100, 2)
            holding_days = psp_info.get("holding_days", 0)
            psp_holding_release_date = (now + timedelta(days=holding_days)).isoformat() if holding_days > 0 else None
        
        tx_doc = {
            "transaction_id": tx_id,
            "client_id": client_id,
            "client_name": doc["client_name"],
            "transaction_type": "deposit",
            "amount": amount,
            "currency": "USD",
            "base_currency": base_currency,
            "base_amount": base_amount if base_currency != "USD" else None,
            "exchange_rate": exchange_rate,
            "destination_type": destination_type,
            "destination_account_id": destination_account_id if destination_type in ["treasury", "usdt"] else None,
            "destination_account_name": destination_account["account_name"] if destination_account else None,
            "destination_bank_name": destination_account.get("bank_name") if destination_account else None,
            "vendor_id": vendor_id,
            "vendor_name": vendor_info["vendor_name"] if vendor_info else None,
            "psp_id": psp_id if destination_type == "psp" else None,
            "psp_name": psp_info["psp_name"] if psp_info else None,
            "psp_commission_rate": psp_info["commission_rate"] if psp_info else None,
            "psp_commission_amount": psp_commission_amount if psp_info else None,
            "psp_net_amount": psp_net_amount if psp_info else None,
            "psp_expected_settlement_date": psp_expected_settlement_date,
            "psp_reserve_fund_rate": psp_info.get("reserve_fund_rate", psp_info.get("chargeback_rate", 0)) if psp_info else None,
            "psp_reserve_fund_amount": psp_reserve_fund_amount if psp_info else None,
            "psp_holding_days": psp_info.get("holding_days", 0) if psp_info else None,
            "psp_holding_release_date": psp_holding_release_date,
            "client_bank_name": client_bank_name,
            "client_bank_account_name": client_bank_account_name,
            "client_bank_account_number": client_bank_account_number,
            "client_bank_swift_iban": client_bank_swift_iban,
            "client_bank_currency": client_bank_currency,
            "client_usdt_address": client_usdt_address,
            "client_usdt_network": client_usdt_network,
            "vendor_commission_rate": v_comm_rate if v_comm_rate > 0 else None,
            "vendor_commission_amount": v_comm_amt if v_comm_amt > 0 else None,
            "vendor_commission_base_amount": v_comm_base if v_comm_base > 0 else None,
            "vendor_commission_base_currency": base_currency if v_comm_base > 0 else None,
            "transaction_mode": "bank",
            "status": TransactionStatus.PENDING,
            "description": description,
            "reference": reference or f"REF{uuid.uuid4().hex[:8].upper()}",
            "crm_reference": crm_reference.strip() if crm_reference else None,
            "transaction_date": transaction_date or now.strftime("%Y-%m-%d"),
            "client_tags": doc.get("client_tags", []),
            "proof_image": proof_url,
            "created_by": user["user_id"],
            "created_by_name": user["name"],
            "processed_by": None,
            "processed_by_name": None,
            "rejection_reason": None,
            "settled": False,
            "settlement_id": None,
            "settlement_status": None,
            "request_id": request_id,
            "created_at": now.isoformat(),
            "processed_at": None
        }
        await db.transactions.insert_one(tx_doc)
        
        # Mark request as processed
        await db.transaction_requests.update_one(
            {"request_id": request_id},
            {"$set": {
                "status": "processed",
                "processed_at": now.isoformat(),
                "processed_by": user["user_id"],
                "processed_by_name": user["name"],
                "transaction_id": tx_id
            }}
        )
        
        # Send notifications
        import asyncio
        if destination_type == "vendor" and vendor_id:
            amt_display = f"{base_amount:,.2f} {base_currency}" if base_currency and base_currency != "USD" and base_amount else f"${amount:,.2f} USD"
            asyncio.create_task(send_exchanger_notification("transaction", vendor_id, {
                "reference": tx_doc["reference"],
                "type": "deposit",
                "client": doc["client_name"],
                "amount_display": amt_display,
            }))
        else:
            asyncio.create_task(send_approval_notification("transaction", {
                "reference": tx_doc["reference"],
                "type": "deposit",
                "client": doc["client_name"],
                "amount": amount,
                "base_amount": base_amount,
                "base_currency": base_currency,
                "destination": vendor_info["vendor_name"] if vendor_info else destination_type,
                "created_by": user["name"]
            }))
        
        await log_activity(request, user, "approve", "transaction_requests", f"Auto-processed deposit → TX {tx_id}")
        
        result = await db.transaction_requests.find_one({"request_id": request_id}, {"_id": 0})
        return result
    
    return await db.transaction_requests.find_one({"request_id": request_id}, {"_id": 0})

@api_router.put("/transaction-requests/{request_id}")
async def update_transaction_request(
    request: Request,
    request_id: str,
    data: dict = Body(...),
    user: dict = Depends(require_permission(Modules.TRANSACTION_REQUESTS, Actions.EDIT))
):
    req = await db.transaction_requests.find_one({"request_id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    if req["status"] != "pending":
        raise HTTPException(status_code=400, detail="Only pending requests can be edited")
    
    allowed = ["transaction_type", "client_id", "amount", "currency", "base_currency", "base_amount", "exchange_rate",
               "description", "reference", "crm_reference", "destination_type", "vendor_id",
               "client_bank_name", "client_bank_account_name", "client_bank_account_number",
               "client_bank_swift_iban", "client_bank_currency", "client_usdt_address", "client_usdt_network",
               "destination_account_id", "psp_id"]
    updates = {k: v for k, v in data.items() if k in allowed}
    if not updates:
        raise HTTPException(status_code=400, detail="No valid fields to update")
    
    # CRM ref uniqueness check
    if "crm_reference" in updates and updates["crm_reference"]:
        existing = await db.transaction_requests.find_one({"crm_reference": updates["crm_reference"], "request_id": {"$ne": request_id}}, {"_id": 0})
        if existing:
            raise HTTPException(status_code=400, detail=f"CRM Reference '{updates['crm_reference']}' already exists")
    
    # Resolve client name if client_id changed
    if "client_id" in updates:
        client = await db.clients.find_one({"client_id": updates["client_id"]}, {"_id": 0})
        if client:
            updates["client_name"] = f"{client.get('first_name', '')} {client.get('last_name', '')}".strip()
    
    # Convert numeric fields
    for nf in ["amount", "base_amount", "exchange_rate"]:
        if nf in updates and updates[nf]:
            try:
                updates[nf] = float(updates[nf])
            except (ValueError, TypeError):
                pass

    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.transaction_requests.update_one({"request_id": request_id}, {"$set": updates})
    await log_activity(request, user, "edit", "transaction_requests", "Updated request")
    return await db.transaction_requests.find_one({"request_id": request_id}, {"_id": 0})

@api_router.post("/transaction-requests/{request_id}/process")
async def process_transaction_request(
    request: Request,
    request_id: str,
    data: dict = Body(...),
    user: dict = Depends(require_permission(Modules.TRANSACTION_REQUESTS, Actions.APPROVE))
):
    """Process a pending request — creates a real transaction in the Transactions collection.
    Only for withdrawals. Requires captcha verification."""
    req = await db.transaction_requests.find_one({"request_id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    if req["status"] != "pending":
        raise HTTPException(status_code=400, detail="Only pending requests can be processed")
    
    # Verify captcha
    captcha_answer = data.get("captcha_answer")
    expected = data.get("captcha_expected")
    if not captcha_answer or str(captcha_answer) != str(expected):
        raise HTTPException(status_code=400, detail="Invalid captcha answer")
    
    now = datetime.now(timezone.utc)
    
    # Create real transaction in the transactions collection
    tx_id = f"tx_{uuid.uuid4().hex[:12]}"
    
    client = await db.clients.find_one({"client_id": req["client_id"]}, {"_id": 0})
    usd_amount = req["amount"]
    base_currency = req.get("base_currency", "USD")
    base_amount = req.get("base_amount")
    
    # Get vendor info if exchanger destination
    vendor_info = None
    psp_info = None
    destination_account = None
    if req.get("vendor_id"):
        vendor_info = await db.vendors.find_one({"vendor_id": req["vendor_id"]}, {"_id": 0})
    if req.get("psp_id"):
        psp_info = await db.psps.find_one({"psp_id": req["psp_id"]}, {"_id": 0})
    if req.get("destination_account_id") and req.get("destination_type") in ["treasury", "usdt"]:
        destination_account = await db.treasury_accounts.find_one({"account_id": req["destination_account_id"]}, {"_id": 0})
    
    # Calculate vendor commission
    vendor_commission_rate = 0.0
    vendor_commission_amount = 0.0
    vendor_commission_base_amount = 0.0
    if req.get("destination_type") == "vendor" and vendor_info:
        tx_mode = "bank"
        if req["transaction_type"] == "deposit":
            vendor_commission_rate = vendor_info.get("deposit_commission", 0)
        else:
            vendor_commission_rate = vendor_info.get("withdrawal_commission", 0)
        if vendor_commission_rate > 0:
            v_base = base_amount if (base_currency and base_currency != "USD" and base_amount) else usd_amount
            vendor_commission_base_amount = round(v_base * vendor_commission_rate / 100, 2)
            vendor_commission_amount = round(usd_amount * vendor_commission_rate / 100, 2)
    
    # Calculate PSP commission if applicable
    psp_commission_amount = 0.0
    psp_net_amount = usd_amount
    psp_expected_settlement_date = None
    psp_reserve_fund_amount = 0.0
    psp_holding_release_date = None
    if req.get("destination_type") == "psp" and psp_info:
        comm_rate = psp_info.get("commission_rate", 0) / 100
        psp_commission_amount = round(usd_amount * comm_rate, 2)
        psp_net_amount = usd_amount - psp_commission_amount
        settlement_days = psp_info.get("settlement_days", 1)
        psp_expected_settlement_date = (now + timedelta(days=settlement_days)).isoformat()
        reserve_fund_rate_pct = psp_info.get("reserve_fund_rate", psp_info.get("chargeback_rate", 0))
        psp_reserve_fund_amount = round(usd_amount * reserve_fund_rate_pct / 100, 2)
        holding_days = psp_info.get("holding_days", 0)
        psp_holding_release_date = (now + timedelta(days=holding_days)).isoformat() if holding_days > 0 else None
    
    tx_doc = {
        "transaction_id": tx_id,
        "client_id": req["client_id"],
        "client_name": req.get("client_name", "Unknown"),
        "transaction_type": req["transaction_type"],
        "amount": usd_amount,
        "currency": "USD",
        "base_currency": base_currency,
        "base_amount": base_amount if base_currency != "USD" else None,
        "exchange_rate": req.get("exchange_rate"),
        "destination_type": req.get("destination_type", "bank"),
        "destination_account_id": req.get("destination_account_id") if req.get("destination_type") in ["treasury", "usdt"] else None,
        "destination_account_name": destination_account["account_name"] if destination_account else None,
        "destination_bank_name": destination_account.get("bank_name") if destination_account else None,
        "vendor_id": req.get("vendor_id"),
        "vendor_name": vendor_info["vendor_name"] if vendor_info else None,
        "psp_id": req.get("psp_id") if req.get("destination_type") == "psp" else None,
        "psp_name": psp_info["psp_name"] if psp_info else None,
        "psp_commission_rate": psp_info["commission_rate"] if psp_info else None,
        "psp_commission_amount": psp_commission_amount if psp_info else None,
        "psp_net_amount": psp_net_amount if psp_info else None,
        "psp_expected_settlement_date": psp_expected_settlement_date,
        "psp_reserve_fund_rate": psp_info.get("reserve_fund_rate", psp_info.get("chargeback_rate", 0)) if psp_info else None,
        "psp_reserve_fund_amount": psp_reserve_fund_amount if psp_info else None,
        "psp_holding_days": psp_info.get("holding_days", 0) if psp_info else None,
        "psp_holding_release_date": psp_holding_release_date,
        "client_bank_name": req.get("client_bank_name"),
        "client_bank_account_name": req.get("client_bank_account_name"),
        "client_bank_account_number": req.get("client_bank_account_number"),
        "client_bank_swift_iban": req.get("client_bank_swift_iban"),
        "client_bank_currency": req.get("client_bank_currency"),
        "client_usdt_address": req.get("client_usdt_address"),
        "client_usdt_network": req.get("client_usdt_network"),
        "vendor_commission_rate": vendor_commission_rate if vendor_commission_rate > 0 else None,
        "vendor_commission_amount": vendor_commission_amount if vendor_commission_amount > 0 else None,
        "vendor_commission_base_amount": vendor_commission_base_amount if vendor_commission_base_amount > 0 else None,
        "vendor_commission_base_currency": base_currency if vendor_commission_base_amount > 0 else None,
        "transaction_mode": "bank",
        "status": TransactionStatus.PENDING,
        "description": req.get("description"),
        "reference": req.get("reference") or f"REF{uuid.uuid4().hex[:8].upper()}",
        "crm_reference": req.get("crm_reference"),
        "transaction_date": req.get("transaction_date") or now.strftime("%Y-%m-%d"),
        "client_tags": req.get("client_tags", []),
        "proof_image": req.get("proof_image"),
        "created_by": req.get("created_by"),
        "created_by_name": req.get("created_by_name"),
        "processed_by": None,
        "processed_by_name": None,
        "rejection_reason": None,
        "settled": False,
        "settlement_id": None,
        "settlement_status": None,
        "request_id": request_id,
        "created_at": now.isoformat(),
        "processed_at": None
    }
    
    await db.transactions.insert_one(tx_doc)
    
    # Update request status
    await db.transaction_requests.update_one(
        {"request_id": request_id},
        {"$set": {
            "status": "processed",
            "processed_at": now.isoformat(),
            "processed_by": user["user_id"],
            "processed_by_name": user["name"],
            "transaction_id": tx_id
        }}
    )
    
    # Send notifications
    import asyncio
    if req.get("destination_type") == "vendor" and req.get("vendor_id"):
        amt_display = f"{base_amount:,.2f} {base_currency}" if base_currency and base_currency != "USD" and base_amount else f"${usd_amount:,.2f} USD"
        asyncio.create_task(send_exchanger_notification("transaction", req["vendor_id"], {
            "reference": tx_doc["reference"],
            "type": req["transaction_type"],
            "client": req.get("client_name", "Unknown"),
            "amount_display": amt_display,
        }))
    else:
        asyncio.create_task(send_approval_notification("transaction", {
            "reference": tx_doc["reference"],
            "type": req["transaction_type"],
            "client": req.get("client_name", "Unknown"),
            "amount": usd_amount,
            "base_amount": base_amount,
            "base_currency": base_currency,
            "destination": vendor_info["vendor_name"] if vendor_info else req.get("destination_type", "-"),
            "created_by": user["name"]
        }))
    
    await log_activity(request, user, "approve", "transaction_requests", f"Processed request → TX {tx_id}")
    
    return {"message": "Request processed", "transaction_id": tx_id, "request_id": request_id}

@api_router.delete("/transaction-requests/{request_id}")
async def delete_transaction_request(request: Request, request_id: str, user: dict = Depends(require_permission(Modules.TRANSACTION_REQUESTS, Actions.DELETE))):
    req = await db.transaction_requests.find_one({"request_id": request_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    if req["status"] == "processed":
        raise HTTPException(status_code=400, detail="Cannot delete processed requests")
    await db.transaction_requests.delete_one({"request_id": request_id})
    await log_activity(request, user, "delete", "transaction_requests", "Deleted request")
    return {"message": "Request deleted"}


# ============== REPORTS/ANALYTICS ROUTES ==============

@api_router.get("/reports/dashboard")
async def get_dashboard_stats(user: dict = Depends(require_permission(Modules.DASHBOARD, Actions.VIEW))):
    # Get client stats
    total_clients = await db.clients.count_documents({})
    approved_clients = await db.clients.count_documents({"kyc_status": ClientStatus.APPROVED})
    pending_clients = await db.clients.count_documents({"kyc_status": ClientStatus.PENDING})
    
    # Get treasury stats
    total_treasury = await db.treasury_accounts.count_documents({})
    active_treasury = await db.treasury_accounts.count_documents({"status": TreasuryAccountStatus.ACTIVE})
    
    # Get total treasury balance using manual FX rates
    all_accounts = await db.treasury_accounts.find({"status": TreasuryAccountStatus.ACTIVE}, {"_id": 0}).to_list(1000)
    fx_settings = await db.app_settings.find_one({"setting_type": "manual_fx_rates"}, {"_id": 0})
    manual_rates = fx_settings.get("rates", {}) if fx_settings else {}
    total_balance_usd = 0
    for acc in all_accounts:
        currency = acc.get("currency", "USD")
        balance = acc.get("balance", 0)
        if currency == "USD":
            total_balance_usd += balance
        elif currency in manual_rates and manual_rates[currency] > 0:
            total_balance_usd += balance * manual_rates[currency]
    
    # Get transaction stats
    total_transactions = await db.transactions.count_documents({})
    pending_transactions = await db.transactions.count_documents({"status": TransactionStatus.PENDING})
    
    # Get deposits/withdrawals totals
    deposit_pipeline = [
        {"$match": {"transaction_type": TransactionType.DEPOSIT, "status": {"$in": [TransactionStatus.APPROVED, TransactionStatus.COMPLETED]}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]
    withdrawal_pipeline = [
        {"$match": {"transaction_type": TransactionType.WITHDRAWAL, "status": {"$in": [TransactionStatus.APPROVED, TransactionStatus.COMPLETED]}}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]
    
    deposit_result = await db.transactions.aggregate(deposit_pipeline).to_list(1)
    withdrawal_result = await db.transactions.aggregate(withdrawal_pipeline).to_list(1)
    
    total_deposits = deposit_result[0]["total"] if deposit_result else 0
    total_withdrawals = withdrawal_result[0]["total"] if withdrawal_result else 0
    
    return {
        "clients": {
            "total": total_clients,
            "approved": approved_clients,
            "pending": pending_clients
        },
        "treasury": {
            "total": total_treasury,
            "active": active_treasury,
            "total_balance": total_balance_usd
        },
        "transactions": {
            "total": total_transactions,
            "pending": pending_transactions,
            "total_deposits": total_deposits,
            "total_withdrawals": total_withdrawals
        }
    }

@api_router.get("/reports/transactions-summary")
async def get_transactions_summary(
    user: dict = Depends(require_permission(Modules.REPORTS, Actions.VIEW)),
    days: int = 30
):
    from_date = datetime.now(timezone.utc) - timedelta(days=days)
    
    pipeline = [
        {"$addFields": {
            "created_date": {"$dateFromString": {"dateString": "$created_at"}}
        }},
        {"$match": {"created_date": {"$gte": from_date}}},
        {"$group": {
            "_id": {
                "date": {"$dateToString": {"format": "%Y-%m-%d", "date": "$created_date"}},
                "type": "$transaction_type"
            },
            "count": {"$sum": 1},
            "total_amount": {"$sum": "$amount"}
        }},
        {"$sort": {"_id.date": 1}}
    ]
    
    results = await db.transactions.aggregate(pipeline).to_list(1000)
    
    summary = {}
    for r in results:
        date = r["_id"]["date"]
        tx_type = r["_id"]["type"]
        if date not in summary:
            summary[date] = {"date": date, "deposits": 0, "withdrawals": 0}
        if tx_type == TransactionType.DEPOSIT:
            summary[date]["deposits"] = r["total_amount"]
        elif tx_type == TransactionType.WITHDRAWAL:
            summary[date]["withdrawals"] = r["total_amount"]
    
    return list(summary.values())

@api_router.get("/reports/client-analytics")
async def get_client_analytics(user: dict = Depends(require_permission(Modules.REPORTS, Actions.VIEW))):
    # KYC status distribution
    kyc_pipeline = [
        {"$group": {"_id": "$kyc_status", "count": {"$sum": 1}}}
    ]
    kyc_stats = await db.clients.aggregate(kyc_pipeline).to_list(10)
    
    # Country distribution
    country_pipeline = [
        {"$match": {"country": {"$ne": None}}},
        {"$group": {"_id": "$country", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    country_stats = await db.clients.aggregate(country_pipeline).to_list(10)
    
    return {
        "kyc_distribution": [{"status": s["_id"], "count": s["count"]} for s in kyc_stats],
        "country_distribution": [{"country": c["_id"], "count": c["count"]} for c in country_stats]
    }

@api_router.get("/reports/recent-activity")
async def get_recent_activity(user: dict = Depends(require_permission(Modules.REPORTS, Actions.VIEW)), limit: int = 10):
    transactions = await db.transactions.find({}, {"_id": 0}).sort("created_at", -1).to_list(limit)
    clients = await db.clients.find({}, {"_id": 0}).sort("created_at", -1).to_list(limit)
    
    return {
        "recent_transactions": transactions,
        "recent_clients": clients
    }

# ============== SEED DATA ROUTE ==============

@api_router.post("/seed")
async def seed_demo_data():
    """Seed demo data for testing"""
    now = datetime.now(timezone.utc)
    
    # Create admin user if not exists
    admin = await db.users.find_one({"email": "admin@fxbroker.com"}, {"_id": 0})
    if not admin:
        admin_doc = {
            "user_id": f"user_{uuid.uuid4().hex[:12]}",
            "email": "admin@fxbroker.com",
            "password_hash": hash_password("admin123"),
            "name": "System Admin",
            "role": UserRole.ADMIN,
            "picture": None,
            "is_active": True,
            "created_at": now.isoformat()
        }
        await db.users.insert_one(admin_doc)
    
    # Create accountant user if not exists
    accountant = await db.users.find_one({"email": "accountant@fxbroker.com"}, {"_id": 0})
    if not accountant:
        accountant_doc = {
            "user_id": f"user_{uuid.uuid4().hex[:12]}",
            "email": "accountant@fxbroker.com",
            "password_hash": hash_password("accountant123"),
            "name": "Finance Manager",
            "role": UserRole.ACCOUNTANT,
            "picture": None,
            "is_active": True,
            "created_at": now.isoformat()
        }
        await db.users.insert_one(accountant_doc)
    
    # Create treasury accounts
    treasury_accounts_data = [
        {"account_name": "Main Operating Account", "account_type": "bank", "bank_name": "Chase Bank", "account_number": "****1234", "currency": "USD"},
        {"account_name": "Client Funds Account", "account_type": "bank", "bank_name": "Bank of America", "account_number": "****5678", "currency": "USD"},
        {"account_name": "EUR Account", "account_type": "bank", "bank_name": "Deutsche Bank", "account_number": "****9012", "currency": "EUR"},
        {"account_name": "AED Account", "account_type": "bank", "bank_name": "Emirates NBD", "account_number": "****3456", "currency": "AED"},
    ]
    
    treasury_ids = []
    for ta in treasury_accounts_data:
        existing = await db.treasury_accounts.find_one({"account_name": ta["account_name"]}, {"_id": 0})
        if not existing:
            account_id = f"treasury_{uuid.uuid4().hex[:12]}"
            ta_doc = {
                "account_id": account_id,
                **ta,
                "balance": 50000.0,
                "status": TreasuryAccountStatus.ACTIVE,
                "created_at": now.isoformat(),
                "updated_at": now.isoformat()
            }
            await db.treasury_accounts.insert_one(ta_doc)
            treasury_ids.append(account_id)
        else:
            treasury_ids.append(existing["account_id"])
    
    # Create sample clients
    sample_clients = [
        {"first_name": "John", "last_name": "Smith", "email": "john.smith@email.com", "phone": "+1234567890", "country": "USA", "kyc_status": ClientStatus.APPROVED},
        {"first_name": "Emma", "last_name": "Wilson", "email": "emma.wilson@email.com", "phone": "+4477889900", "country": "UK", "kyc_status": ClientStatus.APPROVED},
        {"first_name": "Michael", "last_name": "Brown", "email": "michael.brown@email.com", "phone": "+6199887766", "country": "Australia", "kyc_status": ClientStatus.PENDING},
        {"first_name": "Sarah", "last_name": "Davis", "email": "sarah.davis@email.com", "phone": "+4912345678", "country": "Germany", "kyc_status": ClientStatus.APPROVED},
        {"first_name": "James", "last_name": "Miller", "email": "james.miller@email.com", "phone": "+8187654321", "country": "Japan", "kyc_status": ClientStatus.PENDING},
    ]
    
    client_ids = []
    for c in sample_clients:
        existing = await db.clients.find_one({"email": c["email"]}, {"_id": 0})
        if not existing:
            client_id = f"client_{uuid.uuid4().hex[:12]}"
            client_doc = {
                "client_id": client_id,
                **c,
                "kyc_documents": [],
                "notes": None,
                "created_at": now.isoformat(),
                "updated_at": now.isoformat()
            }
            await db.clients.insert_one(client_doc)
            client_ids.append(client_id)
        else:
            client_ids.append(existing["client_id"])
    
    # Create sample transactions
    for i, client_id in enumerate(client_ids[:3]):
        client = await db.clients.find_one({"client_id": client_id}, {"_id": 0})
        if client:
            tx_id = f"tx_{uuid.uuid4().hex[:12]}"
            tx_doc = {
                "transaction_id": tx_id,
                "client_id": client_id,
                "client_name": f"{client['first_name']} {client['last_name']}",
                "transaction_type": TransactionType.DEPOSIT,
                "amount": [5000, 10000, 3000][i],
                "currency": "USD",
                "base_currency": "USD",
                "base_amount": None,
                "destination_type": "treasury",
                "destination_account_id": treasury_ids[0] if treasury_ids else None,
                "destination_account_name": "Main Operating Account",
                "destination_bank_name": "Chase Bank",
                "psp_id": None,
                "psp_name": None,
                "psp_commission_rate": None,
                "status": TransactionStatus.PENDING,
                "description": "Initial deposit",
                "reference": f"DEP{uuid.uuid4().hex[:8].upper()}",
                "proof_image": None,
                "created_by": None,
                "created_by_name": "System",
                "processed_by": None,
                "processed_by_name": None,
                "rejection_reason": None,
                "settled": False,
                "settlement_id": None,
                "settlement_status": None,
                "created_at": (now - timedelta(days=i+1)).isoformat(),
                "processed_at": None
            }
            await db.transactions.insert_one(tx_doc)
    
    # Create sample PSPs
    psp_data = [
        {"psp_name": "Stripe", "commission_rate": 2.9, "settlement_days": 2, "description": "Credit card payments"},
        {"psp_name": "PayPal", "commission_rate": 3.5, "settlement_days": 3, "description": "PayPal payments"},
        {"psp_name": "Skrill", "commission_rate": 2.5, "settlement_days": 1, "description": "E-wallet payments"},
    ]
    
    for psp in psp_data:
        existing = await db.psps.find_one({"psp_name": psp["psp_name"]}, {"_id": 0})
        if not existing and treasury_ids:
            psp_id = f"psp_{uuid.uuid4().hex[:12]}"
            psp_doc = {
                "psp_id": psp_id,
                "psp_name": psp["psp_name"],
                "commission_rate": psp["commission_rate"],
                "settlement_days": psp["settlement_days"],
                "settlement_destination_id": treasury_ids[0],  # Main Operating Account
                "min_settlement_amount": 100,
                "description": psp["description"],
                "total_volume": 0.0,
                "total_commission": 0.0,
                "pending_settlement": 0.0,
                "status": PSPStatus.ACTIVE,
                "created_at": now.isoformat(),
                "updated_at": now.isoformat()
            }
            await db.psps.insert_one(psp_doc)
    
    # Create sample vendors
    vendor_data = [
        {"vendor_name": "MoneyExchange Pro", "email": "vendor1@fxbroker.com", "password": "vendor123", "deposit_commission": 1.5, "withdrawal_commission": 2.0, "bank_settlement_commission": 0.5, "cash_settlement_commission": 1.0, "description": "Premium exchange vendor"},
        {"vendor_name": "FastPay Solutions", "email": "vendor2@fxbroker.com", "password": "vendor123", "deposit_commission": 2.0, "withdrawal_commission": 2.5, "bank_settlement_commission": 0.3, "cash_settlement_commission": 0.8, "description": "Fast payment processing"},
    ]
    
    for v in vendor_data:
        existing = await db.vendors.find_one({"vendor_name": v["vendor_name"]}, {"_id": 0})
        if not existing and treasury_ids:
            vendor_id = f"vendor_{uuid.uuid4().hex[:12]}"
            user_id = f"user_{uuid.uuid4().hex[:12]}"
            
            # Create user account for vendor
            user_doc = {
                "user_id": user_id,
                "email": v["email"],
                "password_hash": hash_password(v["password"]),
                "name": v["vendor_name"],
                "role": UserRole.VENDOR,
                "vendor_id": vendor_id,
                "picture": None,
                "is_active": True,
                "created_at": now.isoformat()
            }
            await db.users.insert_one(user_doc)
            
            vendor_doc = {
                "vendor_id": vendor_id,
                "user_id": user_id,
                "vendor_name": v["vendor_name"],
                "email": v["email"],
                "deposit_commission": v["deposit_commission"],
                "withdrawal_commission": v["withdrawal_commission"],
                "bank_settlement_commission": v["bank_settlement_commission"],
                "cash_settlement_commission": v["cash_settlement_commission"],
                "settlement_destination_id": treasury_ids[0],
                "description": v["description"],
                "total_volume": 0.0,
                "total_commission": 0.0,
                "pending_settlement": 0.0,
                "status": VendorStatus.ACTIVE,
                "created_at": now.isoformat(),
                "updated_at": now.isoformat()
            }
            await db.vendors.insert_one(vendor_doc)
    
    return {"message": "Demo data seeded successfully"}

# ============== INCOME & EXPENSES ROUTES ==============

@api_router.get("/income-expenses")
async def get_income_expenses(
    entry_type: Optional[str] = None,
    category: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    treasury_account_id: Optional[str] = None,
    vendor_id: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.VIEW))
):
    """Get all income and expense entries with optional filters and pagination"""
    query = {}
    
    if entry_type:
        query["entry_type"] = entry_type
    if category:
        query["category"] = category
    if treasury_account_id:
        query["treasury_account_id"] = treasury_account_id
    if vendor_id:
        query["vendor_id"] = vendor_id
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    # Check cache
    cache_key = get_cache_key("ie:list", page=page, page_size=page_size, 
                              entry_type=entry_type, category=category, 
                              start_date=start_date, end_date=end_date,
                              treasury_account_id=treasury_account_id, vendor_id=vendor_id)
    cached = get_cached(cache_key)
    if cached:
        return cached
    
    # Paginated query
    skip = (page - 1) * page_size
    total = await db.income_expenses.count_documents(query)
    # Sort by date descending, then by entry_id descending for stable pagination
    entries = await db.income_expenses.find(query, {"_id": 0}).sort([("date", -1), ("entry_id", -1)]).skip(skip).limit(page_size).to_list(page_size)
    
    # Batch fetch treasury accounts to avoid N+1 queries
    treasury_ids = list(set(e.get("treasury_account_id") for e in entries if e.get("treasury_account_id")))
    treasury_map = {}
    if treasury_ids:
        treasuries = await db.treasury_accounts.find({"account_id": {"$in": treasury_ids}}, {"_id": 0}).to_list(len(treasury_ids))
        treasury_map = {t["account_id"]: t["account_name"] for t in treasuries}
    
    for entry in entries:
        if entry.get("treasury_account_id"):
            entry["treasury_account_name"] = treasury_map.get(entry["treasury_account_id"], "Unknown")
    
    response = {
        "items": entries,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size if total > 0 else 1
    }
    
    # Cache response
    set_cached(cache_key, response, CACHE_TTL['income_expenses'])
    
    return response

@api_router.get("/income-expenses/reports/summary")
async def get_income_expense_summary(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.VIEW))
):
    """Get income vs expense summary report"""
    query = {}
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    # Get all entries (exclude converted to loan to avoid double-counting)
    entries = await db.income_expenses.find({**query, "converted_to_loan": {"$ne": True}}, {"_id": 0}).to_list(10000)
    
    # Use .get() with fallback to handle entries without amount_usd
    total_income = sum(e.get("amount_usd", convert_to_usd(e.get("amount", 0), e.get("currency", "USD"))) for e in entries if e["entry_type"] == IncomeExpenseType.INCOME)
    total_expense = sum(e.get("amount_usd", convert_to_usd(e.get("amount", 0), e.get("currency", "USD"))) for e in entries if e["entry_type"] == IncomeExpenseType.EXPENSE)
    net_profit = total_income - total_expense
    
    # Category breakdown
    income_by_category = {}
    expense_by_category = {}
    
    for entry in entries:
        cat = entry.get("custom_category") or entry.get("category", "other")
        amount = entry.get("amount_usd", convert_to_usd(entry.get("amount", 0), entry.get("currency", "USD")))
        if entry["entry_type"] == IncomeExpenseType.INCOME:
            income_by_category[cat] = income_by_category.get(cat, 0) + amount
        else:
            expense_by_category[cat] = expense_by_category.get(cat, 0) + amount
    
    return {
        "total_income_usd": round(total_income, 2),
        "total_expense_usd": round(total_expense, 2),
        "net_profit_usd": round(net_profit, 2),
        "income_by_category": {k: round(v, 2) for k, v in income_by_category.items()},
        "expense_by_category": {k: round(v, 2) for k, v in expense_by_category.items()},
        "entry_count": len(entries)
    }

@api_router.get("/income-expenses/reports/monthly")
async def get_monthly_report(
    year: Optional[int] = None,
    user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.VIEW))
):
    """Get monthly P&L report"""
    if not year:
        year = datetime.now().year
    
    # Get entries for the year
    start_date = f"{year}-01-01"
    end_date = f"{year}-12-31"
    
    entries = await db.income_expenses.find({
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(10000)
    
    # Group by month
    monthly_data = {}
    for month in range(1, 13):
        month_str = f"{year}-{month:02d}"
        monthly_data[month_str] = {"income": 0, "expense": 0}
    
    for entry in entries:
        month_str = entry.get("date", "")[:7]  # YYYY-MM
        if month_str in monthly_data:
            amt_usd = entry.get("amount_usd") or convert_to_usd(entry.get("amount", 0), entry.get("currency", "USD"))
            if entry["entry_type"] == IncomeExpenseType.INCOME:
                monthly_data[month_str]["income"] += amt_usd
            else:
                monthly_data[month_str]["expense"] += amt_usd
    
    # Calculate net for each month
    result = []
    for month, data in monthly_data.items():
        result.append({
            "month": month,
            "income": round(data["income"], 2),
            "expense": round(data["expense"], 2),
            "net": round(data["income"] - data["expense"], 2)
        })
    
    return result

@api_router.get("/income-expenses/categories")
async def get_categories(user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.VIEW))):
    """Get available categories and custom categories"""
    income_categories = [
        {"value": "commission", "label": "Commission Income"},
        {"value": "service_fee", "label": "Service Fees"},
        {"value": "interest", "label": "Interest Income"},
        {"value": "other", "label": "Other Income"},
    ]
    
    expense_categories = [
        {"value": "bank_fee", "label": "Bank Fees"},
        {"value": "transfer_charge", "label": "Transfer Charges"},
        {"value": "vendor_payment", "label": "Vendor Payments"},
        {"value": "operational", "label": "Operational Costs"},
        {"value": "marketing", "label": "Marketing"},
        {"value": "software", "label": "Software/Subscriptions"},
        {"value": "other", "label": "Other Expenses"},
    ]
    
    # Get custom categories from existing entries
    custom_income = await db.income_expenses.distinct("custom_category", {"entry_type": "income", "custom_category": {"$ne": None}})
    custom_expense = await db.income_expenses.distinct("custom_category", {"entry_type": "expense", "custom_category": {"$ne": None}})
    
    return {
        "income_categories": income_categories,
        "expense_categories": expense_categories,
        "custom_income_categories": custom_income,
        "custom_expense_categories": custom_expense
    }

@api_router.get("/income-expenses/export-template")
async def get_ie_import_template_route(user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.EXPORT))):
    """Download Excel template for bulk importing income/expense entries"""
    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Income_Expenses"
    
    # Headers
    headers = ["Entry Type", "Category", "Amount", "Currency", "Date", "Description", "Reference"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Example rows
    examples = [
        ["income", "commission", 1000, "USD", "2025-01-15", "Monthly commission", "INV-001"],
        ["expense", "bank_fee", 50, "USD", "2025-01-15", "Wire transfer fee", "REF-002"],
        ["income", "service_fee", 500, "EUR", "2025-01-16", "Service charge", "INV-003"],
        ["expense", "operational", 200, "USD", "2025-01-17", "Office supplies", "REF-004"],
    ]
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=income_expenses_template.xlsx"}
    )

@api_router.post("/income-expenses/bulk-import")
async def bulk_import_ie_entries_route(

    request: Request,

    user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.CREATE)),
    file: UploadFile = File(...),
    treasury_account_id: str = Form(...)
):
    """
    Bulk import income/expense entries from Excel file.
    Excel columns expected: entry_type, category, amount, currency, date, description, reference
    """
    import openpyxl
    from io import BytesIO
    
    # Verify treasury account
    treasury = await db.treasury_accounts.find_one({"account_id": treasury_account_id}, {"_id": 0})
    if not treasury:
        raise HTTPException(status_code=404, detail="Treasury account not found")
    
    # Read Excel file
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")
    
    # Get headers from first row
    headers = [str(cell.value).lower().strip() if cell.value else "" for cell in ws[1]]
    
    # Map expected columns
    col_map = {}
    expected = ["entry_type", "type", "category", "amount", "currency", "date", "description", "reference"]
    for idx, h in enumerate(headers):
        for exp in expected:
            if exp in h:
                col_map[exp] = idx
                break
    
    if "amount" not in col_map:
        raise HTTPException(status_code=400, detail="Excel file must have an 'amount' column")
    
    now = datetime.now(timezone.utc)
    imported = 0
    errors = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # Get values
            amount_val = row[col_map.get("amount", 0)]
            if not amount_val or float(amount_val) <= 0:
                continue
            
            amount = float(amount_val)
            entry_type_val = row[col_map.get("entry_type", col_map.get("type", 0))] if "entry_type" in col_map or "type" in col_map else "expense"
            entry_type = "income" if str(entry_type_val).lower().strip() in ["income", "in", "credit", "cr"] else "expense"
            category = str(row[col_map.get("category", 0)] or "other").lower().replace(" ", "_")
            currency = str(row[col_map.get("currency", 0)] or "USD").upper()
            date_val = row[col_map.get("date", 0)]
            description = str(row[col_map.get("description", 0)] or "")
            reference = str(row[col_map.get("reference", 0)] or "")
            
            # Parse date
            if date_val:
                if isinstance(date_val, datetime):
                    entry_date = date_val.strftime("%Y-%m-%d")
                else:
                    entry_date = str(date_val)[:10]
            else:
                entry_date = now.strftime("%Y-%m-%d")
            
            entry_id = f"ie_{uuid.uuid4().hex[:12]}"
            amount_usd = convert_to_usd(amount, currency)
            
            entry_doc = {
                "entry_id": entry_id,
                "entry_type": entry_type,
                "category": category,
                "custom_category": None,
                "amount": amount,
                "currency": currency,
                "amount_usd": amount_usd,
                "treasury_account_id": treasury_account_id,
                "treasury_account_name": treasury["account_name"],
                "vendor_id": None,
                "vendor_name": None,
                "vendor_supplier_id": None,
                "vendor_supplier_name": None,
                "client_id": None,
                "client_name": None,
                "ie_category_id": None,
                "ie_category_name": None,
                "description": description if description != "None" else "",
                "reference": reference if reference != "None" else "",
                "date": entry_date,
                "status": "completed",
                "converted_to_loan": False,
                "loan_id": None,
                "vendor_proof_image": None,
                "invoice_file": None,
                "created_at": now.isoformat(),
                "created_by": user["user_id"],
                "created_by_name": user["name"],
                "imported_from": file.filename
            }
            
            await db.income_expenses.insert_one(entry_doc)
            
            # Update treasury balance
            if entry_type == "income":
                await db.treasury_accounts.update_one(
                    {"account_id": treasury_account_id},
                    {"$inc": {"balance": amount}, "$set": {"updated_at": now.isoformat()}}
                )
            else:
                await db.treasury_accounts.update_one(
                    {"account_id": treasury_account_id},
                    {"$inc": {"balance": -amount}, "$set": {"updated_at": now.isoformat()}}
                )
            
            imported += 1
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
    
    await log_activity(request, user, "create", "income_expenses", "Bulk imported I&E entries")

    return {
        "message": f"Import completed",
        "imported": imported,
        "errors": errors[:10] if errors else []
    }

@api_router.get("/income-expenses/{entry_id}")
async def get_income_expense(entry_id: str, user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.VIEW))):
    """Get a single income/expense entry"""
    entry = await db.income_expenses.find_one({"entry_id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    return entry

@api_router.post("/income-expenses")
async def create_income_expense(entry_data: IncomeExpenseCreate, request: Request, user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.CREATE))):
    """Create a new income or expense entry, optionally linked to a vendor/supplier/client for approval"""
    if entry_data.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be positive")
    
    if not entry_data.category and not entry_data.ie_category_id:
        raise HTTPException(status_code=400, detail="Either a category or custom category is required")
    
    if not entry_data.vendor_id and not entry_data.treasury_account_id:
        raise HTTPException(status_code=400, detail="Either an exchanger or treasury account is required")
    
    # Exchanger-linked entry: auto-create pending approval in vendor portal
    vendor_info = None
    if entry_data.vendor_id:
        vendor_info = await db.vendors.find_one({"vendor_id": entry_data.vendor_id}, {"_id": 0})
        if not vendor_info:
            raise HTTPException(status_code=404, detail="Exchanger not found")
    
    # Vendor Supplier (service supplier like rent, utilities) - lookup for name
    vendor_supplier_info = None
    if entry_data.vendor_supplier_id:
        vendor_supplier_info = await db.vendor_suppliers.find_one({"supplier_id": entry_data.vendor_supplier_id}, {"_id": 0})
        if not vendor_supplier_info:
            raise HTTPException(status_code=404, detail="Vendor supplier not found")
    
    # Client - lookup for name
    client_info = None
    if entry_data.client_id:
        client_info = await db.clients.find_one({"client_id": entry_data.client_id}, {"_id": 0})
        if not client_info:
            raise HTTPException(status_code=404, detail="Client not found")
    
    # IE Category - lookup for name
    ie_category_info = None
    if entry_data.ie_category_id:
        ie_category_info = await db.ie_categories.find_one({"category_id": entry_data.ie_category_id}, {"_id": 0})
        if not ie_category_info:
            raise HTTPException(status_code=404, detail="Category not found")
    
    treasury = None
    if entry_data.treasury_account_id and not entry_data.vendor_id:
        treasury = await db.treasury_accounts.find_one({"account_id": entry_data.treasury_account_id}, {"_id": 0})
        if not treasury:
            raise HTTPException(status_code=404, detail="Treasury account not found")
    
    entry_id = f"ie_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    entry_date = entry_data.date if entry_data.date else now.isoformat()[:10]
    
    # Calculate USD equivalent
    amount_usd = convert_to_usd(entry_data.amount, entry_data.currency)
    
    # Determine initial status: always pending for approval
    status = "pending"
    
    # Calculate vendor commission at creation time
    ie_commission_rate = 0.0
    ie_commission_amount = 0.0
    ie_commission_base_amount = 0.0
    if vendor_info:
        tx_mode = entry_data.transaction_mode or "bank"
        if entry_data.entry_type == "income":
            if tx_mode == "cash":
                ie_commission_rate = vendor_info.get("deposit_commission_cash", vendor_info.get("deposit_commission", 0))
            else:
                ie_commission_rate = vendor_info.get("deposit_commission", 0)
        else:
            if tx_mode == "cash":
                ie_commission_rate = vendor_info.get("withdrawal_commission_cash", vendor_info.get("withdrawal_commission", 0))
            else:
                ie_commission_rate = vendor_info.get("withdrawal_commission", 0)
        if ie_commission_rate > 0:
            # Commission calculated on BASE currency amount (the actual transaction currency)
            # If base_amount exists, use it (actual INR amount), otherwise use amount
            actual_tx_amount = entry_data.base_amount if entry_data.base_amount else entry_data.amount
            actual_tx_currency = entry_data.base_currency if entry_data.base_currency else entry_data.currency
            ie_commission_base_amount = round(actual_tx_amount * ie_commission_rate / 100, 2)
            ie_commission_base_currency = actual_tx_currency
            # USD commission (separate from base currency commission)
            ie_commission_amount = round(amount_usd * ie_commission_rate / 100, 2)

    entry_doc = {
        "entry_id": entry_id,
        "entry_type": entry_data.entry_type,
        "category": entry_data.category,
        "custom_category": entry_data.custom_category,
        "amount": entry_data.amount,
        "currency": entry_data.currency,
        "base_currency": entry_data.base_currency,
        "base_amount": entry_data.base_amount,
        "exchange_rate": entry_data.exchange_rate,
        "amount_usd": amount_usd,
        "treasury_account_id": entry_data.treasury_account_id,
        "treasury_account_name": treasury["account_name"] if treasury else None,
        "vendor_id": entry_data.vendor_id,
        "vendor_name": vendor_info["vendor_name"] if vendor_info else None,
        "vendor_supplier_id": entry_data.vendor_supplier_id,
        "vendor_supplier_name": vendor_supplier_info["name"] if vendor_supplier_info else None,
        "client_id": entry_data.client_id,
        "client_name": f"{client_info['first_name']} {client_info['last_name']}" if client_info else None,
        "ie_category_id": entry_data.ie_category_id,
        "ie_category_name": ie_category_info["name"] if ie_category_info else None,
        "vendor_bank_account_name": entry_data.vendor_bank_account_name,
        "vendor_bank_account_number": entry_data.vendor_bank_account_number,
        "vendor_bank_ifsc": entry_data.vendor_bank_ifsc,
        "vendor_bank_branch": entry_data.vendor_bank_branch,
        "description": entry_data.description,
        "reference": entry_data.reference,
        "date": entry_date,
        "transaction_mode": entry_data.transaction_mode or "bank",
        "collecting_person_name": entry_data.collecting_person_name if entry_data.transaction_mode == "cash" else None,
        "collecting_person_number": entry_data.collecting_person_number if entry_data.transaction_mode == "cash" else None,
        "vendor_commission_rate": ie_commission_rate if ie_commission_rate > 0 else None,
        "vendor_commission_amount": ie_commission_amount if ie_commission_amount > 0 else None,
        "vendor_commission_base_amount": ie_commission_base_amount if ie_commission_base_amount > 0 else None,
        "vendor_commission_base_currency": ie_commission_base_currency if ie_commission_base_amount > 0 else None,
        "status": status,
        "converted_to_loan": False,
        "loan_id": None,
        "vendor_proof_image": None,
        "created_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    
    await db.income_expenses.insert_one(entry_doc)
    
    # Treasury update is deferred to approval step
    
    entry_doc.pop("_id", None)
    if treasury:
        entry_doc["treasury_account_name"] = treasury["account_name"]
    
    # Log activity
    await log_activity(request, user, "create", "income_expenses", f"Created {entry_data.entry_type}: {entry_data.description or entry_data.category} (pending approval)", reference_id=entry_id)
    
    # Notify exchanger if assigned
    if vendor_info and entry_data.vendor_id:
        import asyncio
        amt_display = f"{entry_data.base_amount:,.2f} {entry_data.base_currency}" if entry_data.base_currency and entry_data.base_amount else f"{entry_data.amount:,.2f} {entry_data.currency}"
        asyncio.create_task(send_exchanger_notification("ie", entry_data.vendor_id, {
            "entry_type": entry_data.entry_type,
            "category": (entry_data.category or "").replace("_", " "),
            "amount_display": amt_display,
        }))
    
    return entry_doc

@api_router.put("/income-expenses/{entry_id}")
async def update_income_expense(request: Request, entry_id: str, update_data: IncomeExpenseUpdate, user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.EDIT))):

    """Update an income/expense entry (only description, reference, category - not amount)"""
    entry = await db.income_expenses.find_one({"entry_id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if not update_dict:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    # Don't allow amount changes (would need to reverse treasury changes)
    if "amount" in update_dict:
        del update_dict["amount"]
    
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.income_expenses.update_one({"entry_id": entry_id}, {"$set": update_dict})
    
    updated = await db.income_expenses.find_one({"entry_id": entry_id}, {"_id": 0})
    await log_activity(request, user, "edit", "income_expenses", "Updated I&E entry")

    return updated

@api_router.delete("/income-expenses/{entry_id}")
async def delete_income_expense(entry_id: str, request: Request, user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.DELETE))):
    """Delete an income/expense entry and reverse treasury balance if approved"""
    entry = await db.income_expenses.find_one({"entry_id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    now = datetime.now(timezone.utc)
    
    # Only reverse treasury balance if entry was approved (treasury was actually affected)
    if entry.get("status") == "approved":
        if entry["entry_type"] == IncomeExpenseType.INCOME:
            await db.treasury_accounts.update_one(
                {"account_id": entry["treasury_account_id"]},
                {"$inc": {"balance": -entry["amount"]}, "$set": {"updated_at": now.isoformat()}}
            )
        else:
            await db.treasury_accounts.update_one(
                {"account_id": entry["treasury_account_id"]},
                {"$inc": {"balance": entry["amount"]}, "$set": {"updated_at": now.isoformat()}}
            )
        # Delete related treasury transaction
        await db.treasury_transactions.delete_one({"income_expense_id": entry_id})
    
    # Delete the entry
    await db.income_expenses.delete_one({"entry_id": entry_id})
    
    # Log activity
    await log_activity(request, user, "delete", "income_expenses", f"Deleted {entry['entry_type']}: {entry.get('description', 'N/A')}", reference_id=entry_id)
    
    return {"message": "Entry deleted successfully"}

class ConvertToLoanRequest(BaseModel):
    borrower_name: str
    interest_rate: float = 0
    due_date: str
    treasury_account_id: Optional[str] = None  # Optional - will use expense's treasury if not provided
    notes: Optional[str] = None

@api_router.get("/loans/borrowers")
async def get_loan_borrowers(user: dict = Depends(require_permission(Modules.LOANS, Actions.VIEW))):
    """Get list of borrower companies (vendors) plus unique borrower names from existing loans"""
    # Get all vendor companies as potential borrowers
    vendors = await db.vendors.find({"status": "active"}, {"_id": 0, "vendor_id": 1, "vendor_name": 1, "email": 1}).to_list(length=1000)
    
    # Also get unique borrower names from existing loans (for historical entries)
    loan_borrowers = await db.loans.distinct("borrower_name")
    
    # Combine into borrowers list
    borrowers = []
    seen_names = set()
    
    # Add vendors as borrower companies
    for v in vendors:
        name = v.get("vendor_name") or v.get("name")
        if name and name not in seen_names:
            borrowers.append({
                "name": name,
                "vendor_id": v.get("vendor_id"),
                "email": v.get("email"),
                "type": "company"
            })
            seen_names.add(name)
    
    # Add historical borrower names from loans (if not already in vendor list)
    for b in loan_borrowers:
        if b and b not in seen_names:
            borrowers.append({
                "name": b,
                "vendor_id": None,
                "email": None,
                "type": "historical"
            })
            seen_names.add(b)
    
    return {"borrowers": sorted(borrowers, key=lambda x: x["name"])}

@api_router.post("/income-expenses/{entry_id}/convert-to-loan")
async def convert_expense_to_loan(request: Request, entry_id: str, req: ConvertToLoanRequest, user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.EDIT))):

    """Convert an expense entry to a loan (marks expense as converted, creates loan)"""
    entry = await db.income_expenses.find_one({"entry_id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    if entry["entry_type"] != "expense":
        raise HTTPException(status_code=400, detail="Only expenses can be converted to loans")
    if entry.get("converted_to_loan"):
        raise HTTPException(status_code=400, detail="This expense has already been converted to a loan")
    
    now = datetime.now(timezone.utc)
    loan_id = f"loan_{uuid.uuid4().hex[:12]}"
    
    # Use expense's treasury account if not provided (since money was already taken from vendor)
    treasury_account_id = req.treasury_account_id or entry.get("treasury_account_id")
    
    # Get treasury info for the loan record
    treasury = None
    treasury_name = "N/A"
    if treasury_account_id:
        treasury = await db.treasury_accounts.find_one({"account_id": treasury_account_id}, {"_id": 0})
        if treasury:
            treasury_name = treasury["account_name"]
    
    loan_doc = {
        "loan_id": loan_id,
        "borrower_name": req.borrower_name,
        "amount": entry["amount"],
        "currency": entry.get("currency", "USD"),
        "interest_rate": req.interest_rate,
        "loan_date": entry.get("date", now.isoformat()[:10]),
        "due_date": req.due_date,
        "repayment_mode": "lump_sum",
        "treasury_account_id": treasury_account_id,
        "source_treasury_name": treasury_name,
        "outstanding_balance": entry["amount"],
        "total_repaid": 0,
        "total_interest": 0,
        "status": "active",
        "repayment_count": 0,
        "converted_from_expense": entry_id,
        "notes": req.notes or f"Converted from expense: {entry.get('description', '')}",
        "created_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    
    # Only create treasury disbursement transaction if we have a valid treasury
    if treasury_account_id and treasury:
        # Deduct from treasury for loan disbursement
        await db.treasury_accounts.update_one(
            {"account_id": treasury_account_id},
            {"$inc": {"balance": -entry["amount"]}, "$set": {"updated_at": now.isoformat()}}
        )
        
        # Create treasury transaction for loan
        await db.treasury_transactions.insert_one({
            "treasury_transaction_id": f"ttx_{uuid.uuid4().hex[:12]}",
            "account_id": treasury_account_id,
            "account_name": treasury_name,
            "transaction_type": "loan_disbursement",
            "amount": -entry["amount"],
            "currency": entry.get("currency", "USD"),
            "reference": f"Loan to {req.borrower_name} (converted from expense {entry_id})",
            "loan_id": loan_id,
            "created_at": now.isoformat(),
            "created_by": user["user_id"],
            "created_by_name": user["name"]
        })
    
    await db.loans.insert_one(loan_doc)
    
    # Mark expense as converted
    await db.income_expenses.update_one(
        {"entry_id": entry_id},
        {"$set": {"converted_to_loan": True, "loan_id": loan_id, "status": "converted_to_loan", "updated_at": now.isoformat()}}
    )
    
    loan_doc.pop("_id", None)
    await log_activity(request, user, "edit", "income_expenses", "Converted expense to loan")

    return {"message": "Expense converted to loan", "loan": loan_doc}

@api_router.post("/income-expenses/{entry_id}/vendor-approve")
async def vendor_approve_ie(request: Request, entry_id: str, user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.APPROVE))):

    """Vendor approves a pending income/expense entry with commission calculation"""
    entry = await db.income_expenses.find_one({"entry_id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    if entry.get("status") != "pending_vendor":
        raise HTTPException(status_code=400, detail="Entry is not pending vendor approval")
    
    # Verify this vendor is the assigned vendor
    vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not vendor or vendor.get("vendor_id") != entry.get("vendor_id"):
        raise HTTPException(status_code=403, detail="Not authorized for this entry")
    
    # Require proof upload
    if not entry.get("vendor_proof_image"):
        raise HTTPException(status_code=400, detail="Please upload proof screenshot before approving")
    
    now = datetime.now(timezone.utc)
    
    # Calculate commission (same as deposit/withdrawal)
    # For IE entries: income = deposit-like (vendor receives), expense = withdrawal-like (vendor pays)
    tx_mode = entry.get("transaction_mode", "bank")
    if entry["entry_type"] == "income":
        if tx_mode == "cash":
            commission_rate = vendor.get("deposit_commission_cash", vendor.get("deposit_commission", 0)) / 100
        else:
            commission_rate = vendor.get("deposit_commission", 0) / 100
    else:
        if tx_mode == "cash":
            commission_rate = vendor.get("withdrawal_commission_cash", vendor.get("withdrawal_commission", 0)) / 100
        else:
            commission_rate = vendor.get("withdrawal_commission", 0) / 100
    
    amount = entry.get("amount", 0)
    base_amount = entry.get("base_amount") or amount
    base_currency = entry.get("base_currency") or entry.get("currency", "USD")
    amount_usd = entry.get("amount_usd") or convert_to_usd(amount, entry.get("currency", "USD"))
    commission_amount_usd = round(amount_usd * commission_rate, 2)
    commission_amount_base = round(base_amount * commission_rate, 2)
    
    # Update entry status with commission details
    await db.income_expenses.update_one(
        {"entry_id": entry_id},
        {"$set": {
            "status": "completed",
            "vendor_approved_at": now.isoformat(),
            "vendor_approved_by": user["user_id"],
            "vendor_commission_rate": commission_rate * 100,
            "vendor_commission_amount": commission_amount_usd,
            "vendor_commission_base_amount": commission_amount_base,
            "vendor_commission_base_currency": base_currency,
            "amount_usd": amount_usd,
        }}
    )
    
    # Update vendor's total commission and volume
    await db.vendors.update_one(
        {"vendor_id": vendor["vendor_id"]},
        {
            "$inc": {
                "total_commission": commission_amount_usd,
                "total_volume": amount_usd,
            },
            "$set": {"updated_at": now.isoformat()}
        }
    )
    
    await log_activity(request, user, "approve", "income_expenses", "Vendor approved I&E entry")

    return {
        "message": "Entry approved",
        "status": "completed",
        "vendor_commission_rate": commission_rate * 100,
        "vendor_commission_amount": commission_amount_usd,
        "vendor_commission_base_amount": commission_amount_base,
    }

@api_router.post("/income-expenses/{entry_id}/vendor-reject")
async def vendor_reject_ie(request: Request, entry_id: str, reason: str = "", user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.APPROVE))):

    """Vendor rejects a pending income/expense entry"""
    entry = await db.income_expenses.find_one({"entry_id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    if entry.get("status") != "pending_vendor":
        raise HTTPException(status_code=400, detail="Entry is not pending vendor approval")
    
    vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not vendor or vendor.get("vendor_id") != entry.get("vendor_id"):
        raise HTTPException(status_code=403, detail="Not authorized for this entry")
    
    now = datetime.now(timezone.utc)
    await db.income_expenses.update_one(
        {"entry_id": entry_id},
        {"$set": {"status": "rejected", "rejection_reason": reason, "vendor_rejected_at": now.isoformat()}}
    )
    
    await log_activity(request, user, "reject", "income_expenses", "Vendor rejected I&E entry")

    return {"message": "Entry rejected"}

@api_router.post("/income-expenses/{entry_id}/vendor-upload-proof")
async def vendor_upload_ie_proof(entry_id: str, user: dict = Depends(require_vendor), proof_image: UploadFile = File(...)):
    """Vendor uploads proof screenshot for income/expense entry"""
    entry = await db.income_expenses.find_one({"entry_id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not vendor or vendor.get("vendor_id") != entry.get("vendor_id"):
        raise HTTPException(status_code=403, detail="Not authorized for this entry")
    
    contents = await proof_image.read()
    proof_url = upload_to_r2(contents, proof_image.filename or "proof.png", proof_image.content_type or "image/png", "proofs")
    
    await db.income_expenses.update_one(
        {"entry_id": entry_id},
        {"$set": {"vendor_proof_image": proof_url, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Proof uploaded"}

@api_router.post("/income-expenses/{entry_id}/upload-invoice")
async def upload_ie_invoice(request: Request, entry_id: str, user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.EDIT)), invoice_file: UploadFile = File(...)):

    """Upload invoice/document to an income/expense entry"""
    entry = await db.income_expenses.find_one({"entry_id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    contents = await invoice_file.read()
    invoice_url = upload_to_r2(contents, invoice_file.filename or "invoice.pdf", invoice_file.content_type or "application/pdf", "invoices")
    
    # Store with filename info
    file_info = {
        "url": invoice_url,
        "filename": invoice_file.filename,
        "content_type": invoice_file.content_type,
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "uploaded_by": user["user_id"],
        "uploaded_by_name": user["name"]
    }
    
    await db.income_expenses.update_one(
        {"entry_id": entry_id},
        {"$set": {"invoice_file": file_info, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    await log_activity(request, user, "edit", "income_expenses", "Uploaded I&E invoice")

    return {"message": "Invoice uploaded successfully", "filename": invoice_file.filename}

@api_router.get("/vendor/income-expenses")
async def get_vendor_ie_entries(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    user: dict = Depends(require_vendor)
):
    """Get income/expense entries linked to this vendor"""
    vendor = await db.vendors.find_one({"user_id": user["user_id"]}, {"_id": 0})
    if not vendor:
        raise HTTPException(status_code=404, detail="Vendor not found")
    
    return await paginate_query(db.income_expenses, {"vendor_id": vendor["vendor_id"]}, page, page_size)


# ============== VENDOR SUPPLIERS (Service Suppliers) ROUTES ==============

@api_router.get("/vendor-suppliers")
async def get_vendor_suppliers(
    status: Optional[str] = None,
    search: Optional[str] = None,
    user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.VIEW))
):
    """Get all vendor suppliers (service providers like rent, utilities)"""
    query = {}
    if status:
        query["status"] = status
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"contact_person": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}}
        ]
    
    suppliers = await db.vendor_suppliers.find(query, {"_id": 0}).sort("name", 1).to_list(1000)
    return suppliers

@api_router.get("/vendor-suppliers/{supplier_id}")
async def get_vendor_supplier(supplier_id: str, user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.VIEW))):
    """Get a specific vendor supplier"""
    supplier = await db.vendor_suppliers.find_one({"supplier_id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Vendor supplier not found")
    return supplier

@api_router.post("/vendor-suppliers")
async def create_vendor_supplier(request: Request, data: VendorSupplierCreate, user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.CREATE))):

    """Create a new vendor supplier (for services like rent, utilities)"""
    # Check for duplicate name
    existing = await db.vendor_suppliers.find_one({"name": {"$regex": f"^{data.name}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="A supplier with this name already exists")
    
    supplier_id = f"vs_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    supplier_doc = {
        "supplier_id": supplier_id,
        "name": data.name,
        "contact_person": data.contact_person,
        "email": data.email,
        "phone": data.phone,
        "address": data.address,
        "bank_name": data.bank_name,
        "bank_account_name": data.bank_account_name,
        "bank_account_number": data.bank_account_number,
        "bank_ifsc": data.bank_ifsc,
        "bank_branch": data.bank_branch,
        "notes": data.notes,
        "status": VendorSupplierStatus.ACTIVE,
        "created_at": now.isoformat(),
        "updated_at": now.isoformat(),
        "created_by": user["user_id"]
    }
    
    await db.vendor_suppliers.insert_one(supplier_doc)
    supplier_doc.pop("_id", None)
    await log_activity(request, user, "create", "income_expenses", "Created vendor supplier")

    return supplier_doc

@api_router.put("/vendor-suppliers/{supplier_id}")
async def update_vendor_supplier(request: Request, supplier_id: str, data: VendorSupplierUpdate, user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.EDIT))):

    """Update a vendor supplier"""
    supplier = await db.vendor_suppliers.find_one({"supplier_id": supplier_id})
    if not supplier:
        raise HTTPException(status_code=404, detail="Vendor supplier not found")
    
    update_dict = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_dict:
        update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.vendor_suppliers.update_one({"supplier_id": supplier_id}, {"$set": update_dict})
    
    updated = await db.vendor_suppliers.find_one({"supplier_id": supplier_id}, {"_id": 0})
    await log_activity(request, user, "edit", "income_expenses", "Updated vendor supplier")

    return updated

@api_router.delete("/vendor-suppliers/{supplier_id}")
async def delete_vendor_supplier(request: Request, supplier_id: str, user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.DELETE))):

    """Delete a vendor supplier"""
    # Check if there are linked income/expenses
    linked_entries = await db.income_expenses.count_documents({"vendor_supplier_id": supplier_id})
    if linked_entries > 0:
        # Soft delete - mark as inactive instead
        await db.vendor_suppliers.update_one(
            {"supplier_id": supplier_id},
            {"$set": {"status": VendorSupplierStatus.INACTIVE, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        return {"message": "Supplier marked as inactive (has linked entries)"}
    
    result = await db.vendor_suppliers.delete_one({"supplier_id": supplier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Vendor supplier not found")
    await log_activity(request, user, "delete", "income_expenses", "Deleted vendor supplier")

    return {"message": "Supplier deleted"}

# ============== IE CATEGORIES (Account Categories) ROUTES ==============

@api_router.get("/ie-categories")
async def get_ie_categories(
    category_type: Optional[str] = None,
    active_only: bool = True,
    user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.VIEW))
):
    """Get all income/expense account categories"""
    query = {}
    if category_type:
        query["$or"] = [{"category_type": category_type}, {"category_type": "both"}]
    if active_only:
        query["is_active"] = True
    
    categories = await db.ie_categories.find(query, {"_id": 0}).sort("name", 1).to_list(500)
    return categories

@api_router.get("/ie-categories/{category_id}")
async def get_ie_category(category_id: str, user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.VIEW))):
    """Get a specific category"""
    category = await db.ie_categories.find_one({"category_id": category_id}, {"_id": 0})
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    return category

@api_router.post("/ie-categories")
async def create_ie_category(request: Request, data: IECategoryCreate, user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.CREATE))):

    """Create a new income/expense category"""
    # Check for duplicate name within same type
    existing = await db.ie_categories.find_one({
        "name": {"$regex": f"^{data.name}$", "$options": "i"},
        "$or": [{"category_type": data.category_type}, {"category_type": "both"}]
    })
    if existing:
        raise HTTPException(status_code=400, detail="A category with this name already exists")
    
    category_id = f"iec_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    category_doc = {
        "category_id": category_id,
        "name": data.name,
        "category_type": data.category_type,
        "description": data.description,
        "parent_category_id": data.parent_category_id,
        "is_active": True,
        "created_at": now.isoformat(),
        "updated_at": now.isoformat(),
        "created_by": user["user_id"]
    }
    
    await db.ie_categories.insert_one(category_doc)
    category_doc.pop("_id", None)
    await log_activity(request, user, "create", "income_expenses", "Created I&E category")

    return category_doc

@api_router.put("/ie-categories/{category_id}")
async def update_ie_category(request: Request, category_id: str, data: IECategoryUpdate, user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.EDIT))):

    """Update a category"""
    category = await db.ie_categories.find_one({"category_id": category_id})
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    update_dict = {k: v for k, v in data.model_dump().items() if v is not None}
    if update_dict:
        update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.ie_categories.update_one({"category_id": category_id}, {"$set": update_dict})
    
    updated = await db.ie_categories.find_one({"category_id": category_id}, {"_id": 0})
    await log_activity(request, user, "edit", "income_expenses", "Updated I&E category")

    return updated

@api_router.delete("/ie-categories/{category_id}")
async def delete_ie_category(request: Request, category_id: str, user: dict = Depends(require_permission(Modules.INCOME_EXPENSES, Actions.DELETE))):

    """Delete a category (soft delete if linked to entries)"""
    # Check if there are linked income/expenses
    linked_entries = await db.income_expenses.count_documents({"ie_category_id": category_id})
    if linked_entries > 0:
        # Soft delete - mark as inactive
        await db.ie_categories.update_one(
            {"category_id": category_id},
            {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        return {"message": "Category marked as inactive (has linked entries)"}
    
    result = await db.ie_categories.delete_one({"category_id": category_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    await log_activity(request, user, "delete", "income_expenses", "Deleted I&E category")

    return {"message": "Category deleted"}


# ============== LOAN MANAGEMENT ROUTES ==============

# === Static loan routes (MUST come before /{loan_id} route) ===

@api_router.get("/loans/dashboard")
async def get_loan_dashboard(user: dict = Depends(require_permission(Modules.LOANS, Actions.VIEW))):
    """Get comprehensive loan dashboard data"""
    loans = await db.loans.find({}, {"_id": 0}).to_list(10000)
    now = datetime.now(timezone.utc)
    
    # Basic metrics
    total_disbursed_usd = 0
    total_outstanding_usd = 0
    total_repaid_usd = 0
    total_interest_earned = 0
    
    # Aging analysis
    aging = {"current": 0, "days_1_30": 0, "days_31_60": 0, "days_61_90": 0, "days_90_plus": 0}
    
    # Status counts
    status_counts = {"active": 0, "partially_paid": 0, "fully_paid": 0, "overdue": 0, "written_off": 0}
    
    # Loan type breakdown
    type_breakdown = {"short_term": 0, "long_term": 0, "credit_line": 0}
    
    # Top borrowers
    borrower_data = {}
    
    # Currency exposure
    currency_exposure = {}
    
    # Upcoming dues (next 30 days)
    upcoming_dues = []
    
    for loan in loans:
        amount_usd = loan.get("amount_usd", convert_to_usd(loan["amount"], loan.get("currency", "USD")))
        interest_usd = convert_to_usd(loan.get("total_interest", 0), loan.get("currency", "USD"))
        repaid_usd = convert_to_usd(loan.get("total_repaid", 0), loan.get("currency", "USD"))
        outstanding = amount_usd + interest_usd - repaid_usd
        
        total_disbursed_usd += amount_usd
        if outstanding > 0:
            total_outstanding_usd += outstanding
        total_repaid_usd += repaid_usd
        
        if loan.get("total_repaid", 0) > loan["amount"]:
            total_interest_earned += convert_to_usd(loan["total_repaid"] - loan["amount"], loan.get("currency", "USD"))
        
        # Status counts
        status = loan.get("status", "active")
        if status in status_counts:
            status_counts[status] += 1
        
        # Loan type
        loan_type = loan.get("loan_type", "short_term")
        if loan_type in type_breakdown:
            type_breakdown[loan_type] += amount_usd
        
        # Currency exposure
        currency = loan.get("currency", "USD")
        if currency not in currency_exposure:
            currency_exposure[currency] = {"amount": 0, "outstanding": 0, "count": 0}
        currency_exposure[currency]["amount"] += loan["amount"]
        currency_exposure[currency]["outstanding"] += max(0, loan["amount"] + loan.get("total_interest", 0) - loan.get("total_repaid", 0))
        currency_exposure[currency]["count"] += 1
        
        # Top borrowers
        borrower = loan["borrower_name"]
        if borrower not in borrower_data:
            borrower_data[borrower] = {"total_disbursed": 0, "outstanding": 0, "loan_count": 0, "vendor_id": loan.get("vendor_id")}
        borrower_data[borrower]["total_disbursed"] += amount_usd
        borrower_data[borrower]["outstanding"] += max(0, outstanding)
        borrower_data[borrower]["loan_count"] += 1
        
        # Aging analysis (for active/partially paid loans)
        if status in ["active", "partially_paid"] and loan.get("due_date"):
            due_str = loan["due_date"]
            try:
                if "T" in due_str:
                    due = datetime.fromisoformat(due_str.replace("Z", "+00:00"))
                else:
                    due = datetime.strptime(due_str[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc)
                if due.tzinfo is None:
                    due = due.replace(tzinfo=timezone.utc)
                
                days_overdue = (now - due).days
                
                if days_overdue < 0:
                    aging["current"] += outstanding
                    # Check if due within 30 days
                    if days_overdue >= -30:
                        upcoming_dues.append({
                            "loan_id": loan["loan_id"],
                            "borrower": loan["borrower_name"],
                            "amount": loan["amount"],
                            "currency": loan.get("currency", "USD"),
                            "outstanding": outstanding,
                            "due_date": loan["due_date"],
                            "days_until_due": abs(days_overdue)
                        })
                elif days_overdue <= 30:
                    aging["days_1_30"] += outstanding
                elif days_overdue <= 60:
                    aging["days_31_60"] += outstanding
                elif days_overdue <= 90:
                    aging["days_61_90"] += outstanding
                else:
                    aging["days_90_plus"] += outstanding
            except:
                pass
    
    # Sort upcoming dues by days
    upcoming_dues.sort(key=lambda x: x["days_until_due"])
    
    # Top 5 borrowers by outstanding
    top_borrowers = sorted(borrower_data.items(), key=lambda x: x[1]["outstanding"], reverse=True)[:5]
    
    # Collection rate (this month)
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    this_month_repayments = await db.loan_repayments.find({
        "payment_date": {"$gte": start_of_month.isoformat()}
    }, {"_id": 0}).to_list(10000)
    
    this_month_collected = sum(convert_to_usd(r.get("amount", 0), r.get("currency", "USD")) for r in this_month_repayments)
    
    return {
        "portfolio_overview": {
            "total_disbursed_usd": round(total_disbursed_usd, 2),
            "total_outstanding_usd": round(total_outstanding_usd, 2),
            "total_repaid_usd": round(total_repaid_usd, 2),
            "total_interest_earned_usd": round(total_interest_earned, 2),
            "total_loans": len(loans)
        },
        "status_breakdown": status_counts,
        "loan_type_breakdown": {k: round(v, 2) for k, v in type_breakdown.items()},
        "aging_analysis": {k: round(v, 2) for k, v in aging.items()},
        "currency_exposure": currency_exposure,
        "top_borrowers": [{"name": k, **{kk: round(vv, 2) if isinstance(vv, float) else vv for kk, vv in v.items()}} for k, v in top_borrowers],
        "upcoming_dues": upcoming_dues[:10],
        "collection_this_month": round(this_month_collected, 2)
    }

@api_router.get("/loans/transactions")
async def get_loan_transactions(
    loan_id: Optional[str] = None,
    transaction_type: Optional[str] = None,
    vendor_id: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    user: dict = Depends(require_permission(Modules.LOANS, Actions.VIEW))
):
    """Get loan transactions log"""
    query = {}
    if loan_id:
        query["loan_id"] = loan_id
    if transaction_type:
        query["transaction_type"] = transaction_type
    if vendor_id:
        query["$or"] = [
            {"source_vendor_id": vendor_id},
            {"credit_to_vendor_id": vendor_id}
        ]
    
    skip = (page - 1) * page_size
    total = await db.loan_transactions.count_documents(query)
    total_pages = (total + page_size - 1) // page_size if total > 0 else 1
    
    transactions = await db.loan_transactions.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    
    # Enrich with treasury/vendor names
    treasury_ids = list(set([tx.get("treasury_account_id") for tx in transactions if tx.get("treasury_account_id")]))
    vendor_ids = list(set(
        [tx.get("credit_to_vendor_id") for tx in transactions if tx.get("credit_to_vendor_id")] +
        [tx.get("source_vendor_id") for tx in transactions if tx.get("source_vendor_id")]
    ))
    
    # Get treasury accounts
    treasury_map = {}
    if treasury_ids:
        treasuries = await db.treasury_accounts.find({"account_id": {"$in": treasury_ids}}, {"_id": 0}).to_list(100)
        treasury_map = {t["account_id"]: t["account_name"] for t in treasuries}
    
    # Get vendors
    vendor_map = {}
    if vendor_ids:
        vendors = await db.vendors.find({"vendor_id": {"$in": vendor_ids}}, {"_id": 0}).to_list(100)
        vendor_map = {v["vendor_id"]: v.get("name") or v.get("vendor_name") for v in vendors}
    
    # Also get loan details to find source vendor/treasury
    loan_ids = list(set([tx.get("loan_id") for tx in transactions if tx.get("loan_id")]))
    loan_map = {}
    if loan_ids:
        loans = await db.loans.find({"loan_id": {"$in": loan_ids}}, {"_id": 0}).to_list(500)
        for loan in loans:
            loan_map[loan["loan_id"]] = loan
    
    # Enrich transactions
    for tx in transactions:
        tx["treasury_account_name"] = treasury_map.get(tx.get("treasury_account_id"))
        tx["credit_vendor_name"] = vendor_map.get(tx.get("credit_to_vendor_id"))
        tx["source_vendor_name"] = vendor_map.get(tx.get("source_vendor_id"))
        tx["status"] = tx.get("status", "completed")
        
        # For disbursements, get source from loan
        if tx.get("transaction_type") == "disbursement" and tx.get("loan_id"):
            loan = loan_map.get(tx["loan_id"])
            if loan:
                if loan.get("source_treasury_id"):
                    source_treasury = await db.treasury_accounts.find_one({"account_id": loan["source_treasury_id"]}, {"_id": 0})
                    tx["treasury_account_name"] = source_treasury["account_name"] if source_treasury else None
                elif loan.get("source_vendor_id"):
                    source_vendor = await db.vendors.find_one({"vendor_id": loan["source_vendor_id"]}, {"_id": 0})
                    tx["source_vendor_name"] = source_vendor.get("name") or source_vendor.get("vendor_name") if source_vendor else None
    
    return {"items": transactions, "total": total, "page": page, "page_size": page_size, "total_pages": total_pages}

@api_router.get("/loans/vendors")
async def get_vendor_borrowers(user: dict = Depends(require_permission(Modules.LOANS, Actions.VIEW))):
    """Get all vendors that can be borrowers with their loan stats"""
    vendors = await db.vendors.find({}, {"_id": 0}).to_list(1000)
    
    # Get loan stats for each vendor
    vendor_stats = {}
    loans = await db.loans.find({}, {"_id": 0, "vendor_id": 1, "amount": 1, "currency": 1, "total_repaid": 1, "total_interest": 1, "status": 1}).to_list(10000)
    
    for loan in loans:
        vid = loan.get("vendor_id")
        if vid:
            if vid not in vendor_stats:
                vendor_stats[vid] = {"total_loans": 0, "total_disbursed": 0, "total_outstanding": 0, "active_loans": 0}
            
            amount_usd = convert_to_usd(loan["amount"], loan.get("currency", "USD"))
            outstanding = loan["amount"] + loan.get("total_interest", 0) - loan.get("total_repaid", 0)
            outstanding_usd = convert_to_usd(max(0, outstanding), loan.get("currency", "USD"))
            
            vendor_stats[vid]["total_loans"] += 1
            vendor_stats[vid]["total_disbursed"] += amount_usd
            vendor_stats[vid]["total_outstanding"] += outstanding_usd
            if loan["status"] in ["active", "partially_paid"]:
                vendor_stats[vid]["active_loans"] += 1
    
    # Combine vendor info with stats
    result = []
    for v in vendors:
        stats = vendor_stats.get(v["vendor_id"], {"total_loans": 0, "total_disbursed": 0, "total_outstanding": 0, "active_loans": 0})
        result.append({
            "vendor_id": v["vendor_id"],
            "name": v.get("vendor_name") or v.get("name", "Unknown"),
            "email": v.get("email"),
            "status": v.get("status", "active"),
            "loan_stats": {
                "total_loans": stats["total_loans"],
                "total_disbursed_usd": round(stats["total_disbursed"], 2),
                "total_outstanding_usd": round(stats["total_outstanding"], 2),
                "active_loans": stats["active_loans"]
            }
        })
    
    return result

# === List all loans ===

@api_router.get("/loans")
async def get_loans(
    status: Optional[str] = None,
    borrower: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=200),
    user: dict = Depends(require_permission(Modules.LOANS, Actions.VIEW))
):
    """Get all loans with optional filters"""
    query = {}
    if status:
        query["status"] = status
    if borrower:
        query["borrower_name"] = {"$regex": borrower, "$options": "i"}
    
    skip = (page - 1) * page_size
    total = await db.loans.count_documents(query)
    total_pages = (total + page_size - 1) // page_size if total > 0 else 1
    
    loans = await db.loans.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    
    # Batch fetch treasury accounts to avoid N+1 queries
    treasury_ids = list(set(loan.get("source_treasury_id") for loan in loans if loan.get("source_treasury_id")))
    treasury_map = {}
    if treasury_ids:
        treasuries = await db.treasury_accounts.find({"account_id": {"$in": treasury_ids}}, {"_id": 0}).to_list(len(treasury_ids))
        treasury_map = {t["account_id"]: t["account_name"] for t in treasuries}
    
    # Process loans
    for loan in loans:
        if loan.get("source_treasury_id"):
            loan["source_treasury_name"] = treasury_map.get(loan["source_treasury_id"], "Unknown")
        
        # Calculate outstanding balance
        loan["outstanding_balance"] = loan["amount"] + loan.get("total_interest", 0) - loan.get("total_repaid", 0)
        
        # Check if overdue
        if loan["status"] == LoanStatus.ACTIVE:
            if loan.get("due_date"):
                due_str = loan["due_date"]
                try:
                    if "T" in due_str:
                        due = datetime.fromisoformat(due_str.replace("Z", "+00:00"))
                    else:
                        due = datetime.strptime(due_str[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc)
                    if due.tzinfo is None:
                        due = due.replace(tzinfo=timezone.utc)
                    if due < datetime.now(timezone.utc):
                        loan["is_overdue"] = True
                except (ValueError, TypeError):
                    # Invalid date format - skip overdue check
                    loan["is_overdue"] = False
    
    return {"items": loans, "total": total, "page": page, "page_size": page_size, "total_pages": total_pages}

@api_router.get("/loans/{loan_id}")
async def get_loan(loan_id: str, user: dict = Depends(require_permission(Modules.LOANS, Actions.VIEW))):
    """Get a single loan with repayment history"""
    loan = await db.loans.find_one({"loan_id": loan_id}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    # Get treasury account name
    if loan.get("source_treasury_id"):
        acc = await db.treasury_accounts.find_one({"account_id": loan["source_treasury_id"]}, {"_id": 0})
        loan["source_treasury_name"] = acc["account_name"] if acc else "Unknown"
    
    # Get repayment history
    repayments = await db.loan_repayments.find({"loan_id": loan_id}, {"_id": 0}).sort("payment_date", -1).to_list(1000)
    for rep in repayments:
        if rep.get("treasury_account_id"):
            acc = await db.treasury_accounts.find_one({"account_id": rep["treasury_account_id"]}, {"_id": 0})
            rep["treasury_account_name"] = acc["account_name"] if acc else "Unknown"
    
    loan["repayments"] = repayments
    loan["outstanding_balance"] = loan["amount"] + loan.get("total_interest", 0) - loan.get("total_repaid", 0)
    
    return loan

@api_router.post("/loans")
async def create_loan(loan_data: LoanCreate, request: Request, user: dict = Depends(require_permission(Modules.LOANS, Actions.CREATE))):
    """Create a new loan and deduct from treasury or vendor"""
    
    # Validate that at least one source is provided
    if not loan_data.treasury_account_id and not loan_data.disburse_from_vendor_id:
        raise HTTPException(status_code=400, detail="Please select Treasury or Exchanger to disburse from")
    
    treasury = None
    disburse_vendor = None
    
    # Check treasury source (balance check deferred to approval)
    if loan_data.treasury_account_id:
        treasury = await db.treasury_accounts.find_one({"account_id": loan_data.treasury_account_id}, {"_id": 0})
        if not treasury:
            raise HTTPException(status_code=404, detail="Treasury account not found")
    
    # Check vendor source
    if loan_data.disburse_from_vendor_id:
        disburse_vendor = await db.vendors.find_one({"vendor_id": loan_data.disburse_from_vendor_id}, {"_id": 0})
        if not disburse_vendor:
            raise HTTPException(status_code=404, detail="Exchanger not found")
    
    if loan_data.amount <= 0:
        raise HTTPException(status_code=400, detail="Loan amount must be positive")
    
    # If vendor_id provided, get vendor details
    vendor = None
    if loan_data.vendor_id:
        vendor = await db.vendors.find_one({"vendor_id": loan_data.vendor_id}, {"_id": 0})
    
    loan_id = f"loan_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    # Calculate interest (simple interest)
    # Interest = Principal * Rate * Time (in years)
    loan_date = datetime.fromisoformat(loan_data.loan_date.replace("Z", "+00:00"))
    due_date = datetime.fromisoformat(loan_data.due_date.replace("Z", "+00:00"))
    days_diff = (due_date - loan_date).days
    years = days_diff / 365
    total_interest = round(loan_data.amount * (loan_data.interest_rate / 100) * years, 2)
    
    # Calculate USD equivalent
    amount_usd = convert_to_usd(loan_data.amount, loan_data.currency)
    
    loan_doc = {
        "loan_id": loan_id,
        "vendor_id": loan_data.vendor_id,
        "vendor_name": vendor.get("vendor_name") or vendor.get("name") if vendor else None,
        "borrower_name": loan_data.borrower_name,
        "amount": loan_data.amount,
        "currency": loan_data.currency,
        "amount_usd": amount_usd,
        "interest_rate": loan_data.interest_rate,
        "loan_type": loan_data.loan_type,
        "total_interest": total_interest,
        "loan_date": loan_data.loan_date,
        "due_date": loan_data.due_date,
        "repayment_mode": loan_data.repayment_mode,
        "installment_amount": loan_data.installment_amount,
        "installment_frequency": loan_data.installment_frequency,
        "num_installments": loan_data.num_installments,
        "collateral": loan_data.collateral,
        "source_treasury_id": loan_data.treasury_account_id,
        "source_vendor_id": loan_data.disburse_from_vendor_id,
        "source_vendor_name": disburse_vendor.get("name") or disburse_vendor.get("vendor_name") if disburse_vendor else None,
        "bank_details": loan_data.bank_details,
        "total_repaid": 0,
        "repayment_count": 0,
        "status": "pending_approval",
        "notes": loan_data.notes,
        "created_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    
    await db.loans.insert_one(loan_doc)
    
    # Treasury deduction is deferred to approval step
    
    # Calculate commission for vendor disbursement (OUT = withdrawal type)
    vendor_commission_rate = 0.0
    vendor_commission_amount = 0.0
    vendor_commission_base_amount = 0.0
    vendor_commission_base_currency = None
    
    if disburse_vendor:
        # Use withdrawal commission rate for disbursement (OUT)
        # Default to bank transfer commission rate
        vendor_commission_rate = disburse_vendor.get("withdrawal_commission", 0)
        if vendor_commission_rate > 0:
            # Calculate commission in payment currency (loan_data.currency)
            vendor_commission_amount = round(loan_data.amount * vendor_commission_rate / 100, 2)
            vendor_commission_base_amount = vendor_commission_amount
            vendor_commission_base_currency = loan_data.currency
    
    # Record loan transaction - always pending_approval
    await db.loan_transactions.insert_one({
        "transaction_id": f"ltx_{uuid.uuid4().hex[:12]}",
        "loan_id": loan_id,
        "transaction_type": LoanTransactionType.DISBURSEMENT,
        "amount": loan_data.amount,
        "currency": loan_data.currency,
        "treasury_account_id": loan_data.treasury_account_id,
        "source_vendor_id": loan_data.disburse_from_vendor_id,
        "source_vendor_name": disburse_vendor.get("name") or disburse_vendor.get("vendor_name") if disburse_vendor else None,
        "bank_details": loan_data.bank_details,
        "borrower_name": loan_data.borrower_name,
        "status": "pending_approval",
        "description": f"Loan disbursement to {loan_data.borrower_name}",
        "vendor_commission_rate": vendor_commission_rate if vendor_commission_rate > 0 else None,
        "vendor_commission_amount": vendor_commission_amount if vendor_commission_amount > 0 else None,
        "vendor_commission_base_amount": vendor_commission_base_amount if vendor_commission_base_amount > 0 else None,
        "vendor_commission_base_currency": vendor_commission_base_currency,
        "created_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    })
    
    loan_doc.pop("_id", None)
    if treasury:
        loan_doc["source_treasury_name"] = treasury["account_name"]
    elif disburse_vendor:
        loan_doc["source_vendor_name"] = disburse_vendor.get("name") or disburse_vendor.get("vendor_name")
    loan_doc["outstanding_balance"] = loan_data.amount + total_interest
    
    # Log activity
    await log_activity(request, user, "create", "loans", f"Created loan to {loan_data.borrower_name}: {loan_data.amount} {loan_data.currency}", reference_id=loan_id)
    
    # Notify exchanger if disbursing from vendor
    if loan_data.disburse_from_vendor_id:
        import asyncio
        asyncio.create_task(send_exchanger_notification("loan", loan_data.disburse_from_vendor_id, {
            "loan_type": "disbursement",
            "borrower": loan_data.borrower_name,
            "amount_display": f"{loan_data.amount:,.2f} {loan_data.currency}",
        }))
    
    return loan_doc

@api_router.put("/loans/{loan_id}")
async def update_loan(request: Request, loan_id: str, update_data: LoanUpdate, user: dict = Depends(require_permission(Modules.LOANS, Actions.EDIT))):

    """Update loan details (not amount)"""
    loan = await db.loans.find_one({"loan_id": loan_id}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if not update_dict:
        raise HTTPException(status_code=400, detail="No fields to update")
    
    update_dict["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.loans.update_one({"loan_id": loan_id}, {"$set": update_dict})
    
    updated = await db.loans.find_one({"loan_id": loan_id}, {"_id": 0})
    await log_activity(request, user, "edit", "loans", "Updated loan")

    return updated

@api_router.post("/loans/{loan_id}/attachments")
async def upload_loan_attachments(
    request: Request,
    loan_id: str,
    files: list[UploadFile] = File(...),
    user: dict = Depends(require_permission(Modules.LOANS, Actions.EDIT))
):
    """Upload attachments (xlsx, pdf, images) to a loan"""
    loan = await db.loans.find_one({"loan_id": loan_id}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    ALLOWED_EXTENSIONS = {'pdf', 'xlsx', 'xls', 'csv', 'png', 'jpg', 'jpeg', 'gif', 'webp'}
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    
    uploaded = []
    now = datetime.now(timezone.utc).isoformat()
    
    for file in files:
        ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
        if ext not in ALLOWED_EXTENSIONS:
            raise HTTPException(status_code=400, detail=f"File type .{ext} not allowed. Allowed: {', '.join(ALLOWED_EXTENSIONS)}")
        
        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            raise HTTPException(status_code=400, detail=f"File {file.filename} exceeds 10MB limit")
        
        url = upload_to_r2(content, file.filename, file.content_type or "application/octet-stream", folder=f"loans/{loan_id}")
        
        attachment = {
            "attachment_id": f"att_{uuid.uuid4().hex[:10]}",
            "filename": file.filename,
            "url": url,
            "file_type": ext,
            "size": len(content),
            "uploaded_at": now,
            "uploaded_by": user["user_id"],
            "uploaded_by_name": user["name"]
        }
        uploaded.append(attachment)
    
    await db.loans.update_one(
        {"loan_id": loan_id},
        {"$push": {"attachments": {"$each": uploaded}}}
    )
    
    # Also log in loan_transactions
    await db.loan_transactions.insert_one({
        "transaction_id": f"ltx_{uuid.uuid4().hex[:12]}",
        "loan_id": loan_id,
        "transaction_type": "attachment",
        "description": f"Uploaded {len(uploaded)} file(s): {', '.join(f.filename for f in files)}",
        "amount": 0,
        "amount_usd": 0,
        "currency": loan.get("currency", "USD"),
        "status": "completed",
        "attachments": uploaded,
        "created_at": now,
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    })
    
    await log_activity(request, user, "edit", "loans", f"Uploaded {len(uploaded)} attachment(s) to loan {loan_id}")
    return {"message": f"{len(uploaded)} file(s) uploaded", "attachments": uploaded}

@api_router.delete("/loans/{loan_id}/attachments/{attachment_id}")
async def delete_loan_attachment(
    request: Request,
    loan_id: str,
    attachment_id: str,
    user: dict = Depends(require_permission(Modules.LOANS, Actions.EDIT))
):
    """Delete an attachment from a loan"""
    loan = await db.loans.find_one({"loan_id": loan_id}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    await db.loans.update_one(
        {"loan_id": loan_id},
        {"$pull": {"attachments": {"attachment_id": attachment_id}}}
    )
    await log_activity(request, user, "delete", "loans", f"Removed attachment from loan {loan_id}")
    return {"message": "Attachment removed"}

@api_router.post("/loans/{loan_id}/repayment")
async def record_loan_repayment(request: Request, loan_id: str, repayment: LoanRepaymentCreate, user: dict = Depends(require_permission(Modules.LOANS, Actions.CREATE))):

    """Record a loan repayment"""
    loan = await db.loans.find_one({"loan_id": loan_id}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    if loan["status"] == LoanStatus.FULLY_PAID:
        raise HTTPException(status_code=400, detail="Loan is already fully paid")
    
    # Validate that at least one destination is provided
    if not repayment.treasury_account_id and not repayment.credit_to_vendor_id:
        raise HTTPException(status_code=400, detail="Please select Treasury or Exchanger to credit")
    
    treasury = None
    credit_vendor = None
    
    # Verify treasury account exists if provided
    if repayment.treasury_account_id:
        treasury = await db.treasury_accounts.find_one({"account_id": repayment.treasury_account_id}, {"_id": 0})
        if not treasury:
            raise HTTPException(status_code=404, detail="Treasury account not found")
    
    # Verify vendor exists if provided
    if repayment.credit_to_vendor_id:
        credit_vendor = await db.vendors.find_one({"vendor_id": repayment.credit_to_vendor_id}, {"_id": 0})
        if not credit_vendor:
            raise HTTPException(status_code=404, detail="Exchanger not found")
    
    if repayment.amount <= 0:
        raise HTTPException(status_code=400, detail="Repayment amount must be positive")
    
    repayment_id = f"rep_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    payment_date = repayment.payment_date if repayment.payment_date else now.isoformat()[:10]
    
    # Convert to loan currency if different
    repayment_amount_in_loan_currency = repayment.amount
    exchange_rate_used = 1.0
    if repayment.currency != loan["currency"]:
        if repayment.amount_in_loan_currency and repayment.amount_in_loan_currency > 0:
            # Use pre-calculated amount from frontend
            repayment_amount_in_loan_currency = repayment.amount_in_loan_currency
            exchange_rate_used = repayment.exchange_rate or (repayment_amount_in_loan_currency / repayment.amount if repayment.amount else 1.0)
        elif repayment.exchange_rate and repayment.exchange_rate > 0:
            # Use provided exchange rate (payment currency -> loan currency)
            repayment_amount_in_loan_currency = round(repayment.amount * repayment.exchange_rate, 2)
            exchange_rate_used = repayment.exchange_rate
        else:
            # Convert using live FX rates
            amount_usd = convert_to_usd(repayment.amount, repayment.currency)
            rates = _fx_cache.get("rates") or FALLBACK_RATES_TO_USD
            loan_rate = rates.get(loan["currency"].upper(), 1.0)
            repayment_amount_in_loan_currency = round(amount_usd / loan_rate, 2) if loan_rate else repayment.amount
            exchange_rate_used = round(repayment_amount_in_loan_currency / repayment.amount, 6) if repayment.amount else 1.0
    
    repayment_doc = {
        "repayment_id": repayment_id,
        "loan_id": loan_id,
        "amount": repayment.amount,
        "currency": repayment.currency,
        "amount_in_loan_currency": repayment_amount_in_loan_currency,
        "loan_currency": loan["currency"],
        "exchange_rate": exchange_rate_used,
        "treasury_account_id": repayment.treasury_account_id,
        "credit_to_vendor_id": repayment.credit_to_vendor_id,
        "payment_date": payment_date,
        "reference": repayment.reference,
        "notes": repayment.notes,
        "status": "pending_approval",
        "created_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    
    await db.loan_repayments.insert_one(repayment_doc)
    
    # Loan totals and treasury credit are deferred to approval step
    
    # Calculate commission for vendor repayment (IN = deposit type)
    vendor_commission_rate = 0.0
    vendor_commission_amount = 0.0
    vendor_commission_base_amount = 0.0
    vendor_commission_base_currency = None
    
    if credit_vendor:
        # Use deposit commission rate for repayment (IN)
        # Default to bank transfer commission rate
        vendor_commission_rate = credit_vendor.get("deposit_commission", 0)
        if vendor_commission_rate > 0:
            # Calculate commission in payment currency (repayment.currency)
            vendor_commission_amount = round(repayment.amount * vendor_commission_rate / 100, 2)
            vendor_commission_base_amount = vendor_commission_amount
            vendor_commission_base_currency = repayment.currency
    
    # Record loan transaction - always pending_approval
    await db.loan_transactions.insert_one({
        "transaction_id": f"ltx_{uuid.uuid4().hex[:12]}",
        "loan_id": loan_id,
        "transaction_type": LoanTransactionType.REPAYMENT,
        "amount": repayment.amount,
        "currency": repayment.currency,
        "treasury_account_id": repayment.treasury_account_id,
        "credit_to_vendor_id": repayment.credit_to_vendor_id,
        "credit_vendor_name": credit_vendor.get("name") or credit_vendor.get("vendor_name") if credit_vendor else None,
        "borrower_name": loan.get("borrower_name"),
        "status": "pending_approval",
        "description": f"Repayment from {loan['borrower_name']}",
        "vendor_commission_rate": vendor_commission_rate if vendor_commission_rate > 0 else None,
        "vendor_commission_amount": vendor_commission_amount if vendor_commission_amount > 0 else None,
        "vendor_commission_base_amount": vendor_commission_base_amount if vendor_commission_base_amount > 0 else None,
        "vendor_commission_base_currency": vendor_commission_base_currency,
        "created_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    })
    
    repayment_doc.pop("_id", None)
    if treasury:
        repayment_doc["treasury_account_name"] = treasury["account_name"]
    elif credit_vendor:
        repayment_doc["vendor_name"] = credit_vendor.get("name") or credit_vendor.get("vendor_name")
    repayment_doc["new_outstanding"] = max(0, loan["amount"] + loan.get("total_interest", 0) - loan.get("total_repaid", 0))
    repayment_doc["loan_status"] = loan["status"]
    await log_activity(request, user, "create", "loans", "Recorded loan repayment (pending approval)")

    # Notify exchanger if crediting to vendor
    if repayment.credit_to_vendor_id:
        import asyncio
        asyncio.create_task(send_exchanger_notification("loan", repayment.credit_to_vendor_id, {
            "loan_type": "repayment",
            "borrower": loan.get("borrower_name", "-"),
            "amount_display": f"{repayment.amount:,.2f} {repayment.currency}",
        }))

    return repayment_doc

@api_router.get("/loans/{loan_id}/repayments")
async def get_loan_repayments(loan_id: str, user: dict = Depends(require_permission(Modules.LOANS, Actions.VIEW))):
    """Get repayment history for a loan"""
    loan = await db.loans.find_one({"loan_id": loan_id}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    repayments = await db.loan_repayments.find({"loan_id": loan_id}, {"_id": 0}).sort("payment_date", -1).to_list(1000)
    
    for rep in repayments:
        if rep.get("treasury_account_id"):
            acc = await db.treasury_accounts.find_one({"account_id": rep["treasury_account_id"]}, {"_id": 0})
            rep["treasury_account_name"] = acc["account_name"] if acc else "Unknown"
    
    return repayments

@api_router.delete("/loans/{loan_id}")
async def delete_loan(request: Request, loan_id: str, user: dict = Depends(require_permission(Modules.LOANS, Actions.DELETE))):

    """Delete a loan (admin only) - reverses treasury if no repayments"""
    loan = await db.loans.find_one({"loan_id": loan_id}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan not found")
    
    # Check if there are repayments
    repayment_count = await db.loan_repayments.count_documents({"loan_id": loan_id})
    if repayment_count > 0:
        raise HTTPException(status_code=400, detail="Cannot delete loan with repayments. Delete repayments first.")
    
    now = datetime.now(timezone.utc)
    
    # Reverse treasury balance (credit back the loan amount)
    await db.treasury_accounts.update_one(
        {"account_id": loan["source_treasury_id"]},
        {"$inc": {"balance": loan["amount"]}, "$set": {"updated_at": now.isoformat()}}
    )
    
    # Delete the loan
    await db.loans.delete_one({"loan_id": loan_id})
    
    # Delete related treasury transaction
    await db.treasury_transactions.delete_one({"loan_id": loan_id})
    
    await log_activity(request, user, "delete", "loans", "Deleted loan")

    return {"message": "Loan deleted successfully"}

@api_router.get("/loans/reports/summary")
async def get_loans_summary(user: dict = Depends(require_permission(Modules.LOANS, Actions.VIEW))):
    """Get loan portfolio summary"""
    loans = await db.loans.find({}, {"_id": 0}).to_list(10000)
    
    total_disbursed = 0
    total_outstanding = 0
    total_repaid = 0
    total_interest_expected = 0
    total_interest_earned = 0
    
    active_count = 0
    partially_paid_count = 0
    fully_paid_count = 0
    overdue_count = 0
    
    by_borrower = {}
    
    for loan in loans:
        amount_usd = loan.get("amount_usd", convert_to_usd(loan["amount"], loan["currency"]))
        interest_usd = convert_to_usd(loan.get("total_interest", 0), loan["currency"])
        repaid_usd = convert_to_usd(loan.get("total_repaid", 0), loan["currency"])
        
        total_disbursed += amount_usd
        total_interest_expected += interest_usd
        total_repaid += repaid_usd
        
        outstanding = amount_usd + interest_usd - repaid_usd
        if outstanding > 0:
            total_outstanding += outstanding
        
        # Interest earned is the interest portion of repayments
        if loan.get("total_repaid", 0) > loan["amount"]:
            total_interest_earned += convert_to_usd(loan["total_repaid"] - loan["amount"], loan["currency"])
        
        # Count by status
        if loan["status"] == LoanStatus.ACTIVE:
            active_count += 1
            # Check if overdue
            if loan.get("due_date"):
                due_str = loan["due_date"]
                if "T" in due_str:
                    due = datetime.fromisoformat(due_str.replace("Z", "+00:00"))
                else:
                    # Date only string like "2026-03-15"
                    due = datetime.strptime(due_str[:10], "%Y-%m-%d").replace(tzinfo=timezone.utc)
                if due.tzinfo is None:
                    due = due.replace(tzinfo=timezone.utc)
                if due < datetime.now(timezone.utc):
                    overdue_count += 1
        elif loan["status"] == LoanStatus.PARTIALLY_PAID:
            partially_paid_count += 1
        elif loan["status"] == LoanStatus.FULLY_PAID:
            fully_paid_count += 1
        
        # Group by borrower
        borrower = loan["borrower_name"]
        if borrower not in by_borrower:
            by_borrower[borrower] = {"disbursed": 0, "outstanding": 0, "count": 0}
        by_borrower[borrower]["disbursed"] += amount_usd
        by_borrower[borrower]["outstanding"] += max(0, outstanding)
        by_borrower[borrower]["count"] += 1
    
    return {
        "total_loans": len(loans),
        "total_disbursed_usd": round(total_disbursed, 2),
        "total_outstanding_usd": round(total_outstanding, 2),
        "total_repaid_usd": round(total_repaid, 2),
        "total_interest_expected_usd": round(total_interest_expected, 2),
        "total_interest_earned_usd": round(total_interest_earned, 2),
        "status_breakdown": {
            "active": active_count,
            "partially_paid": partially_paid_count,
            "fully_paid": fully_paid_count,
            "overdue": overdue_count
        },
        "by_borrower": {k: {**v, "disbursed": round(v["disbursed"], 2), "outstanding": round(v["outstanding"], 2)} for k, v in by_borrower.items()}
    }

@api_router.get("/loans/export/csv")
async def export_loans_csv(user: dict = Depends(require_permission(Modules.LOANS, Actions.EXPORT))):
    """Export all loans as CSV"""
    from io import StringIO
    import csv
    
    loans = await db.loans.find({}, {"_id": 0}).sort("loan_date", -1).to_list(50000)
    
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Loan ID", "Borrower", "Amount", "Currency", "Interest Rate (%)",
        "Loan Date", "Due Date", "Outstanding Balance", "Total Repaid",
        "Status", "Repayment Mode", "Notes", "Created At"
    ])
    
    for loan in loans:
        outstanding = loan.get("amount", 0) + loan.get("total_interest", 0) - loan.get("total_repaid", 0)
        writer.writerow([
            loan.get("loan_id", ""),
            loan.get("borrower_name", ""),
            loan.get("amount", 0),
            loan.get("currency", "USD"),
            loan.get("interest_rate", 0),
            loan.get("loan_date", ""),
            loan.get("due_date", ""),
            outstanding,
            loan.get("total_repaid", 0),
            loan.get("status", ""),
            loan.get("repayment_mode", ""),
            loan.get("notes", ""),
            loan.get("created_at", "")
        ])
    
    from fastapi.responses import StreamingResponse
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=loans_export.csv"}
    )


@api_router.get("/loans/export/excel")
async def export_loans_excel(user: dict = Depends(require_permission(Modules.LOANS, Actions.EXPORT))):
    """Export all loans as Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    
    loans = await db.loans.find({}, {"_id": 0}).sort("loan_date", -1).to_list(50000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Loans"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0B3D91", end_color="0B3D91", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["Loan ID", "Borrower", "Amount", "Currency", "Interest Rate (%)",
               "Loan Date", "Due Date", "Outstanding", "Total Repaid", "Status"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    for row, loan in enumerate(loans, 2):
        outstanding = loan.get("amount", 0) + loan.get("total_interest", 0) - loan.get("total_repaid", 0)
        ws.cell(row=row, column=1, value=loan.get("loan_id", "")).border = thin_border
        ws.cell(row=row, column=2, value=loan.get("borrower_name", "")).border = thin_border
        ws.cell(row=row, column=3, value=loan.get("amount", 0)).border = thin_border
        ws.cell(row=row, column=4, value=loan.get("currency", "USD")).border = thin_border
        ws.cell(row=row, column=5, value=loan.get("interest_rate", 0)).border = thin_border
        ws.cell(row=row, column=6, value=str(loan.get("loan_date", ""))).border = thin_border
        ws.cell(row=row, column=7, value=str(loan.get("due_date", ""))).border = thin_border
        ws.cell(row=row, column=8, value=outstanding).border = thin_border
        ws.cell(row=row, column=9, value=loan.get("total_repaid", 0)).border = thin_border
        ws.cell(row=row, column=10, value=loan.get("status", "")).border = thin_border
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=loans_export.xlsx"}
    )


@api_router.get("/loans/export/pdf")
async def export_loans_pdf(user: dict = Depends(require_permission(Modules.LOANS, Actions.EXPORT))):
    """Export all loans as PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from io import BytesIO
    
    loans = await db.loans.find({}, {"_id": 0}).sort("loan_date", -1).to_list(50000)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=1, spaceAfter=20)
    
    elements = []
    elements.append(Paragraph("Loans Report", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [["Loan ID", "Borrower", "Amount", "Currency", "Interest", "Due Date", "Outstanding", "Status"]]
    for loan in loans:
        outstanding = loan.get("amount", 0) + loan.get("total_interest", 0) - loan.get("total_repaid", 0)
        data.append([
            loan.get("loan_id", "")[:12],
            loan.get("borrower_name", "")[:20],
            f"${loan.get('amount', 0):,.2f}",
            loan.get("currency", "USD"),
            f"{loan.get('interest_rate', 0)}%",
            str(loan.get("due_date", ""))[:10],
            f"${outstanding:,.2f}",
            loan.get("status", "").replace("_", " ").title()
        ])
    
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0B3D91')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=loans_export.pdf"}
    )

# ============== DEBT MANAGEMENT ==============

class DebtType:
    RECEIVABLE = "receivable"  # Money owed TO us (Debtors)
    PAYABLE = "payable"  # Money owed BY us (Creditors)

class DebtStatus:
    PENDING = "pending"
    PARTIALLY_PAID = "partially_paid"
    FULLY_PAID = "fully_paid"
    OVERDUE = "overdue"

class DebtPartyType:
    CLIENT = "client"
    VENDOR = "vendor"
    OTHER = "other"

class DebtCreate(BaseModel):
    debt_type: str  # receivable or payable
    party_type: str  # client, vendor, or other
    party_id: Optional[str] = None  # client_id or vendor_id if linked
    party_name: str  # Name (auto-filled if linked, manual if other)
    amount: float
    currency: str = "USD"
    due_date: str  # ISO date
    interest_rate: float = 0  # Annual interest rate for overdue
    description: Optional[str] = None
    reference: Optional[str] = None
    treasury_account_id: Optional[str] = None  # For tracking payments

class DebtUpdate(BaseModel):
    party_name: Optional[str] = None
    amount: Optional[float] = None
    due_date: Optional[str] = None
    interest_rate: Optional[float] = None
    description: Optional[str] = None
    reference: Optional[str] = None
    status: Optional[str] = None

class DebtPaymentCreate(BaseModel):
    amount: float
    currency: str = "USD"
    payment_date: Optional[str] = None
    treasury_account_id: str
    reference: Optional[str] = None
    notes: Optional[str] = None

def calculate_debt_interest(principal: float, annual_rate: float, days_overdue: int) -> float:
    """Calculate simple interest on overdue debt"""
    if days_overdue <= 0 or annual_rate <= 0:
        return 0
    daily_rate = annual_rate / 100 / 365
    return round(principal * daily_rate * days_overdue, 2)

def get_debt_status(debt: dict) -> str:
    """Calculate current debt status based on payments and due date"""
    total_paid = debt.get("total_paid", 0)
    amount = debt.get("amount", 0)
    due_date_str = debt.get("due_date")
    
    if total_paid >= amount:
        return DebtStatus.FULLY_PAID
    
    if due_date_str:
        try:
            due_date = datetime.fromisoformat(due_date_str.replace('Z', '+00:00'))
            if due_date.tzinfo is None:
                due_date = due_date.replace(tzinfo=timezone.utc)
            if datetime.now(timezone.utc) > due_date and total_paid < amount:
                return DebtStatus.OVERDUE
        except:
            pass
    
    if total_paid > 0:
        return DebtStatus.PARTIALLY_PAID
    
    return DebtStatus.PENDING

@api_router.get("/debts")
async def get_debts(
    user: dict = Depends(require_permission(Modules.DEBTS, Actions.VIEW)),
    debt_type: Optional[str] = None,
    status: Optional[str] = None,
    party_type: Optional[str] = None
):
    """Get all debts with calculated interest and status"""
    query = {}
    if debt_type:
        query["debt_type"] = debt_type
    if party_type:
        query["party_type"] = party_type
    
    debts = await db.debts.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    now = datetime.now(timezone.utc)
    
    for debt in debts:
        # Calculate current status
        debt["calculated_status"] = get_debt_status(debt)
        
        # Calculate days overdue and interest
        due_date_str = debt.get("due_date")
        if due_date_str:
            try:
                due_date = datetime.fromisoformat(due_date_str.replace('Z', '+00:00'))
                if due_date.tzinfo is None:
                    due_date = due_date.replace(tzinfo=timezone.utc)
                
                if now > due_date:
                    debt["days_overdue"] = (now - due_date).days
                    outstanding = debt.get("amount", 0) - debt.get("total_paid", 0)
                    debt["accrued_interest"] = calculate_debt_interest(
                        outstanding,
                        debt.get("interest_rate", 0),
                        debt["days_overdue"]
                    )
                else:
                    debt["days_overdue"] = 0
                    debt["accrued_interest"] = 0
                    debt["days_until_due"] = (due_date - now).days
            except:
                debt["days_overdue"] = 0
                debt["accrued_interest"] = 0
        
        # Calculate outstanding balance
        debt["outstanding_balance"] = debt.get("amount", 0) - debt.get("total_paid", 0)
        debt["total_due"] = debt["outstanding_balance"] + debt.get("accrued_interest", 0)
    
    # Filter by status if requested (after calculation)
    if status:
        debts = [d for d in debts if d.get("calculated_status") == status]
    
    return debts

@api_router.get("/debts/{debt_id}")
async def get_debt(debt_id: str, user: dict = Depends(require_permission(Modules.DEBTS, Actions.VIEW))):
    """Get single debt with full details and payment history"""
    debt = await db.debts.find_one({"debt_id": debt_id}, {"_id": 0})
    if not debt:
        raise HTTPException(status_code=404, detail="Debt not found")
    
    now = datetime.now(timezone.utc)
    
    # Calculate status and interest
    debt["calculated_status"] = get_debt_status(debt)
    
    due_date_str = debt.get("due_date")
    if due_date_str:
        try:
            due_date = datetime.fromisoformat(due_date_str.replace('Z', '+00:00'))
            if due_date.tzinfo is None:
                due_date = due_date.replace(tzinfo=timezone.utc)
            
            if now > due_date:
                debt["days_overdue"] = (now - due_date).days
                outstanding = debt.get("amount", 0) - debt.get("total_paid", 0)
                debt["accrued_interest"] = calculate_debt_interest(
                    outstanding,
                    debt.get("interest_rate", 0),
                    debt["days_overdue"]
                )
            else:
                debt["days_overdue"] = 0
                debt["accrued_interest"] = 0
                debt["days_until_due"] = (due_date - now).days
        except:
            debt["days_overdue"] = 0
            debt["accrued_interest"] = 0
    
    debt["outstanding_balance"] = debt.get("amount", 0) - debt.get("total_paid", 0)
    debt["total_due"] = debt["outstanding_balance"] + debt.get("accrued_interest", 0)
    
    # Get payment history
    payments = await db.debt_payments.find({"debt_id": debt_id}, {"_id": 0}).sort("payment_date", -1).to_list(100)
    debt["payments"] = payments
    
    return debt

@api_router.post("/debts")
async def create_debt(request: Request, debt_data: DebtCreate, user: dict = Depends(require_permission(Modules.DEBTS, Actions.CREATE))):

    """Create a new debt record"""
    debt_id = f"debt_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    # If linked to client/vendor, verify and get name
    party_name = debt_data.party_name
    if debt_data.party_type == DebtPartyType.CLIENT and debt_data.party_id:
        client = await db.clients.find_one({"client_id": debt_data.party_id}, {"_id": 0})
        if client:
            party_name = f"{client.get('first_name', '')} {client.get('last_name', '')}".strip()
    elif debt_data.party_type == DebtPartyType.VENDOR and debt_data.party_id:
        vendor = await db.vendors.find_one({"vendor_id": debt_data.party_id}, {"_id": 0})
        if vendor:
            party_name = vendor.get("vendor_name", party_name)
    
    debt_doc = {
        "debt_id": debt_id,
        "debt_type": debt_data.debt_type,
        "party_type": debt_data.party_type,
        "party_id": debt_data.party_id,
        "party_name": party_name,
        "amount": debt_data.amount,
        "currency": debt_data.currency,
        "amount_usd": convert_to_usd(debt_data.amount, debt_data.currency),
        "due_date": debt_data.due_date,
        "interest_rate": debt_data.interest_rate,
        "description": debt_data.description,
        "reference": debt_data.reference,
        "treasury_account_id": debt_data.treasury_account_id,
        "total_paid": 0,
        "total_paid_usd": 0,
        "payment_count": 0,
        "status": DebtStatus.PENDING,
        "created_at": now.isoformat(),
        "updated_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    
    await db.debts.insert_one(debt_doc)
    await log_activity(request, user, "create", "debts", "Created debt record")

    return await db.debts.find_one({"debt_id": debt_id}, {"_id": 0})

@api_router.put("/debts/{debt_id}")
async def update_debt(request: Request, debt_id: str, update_data: DebtUpdate, user: dict = Depends(require_permission(Modules.DEBTS, Actions.EDIT))):

    """Update a debt record"""
    debt = await db.debts.find_one({"debt_id": debt_id}, {"_id": 0})
    if not debt:
        raise HTTPException(status_code=404, detail="Debt not found")
    
    updates = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No updates provided")
    
    # Recalculate USD if amount changed
    if "amount" in updates:
        updates["amount_usd"] = convert_to_usd(updates["amount"], debt.get("currency", "USD"))
    
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.debts.update_one({"debt_id": debt_id}, {"$set": updates})
    await log_activity(request, user, "edit", "debts", "Updated debt")

    return await db.debts.find_one({"debt_id": debt_id}, {"_id": 0})

@api_router.delete("/debts/{debt_id}")
async def delete_debt(request: Request, debt_id: str, user: dict = Depends(require_permission(Modules.DEBTS, Actions.DELETE))):

    """Delete a debt record (admin only)"""
    debt = await db.debts.find_one({"debt_id": debt_id}, {"_id": 0})
    if not debt:
        raise HTTPException(status_code=404, detail="Debt not found")
    
    if debt.get("total_paid", 0) > 0:
        raise HTTPException(status_code=400, detail="Cannot delete debt with payments. Mark as cancelled instead.")
    
    await db.debts.delete_one({"debt_id": debt_id})
    await log_activity(request, user, "delete", "debts", "Deleted debt")

    return {"message": "Debt deleted"}

@api_router.post("/debts/{debt_id}/payments")
async def record_debt_payment(

    request: Request,

    debt_id: str,
    payment_data: DebtPaymentCreate,
    user: dict = Depends(require_permission(Modules.DEBTS, Actions.CREATE))
):
    """Record a payment against a debt"""
    debt = await db.debts.find_one({"debt_id": debt_id}, {"_id": 0})
    if not debt:
        raise HTTPException(status_code=404, detail="Debt not found")
    
    # Verify treasury account
    treasury = await db.treasury_accounts.find_one({"account_id": payment_data.treasury_account_id}, {"_id": 0})
    if not treasury:
        raise HTTPException(status_code=404, detail="Treasury account not found")
    
    payment_id = f"dpay_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    payment_date = payment_data.payment_date or now.isoformat()
    
    # Convert payment to USD
    payment_usd = convert_to_usd(payment_data.amount, payment_data.currency)
    
    # Convert payment amount to treasury currency for balance update
    treasury_currency = treasury.get("currency", "USD")
    if payment_data.currency == treasury_currency:
        amount_in_treasury_currency = payment_data.amount
    else:
        # Convert payment to treasury currency via USD
        amount_in_treasury_currency = convert_from_usd(payment_usd, treasury_currency)
    
    payment_doc = {
        "payment_id": payment_id,
        "debt_id": debt_id,
        "debt_type": debt["debt_type"],
        "party_name": debt.get("party_name"),
        "amount": payment_data.amount,
        "currency": payment_data.currency,
        "amount_usd": payment_usd,
        "amount_in_treasury_currency": amount_in_treasury_currency,
        "treasury_currency": treasury_currency,
        "payment_date": payment_date,
        "treasury_account_id": payment_data.treasury_account_id,
        "treasury_account_name": treasury.get("account_name"),
        "reference": payment_data.reference,
        "notes": payment_data.notes,
        "created_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    
    await db.debt_payments.insert_one(payment_doc)
    
    # Determine balance change direction
    if debt["debt_type"] == DebtType.RECEIVABLE:
        # We received money - increase treasury (positive)
        balance_change = amount_in_treasury_currency
        tx_type = "debt_collection"
        tx_description = f"Debt collection from {debt.get('party_name', 'Unknown')}"
    else:
        # We paid money (creditor/payable) - decrease treasury (negative)
        balance_change = -amount_in_treasury_currency
        tx_type = "debt_payment"
        tx_description = f"Debt payment to {debt.get('party_name', 'Unknown')}"
    
    # Update treasury balance
    await db.treasury_accounts.update_one(
        {"account_id": payment_data.treasury_account_id},
        {"$inc": {"balance": balance_change}, "$set": {"updated_at": now.isoformat()}}
    )
    
    # Record treasury transaction for history
    treasury_tx_id = f"ttx_{uuid.uuid4().hex[:12]}"
    treasury_tx = {
        "transaction_id": treasury_tx_id,
        "account_id": payment_data.treasury_account_id,
        "account_name": treasury.get("account_name"),
        "transaction_type": tx_type,
        "amount": abs(amount_in_treasury_currency),
        "currency": treasury_currency,
        "amount_usd": payment_usd,
        "balance_after": treasury.get("balance", 0) + balance_change,
        "description": tx_description,
        "reference": payment_data.reference or payment_id,
        "related_debt_id": debt_id,
        "related_payment_id": payment_id,
        "created_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    await db.treasury_transactions.insert_one(treasury_tx)
    
    # AUTO-CREATE INCOME/EXPENSE ENTRY
    # Receivable payment → Income entry
    # Payable payment → Expense entry
    ie_entry_id = f"ie_{uuid.uuid4().hex[:12]}"
    if debt["debt_type"] == DebtType.RECEIVABLE:
        # Create Income entry when collecting receivable
        income_entry = {
            "entry_id": ie_entry_id,
            "entry_type": "income",
            "category": "Debt Collection",
            "amount": payment_usd,  # Store in USD
            "currency": "USD",
            "date": payment_date.split('T')[0] if 'T' in str(payment_date) else payment_date,
            "description": f"Collection from {debt.get('party_name', 'Unknown')} - {debt.get('description', '')}".strip(),
            "reference": payment_data.reference or payment_id,
            "treasury_account_id": payment_data.treasury_account_id,
            "related_debt_id": debt_id,
            "related_payment_id": payment_id,
            "auto_generated": True,
            "created_at": now.isoformat(),
            "created_by": user["user_id"],
            "created_by_name": user["name"]
        }
        await db.income_expenses.insert_one(income_entry)
    else:
        # Create Expense entry when paying payable
        expense_entry = {
            "entry_id": ie_entry_id,
            "entry_type": "expense",
            "category": "Debt Payment",
            "amount": payment_usd,  # Store in USD
            "currency": "USD",
            "date": payment_date.split('T')[0] if 'T' in str(payment_date) else payment_date,
            "description": f"Payment to {debt.get('party_name', 'Unknown')} - {debt.get('description', '')}".strip(),
            "reference": payment_data.reference or payment_id,
            "treasury_account_id": payment_data.treasury_account_id,
            "related_debt_id": debt_id,
            "related_payment_id": payment_id,
            "auto_generated": True,
            "created_at": now.isoformat(),
            "created_by": user["user_id"],
            "created_by_name": user["name"]
        }
        await db.income_expenses.insert_one(expense_entry)
    
    # Update debt totals
    new_total_paid = debt.get("total_paid", 0) + payment_data.amount
    new_total_paid_usd = debt.get("total_paid_usd", 0) + payment_usd
    new_payment_count = debt.get("payment_count", 0) + 1
    
    # Determine new status
    new_status = DebtStatus.PARTIALLY_PAID
    if new_total_paid >= debt["amount"]:
        new_status = DebtStatus.FULLY_PAID
    
    await db.debts.update_one(
        {"debt_id": debt_id},
        {"$set": {
            "total_paid": new_total_paid,
            "total_paid_usd": new_total_paid_usd,
            "payment_count": new_payment_count,
            "status": new_status,
            "updated_at": now.isoformat(),
            "last_payment_date": payment_date
        }}
    )
    
    await log_activity(request, user, "create", "debts", "Recorded debt payment")

    return await db.debts.find_one({"debt_id": debt_id}, {"_id": 0})

@api_router.get("/debts/{debt_id}/payments")
async def get_debt_payments(debt_id: str, user: dict = Depends(require_permission(Modules.DEBTS, Actions.VIEW))):
    """Get all payments for a debt"""
    payments = await db.debt_payments.find({"debt_id": debt_id}, {"_id": 0}).sort("payment_date", -1).to_list(100)
    return payments

@api_router.get("/debts/summary/overview")
async def get_debts_summary(user: dict = Depends(require_permission(Modules.DEBTS, Actions.VIEW))):
    """Get summary of all debts"""
    now = datetime.now(timezone.utc)
    
    # Get all debts
    debts = await db.debts.find({}, {"_id": 0}).to_list(10000)
    
    summary = {
        "receivables": {
            "total_amount": 0,
            "total_paid": 0,
            "outstanding": 0,
            "overdue_amount": 0,
            "accrued_interest": 0,
            "count": 0,
            "overdue_count": 0
        },
        "payables": {
            "total_amount": 0,
            "total_paid": 0,
            "outstanding": 0,
            "overdue_amount": 0,
            "accrued_interest": 0,
            "count": 0,
            "overdue_count": 0
        },
        "aging": {
            "current": 0,
            "days_1_30": 0,
            "days_31_60": 0,
            "days_61_90": 0,
            "days_over_90": 0
        }
    }
    
    for debt in debts:
        debt_type = debt.get("debt_type", "receivable")
        category = "receivables" if debt_type == "receivable" else "payables"
        amount_usd = debt.get("amount_usd", convert_to_usd(debt.get("amount", 0), debt.get("currency", "USD")))
        paid_usd = debt.get("total_paid_usd", 0)
        outstanding = amount_usd - paid_usd
        
        summary[category]["total_amount"] += amount_usd
        summary[category]["total_paid"] += paid_usd
        summary[category]["outstanding"] += outstanding
        summary[category]["count"] += 1
        
        # Check if overdue
        due_date_str = debt.get("due_date")
        if due_date_str and outstanding > 0:
            try:
                due_date = datetime.fromisoformat(due_date_str.replace('Z', '+00:00'))
                if due_date.tzinfo is None:
                    due_date = due_date.replace(tzinfo=timezone.utc)
                
                if now > due_date:
                    days_overdue = (now - due_date).days
                    summary[category]["overdue_amount"] += outstanding
                    summary[category]["overdue_count"] += 1
                    
                    # Calculate interest
                    interest = calculate_debt_interest(outstanding, debt.get("interest_rate", 0), days_overdue)
                    summary[category]["accrued_interest"] += interest
                    
                    # Aging buckets
                    if days_overdue <= 30:
                        summary["aging"]["days_1_30"] += outstanding
                    elif days_overdue <= 60:
                        summary["aging"]["days_31_60"] += outstanding
                    elif days_overdue <= 90:
                        summary["aging"]["days_61_90"] += outstanding
                    else:
                        summary["aging"]["days_over_90"] += outstanding
                else:
                    summary["aging"]["current"] += outstanding
            except:
                summary["aging"]["current"] += outstanding
    
    # Net position
    summary["net_position"] = summary["receivables"]["outstanding"] - summary["payables"]["outstanding"]
    
    return summary

# ============== COMPREHENSIVE REPORTS ==============

@api_router.get("/reports/transactions-detailed")
async def get_transactions_detailed_report(
    user: dict = Depends(require_permission(Modules.REPORTS, Actions.VIEW)),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    transaction_type: Optional[str] = None,
    currency: Optional[str] = None
):
    """Detailed transaction report with base currency breakdown"""
    query = {"status": {"$in": ["approved", "completed"]}}
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    if transaction_type:
        query["transaction_type"] = transaction_type
    if currency:
        query["$or"] = [{"currency": currency}, {"base_currency": currency}]
    
    # Aggregation for summary by currency
    currency_pipeline = [
        {"$match": query},
        {"$group": {
            "_id": {
                "type": "$transaction_type",
                "currency": {"$ifNull": ["$base_currency", "$currency"]}
            },
            "total_base_amount": {"$sum": {"$ifNull": ["$base_amount", "$amount"]}},
            "total_usd_amount": {"$sum": "$amount"},
            "count": {"$sum": 1}
        }}
    ]
    
    currency_summary = await db.transactions.aggregate(currency_pipeline).to_list(100)
    
    # Build response
    deposits_by_currency = []
    withdrawals_by_currency = []
    total_deposits_usd = 0
    total_withdrawals_usd = 0
    deposit_count = 0
    withdrawal_count = 0
    
    for item in currency_summary:
        entry = {
            "currency": item["_id"]["currency"] or "USD",
            "amount": item["total_base_amount"],
            "usd_equivalent": item["total_usd_amount"],
            "count": item["count"]
        }
        if item["_id"]["type"] == "deposit":
            deposits_by_currency.append(entry)
            total_deposits_usd += item["total_usd_amount"]
            deposit_count += item["count"]
        elif item["_id"]["type"] == "withdrawal":
            withdrawals_by_currency.append(entry)
            total_withdrawals_usd += item["total_usd_amount"]
            withdrawal_count += item["count"]
    
    # Get recent transactions for the table
    transactions = await db.transactions.find(query, {"_id": 0}).sort("created_at", -1).limit(500).to_list(500)
    
    return {
        "summary": {
            "total_deposits_usd": total_deposits_usd,
            "total_withdrawals_usd": total_withdrawals_usd,
            "net_flow_usd": total_deposits_usd - total_withdrawals_usd,
            "deposit_count": deposit_count,
            "withdrawal_count": withdrawal_count,
            "total_count": deposit_count + withdrawal_count
        },
        "deposits_by_currency": deposits_by_currency,
        "withdrawals_by_currency": withdrawals_by_currency,
        "transactions": transactions
    }

@api_router.get("/reports/vendor-summary")
async def get_vendor_summary_report(
    user: dict = Depends(require_permission(Modules.REPORTS, Actions.VIEW)),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    vendor_id: Optional[str] = None
):
    """Vendor settlement report with base currency breakdown (no live FX conversion)"""
    date_filter = {}
    if start_date:
        date_filter["$gte"] = start_date
    if end_date:
        date_filter["$lte"] = end_date

    # Get all vendors
    vendors = await db.vendors.find({}, {"_id": 0}).to_list(1000)
    vendor_map = {v["vendor_id"]: v for v in vendors}

    # --- Transactions: group by vendor + base_currency + type ---
    tx_query = {"vendor_id": {"$exists": True, "$ne": None}, "status": {"$in": ["approved", "completed"]}}
    if date_filter:
        tx_query["created_at"] = date_filter
    if vendor_id:
        tx_query["vendor_id"] = vendor_id

    tx_results = await db.transactions.aggregate([
        {"$match": tx_query},
        {"$group": {
            "_id": {"vendor_id": "$vendor_id", "type": "$transaction_type", "currency": {"$ifNull": ["$base_currency", "$currency"]}},
            "total_base": {"$sum": {"$ifNull": ["$base_amount", "$amount"]}},
            "commission_base": {"$sum": {"$ifNull": ["$vendor_commission_base_amount", 0]}},
            "count": {"$sum": 1}
        }}
    ]).to_list(5000)

    # --- I&E: group by vendor + base_currency + entry_type ---
    ie_query = {"vendor_id": {"$exists": True, "$ne": None}, "status": {"$in": ["approved", "completed"]}, "converted_to_loan": {"$ne": True}}
    if date_filter:
        ie_query["created_at"] = date_filter
    if vendor_id:
        ie_query["vendor_id"] = vendor_id

    ie_results = await db.income_expenses.aggregate([
        {"$match": ie_query},
        {"$group": {
            "_id": {"vendor_id": "$vendor_id", "entry_type": "$entry_type", "currency": {"$ifNull": ["$base_currency", "$currency"]}},
            "total_base": {"$sum": {"$ifNull": ["$base_amount", "$amount"]}},
            "commission_base": {"$sum": {"$ifNull": ["$vendor_commission_base_amount", 0]}},
            "count": {"$sum": 1}
        }}
    ]).to_list(5000)

    # --- Loans: individual docs (need vendor_id logic) ---
    loan_query = {"status": {"$in": ["approved", "completed"]}}
    if date_filter:
        loan_query["created_at"] = date_filter
    loan_txs = await db.loan_transactions.find(loan_query, {"_id": 0, "source_vendor_id": 1, "credit_to_vendor_id": 1, "amount": 1, "currency": 1, "vendor_commission_base_amount": 1}).to_list(5000)

    # Build per-vendor per-currency data
    # Structure: vendor_data[vid][currency] = {deposits, withdrawals, tx_comm, ie_in, ie_out, ie_comm, loan_in, loan_out, loan_comm}
    vendor_data = {}

    def ensure_entry(vid, curr):
        if vid not in vendor_data:
            vendor_data[vid] = {}
        if curr not in vendor_data[vid]:
            vendor_data[vid][curr] = {"deposits": 0, "withdrawals": 0, "tx_comm": 0, "ie_in": 0, "ie_out": 0, "ie_comm": 0, "loan_in": 0, "loan_out": 0, "loan_comm": 0, "dep_count": 0, "wdr_count": 0, "ie_in_count": 0, "ie_out_count": 0, "loan_in_count": 0, "loan_out_count": 0}

    for r in tx_results:
        vid = r["_id"]["vendor_id"]
        curr = r["_id"]["currency"] or "USD"
        ensure_entry(vid, curr)
        d = vendor_data[vid][curr]
        if r["_id"]["type"] == "deposit":
            d["deposits"] += r["total_base"]
            d["dep_count"] += r["count"]
        elif r["_id"]["type"] == "withdrawal":
            d["withdrawals"] += r["total_base"]
            d["wdr_count"] += r["count"]
        d["tx_comm"] += r["commission_base"]

    for r in ie_results:
        vid = r["_id"]["vendor_id"]
        curr = r["_id"]["currency"] or "USD"
        ensure_entry(vid, curr)
        d = vendor_data[vid][curr]
        if r["_id"]["entry_type"] == "income":
            d["ie_in"] += r["total_base"]
            d["ie_in_count"] += r["count"]
        else:
            d["ie_out"] += r["total_base"]
            d["ie_out_count"] += r["count"]
        d["ie_comm"] += r["commission_base"]

    for ltx in loan_txs:
        vid = None
        is_in = False
        if ltx.get("credit_to_vendor_id"):
            vid = ltx["credit_to_vendor_id"]
            is_in = True
        elif ltx.get("source_vendor_id"):
            vid = ltx["source_vendor_id"]
        if not vid:
            continue
        if vendor_id and vid != vendor_id:
            continue
        curr = ltx.get("currency", "USD")
        ensure_entry(vid, curr)
        d = vendor_data[vid][curr]
        amt = ltx.get("amount", 0)
        comm = ltx.get("vendor_commission_base_amount", 0) or 0
        if is_in:
            d["loan_in"] += amt
            d["loan_in_count"] += 1
        else:
            d["loan_out"] += amt
            d["loan_out_count"] += 1
        d["loan_comm"] += comm

    # Build response: one row per vendor per currency
    vendor_list = []
    for vid, currencies in vendor_data.items():
        vi = vendor_map.get(vid, {})
        rows = []
        for curr, d in currencies.items():
            money_in = d["deposits"] + d["ie_in"] + d["loan_in"]
            money_out = d["withdrawals"] + d["ie_out"] + d["loan_out"]
            total_comm = d["tx_comm"] + d["ie_comm"] + d["loan_comm"]
            net = money_in - money_out - total_comm
            rows.append({
                "currency": curr,
                "deposits": round(d["deposits"], 2),
                "withdrawals": round(d["withdrawals"], 2),
                "tx_commission": round(d["tx_comm"], 2),
                "ie_in": round(d["ie_in"], 2),
                "ie_out": round(d["ie_out"], 2),
                "ie_commission": round(d["ie_comm"], 2),
                "loan_in": round(d["loan_in"], 2),
                "loan_out": round(d["loan_out"], 2),
                "loan_commission": round(d["loan_comm"], 2),
                "money_in": round(money_in, 2),
                "money_out": round(money_out, 2),
                "total_commission": round(total_comm, 2),
                "net_settlement": round(net, 2),
                "tx_count": d["dep_count"] + d["wdr_count"] + d["ie_in_count"] + d["ie_out_count"] + d["loan_in_count"] + d["loan_out_count"]
            })
        vendor_list.append({
            "vendor_id": vid,
            "vendor_name": vi.get("vendor_name", "Unknown"),
            "deposit_commission_rate": vi.get("deposit_commission", 0),
            "withdrawal_commission_rate": vi.get("withdrawal_commission", 0),
            "currency_rows": rows
        })

    # Grand totals per currency
    grand_by_currency = {}
    for v in vendor_list:
        for r in v["currency_rows"]:
            c = r["currency"]
            if c not in grand_by_currency:
                grand_by_currency[c] = {"money_in": 0, "money_out": 0, "total_commission": 0, "net_settlement": 0}
            grand_by_currency[c]["money_in"] += r["money_in"]
            grand_by_currency[c]["money_out"] += r["money_out"]
            grand_by_currency[c]["total_commission"] += r["total_commission"]
            grand_by_currency[c]["net_settlement"] += r["net_settlement"]

    return {
        "vendors": vendor_list,
        "grand_totals_by_currency": grand_by_currency,
        "grand_totals": {
            "total_exchangers": len(vendor_list),
            "total_vendors": len(vendor_list)
        }
    }

@api_router.get("/reports/vendor-commissions")
async def get_vendor_commissions_report(
    user: dict = Depends(require_permission(Modules.REPORTS, Actions.VIEW)),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """Detailed vendor commission report"""
    query = {
        "vendor_id": {"$exists": True, "$ne": None},
        "vendor_commission_amount": {"$gt": 0},
        "status": {"$in": ["approved", "completed"]}
    }
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    # Get all vendors
    vendors = await db.vendors.find({}, {"_id": 0}).to_list(1000)
    vendor_map = {v["vendor_id"]: v for v in vendors}
    
    # Get commission transactions
    transactions = await db.transactions.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Aggregate by vendor
    vendor_commissions = {}
    for tx in transactions:
        vid = tx.get("vendor_id")
        if vid not in vendor_commissions:
            vendor_info = vendor_map.get(vid, {})
            vendor_commissions[vid] = {
                "vendor_id": vid,
                "vendor_name": vendor_info.get("vendor_name", "Unknown"),
                "total_commission_usd": 0,
                "commission_by_currency": {},
                "deposit_commissions": 0,
                "withdrawal_commissions": 0,
                "transaction_count": 0
            }
        
        comm_usd = tx.get("vendor_commission_amount", 0)
        comm_base = tx.get("vendor_commission_base_amount", 0)
        currency = tx.get("base_currency") or tx.get("currency", "USD")
        
        vendor_commissions[vid]["total_commission_usd"] += comm_usd
        vendor_commissions[vid]["transaction_count"] += 1
        
        if tx.get("transaction_type") == "deposit":
            vendor_commissions[vid]["deposit_commissions"] += comm_usd
        else:
            vendor_commissions[vid]["withdrawal_commissions"] += comm_usd
        
        if currency not in vendor_commissions[vid]["commission_by_currency"]:
            vendor_commissions[vid]["commission_by_currency"][currency] = {"base": 0, "usd": 0}
        vendor_commissions[vid]["commission_by_currency"][currency]["base"] += comm_base
        vendor_commissions[vid]["commission_by_currency"][currency]["usd"] += comm_usd
    
    return {
        "vendors": list(vendor_commissions.values()),
        "total_commission_usd": sum(v["total_commission_usd"] for v in vendor_commissions.values()),
        "transactions": transactions[:100]  # Return last 100 commission transactions
    }

@api_router.get("/reports/client-balances")
async def get_client_balances_report(
    user: dict = Depends(require_permission(Modules.REPORTS, Actions.VIEW)),
    min_balance: Optional[float] = None,
    max_balance: Optional[float] = None,
    sort_by: str = "net_balance",
    sort_order: str = "desc"
):
    """Client balance summary report"""
    clients = await db.clients.find({}, {"_id": 0}).to_list(10000)
    
    # Get transaction summaries for all clients
    pipeline = [
        {"$match": {"status": {"$in": ["approved", "completed"]}}},
        {"$group": {
            "_id": {
                "client_id": "$client_id",
                "type": "$transaction_type",
                "currency": {"$ifNull": ["$base_currency", "$currency"]}
            },
            "total_base": {"$sum": {"$ifNull": ["$base_amount", "$amount"]}},
            "total_usd": {"$sum": "$amount"},
            "count": {"$sum": 1}
        }}
    ]
    
    tx_data = await db.transactions.aggregate(pipeline).to_list(100000)
    
    # Build client map
    client_tx_map = {}
    for item in tx_data:
        cid = item["_id"]["client_id"]
        if cid not in client_tx_map:
            client_tx_map[cid] = {
                "deposits_usd": 0,
                "withdrawals_usd": 0,
                "deposit_count": 0,
                "withdrawal_count": 0,
                "by_currency": {}
            }
        
        tx_type = item["_id"]["type"]
        currency = item["_id"]["currency"] or "USD"
        
        if currency not in client_tx_map[cid]["by_currency"]:
            client_tx_map[cid]["by_currency"][currency] = {"deposits": 0, "withdrawals": 0}
        
        if tx_type == "deposit":
            client_tx_map[cid]["deposits_usd"] += item["total_usd"]
            client_tx_map[cid]["deposit_count"] += item["count"]
            client_tx_map[cid]["by_currency"][currency]["deposits"] += item["total_base"]
        elif tx_type == "withdrawal":
            client_tx_map[cid]["withdrawals_usd"] += item["total_usd"]
            client_tx_map[cid]["withdrawal_count"] += item["count"]
            client_tx_map[cid]["by_currency"][currency]["withdrawals"] += item["total_base"]
    
    # Enrich clients with transaction data
    results = []
    for client in clients:
        cid = client["client_id"]
        tx_info = client_tx_map.get(cid, {
            "deposits_usd": 0, "withdrawals_usd": 0,
            "deposit_count": 0, "withdrawal_count": 0, "by_currency": {}
        })
        
        net_balance = tx_info["deposits_usd"] - tx_info["withdrawals_usd"]
        
        # Apply filters
        if min_balance is not None and net_balance < min_balance:
            continue
        if max_balance is not None and net_balance > max_balance:
            continue
        
        results.append({
            "client_id": cid,
            "name": f"{client.get('first_name', '')} {client.get('last_name', '')}".strip(),
            "email": client.get("email"),
            "country": client.get("country"),
            "kyc_status": client.get("kyc_status"),
            "total_deposits_usd": tx_info["deposits_usd"],
            "total_withdrawals_usd": tx_info["withdrawals_usd"],
            "net_balance": net_balance,
            "deposit_count": tx_info["deposit_count"],
            "withdrawal_count": tx_info["withdrawal_count"],
            "transaction_count": tx_info["deposit_count"] + tx_info["withdrawal_count"],
            "by_currency": tx_info["by_currency"]
        })
    
    # Sort
    reverse = sort_order == "desc"
    results.sort(key=lambda x: x.get(sort_by, 0) or 0, reverse=reverse)
    
    # Summary stats
    total_deposits = sum(r["total_deposits_usd"] for r in results)
    total_withdrawals = sum(r["total_withdrawals_usd"] for r in results)
    
    return {
        "clients": results,
        "summary": {
            "total_clients": len(results),
            "total_deposits_usd": total_deposits,
            "total_withdrawals_usd": total_withdrawals,
            "total_net_balance": total_deposits - total_withdrawals,
            "active_clients": len([r for r in results if r["transaction_count"] > 0])
        }
    }

@api_router.get("/reports/treasury-summary")
async def get_treasury_summary_report(
    user: dict = Depends(require_permission(Modules.REPORTS, Actions.VIEW)),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """Treasury balances and transaction summary"""
    # Get all treasury accounts
    accounts = await db.treasury_accounts.find({}, {"_id": 0}).to_list(1000)
    
    # Calculate USD equivalents
    total_balance_usd = 0
    accounts_summary = []
    
    for acc in accounts:
        balance = acc.get("balance", 0)
        currency = acc.get("currency", "USD")
        balance_usd = convert_to_usd(balance, currency)
        total_balance_usd += balance_usd
        
        accounts_summary.append({
            "account_id": acc["account_id"],
            "account_name": acc.get("account_name"),
            "account_type": acc.get("account_type"),
            "bank_name": acc.get("bank_name"),
            "currency": currency,
            "balance": balance,
            "balance_usd": balance_usd,
            "status": acc.get("status")
        })
    
    # Group by currency
    balance_by_currency = {}
    for acc in accounts:
        currency = acc.get("currency", "USD")
        if currency not in balance_by_currency:
            balance_by_currency[currency] = {"total": 0, "count": 0}
        balance_by_currency[currency]["total"] += acc.get("balance", 0)
        balance_by_currency[currency]["count"] += 1
    
    # Get transfer history
    query = {}
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    transfers = await db.treasury_transactions.find(
        {**query, "transaction_type": {"$in": ["transfer_in", "transfer_out"]}},
        {"_id": 0}
    ).sort("created_at", -1).limit(100).to_list(100)
    
    return {
        "accounts": accounts_summary,
        "total_balance_usd": total_balance_usd,
        "balance_by_currency": [
            {"currency": k, "total": v["total"], "account_count": v["count"]}
            for k, v in balance_by_currency.items()
        ],
        "recent_transfers": transfers
    }

@api_router.get("/reports/psp-summary")
async def get_psp_summary_report(
    user: dict = Depends(require_permission(Modules.REPORTS, Actions.VIEW)),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """PSP transaction and settlement summary"""
    # Get all PSPs
    psps = await db.psps.find({}, {"_id": 0}).to_list(1000)
    
    # Build query for transactions
    query = {"destination_type": "psp"}
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = end_date
        else:
            query["created_at"] = {"$lte": end_date}
    
    # Aggregate by PSP
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": "$psp_id",
            "total_volume": {"$sum": "$amount"},
            "total_commission": {"$sum": {"$ifNull": ["$psp_commission_amount", 0]}},
            "total_extra_charges": {"$sum": {"$ifNull": ["$psp_extra_charges", 0]}},
            "total_reserve": {"$sum": {"$ifNull": ["$psp_reserve_fund_amount", 0]}},
            "total_net": {"$sum": {"$ifNull": ["$psp_net_amount", "$amount"]}},
            "total_base_amount": {"$sum": {"$ifNull": ["$base_amount", 0]}},
            "base_currencies": {"$addToSet": "$base_currency"},
            "settled_count": {"$sum": {"$cond": [{"$eq": ["$settled", True]}, 1, 0]}},
            "pending_count": {"$sum": {"$cond": [{"$ne": ["$settled", True]}, 1, 0]}},
            "transaction_count": {"$sum": 1}
        }}
    ]
    
    psp_stats = await db.transactions.aggregate(pipeline).to_list(100)
    psp_map = {p["psp_id"]: p for p in psps}
    
    results = []
    for stat in psp_stats:
        psp_id = stat["_id"]
        psp_info = psp_map.get(psp_id, {})
        
        # Determine primary payment currency (exclude USD and None)
        base_currencies = [c for c in stat.get("base_currencies", []) if c and c != "USD"]
        pay_currency = base_currencies[0] if len(base_currencies) == 1 else (", ".join(base_currencies) if base_currencies else None)
        total_base = stat.get("total_base_amount", 0) if base_currencies else None
        
        results.append({
            "psp_id": psp_id,
            "psp_name": psp_info.get("psp_name", "Unknown"),
            "commission_rate": psp_info.get("commission_rate", 0),
            "total_volume": stat["total_volume"],
            "total_commission": stat["total_commission"],
            "total_extra_charges": stat.get("total_extra_charges", 0),
            "total_reserve": stat.get("total_reserve", 0),
            "total_net": stat["total_net"],
            "payment_currency": pay_currency,
            "total_base_volume": round(total_base, 2) if total_base else None,
            "settled_count": stat["settled_count"],
            "pending_count": stat["pending_count"],
            "transaction_count": stat["transaction_count"]
        })
    
    return {
        "psps": results,
        "grand_totals": {
            "total_volume": sum(r["total_volume"] for r in results),
            "total_commission": sum(r["total_commission"] for r in results),
            "total_net": sum(r["total_net"] for r in results),
            "total_transactions": sum(r["transaction_count"] for r in results)
        }
    }

@api_router.get("/reports/financial-summary")
async def get_financial_summary_report(
    user: dict = Depends(require_permission(Modules.REPORTS, Actions.VIEW)),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    """P&L and financial summary combining all data sources"""
    query = {}
    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        if "date" in query:
            query["date"]["$lte"] = end_date
        else:
            query["date"] = {"$lte": end_date}
    
    # Get income/expense summary
    ie_pipeline = [
        {"$match": query} if query else {"$match": {}},
        {"$group": {
            "_id": {"type": "$entry_type", "category": "$category"},
            "total": {"$sum": "$amount"},
            "count": {"$sum": 1}
        }}
    ]
    
    ie_data = await db.income_expenses.aggregate(ie_pipeline).to_list(100)
    
    income_by_category = {}
    expense_by_category = {}
    total_income = 0
    total_expense = 0
    
    for item in ie_data:
        category = item["_id"]["category"]
        entry_type = item["_id"]["type"]
        
        if entry_type == "income":
            income_by_category[category] = item["total"]
            total_income += item["total"]
        else:
            expense_by_category[category] = item["total"]
            total_expense += item["total"]
    
    # Get loan summary
    loans = await db.loans.find({}, {"_id": 0}).to_list(1000)
    total_loan_disbursed = sum(l.get("amount", 0) for l in loans)
    total_loan_outstanding = sum(l.get("outstanding_balance", l.get("amount", 0)) for l in loans)
    total_loan_repaid = sum(l.get("total_repaid", 0) for l in loans)
    
    # Get treasury total
    accounts = await db.treasury_accounts.find({"status": "active"}, {"_id": 0}).to_list(1000)
    total_treasury_usd = sum(convert_to_usd(a.get("balance", 0), a.get("currency", "USD")) for a in accounts)
    
    # Get vendor commission totals
    vendor_comm_pipeline = [
        {"$match": {"vendor_commission_amount": {"$gt": 0}}},
        {"$group": {"_id": None, "total": {"$sum": "$vendor_commission_amount"}}}
    ]
    vendor_comm = await db.transactions.aggregate(vendor_comm_pipeline).to_list(1)
    total_vendor_commission = vendor_comm[0]["total"] if vendor_comm else 0
    
    return {
        "income": {
            "total": total_income,
            "by_category": [{"category": k, "amount": v} for k, v in income_by_category.items()]
        },
        "expenses": {
            "total": total_expense,
            "by_category": [{"category": k, "amount": v} for k, v in expense_by_category.items()]
        },
        "net_profit_loss": total_income - total_expense,
        "loans": {
            "total_disbursed": total_loan_disbursed,
            "total_outstanding": total_loan_outstanding,
            "total_repaid": total_loan_repaid,
            "active_loans": len([l for l in loans if l.get("status") != "fully_paid"])
        },
        "treasury": {
            "total_balance_usd": total_treasury_usd,
            "account_count": len(accounts)
        },
        "vendor_commissions": {
            "total_paid": total_vendor_commission
        }
    }

# ============== RECONCILIATION MODULE ==============

import csv
import io
from openpyxl import load_workbook

class ReconciliationStatus(str, Enum):
    PENDING = "pending"
    MATCHED = "matched"
    PARTIAL = "partial"
    UNMATCHED = "unmatched"
    DISCREPANCY = "discrepancy"

class ReconciliationType(str, Enum):
    BANK = "bank"
    PSP = "psp"
    CLIENT = "client"
    VENDOR = "vendor"

# Bank Statement Upload and Reconciliation
@api_router.post("/reconciliation/bank/upload")
async def upload_bank_statement(

    request: Request,

    account_id: str = Form(...),
    file: UploadFile = File(...),
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.CREATE))
):
    """Upload bank statement CSV/Excel/PDF for reconciliation"""
    import pdfplumber
    
    # Verify treasury account exists
    account = await db.treasury_accounts.find_one({"account_id": account_id}, {"_id": 0})
    if not account:
        raise HTTPException(status_code=404, detail="Treasury account not found")
    
    # Read file content
    content = await file.read()
    filename = file.filename.lower()
    
    parsed_rows = []
    
    try:
        if filename.endswith('.csv'):
            # Parse CSV
            decoded = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(decoded))
            for row in reader:
                parsed_rows.append(dict(row))
        elif filename.endswith(('.xlsx', '.xls')):
            # Parse Excel
            wb = load_workbook(filename=io.BytesIO(content), read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                parsed_rows.append(row_dict)
        elif filename.endswith('.pdf'):
            # Parse PDF using pdfplumber
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                for page in pdf.pages:
                    tables = page.extract_tables()
                    for table in tables:
                        if table and len(table) > 1:
                            # First row is header
                            headers = table[0]
                            for row in table[1:]:
                                if row and any(cell for cell in row):
                                    row_dict = {}
                                    for i, cell in enumerate(row):
                                        if i < len(headers) and headers[i]:
                                            row_dict[headers[i]] = cell
                                    if row_dict:
                                        parsed_rows.append(row_dict)
                    
                    # If no tables found, try to extract text
                    if not tables:
                        text = page.extract_text()
                        if text:
                            lines = text.split('\n')
                            for line in lines:
                                parts = line.split()
                                if len(parts) >= 3:
                                    # Try to detect amount-like values
                                    for i, part in enumerate(parts):
                                        try:
                                            amount = float(part.replace(',', '').replace('$', ''))
                                            parsed_rows.append({
                                                "description": ' '.join(parts[:i]),
                                                "amount": amount,
                                                "raw_line": line
                                            })
                                            break
                                        except ValueError:
                                            continue
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Use CSV, Excel, or PDF.")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")
    
    if not parsed_rows:
        raise HTTPException(status_code=400, detail="No data found in file")
    
    now = datetime.now(timezone.utc)
    batch_id = f"recon_{uuid.uuid4().hex[:12]}"
    
    # Create reconciliation batch record
    batch_doc = {
        "batch_id": batch_id,
        "type": "bank",
        "account_id": account_id,
        "account_name": account["account_name"],
        "filename": file.filename,
        "total_rows": len(parsed_rows),
        "matched": 0,
        "unmatched": 0,
        "discrepancies": 0,
        "status": "processing",
        "created_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    await db.reconciliation_batches.insert_one(batch_doc)
    
    # Process each row and try to match with treasury transactions
    matched_count = 0
    unmatched_count = 0
    discrepancy_count = 0
    
    for idx, row in enumerate(parsed_rows):
        # Try to extract amount, date, reference from common column names
        amount = None
        date = None
        reference = None
        description = None
        
        for key, value in row.items():
            key_lower = key.lower() if key else ""
            if any(x in key_lower for x in ['amount', 'value', 'debit', 'credit', 'sum']):
                try:
                    # Clean the amount string
                    clean_val = str(value).replace(',', '').replace('$', '').replace(' ', '')
                    amount = float(clean_val) if clean_val and clean_val != 'None' else None
                except:
                    pass
            elif any(x in key_lower for x in ['date', 'time', 'posted']):
                date = str(value) if value else None
            elif any(x in key_lower for x in ['reference', 'ref', 'txn', 'transaction', 'id']):
                reference = str(value) if value else None
            elif any(x in key_lower for x in ['description', 'desc', 'narration', 'details', 'memo']):
                description = str(value) if value else None
        
        # Create statement entry
        entry_id = f"stmt_{uuid.uuid4().hex[:12]}"
        entry_doc = {
            "entry_id": entry_id,
            "batch_id": batch_id,
            "account_id": account_id,
            "row_number": idx + 1,
            "raw_data": row,
            "parsed_amount": amount,
            "parsed_date": date,
            "parsed_reference": reference,
            "parsed_description": description,
            "status": ReconciliationStatus.PENDING,
            "matched_transaction_id": None,
            "variance": None,
            "created_at": now.isoformat()
        }
        
        # Try to find matching treasury transaction
        if amount is not None:
            # Look for matching transaction by amount and optionally reference
            match_query = {
                "account_id": account_id,
                "amount": {"$gte": abs(amount) - 0.01, "$lte": abs(amount) + 0.01}  # Allow small variance
            }
            if reference:
                match_query["$or"] = [
                    {"reference": {"$regex": reference, "$options": "i"}},
                    {"treasury_transaction_id": {"$regex": reference, "$options": "i"}}
                ]
            
            potential_match = await db.treasury_transactions.find_one(match_query, {"_id": 0})
            
            if potential_match:
                # Calculate variance
                variance = abs(amount) - potential_match.get("amount", 0)
                
                if abs(variance) < 0.01:
                    entry_doc["status"] = ReconciliationStatus.MATCHED
                    entry_doc["matched_transaction_id"] = potential_match["treasury_transaction_id"]
                    matched_count += 1
                else:
                    entry_doc["status"] = ReconciliationStatus.DISCREPANCY
                    entry_doc["matched_transaction_id"] = potential_match["treasury_transaction_id"]
                    entry_doc["variance"] = variance
                    discrepancy_count += 1
            else:
                entry_doc["status"] = ReconciliationStatus.UNMATCHED
                unmatched_count += 1
        else:
            entry_doc["status"] = ReconciliationStatus.UNMATCHED
            unmatched_count += 1
        
        await db.reconciliation_entries.insert_one(entry_doc)
    
    # Update batch with results
    await db.reconciliation_batches.update_one(
        {"batch_id": batch_id},
        {"$set": {
            "matched": matched_count,
            "unmatched": unmatched_count,
            "discrepancies": discrepancy_count,
            "status": "completed",
            "completed_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    await log_activity(request, user, "create", "reconciliation", "Uploaded bank statement")

    return {
        "batch_id": batch_id,
        "total_rows": len(parsed_rows),
        "matched": matched_count,
        "unmatched": unmatched_count,
        "discrepancies": discrepancy_count,
        "columns_detected": list(parsed_rows[0].keys()) if parsed_rows else []
    }

@api_router.get("/reconciliation/batches")
async def get_reconciliation_batches(
    type: Optional[str] = None,
    limit: int = 50,
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))
):
    """Get reconciliation batch history"""
    query = {}
    if type:
        query["type"] = type
    
    batches = await db.reconciliation_batches.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return batches

@api_router.get("/reconciliation/batch/{batch_id}")
async def get_reconciliation_batch(batch_id: str, user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))):
    """Get reconciliation batch details with entries"""
    batch = await db.reconciliation_batches.find_one({"batch_id": batch_id}, {"_id": 0})
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    
    entries = await db.reconciliation_entries.find({"batch_id": batch_id}, {"_id": 0}).to_list(10000)
    
    return {
        "batch": batch,
        "entries": entries
    }

@api_router.put("/reconciliation/entry/{entry_id}/match")
async def manually_match_entry(
    entry_id: str,
    transaction_id: str,
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.EDIT))
):
    """Manually match a reconciliation entry to a transaction"""
    entry = await db.reconciliation_entries.find_one({"entry_id": entry_id}, {"_id": 0})
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    # Verify transaction exists
    tx = await db.treasury_transactions.find_one({"treasury_transaction_id": transaction_id}, {"_id": 0})
    if not tx:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    variance = (entry.get("parsed_amount") or 0) - tx.get("amount", 0)
    
    await db.reconciliation_entries.update_one(
        {"entry_id": entry_id},
        {"$set": {
            "status": ReconciliationStatus.MATCHED if abs(variance) < 0.01 else ReconciliationStatus.DISCREPANCY,
            "matched_transaction_id": transaction_id,
            "variance": variance,
            "matched_by": user["user_id"],
            "matched_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # Update batch counts
    batch = await db.reconciliation_batches.find_one({"batch_id": entry["batch_id"]}, {"_id": 0})
    if batch:
        await db.reconciliation_batches.update_one(
            {"batch_id": entry["batch_id"]},
            {"$inc": {"matched": 1, "unmatched": -1}}
        )
    
    return {"message": "Entry matched successfully"}

@api_router.put("/reconciliation/entry/{entry_id}/ignore")
async def ignore_reconciliation_entry(entry_id: str, reason: str = "", user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.EDIT))):
    """Mark an entry as ignored/resolved"""
    await db.reconciliation_entries.update_one(
        {"entry_id": entry_id},
        {"$set": {
            "status": "ignored",
            "ignore_reason": reason,
            "ignored_by": user["user_id"],
            "ignored_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "Entry ignored"}

# PSP Reconciliation
@api_router.get("/reconciliation/psp")
async def get_psp_reconciliation(
    psp_id: Optional[str] = None,
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))
):
    """Get PSP reconciliation summary - expected vs actual settlements"""
    query = {"destination_type": "psp", "settled": True}
    if psp_id:
        query["psp_id"] = psp_id
    
    settled_txs = await db.transactions.find(query, {"_id": 0}).to_list(10000)
    
    # Group by PSP
    psp_summary = {}
    for tx in settled_txs:
        pid = tx.get("psp_id", "unknown")
        if pid not in psp_summary:
            psp_summary[pid] = {
                "psp_id": pid,
                "psp_name": tx.get("psp_name", "Unknown"),
                "total_transactions": 0,
                "expected_amount": 0,
                "actual_amount": 0,
                "total_variance": 0,
                "transactions_with_variance": 0
            }
        
        psp_summary[pid]["total_transactions"] += 1
        expected = tx.get("psp_net_amount", tx.get("amount", 0))
        actual = tx.get("psp_actual_amount_received", expected)
        variance = tx.get("psp_settlement_variance", 0)
        
        psp_summary[pid]["expected_amount"] += expected
        psp_summary[pid]["actual_amount"] += actual
        psp_summary[pid]["total_variance"] += variance
        if abs(variance) > 0.01:
            psp_summary[pid]["transactions_with_variance"] += 1
    
    return list(psp_summary.values())

@api_router.get("/reconciliation/psp/{psp_id}/details")
async def get_psp_reconciliation_details(psp_id: str, user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))):
    """Get detailed PSP reconciliation with individual transaction variances"""
    txs = await db.transactions.find({
        "destination_type": "psp",
        "psp_id": psp_id,
        "settled": True
    }, {"_id": 0}).sort("settled_at", -1).to_list(1000)
    
    result = []
    for tx in txs:
        expected = tx.get("psp_net_amount", tx.get("amount", 0))
        actual = tx.get("psp_actual_amount_received", expected)
        variance = tx.get("psp_settlement_variance", actual - expected)
        
        result.append({
            "transaction_id": tx.get("transaction_id"),
            "reference": tx.get("reference"),
            "client_name": tx.get("client_name"),
            "gross_amount": tx.get("amount"),
            "commission": tx.get("psp_commission_amount", 0),
            "chargeback": tx.get("psp_reserve_fund_amount", tx.get("psp_chargeback_amount", 0)),
            "extra_charges": tx.get("psp_extra_charges", 0),
            "expected_net": expected,
            "actual_received": actual,
            "variance": variance,
            "settled_at": tx.get("settled_at"),
            "status": "matched" if abs(variance) < 0.01 else "variance"
        })
    
    return result

# Client Account Reconciliation
@api_router.get("/reconciliation/clients")
async def get_client_reconciliation(user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))):
    """Get client account reconciliation - verify balances match transactions"""
    clients = await db.clients.find({}, {"_id": 0}).to_list(10000)
    
    result = []
    for client in clients:
        client_id = client.get("client_id")
        
        # Get all transactions for this client
        txs = await db.transactions.find({
            "client_id": client_id,
            "status": {"$in": ["approved", "completed"]}
        }, {"_id": 0}).to_list(10000)
        
        # Calculate expected balance from transactions
        calculated_balance = 0
        for tx in txs:
            amount = tx.get("amount", 0)
            if tx.get("transaction_type") == "deposit":
                calculated_balance += amount
            elif tx.get("transaction_type") == "withdrawal":
                calculated_balance -= amount
        
        recorded_balance = client.get("balance", 0)
        variance = recorded_balance - calculated_balance
        
        result.append({
            "client_id": client_id,
            "client_name": f"{client.get('first_name', '')} {client.get('last_name', '')}".strip(),
            "recorded_balance": recorded_balance,
            "calculated_balance": calculated_balance,
            "variance": variance,
            "transaction_count": len(txs),
            "status": "matched" if abs(variance) < 0.01 else "discrepancy"
        })
    
    # Sort by variance (largest first)
    result.sort(key=lambda x: abs(x["variance"]), reverse=True)
    return result

@api_router.get("/reconciliation/client/{client_id}/details")
async def get_client_reconciliation_details(client_id: str, user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))):
    """Get detailed transaction history for client reconciliation"""
    client = await db.clients.find_one({"client_id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    txs = await db.transactions.find({
        "client_id": client_id,
        "status": {"$in": ["approved", "completed"]}
    }, {"_id": 0}).sort("created_at", 1).to_list(10000)
    
    # Build running balance history
    running_balance = 0
    history = []
    for tx in txs:
        amount = tx.get("amount", 0)
        if tx.get("transaction_type") == "deposit":
            running_balance += amount
        elif tx.get("transaction_type") == "withdrawal":
            running_balance -= amount
        
        history.append({
            "transaction_id": tx.get("transaction_id"),
            "date": tx.get("created_at"),
            "type": tx.get("transaction_type"),
            "amount": amount,
            "running_balance": running_balance,
            "reference": tx.get("reference")
        })
    
    return {
        "client": {
            "client_id": client_id,
            "name": f"{client.get('first_name', '')} {client.get('last_name', '')}".strip(),
            "recorded_balance": client.get("balance", 0),
            "calculated_balance": running_balance,
            "variance": client.get("balance", 0) - running_balance
        },
        "transactions": history
    }

# Vendor Reconciliation
@api_router.get("/reconciliation/vendors")
async def get_vendor_reconciliation(user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))):
    """Get vendor reconciliation - commission calculations vs payments"""
    vendors = await db.vendors.find({}, {"_id": 0}).to_list(1000)
    
    result = []
    for vendor in vendors:
        vendor_id = vendor.get("vendor_id")
        
        # Get all transactions through this vendor
        txs = await db.transactions.find({
            "vendor_id": vendor_id,
            "status": {"$in": ["approved", "completed"]}
        }, {"_id": 0}).to_list(10000)
        
        # Calculate expected commission
        total_volume = sum(tx.get("amount", 0) for tx in txs)
        commission_rate = vendor.get("commission_rate", 0) / 100
        expected_commission = total_volume * commission_rate
        
        # Get recorded commission
        paid_commission = vendor.get("total_commission_paid", 0)
        pending_commission = vendor.get("pending_commission", 0)
        total_recorded = paid_commission + pending_commission
        
        variance = expected_commission - total_recorded
        
        result.append({
            "vendor_id": vendor_id,
            "vendor_name": vendor.get("vendor_name") or vendor.get("name", "Unknown"),
            "commission_rate": vendor.get("commission_rate", 0),
            "total_volume": total_volume,
            "transaction_count": len(txs),
            "expected_commission": expected_commission,
            "paid_commission": paid_commission,
            "pending_commission": pending_commission,
            "total_recorded": total_recorded,
            "variance": variance,
            "status": "matched" if abs(variance) < 0.01 else "discrepancy"
        })
    
    result.sort(key=lambda x: abs(x["variance"]), reverse=True)
    return result

# Reconciliation Dashboard Summary
@api_router.get("/reconciliation/summary")
async def get_reconciliation_summary(user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))):
    """Get overall reconciliation status summary"""
    
    # Bank reconciliation status
    bank_batches = await db.reconciliation_batches.find({"type": "bank"}, {"_id": 0}).sort("created_at", -1).to_list(10)
    total_unmatched_bank = sum(b.get("unmatched", 0) for b in bank_batches)
    total_discrepancies_bank = sum(b.get("discrepancies", 0) for b in bank_batches)
    
    # PSP reconciliation
    psp_recon = await get_psp_reconciliation(None, user)
    total_psp_variance = sum(p.get("total_variance", 0) for p in psp_recon)
    psp_with_variance = sum(1 for p in psp_recon if abs(p.get("total_variance", 0)) > 0.01)
    
    # Client reconciliation
    client_recon = await get_client_reconciliation(user)
    clients_with_discrepancy = sum(1 for c in client_recon if c["status"] == "discrepancy")
    total_client_variance = sum(abs(c["variance"]) for c in client_recon)
    
    # Vendor reconciliation
    vendor_recon = await get_vendor_reconciliation(user)
    vendors_with_discrepancy = sum(1 for v in vendor_recon if v["status"] == "discrepancy")
    total_vendor_variance = sum(abs(v["variance"]) for v in vendor_recon)
    
    return {
        "bank": {
            "recent_batches": len(bank_batches),
            "unmatched_entries": total_unmatched_bank,
            "discrepancies": total_discrepancies_bank,
            "status": "attention" if total_unmatched_bank > 0 or total_discrepancies_bank > 0 else "ok"
        },
        "psp": {
            "total_psps": len(psp_recon),
            "psps_with_variance": psp_with_variance,
            "total_variance": total_psp_variance,
            "status": "attention" if psp_with_variance > 0 else "ok"
        },
        "clients": {
            "total_clients": len(client_recon),
            "clients_with_discrepancy": clients_with_discrepancy,
            "total_variance": total_client_variance,
            "status": "attention" if clients_with_discrepancy > 0 else "ok"
        },
        "vendors": {
            "total_vendors": len(vendor_recon),
            "vendors_with_discrepancy": vendors_with_discrepancy,
            "total_variance": total_vendor_variance,
            "status": "attention" if vendors_with_discrepancy > 0 else "ok"
        }
    }

# ============== CALENDAR-BASED RECONCILIATION ==============

@api_router.get("/reconciliation/dates-with-transactions")
async def get_dates_with_transactions(user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))):
    """Get all dates that have transactions for calendar display.
    Returns status breakdown by type (treasury, psp, exchanger) for each date.
    """
    from datetime import datetime, timezone, timedelta
    
    # Get dates with transactions in the last 90 days
    ninety_days_ago = (datetime.now(timezone.utc) - timedelta(days=90)).isoformat()
    
    # Aggregate dates from transactions
    tx_dates_pipeline = [
        {"$match": {"created_at": {"$gte": ninety_days_ago}}},
        {"$project": {"date": {"$substr": ["$created_at", 0, 10]}}},
        {"$group": {"_id": "$date"}}
    ]
    
    tx_dates = await db.transactions.aggregate(tx_dates_pipeline).to_list(100)
    
    # Get reconciliation records with account_type
    recon_records = await db.reconciliations.find(
        {"date": {"$gte": ninety_days_ago[:10]}},
        {"_id": 0, "date": 1, "status": 1, "account_type": 1}
    ).to_list(1000)
    
    # Build status map with type breakdown
    status_map = {}
    for rec in recon_records:
        date = rec.get("date")
        if not date:
            continue
            
        if date not in status_map:
            status_map[date] = {
                "byType": {},
                "overall": None
            }
        
        account_type = rec.get("account_type", "").lower()
        rec_status = rec.get("status", "pending")
        
        # Set status for this type
        if account_type in ["treasury", "psp", "exchanger"]:
            # If already exists, keep the worst status (flagged > pending > completed)
            existing = status_map[date]["byType"].get(account_type)
            if existing == "flagged" or rec_status == "flagged":
                status_map[date]["byType"][account_type] = "flagged"
            elif existing == "pending" or rec_status == "pending":
                status_map[date]["byType"][account_type] = "pending"
            else:
                status_map[date]["byType"][account_type] = "completed"
        
        # Determine overall status (worst case: flagged > pending > completed)
        current_overall = status_map[date]["overall"]
        if rec_status == "flagged" or current_overall == "flagged":
            status_map[date]["overall"] = "flagged"
        elif rec_status == "pending" or current_overall == "pending":
            status_map[date]["overall"] = "pending"
        elif rec_status == "completed" and current_overall is None:
            status_map[date]["overall"] = "completed"
    
    dates = [d["_id"] for d in tx_dates if d["_id"]]
    
    return {"dates": dates, "status": status_map}


@api_router.get("/reconciliation/daily-summary")
async def get_reconciliation_daily_summary(
    date: str,
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))
):
    """Get reconciliation summary for a specific date"""
    # Get all reconciliation records for this date
    recon_records = await db.reconciliations.find({"date": date}, {"_id": 0}).to_list(100)
    
    reconciled = sum(1 for r in recon_records if r.get("status") == "completed")
    pending = sum(1 for r in recon_records if r.get("status") == "pending")
    flagged = sum(r.get("flagged_count", 0) for r in recon_records)
    
    # Get transaction count for this date
    tx_count = await db.transactions.count_documents({
        "created_at": {"$gte": f"{date}T00:00:00", "$lt": f"{date}T23:59:59"}
    })
    
    return {
        "date": date,
        "reconciled": reconciled,
        "pending": pending,
        "flagged": flagged,
        "total": max(tx_count, len(recon_records))
    }


@api_router.get("/reconciliation/pending")
async def get_pending_reconciliation(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    account_type: Optional[str] = None,
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))
):
    """Get all dates/accounts with pending (unreconciled) transactions"""
    from collections import defaultdict
    
    days_back = 180
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days_back)).isoformat()[:10]
    start = date_from or cutoff
    end = date_to or datetime.now(timezone.utc).isoformat()[:10]
    start_iso = f"{start}T00:00:00"
    end_iso = f"{end}T23:59:59"
    
    # Get all reconciliation records to know what's been reconciled
    recon_records = await db.reconciliations.find(
        {"date": {"$gte": start, "$lte": end}},
        {"_id": 0, "date": 1, "account_type": 1, "account_id": 1, "status": 1, "matched_count": 1, "flagged_count": 1}
    ).to_list(5000)
    reconciled_keys = set()
    for r in recon_records:
        if r.get("status") == "completed":
            reconciled_keys.add(f"{r.get('date')}_{r.get('account_type')}_{r.get('account_id')}")
    
    pending_items = []
    
    # --- Treasury transactions ---
    if not account_type or account_type == "treasury":
        treasury_accounts = await db.treasury_accounts.find({}, {"_id": 0, "account_id": 1, "account_name": 1, "currency": 1}).to_list(100)
        acc_map = {a["account_id"]: a for a in treasury_accounts}
        
        tx_pipeline = [
            {"$match": {"created_at": {"$gte": start_iso, "$lte": end_iso}}},
            {"$project": {"date": {"$substr": ["$created_at", 0, 10]}, "account_id": 1, "amount": 1}},
            {"$group": {"_id": {"date": "$date", "account_id": "$account_id"}, "count": {"$sum": 1}, "total_amount": {"$sum": {"$abs": "$amount"}}}}
        ]
        treasury_txs = await db.treasury_transactions.aggregate(tx_pipeline).to_list(5000)
        
        for item in treasury_txs:
            dt = item["_id"]["date"]
            aid = item["_id"]["account_id"]
            key = f"{dt}_treasury_{aid}"
            if key not in reconciled_keys:
                acc = acc_map.get(aid, {})
                pending_items.append({
                    "date": dt,
                    "account_type": "treasury",
                    "account_id": aid,
                    "account_name": acc.get("account_name", "Unknown"),
                    "currency": acc.get("currency", "USD"),
                    "pending_count": item["count"],
                    "total_amount": round(item["total_amount"], 2),
                    "status": "pending"
                })
    
    # --- PSP transactions ---
    if not account_type or account_type == "psp":
        psps = await db.psps.find({}, {"_id": 0, "psp_id": 1, "psp_name": 1}).to_list(100)
        psp_map = {p["psp_id"]: p for p in psps}
        
        psp_pipeline = [
            {"$match": {"psp_id": {"$exists": True, "$ne": None}, "created_at": {"$gte": start_iso, "$lte": end_iso}, "status": {"$in": ["approved", "completed", "pending"]}}},
            {"$project": {"date": {"$substr": ["$created_at", 0, 10]}, "psp_id": 1, "amount": 1}},
            {"$group": {"_id": {"date": "$date", "psp_id": "$psp_id"}, "count": {"$sum": 1}, "total_amount": {"$sum": "$amount"}}}
        ]
        psp_txs = await db.transactions.aggregate(psp_pipeline).to_list(5000)
        
        for item in psp_txs:
            dt = item["_id"]["date"]
            pid = item["_id"]["psp_id"]
            key = f"{dt}_psp_{pid}"
            if key not in reconciled_keys:
                psp = psp_map.get(pid, {})
                pending_items.append({
                    "date": dt,
                    "account_type": "psp",
                    "account_id": pid,
                    "account_name": psp.get("psp_name", "Unknown"),
                    "currency": "USD",
                    "pending_count": item["count"],
                    "total_amount": round(item["total_amount"], 2),
                    "status": "pending"
                })
    
    # --- Exchanger transactions ---
    if not account_type or account_type == "exchanger":
        vendors = await db.vendors.find({}, {"_id": 0, "vendor_id": 1, "vendor_name": 1}).to_list(100)
        vendor_map = {v["vendor_id"]: v for v in vendors}
        
        vendor_pipeline = [
            {"$match": {"vendor_id": {"$exists": True, "$ne": None}, "created_at": {"$gte": start_iso, "$lte": end_iso}, "status": {"$in": ["approved", "completed", "pending"]}}},
            {"$project": {"date": {"$substr": ["$created_at", 0, 10]}, "vendor_id": 1, "amount": 1}},
            {"$group": {"_id": {"date": "$date", "vendor_id": "$vendor_id"}, "count": {"$sum": 1}, "total_amount": {"$sum": "$amount"}}}
        ]
        vendor_txs = await db.transactions.aggregate(vendor_pipeline).to_list(5000)
        
        for item in vendor_txs:
            dt = item["_id"]["date"]
            vid = item["_id"]["vendor_id"]
            key = f"{dt}_exchanger_{vid}"
            if key not in reconciled_keys:
                v = vendor_map.get(vid, {})
                pending_items.append({
                    "date": dt,
                    "account_type": "exchanger",
                    "account_id": vid,
                    "account_name": v.get("vendor_name", "Unknown"),
                    "currency": "USD",
                    "pending_count": item["count"],
                    "total_amount": round(item["total_amount"], 2),
                    "status": "pending"
                })
    
    # Sort by date descending
    pending_items.sort(key=lambda x: x["date"], reverse=True)
    
    # Summary
    total_pending = sum(p["pending_count"] for p in pending_items)
    unique_dates = len(set(p["date"] for p in pending_items))
    
    return {
        "items": pending_items,
        "total_pending_transactions": total_pending,
        "unique_dates": unique_dates,
        "total_items": len(pending_items)
    }



@api_router.get("/reconciliation/account-history")
async def get_account_history_for_reconciliation(
    type: str,
    account_id: str,
    date: str,
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))
):
    """Get transaction history for an account on a specific date"""
    date_start = f"{date}T00:00:00"
    date_end = f"{date}T23:59:59"
    # Also match date-only strings (e.g., "2026-03-01" without time component)
    date_only = date
    
    transactions = []
    
    if type == "treasury":
        # Get treasury transactions - match both ISO datetime and date-only formats
        txs = await db.treasury_transactions.find({
            "account_id": account_id,
            "$or": [
                {"created_at": {"$gte": date_start, "$lte": date_end}},
                {"created_at": date_only}
            ]
        }, {"_id": 0}).to_list(500)
        
        for tx in txs:
            transactions.append({
                "transaction_id": tx.get("treasury_transaction_id"),
                "reference": tx.get("reference", ""),
                "amount": tx.get("amount", 0),
                "currency": tx.get("currency", "USD"),
                "description": tx.get("reference", "Treasury Transaction"),
                "created_at": tx.get("created_at"),
                "type": tx.get("transaction_type")
            })
        
        # Also get regular transactions to this account
        reg_txs = await db.transactions.find({
            "destination_account_id": account_id,
            "created_at": {"$gte": date_start, "$lte": date_end},
            "status": {"$in": ["approved", "completed"]}
        }, {"_id": 0}).to_list(500)
        
        for tx in reg_txs:
            transactions.append({
                "transaction_id": tx.get("transaction_id"),
                "reference": tx.get("reference", ""),
                "amount": tx.get("amount", 0),
                "currency": tx.get("currency", "USD"),
                "description": f"{tx.get('transaction_type', 'Transaction')} - {tx.get('client_name', '')}",
                "created_at": tx.get("created_at"),
                "type": tx.get("transaction_type")
            })
            
    elif type == "psp":
        # Get PSP transactions
        txs = await db.transactions.find({
            "psp_id": account_id,
            "created_at": {"$gte": date_start, "$lte": date_end},
            "status": {"$in": ["approved", "completed", "pending"]}
        }, {"_id": 0}).to_list(500)
        
        for tx in txs:
            transactions.append({
                "transaction_id": tx.get("transaction_id"),
                "reference": tx.get("reference", ""),
                "amount": tx.get("amount", 0),
                "currency": tx.get("currency", "USD"),
                "description": f"{tx.get('transaction_type', 'Transaction')} - {tx.get('client_name', '')}",
                "created_at": tx.get("created_at"),
                "type": tx.get("transaction_type")
            })
            
    elif type == "exchanger":
        # Get vendor transactions
        txs = await db.transactions.find({
            "vendor_id": account_id,
            "created_at": {"$gte": date_start, "$lte": date_end},
            "status": {"$in": ["approved", "completed", "pending"]}
        }, {"_id": 0}).to_list(500)
        
        for tx in txs:
            transactions.append({
                "transaction_id": tx.get("transaction_id"),
                "reference": tx.get("reference", ""),
                "amount": tx.get("amount", 0),
                "currency": tx.get("currency", "USD"),
                "description": f"{tx.get('transaction_type', 'Transaction')} - {tx.get('client_name', '')}",
                "created_at": tx.get("created_at"),
                "type": tx.get("transaction_type")
            })
    
    # Sort by created_at
    transactions.sort(key=lambda x: x.get("created_at", ""), reverse=True)
    
    return transactions


@api_router.post("/reconciliation/upload-statement")
async def upload_statement_for_reconciliation(
    request: Request,
    file: UploadFile = File(...),
    account_type: str = Form(...),
    account_id: str = Form(...),
    date: str = Form(...),
    statement_type: str = Form(default="auto"),  # auto, bank, or psp
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.CREATE))
):
    """Upload and parse a statement file for reconciliation.
    Supports:
    - Banks: Emirates NBD, ADCB, FAB, Mashreq, RAK Bank, DIB, CBD, and more
    - PSPs: PayTabs, Telr, Network International, Stripe, PayPal, and more
    """
    from bank_parsers import (
        parse_bank_statement_pdf, 
        parse_bank_statement_csv, 
        parse_bank_statement_excel,
        parse_psp_statement_pdf,
        parse_psp_statement_csv,
        detect_statement_type,
        SUPPORTED_BANKS,
        SUPPORTED_PSPS
    )
    
    content = await file.read()
    filename = file.filename.lower() if file.filename else ""
    
    parsed_entries = []
    detected_source = "unknown"
    source_type = "unknown"
    
    try:
        # Auto-detect statement type if not specified
        if statement_type == "auto":
            # Try to detect from filename first
            if filename.endswith('.pdf'):
                # Need to read content for detection
                try:
                    from pdf2image import convert_from_bytes
                    import pytesseract
                    images = convert_from_bytes(content, dpi=100, first_page=1, last_page=1)
                    sample_text = pytesseract.image_to_string(images[0]) if images else ""
                    detected_type, detected_name = detect_statement_type(sample_text, filename)
                    statement_type = detected_type if detected_type != "unknown" else "bank"
                except:
                    statement_type = "bank"  # Default to bank
            else:
                statement_type = "bank"  # Default for non-PDF
        
        if filename.endswith('.csv'):
            decoded = content.decode('utf-8')
            if statement_type == "psp":
                parsed_entries, detected_source = parse_psp_statement_csv(decoded, filename)
                source_type = "psp"
            else:
                parsed_entries, detected_source = parse_bank_statement_csv(decoded, filename)
                source_type = "bank"
                
        elif filename.endswith(('.xlsx', '.xls')):
            parsed_entries, detected_source = parse_bank_statement_excel(content, filename)
            source_type = "bank"
                
        elif filename.endswith('.pdf'):
            if statement_type == "psp":
                parsed_entries, detected_source = parse_psp_statement_pdf(content, filename, date)
                source_type = "psp"
            else:
                parsed_entries, detected_source = parse_bank_statement_pdf(content, filename, date)
                source_type = "bank"
            
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Use CSV, XLSX, or PDF.")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Statement parse error: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse statement: {str(e)}")
    
    return {
        "entries": parsed_entries, 
        "count": len(parsed_entries),
        "detected_source": detected_source,
        "source_type": source_type,
        "supported_banks": [b["name"] for b in SUPPORTED_BANKS],
        "supported_psps": [p["name"] for p in SUPPORTED_PSPS]
    }


@api_router.get("/reconciliation/supported-banks")
async def get_supported_banks(user: dict = Depends(get_current_user)):
    """Get list of supported bank statement formats"""
    from bank_parsers import SUPPORTED_BANKS
    return {"banks": SUPPORTED_BANKS}


@api_router.get("/reconciliation/supported-psps")
async def get_supported_psps(user: dict = Depends(get_current_user)):
    """Get list of supported PSP statement formats"""
    from bank_parsers import SUPPORTED_PSPS
    return {"psps": SUPPORTED_PSPS}


@api_router.get("/reconciliation/supported-sources")
async def get_supported_sources(user: dict = Depends(get_current_user)):
    """Get all supported statement sources (banks and PSPs)"""
    from bank_parsers import SUPPORTED_BANKS, SUPPORTED_PSPS
    return {
        "banks": SUPPORTED_BANKS,
        "psps": SUPPORTED_PSPS,
        "total_count": len(SUPPORTED_BANKS) + len(SUPPORTED_PSPS)
    }


@api_router.post("/reconciliation/submit")
async def submit_reconciliation(
    request: Request,
    data: dict = Body(...),
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.CREATE))
):
    """Submit reconciliation results"""
    now = datetime.now(timezone.utc)
    
    recon_id = f"recon_{uuid.uuid4().hex[:12]}"
    
    matched_pairs = data.get("matched_pairs", [])
    flagged_entries = data.get("flagged_entries", [])
    
    # Determine status
    status = "completed"
    if flagged_entries:
        status = "flagged"
    elif data.get("unmatched_system", 0) > 0 or data.get("unmatched_statement", 0) > 0:
        status = "pending"
    
    recon_doc = {
        "recon_id": recon_id,
        "date": data.get("date"),
        "account_type": data.get("account_type"),
        "account_id": data.get("account_id"),
        "matched_pairs": matched_pairs,
        "flagged_entries": flagged_entries,
        "matched_count": len(matched_pairs),
        "flagged_count": len(flagged_entries),
        "unmatched_system": data.get("unmatched_system", 0),
        "unmatched_statement": data.get("unmatched_statement", 0),
        "remarks": data.get("remarks", ""),
        "status": status,
        "created_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    
    await db.reconciliations.insert_one(recon_doc)
    
    await log_activity(request, user, "create", "reconciliation", f"Submitted reconciliation for {data.get('date')}")
    
    return {"message": "Reconciliation submitted successfully", "recon_id": recon_id, "status": status}


@api_router.get("/reconciliation/calendar-history")
async def get_calendar_reconciliation_history(
    limit: int = 50,
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))
):
    """Get reconciliation history for calendar view"""
    records = await db.reconciliations.find({}, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return records


@api_router.get("/reconciliation/calendar-status")
async def get_calendar_status(
    month: str = None,  # Format: YYYY-MM
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))
):
    """Get reconciliation status by date and type for calendar view.
    Returns status breakdown by type (treasury, psp, exchanger) for each date.
    """
    from datetime import datetime, timedelta
    from calendar import monthrange
    
    # Default to current month if not specified
    if month:
        year, month_num = map(int, month.split('-'))
    else:
        now = datetime.now()
        year, month_num = now.year, now.month
    
    # Get date range for the month
    _, last_day = monthrange(year, month_num)
    start_date = f"{year}-{month_num:02d}-01"
    end_date = f"{year}-{month_num:02d}-{last_day:02d}"
    
    # Query reconciliations for this month
    reconciliations = await db.reconciliations.find({
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(1000)
    
    # Build status by date and type
    status_by_date = {}
    
    for recon in reconciliations:
        date = recon.get("date")
        account_type = recon.get("account_type", "").lower()
        recon_status = recon.get("status", "pending")
        
        if not date:
            continue
        
        if date not in status_by_date:
            status_by_date[date] = {
                "byType": {},
                "overall": None
            }
        
        # Set status for this type
        if account_type in ["treasury", "psp", "exchanger"]:
            status_by_date[date]["byType"][account_type] = recon_status
        
        # Determine overall status (worst case: flagged > pending > completed)
        current_overall = status_by_date[date]["overall"]
        if recon_status == "flagged" or current_overall == "flagged":
            status_by_date[date]["overall"] = "flagged"
        elif recon_status == "pending" or current_overall == "pending":
            status_by_date[date]["overall"] = "pending"
        elif recon_status == "completed":
            if current_overall is None:
                status_by_date[date]["overall"] = "completed"
    
    return {
        "month": f"{year}-{month_num:02d}",
        "status": status_by_date
    }


# ============== INTERNAL MESSAGING SYSTEM ==============


@api_router.get("/messages/users")
async def get_users_for_messaging(user: dict = Depends(get_current_user)):
    """Get all users for message recipient selection (lightweight, no admin permission needed)"""
    users = await db.users.find({}, {"_id": 0, "user_id": 1, "name": 1, "email": 1, "role": 1}).to_list(500)
    # Filter out current user
    return [u for u in users if u.get("user_id") != user["user_id"]]


@api_router.get("/messages/unread-count")
async def get_unread_messages_count(user: dict = Depends(get_current_user)):
    """Get total count of unread messages for the current user"""
    user_id = user["user_id"]
    
    # Count unread messages where user is the recipient
    unread_count = await db.user_messages.count_documents({
        "recipient_id": user_id,
        "read": {"$ne": True}
    })
    
    return {"count": unread_count}


@api_router.get("/messages/conversations")
async def get_conversations(user: dict = Depends(get_current_user)):
    """Get all conversations for the current user"""
    user_id = user["user_id"]
    
    # Get all messages where user is sender or recipient
    messages = await db.user_messages.find({
        "$or": [{"sender_id": user_id}, {"recipient_id": user_id}]
    }, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Group by conversation partner
    conversations = {}
    for msg in messages:
        partner_id = msg["recipient_id"] if msg["sender_id"] == user_id else msg["sender_id"]
        if partner_id not in conversations:
            # Get partner info
            partner = await db.users.find_one({"user_id": partner_id}, {"_id": 0, "user_id": 1, "name": 1, "email": 1, "role": 1})
            if partner:
                conversations[partner_id] = {
                    "user_id": partner_id,
                    "name": partner.get("name", "Unknown"),
                    "email": partner.get("email", ""),
                    "role": partner.get("role", "user"),
                    "last_message": msg.get("content", "")[:50],
                    "last_message_at": msg.get("created_at"),
                    "unread_count": 0
                }
        
        # Count unread messages
        if msg["recipient_id"] == user_id and not msg.get("read"):
            if partner_id in conversations:
                conversations[partner_id]["unread_count"] += 1
    
    return list(conversations.values())


@api_router.get("/messages/conversation/{recipient_id}")
async def get_conversation_messages(
    recipient_id: str,
    limit: int = 100,
    user: dict = Depends(get_current_user)
):
    """Get messages between current user and recipient"""
    user_id = user["user_id"]
    
    messages = await db.user_messages.find({
        "$or": [
            {"sender_id": user_id, "recipient_id": recipient_id},
            {"sender_id": recipient_id, "recipient_id": user_id}
        ]
    }, {"_id": 0}).sort("created_at", 1).to_list(limit)
    
    return messages


@api_router.post("/messages/send")
async def send_user_message(
    request: Request,
    recipient_id: str = Form(...),
    content: str = Form(""),
    attachment: Optional[UploadFile] = File(None),
    user: dict = Depends(get_current_user)
):
    """Send a message to another user, optionally with a file attachment"""
    now = datetime.now(timezone.utc)
    
    if not recipient_id:
        raise HTTPException(status_code=400, detail="Recipient ID is required")
    
    # Verify recipient exists
    recipient = await db.users.find_one({"user_id": recipient_id})
    if not recipient:
        raise HTTPException(status_code=404, detail="Recipient not found")
    
    if not content.strip() and not attachment:
        raise HTTPException(status_code=400, detail="Message content or attachment is required")
    
    message_id = f"msg_{uuid.uuid4().hex[:12]}"
    
    attachment_data = None
    if attachment and attachment.filename:
        file_content = await attachment.read()
        if len(file_content) > 10 * 1024 * 1024:  # 10MB limit
            raise HTTPException(status_code=400, detail="File size exceeds 10MB limit")
        attachment_url = upload_to_r2(file_content, attachment.filename, attachment.content_type or "application/octet-stream", "attachments")
        attachment_data = {
            "filename": attachment.filename,
            "content_type": attachment.content_type or "application/octet-stream",
            "size": len(file_content),
            "url": attachment_url
        }
    
    message_doc = {
        "message_id": message_id,
        "sender_id": user["user_id"],
        "sender_name": user["name"],
        "recipient_id": recipient_id,
        "recipient_name": recipient.get("name", "Unknown"),
        "content": content,
        "attachment": {
            "filename": attachment_data["filename"],
            "content_type": attachment_data["content_type"],
            "size": attachment_data["size"],
        } if attachment_data else None,
        "read": False,
        "created_at": now.isoformat()
    }
    
    # Store attachment data separately to keep message doc lightweight
    if attachment_data:
        await db.message_attachments.insert_one({
            "message_id": message_id,
            "filename": attachment_data["filename"],
            "content_type": attachment_data["content_type"],
            "size": attachment_data["size"],
            "data": attachment_data["data"],
            "created_at": now.isoformat()
        })
    
    await db.user_messages.insert_one(message_doc)
    
    return {"message": "Message sent", "message_id": message_id}


@api_router.get("/messages/attachment/{message_id}")
async def get_message_attachment(message_id: str, user: dict = Depends(get_current_user)):
    """Download a message attachment"""
    # Verify user has access to this message
    msg = await db.user_messages.find_one({"message_id": message_id}, {"_id": 0})
    if not msg:
        raise HTTPException(status_code=404, detail="Message not found")
    
    if msg["sender_id"] != user["user_id"] and msg["recipient_id"] != user["user_id"]:
        # Allow admin to access any attachment
        if user.get("role") != "admin":
            raise HTTPException(status_code=403, detail="Access denied")
    
    att = await db.message_attachments.find_one({"message_id": message_id}, {"_id": 0})
    if not att:
        raise HTTPException(status_code=404, detail="Attachment not found")
    
    # Support both new R2 URLs and legacy base64 data
    if att.get("url"):
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=att["url"])
    
    file_data = base64.b64decode(att["data"])
    
    from fastapi.responses import Response
    return Response(
        content=file_data,
        media_type=att["content_type"],
        headers={"Content-Disposition": f'attachment; filename="{att["filename"]}"'}
    )


@api_router.put("/messages/mark-read/{recipient_id}")
async def mark_conversation_read(
    recipient_id: str,
    user: dict = Depends(get_current_user)
):
    """Mark all messages from a sender as read"""
    await db.user_messages.update_many(
        {"sender_id": recipient_id, "recipient_id": user["user_id"], "read": False},
        {"$set": {"read": True}}
    )
    return {"message": "Messages marked as read"}


@api_router.get("/messages")
async def get_messages(
    limit: int = 100,
    context_type: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get internal messages (legacy endpoint)"""
    query = {}
    if context_type:
        query["context.type"] = context_type
    
    messages = await db.internal_messages.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return messages


@api_router.post("/messages")
async def send_message(
    request: Request,
    data: dict = Body(...),
    user: dict = Depends(get_current_user)
):
    """Send an internal message (legacy endpoint)"""
    now = datetime.now(timezone.utc)
    
    message_id = f"msg_{uuid.uuid4().hex[:12]}"
    
    message_doc = {
        "message_id": message_id,
        "content": data.get("content", ""),
        "context": data.get("context"),
        "sender_id": user["user_id"],
        "sender_name": user["name"],
        "created_at": now.isoformat(),
        "read_by": []
    }
    
    await db.internal_messages.insert_one(message_doc)
    
    return {"message": "Message sent", "message_id": message_id}


@api_router.put("/messages/{message_id}/read")
async def mark_message_read(
    message_id: str,
    user: dict = Depends(get_current_user)
):
    """Mark a message as read"""
    await db.internal_messages.update_one(
        {"message_id": message_id},
        {"$addToSet": {"read_by": user["user_id"]}}
    )
    return {"message": "Message marked as read"}


# Admin: View all conversations in the system
@api_router.get("/messages/admin/all-conversations")
async def get_all_conversations_admin(user: dict = Depends(get_current_user)):
    """Get all conversations in the system (admin only)"""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Aggregate all unique conversation pairs
    pipeline = [
        {
            "$group": {
                "_id": {
                    "pair": {
                        "$cond": [
                            {"$lt": ["$sender_id", "$recipient_id"]},
                            ["$sender_id", "$recipient_id"],
                            ["$recipient_id", "$sender_id"]
                        ]
                    }
                },
                "message_count": {"$sum": 1},
                "last_message_at": {"$max": "$created_at"},
                "first_message_at": {"$min": "$created_at"}
            }
        },
        {"$sort": {"last_message_at": -1}}
    ]
    
    conversations = await db.user_messages.aggregate(pipeline).to_list(100)
    
    # Get user details for each conversation
    result = []
    for conv in conversations:
        pair = conv["_id"]["pair"]
        if len(pair) != 2:
            continue
            
        user1_id, user2_id = pair
        
        # Fetch user names
        user1 = await db.users.find_one({"user_id": user1_id}, {"_id": 0, "name": 1, "email": 1})
        user2 = await db.users.find_one({"user_id": user2_id}, {"_id": 0, "name": 1, "email": 1})
        
        result.append({
            "user1_id": user1_id,
            "user1_name": user1.get("name") if user1 else "Unknown",
            "user1_email": user1.get("email") if user1 else "",
            "user2_id": user2_id,
            "user2_name": user2.get("name") if user2 else "Unknown",
            "user2_email": user2.get("email") if user2 else "",
            "message_count": conv["message_count"],
            "last_message_at": conv["last_message_at"],
            "first_message_at": conv["first_message_at"]
        })
    
    return result


@api_router.get("/messages/admin/conversation/{user1_id}/{user2_id}")
async def get_conversation_messages_admin(
    user1_id: str,
    user2_id: str,
    user: dict = Depends(get_current_user)
):
    """Get all messages between two users (admin only)"""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Fetch messages between the two users
    messages = await db.user_messages.find({
        "$or": [
            {"sender_id": user1_id, "recipient_id": user2_id},
            {"sender_id": user2_id, "recipient_id": user1_id}
        ]
    }, {"_id": 0}).sort("created_at", 1).to_list(500)
    
    # Add sender names
    user_cache = {}
    for msg in messages:
        sender_id = msg.get("sender_id")
        if sender_id not in user_cache:
            sender = await db.users.find_one({"user_id": sender_id}, {"_id": 0, "name": 1})
            user_cache[sender_id] = sender.get("name") if sender else "Unknown"
        msg["sender_name"] = user_cache[sender_id]
    
    return messages


# ============== ENHANCED RECONCILIATION FEATURES ==============

# Daily Reconciliation Dashboard
@api_router.get("/reconciliation/daily")
async def get_daily_reconciliation(user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))):
    """Get today's transactions pending reconciliation"""
    from datetime import datetime, timezone, timedelta
    
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = today + timedelta(days=1)
    
    # Get today's transactions
    transactions = await db.transactions.find({
        "created_at": {"$gte": today.isoformat(), "$lt": tomorrow.isoformat()}
    }, {"_id": 0}).to_list(1000)
    
    # Get today's income/expenses
    ie_entries = await db.income_expenses.find({
        "created_at": {"$gte": today.isoformat(), "$lt": tomorrow.isoformat()}
    }, {"_id": 0}).to_list(1000)
    
    # Get today's treasury transactions
    treasury_txs = await db.treasury_transactions.find({
        "created_at": {"$gte": today.isoformat(), "$lt": tomorrow.isoformat()}
    }, {"_id": 0}).to_list(1000)
    
    # Get reconciliation items for today
    recon_items = await db.reconciliation_items.find({
        "date": {"$gte": today.isoformat(), "$lt": tomorrow.isoformat()}
    }, {"_id": 0}).to_list(1000)
    
    reconciled_ids = {item["reference_id"] for item in recon_items if item.get("status") == "reconciled"}
    flagged_ids = {item["reference_id"] for item in recon_items if item.get("status") == "flagged"}
    
    # Build pending items list
    pending_items = []
    
    for tx in transactions:
        tx_id = tx.get("transaction_id")
        status = "reconciled" if tx_id in reconciled_ids else ("flagged" if tx_id in flagged_ids else "pending")
        pending_items.append({
            "id": tx_id,
            "type": "transaction",
            "category": tx.get("transaction_type", "unknown"),
            "description": f"{tx.get('transaction_type', 'Transaction')} - {tx.get('client_name', 'N/A')}",
            "amount": tx.get("amount", 0),
            "currency": tx.get("currency", "USD"),
            "date": tx.get("created_at"),
            "status": status,
            "reference": tx.get("reference_number")
        })
    
    for ie in ie_entries:
        ie_id = ie.get("entry_id")
        status = "reconciled" if ie_id in reconciled_ids else ("flagged" if ie_id in flagged_ids else "pending")
        pending_items.append({
            "id": ie_id,
            "type": "income_expense",
            "category": ie.get("entry_type", "unknown"),
            "description": ie.get("description", "Income/Expense"),
            "amount": ie.get("amount", 0),
            "currency": ie.get("currency", "USD"),
            "date": ie.get("created_at"),
            "status": status,
            "reference": ie.get("reference_number")
        })
    
    for ttx in treasury_txs:
        ttx_id = ttx.get("transaction_id")
        status = "reconciled" if ttx_id in reconciled_ids else ("flagged" if ttx_id in flagged_ids else "pending")
        pending_items.append({
            "id": ttx_id,
            "type": "treasury",
            "category": ttx.get("transaction_type", "unknown"),
            "description": ttx.get("description", "Treasury Transaction"),
            "amount": ttx.get("amount", 0),
            "currency": ttx.get("currency", "USD"),
            "date": ttx.get("created_at"),
            "status": status,
            "reference": ttx.get("reference")
        })
    
    # Calculate stats
    total = len(pending_items)
    reconciled = sum(1 for item in pending_items if item["status"] == "reconciled")
    flagged = sum(1 for item in pending_items if item["status"] == "flagged")
    pending = total - reconciled - flagged
    
    return {
        "date": today.isoformat(),
        "items": pending_items,
        "stats": {
            "total": total,
            "reconciled": reconciled,
            "pending": pending,
            "flagged": flagged,
            "reconciled_percent": round((reconciled / total * 100) if total > 0 else 100, 1)
        }
    }


# Quick Reconcile - Mark item as reconciled
@api_router.post("/reconciliation/quick-reconcile")
async def quick_reconcile(

    request: Request,

    reference_id: str,
    item_type: str,
    notes: str = "",
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.EDIT))
):
    """Quick reconcile a single item"""
    from datetime import datetime, timezone
    
    now = datetime.now(timezone.utc)
    
    # Check if item already exists
    existing = await db.reconciliation_items.find_one({"reference_id": reference_id})
    
    if existing:
        await db.reconciliation_items.update_one(
            {"reference_id": reference_id},
            {"$set": {
                "status": "reconciled",
                "reconciled_by": user.get("user_id"),
                "reconciled_by_name": user.get("name"),
                "reconciled_at": now.isoformat(),
                "notes": notes,
                "updated_at": now.isoformat()
            }}
        )
    else:
        await db.reconciliation_items.insert_one({
            "item_id": f"recon_{uuid.uuid4().hex[:12]}",
            "reference_id": reference_id,
            "item_type": item_type,
            "status": "reconciled",
            "reconciled_by": user.get("user_id"),
            "reconciled_by_name": user.get("name"),
            "reconciled_at": now.isoformat(),
            "notes": notes,
            "date": now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat(),
            "created_at": now.isoformat(),
            "updated_at": now.isoformat()
        })
    
    # Log to history
    await db.reconciliation_history.insert_one({
        "history_id": f"rh_{uuid.uuid4().hex[:12]}",
        "reference_id": reference_id,
        "item_type": item_type,
        "action": "reconciled",
        "performed_by": user.get("user_id"),
        "performed_by_name": user.get("name"),
        "notes": notes,
        "created_at": now.isoformat()
    })
    
    await log_activity(request, user, "edit", "reconciliation", "Quick reconciled item")

    return {"message": "Item reconciled successfully", "reference_id": reference_id}


# Bulk Reconcile - Mark multiple items as reconciled
@api_router.post("/reconciliation/bulk-reconcile")
async def bulk_reconcile(

    request: Request,

    items: List[dict] = Body(...),  # List of {reference_id, item_type}
    notes: str = "",
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.EDIT))
):
    """Bulk reconcile multiple items"""
    from datetime import datetime, timezone
    
    now = datetime.now(timezone.utc)
    reconciled_count = 0
    
    for item in items:
        reference_id = item.get("reference_id")
        item_type = item.get("item_type", "unknown")
        
        existing = await db.reconciliation_items.find_one({"reference_id": reference_id})
        
        if existing:
            await db.reconciliation_items.update_one(
                {"reference_id": reference_id},
                {"$set": {
                    "status": "reconciled",
                    "reconciled_by": user.get("user_id"),
                    "reconciled_by_name": user.get("name"),
                    "reconciled_at": now.isoformat(),
                    "notes": notes,
                    "updated_at": now.isoformat()
                }}
            )
        else:
            await db.reconciliation_items.insert_one({
                "item_id": f"recon_{uuid.uuid4().hex[:12]}",
                "reference_id": reference_id,
                "item_type": item_type,
                "status": "reconciled",
                "reconciled_by": user.get("user_id"),
                "reconciled_by_name": user.get("name"),
                "reconciled_at": now.isoformat(),
                "notes": notes,
                "date": now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat(),
                "created_at": now.isoformat(),
                "updated_at": now.isoformat()
            })
        
        reconciled_count += 1
    
    # Log bulk action to history
    await db.reconciliation_history.insert_one({
        "history_id": f"rh_{uuid.uuid4().hex[:12]}",
        "reference_id": "bulk_action",
        "item_type": "bulk",
        "action": "bulk_reconciled",
        "item_count": reconciled_count,
        "performed_by": user.get("user_id"),
        "performed_by_name": user.get("name"),
        "notes": notes,
        "created_at": now.isoformat()
    })
    
    await log_activity(request, user, "edit", "reconciliation", "Bulk reconciled items")

    return {"message": f"Reconciled {reconciled_count} items", "count": reconciled_count}


# Flag for Review
@api_router.post("/reconciliation/flag")
async def flag_for_review(

    request: Request,

    reference_id: str,
    item_type: str,
    reason: str,
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.EDIT))
):
    """Flag an item for supervisor review"""
    from datetime import datetime, timezone
    
    now = datetime.now(timezone.utc)
    
    existing = await db.reconciliation_items.find_one({"reference_id": reference_id})
    
    if existing:
        await db.reconciliation_items.update_one(
            {"reference_id": reference_id},
            {"$set": {
                "status": "flagged",
                "flagged_by": user.get("user_id"),
                "flagged_by_name": user.get("name"),
                "flagged_at": now.isoformat(),
                "flag_reason": reason,
                "updated_at": now.isoformat()
            }}
        )
    else:
        await db.reconciliation_items.insert_one({
            "item_id": f"recon_{uuid.uuid4().hex[:12]}",
            "reference_id": reference_id,
            "item_type": item_type,
            "status": "flagged",
            "flagged_by": user.get("user_id"),
            "flagged_by_name": user.get("name"),
            "flagged_at": now.isoformat(),
            "flag_reason": reason,
            "date": now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat(),
            "created_at": now.isoformat(),
            "updated_at": now.isoformat()
        })
    
    # Log to history
    await db.reconciliation_history.insert_one({
        "history_id": f"rh_{uuid.uuid4().hex[:12]}",
        "reference_id": reference_id,
        "item_type": item_type,
        "action": "flagged",
        "performed_by": user.get("user_id"),
        "performed_by_name": user.get("name"),
        "notes": reason,
        "created_at": now.isoformat()
    })
    
    await log_activity(request, user, "edit", "reconciliation", "Flagged for review")

    return {"message": "Item flagged for review", "reference_id": reference_id}


# Create Adjustment Entry
@api_router.post("/reconciliation/adjustment")
async def create_adjustment(

    request: Request,

    reference_id: str,
    item_type: str,
    adjustment_amount: float,
    currency: str,
    reason: str,
    treasury_account_id: str = None,
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.CREATE))
):
    """Create an adjustment entry for reconciliation discrepancy"""
    from datetime import datetime, timezone
    
    now = datetime.now(timezone.utc)
    adjustment_id = f"adj_{uuid.uuid4().hex[:12]}"
    
    adjustment_doc = {
        "adjustment_id": adjustment_id,
        "reference_id": reference_id,
        "item_type": item_type,
        "amount": adjustment_amount,
        "currency": currency,
        "reason": reason,
        "treasury_account_id": treasury_account_id,
        "created_by": user.get("user_id"),
        "created_by_name": user.get("name"),
        "status": "approved",
        "created_at": now.isoformat()
    }
    
    await db.reconciliation_adjustments.insert_one(adjustment_doc)
    
    # If treasury account specified, create treasury transaction
    if treasury_account_id:
        tx_type = "credit" if adjustment_amount > 0 else "debit"
        await db.treasury_transactions.insert_one({
            "transaction_id": f"ttx_{uuid.uuid4().hex[:12]}",
            "account_id": treasury_account_id,
            "type": tx_type,
            "amount": abs(adjustment_amount),
            "currency": currency,
            "description": f"Reconciliation Adjustment: {reason}",
            "reference": adjustment_id,
            "category": "adjustment",
            "created_by": user.get("user_id"),
            "created_at": now.isoformat()
        })
        
        # Update treasury balance
        update_op = {"$inc": {"balance": adjustment_amount}}
        await db.treasury_accounts.update_one({"account_id": treasury_account_id}, update_op)
    
    # Mark original item as reconciled with adjustment
    await db.reconciliation_items.update_one(
        {"reference_id": reference_id},
        {"$set": {
            "status": "reconciled",
            "adjustment_id": adjustment_id,
            "reconciled_by": user.get("user_id"),
            "reconciled_at": now.isoformat(),
            "updated_at": now.isoformat()
        }},
        upsert=True
    )
    
    # Log to history
    await db.reconciliation_history.insert_one({
        "history_id": f"rh_{uuid.uuid4().hex[:12]}",
        "reference_id": reference_id,
        "item_type": item_type,
        "action": "adjustment_created",
        "adjustment_id": adjustment_id,
        "amount": adjustment_amount,
        "performed_by": user.get("user_id"),
        "performed_by_name": user.get("name"),
        "notes": reason,
        "created_at": now.isoformat()
    })
    
    adjustment_doc.pop("_id", None)
    await log_activity(request, user, "create", "reconciliation", "Created adjustment")

    return adjustment_doc


# Get Reconciliation History / Audit Trail
@api_router.get("/reconciliation/history")
async def get_reconciliation_history(
    date_from: str = None,
    date_to: str = None,
    account_type: str = None,
    account_id: str = None,
    status: str = None,
    has_matched: str = None,
    has_flagged: str = None,
    limit: int = 100,
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))
):
    """Get reconciliation submissions history with filters"""
    query = {}
    
    if date_from:
        query["date"] = {"$gte": date_from}
    if date_to:
        if "date" in query:
            query["date"]["$lte"] = date_to
        else:
            query["date"] = {"$lte": date_to}
    if account_type:
        query["account_type"] = account_type
    if account_id:
        query["account_id"] = account_id
    if status:
        query["status"] = status
    if has_matched == "true":
        query["matched_count"] = {"$gt": 0}
    elif has_matched == "false":
        query["$or"] = [{"matched_count": 0}, {"matched_count": {"$exists": False}}]
    if has_flagged == "true":
        query["flagged_count"] = {"$gt": 0}
    elif has_flagged == "false":
        query["$or"] = [{"flagged_count": 0}, {"flagged_count": {"$exists": False}}]
    
    history = await db.reconciliations.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)
    return history


@api_router.get("/reconciliation/history/export")
async def export_reconciliation_history(
    format: str = "xlsx",
    date_from: str = None,
    date_to: str = None,
    account_type: str = None,
    account_id: str = None,
    status: str = None,
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.EXPORT))
):
    """Export reconciliation history to PDF or Excel"""
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Build query
    query = {}
    if date_from:
        query["date"] = {"$gte": date_from}
    if date_to:
        if "date" in query:
            query["date"]["$lte"] = date_to
        else:
            query["date"] = {"$lte": date_to}
    if account_type:
        query["account_type"] = account_type
    if account_id:
        query["account_id"] = account_id
    if status:
        query["status"] = status
    
    # Fetch data
    history = await db.reconciliations.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    if format == "pdf":
        # Generate PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1  # Center
        )
        elements.append(Paragraph("Reconciliation History Report", title_style))
        
        # Subtitle with filters
        filter_text = f"Generated on: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        if date_from or date_to:
            filter_text += f" | Period: {date_from or 'Start'} to {date_to or 'Now'}"
        elements.append(Paragraph(filter_text, styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Table data
        table_data = [["Date", "Type", "Account ID", "Matched", "Flagged", "Status", "Created By", "Created At"]]
        for item in history:
            table_data.append([
                item.get("date", ""),
                item.get("account_type", "").title(),
                item.get("account_id", "")[:15],
                str(item.get("matched_count", 0)),
                str(item.get("flagged_count", 0)),
                item.get("status", "").title(),
                item.get("created_by_name", ""),
                item.get("created_at", "")[:16]
            ])
        
        # Create table
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ]))
        elements.append(table)
        
        # Summary
        elements.append(Spacer(1, 20))
        total = len(history)
        completed = sum(1 for h in history if h.get("status") == "completed")
        pending = sum(1 for h in history if h.get("status") == "pending")
        elements.append(Paragraph(f"Total Records: {total} | Completed: {completed} | Pending: {pending}", styles['Normal']))
        
        doc.build(elements)
        buffer.seek(0)
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=reconciliation_history_{datetime.now().strftime('%Y%m%d')}.pdf"}
        )
    
    else:  # Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reconciliation History"
        
        # Header style
        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["Date", "Type", "Account ID", "Matched", "Flagged", "Status", "Remarks", "Created By", "Created At"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Data
        for row_num, item in enumerate(history, 2):
            data = [
                item.get("date", ""),
                item.get("account_type", "").title(),
                item.get("account_id", ""),
                item.get("matched_count", 0),
                item.get("flagged_count", 0),
                item.get("status", "").title(),
                item.get("remarks", ""),
                item.get("created_by_name", ""),
                item.get("created_at", "")[:19]
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center' if col in [4, 5] else 'left')
        
        # Auto-width columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=reconciliation_history_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )


# Get Flagged Items
@api_router.get("/reconciliation/flagged")
async def get_flagged_items(user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))):
    """Get all items flagged for review"""
    items = await db.reconciliation_items.find({"status": "flagged"}, {"_id": 0}).sort("flagged_at", -1).to_list(1000)
    return items


# Export Unmatched Items
@api_router.get("/reconciliation/export-unmatched")
async def export_unmatched(
    recon_type: str = "all",
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.EXPORT))
):
    """Export unmatched reconciliation items"""
    query = {"status": {"$in": ["pending", "unmatched", "discrepancy"]}}
    if recon_type != "all":
        query["item_type"] = recon_type
    
    items = await db.reconciliation_items.find(query, {"_id": 0}).to_list(10000)
    
    # Also get unmatched bank entries
    bank_entries = await db.reconciliation_entries.find(
        {"status": {"$in": ["pending", "unmatched"]}}, {"_id": 0}
    ).to_list(10000)
    
    return {
        "reconciliation_items": items,
        "bank_entries": bank_entries,
        "total_count": len(items) + len(bank_entries)
    }


# Write-off Small Variance
@api_router.post("/reconciliation/write-off")
async def write_off_variance(

    request: Request,

    reference_id: str,
    item_type: str,
    variance_amount: float,
    reason: str,
    user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.EDIT))
):
    """Write off small variance and mark as reconciled"""
    from datetime import datetime, timezone
    
    now = datetime.now(timezone.utc)
    
    await db.reconciliation_items.update_one(
        {"reference_id": reference_id},
        {"$set": {
            "status": "reconciled",
            "write_off_amount": variance_amount,
            "write_off_reason": reason,
            "reconciled_by": user.get("user_id"),
            "reconciled_by_name": user.get("name"),
            "reconciled_at": now.isoformat(),
            "updated_at": now.isoformat()
        }},
        upsert=True
    )
    
    # Log to history
    await db.reconciliation_history.insert_one({
        "history_id": f"rh_{uuid.uuid4().hex[:12]}",
        "reference_id": reference_id,
        "item_type": item_type,
        "action": "write_off",
        "amount": variance_amount,
        "performed_by": user.get("user_id"),
        "performed_by_name": user.get("name"),
        "notes": reason,
        "created_at": now.isoformat()
    })
    
    await log_activity(request, user, "edit", "reconciliation", "Wrote off variance")

    return {"message": "Variance written off successfully", "reference_id": reference_id}


# Get Daily Reconciliation Summary for Reports
@api_router.get("/reconciliation/daily-summary")
async def get_daily_reconciliation_summary(user: dict = Depends(require_permission(Modules.RECONCILIATION, Actions.VIEW))):
    """Get daily reconciliation summary for reports"""
    from datetime import datetime, timezone, timedelta
    
    today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = today + timedelta(days=1)
    
    # Get today's reconciliation stats
    today_items = await db.reconciliation_items.find({
        "date": {"$gte": today.isoformat(), "$lt": tomorrow.isoformat()}
    }, {"_id": 0}).to_list(10000)
    
    bank_entries = await db.reconciliation_entries.find({
        "created_at": {"$gte": today.isoformat(), "$lt": tomorrow.isoformat()}
    }, {"_id": 0}).to_list(10000)
    
    # Calculate by category
    by_category = {
        "bank": {"reconciled": 0, "pending": 0, "flagged": 0},
        "psp": {"reconciled": 0, "pending": 0, "flagged": 0},
        "transaction": {"reconciled": 0, "pending": 0, "flagged": 0},
        "income_expense": {"reconciled": 0, "pending": 0, "flagged": 0},
        "treasury": {"reconciled": 0, "pending": 0, "flagged": 0},
    }
    
    for item in today_items:
        cat = item.get("item_type", "other")
        if cat not in by_category:
            by_category[cat] = {"reconciled": 0, "pending": 0, "flagged": 0}
        
        status = item.get("status", "pending")
        if status == "reconciled":
            by_category[cat]["reconciled"] += 1
        elif status == "flagged":
            by_category[cat]["flagged"] += 1
        else:
            by_category[cat]["pending"] += 1
    
    # Add bank entries
    for entry in bank_entries:
        status = entry.get("status", "pending")
        if status == "matched":
            by_category["bank"]["reconciled"] += 1
        elif status in ["unmatched", "pending", "discrepancy"]:
            by_category["bank"]["pending"] += 1
    
    total_reconciled = sum(c["reconciled"] for c in by_category.values())
    total_pending = sum(c["pending"] for c in by_category.values())
    total_flagged = sum(c["flagged"] for c in by_category.values())
    
    # Get flagged items details
    flagged_items = await db.reconciliation_items.find(
        {"status": "flagged", "date": {"$gte": today.isoformat(), "$lt": tomorrow.isoformat()}},
        {"_id": 0, "reference_id": 1, "item_type": 1, "flag_reason": 1}
    ).to_list(100)
    
    return {
        "date": today.strftime("%Y-%m-%d"),
        "total": {
            "reconciled": total_reconciled,
            "pending": total_pending,
            "flagged": total_flagged,
            "total": total_reconciled + total_pending + total_flagged
        },
        "by_category": by_category,
        "flagged_items": flagged_items,
        "attention_required": total_pending > 0 or total_flagged > 0
    }


# ============== ROLES & PERMISSIONS MANAGEMENT ==============

@api_router.get("/roles")
async def get_roles(user: dict = Depends(require_permission(Modules.ROLES, Actions.VIEW))):
    """Get all roles"""
    roles = await db.roles.find({"is_active": {"$ne": False}}, {"_id": 0}).to_list(100)
    return roles

@api_router.get("/roles/{role_id}")
async def get_role(role_id: str, user: dict = Depends(require_permission(Modules.ROLES, Actions.VIEW))):
    """Get a specific role"""
    role = await db.roles.find_one({"role_id": role_id}, {"_id": 0})
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    return role

@api_router.post("/roles")
async def create_role(request: Request, role_data: RoleCreate, user: dict = Depends(require_permission(Modules.ROLES, Actions.CREATE))):

    """Create a new role"""
    # Check if role name already exists
    existing = await db.roles.find_one({"name": role_data.name}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Role name already exists")
    
    now = datetime.now(timezone.utc)
    role_id = f"role_{uuid.uuid4().hex[:12]}"
    
    role_doc = {
        "role_id": role_id,
        "name": role_data.name,
        "display_name": role_data.display_name,
        "description": role_data.description,
        "permissions": role_data.permissions,
        "is_system_role": False,
        "hierarchy_level": role_data.hierarchy_level,
        "is_active": True,
        "created_at": now.isoformat(),
        "updated_at": now.isoformat(),
        "created_by": user["user_id"],
        "created_by_name": user["name"]
    }
    
    await db.roles.insert_one(role_doc)
    
    # Log the action
    await db.audit_logs.insert_one({
        "log_id": f"log_{uuid.uuid4().hex[:12]}",
        "action": "role_created",
        "module": Modules.ROLES,
        "details": f"Created role: {role_data.display_name}",
        "user_id": user["user_id"],
        "user_name": user["name"],
        "created_at": now.isoformat()
    })
    
    await log_activity(request, user, "create", "roles", "Created role")

    return await db.roles.find_one({"role_id": role_id}, {"_id": 0})

@api_router.put("/roles/{role_id}")
async def update_role(request: Request, role_id: str, role_data: RoleUpdate, user: dict = Depends(require_permission(Modules.ROLES, Actions.EDIT))):

    """Update a role"""
    role = await db.roles.find_one({"role_id": role_id}, {"_id": 0})
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    
    # Prevent modifying system roles (except permissions)
    if role.get("is_system_role") and role_data.is_active is False:
        raise HTTPException(status_code=400, detail="Cannot deactivate system roles")
    
    now = datetime.now(timezone.utc)
    updates = {k: v for k, v in role_data.model_dump().items() if v is not None}
    updates["updated_at"] = now.isoformat()
    
    await db.roles.update_one({"role_id": role_id}, {"$set": updates})
    
    # Log the action
    await db.audit_logs.insert_one({
        "log_id": f"log_{uuid.uuid4().hex[:12]}",
        "action": "role_updated",
        "module": Modules.ROLES,
        "details": f"Updated role: {role.get('display_name')}",
        "changes": updates,
        "user_id": user["user_id"],
        "user_name": user["name"],
        "created_at": now.isoformat()
    })
    
    await log_activity(request, user, "edit", "roles", "Updated role")

    return await db.roles.find_one({"role_id": role_id}, {"_id": 0})

@api_router.delete("/roles/{role_id}")
async def delete_role(request: Request, role_id: str, user: dict = Depends(require_permission(Modules.ROLES, Actions.DELETE))):

    """Delete a role (soft delete)"""
    role = await db.roles.find_one({"role_id": role_id}, {"_id": 0})
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    
    if role.get("is_system_role"):
        raise HTTPException(status_code=400, detail="Cannot delete system roles")
    
    # Check if any users have this role
    users_with_role = await db.users.count_documents({"role_id": role_id})
    if users_with_role > 0:
        raise HTTPException(status_code=400, detail=f"Cannot delete role: {users_with_role} users have this role")
    
    await db.roles.update_one({"role_id": role_id}, {"$set": {"is_active": False}})
    
    await log_activity(request, user, "delete", "roles", "Deleted role")

    return {"message": "Role deleted successfully"}

@api_router.get("/permissions/modules")
async def get_modules_and_actions(user: dict = Depends(require_permission(Modules.ROLES, Actions.VIEW))):
    """Get all available modules and actions for permission configuration"""
    return {
        "modules": [{"id": m, "name": MODULE_DISPLAY_NAMES.get(m, m)} for m in ALL_MODULES],
        "actions": ALL_ACTIONS
    }

@api_router.get("/permissions/my")
async def get_my_permissions(user: dict = Depends(get_current_user)):
    """Get current user's permissions"""
    permissions = await get_user_permissions(user["user_id"])
    
    # Get role info
    role_id = user.get("role_id") or user.get("role")
    role = await db.roles.find_one({"$or": [{"role_id": role_id}, {"name": role_id}]}, {"_id": 0})
    
    return {
        "user_id": user["user_id"],
        "role": role.get("display_name") if role else user.get("role"),
        "role_id": role_id,
        "permissions": permissions,
        "modules": {m: MODULE_DISPLAY_NAMES.get(m, m) for m in ALL_MODULES}
    }

@api_router.put("/users/{user_id}/role")
async def assign_user_role(request: Request, user_id: str, role_id: str = Form(...), user: dict = Depends(require_permission(Modules.USERS, Actions.EDIT))):

    """Assign a role to a user"""
    target_user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    role = await db.roles.find_one({"role_id": role_id}, {"_id": 0})
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    
    now = datetime.now(timezone.utc)
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"role_id": role_id, "role": role.get("name"), "updated_at": now.isoformat()}}
    )
    
    # Log the action
    await db.audit_logs.insert_one({
        "log_id": f"log_{uuid.uuid4().hex[:12]}",
        "action": "user_role_assigned",
        "module": Modules.USERS,
        "details": f"Assigned role '{role.get('display_name')}' to user '{target_user.get('name')}'",
        "target_user_id": user_id,
        "user_id": user["user_id"],
        "user_name": user["name"],
        "created_at": now.isoformat()
    })
    
    await log_activity(request, user, "edit", "users", "Assigned role to user")

    return {"message": "Role assigned successfully"}

@api_router.put("/users/{user_id}/permissions")
async def set_user_permission_overrides(

    request: Request,

    user_id: str, 
    permissions: dict,
    user: dict = Depends(require_permission(Modules.USERS, Actions.EDIT))
):
    """Set custom permission overrides for a user"""
    target_user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    now = datetime.now(timezone.utc)
    await db.users.update_one(
        {"user_id": user_id},
        {"$set": {"permission_overrides": permissions, "updated_at": now.isoformat()}}
    )
    
    await log_activity(request, user, "edit", "users", "Updated user permissions")

    return {"message": "Permission overrides updated successfully"}


# ============== APPROVAL EMAIL NOTIFICATIONS ==============

async def send_approval_notification(notification_type: str, details: dict):
    """Send email notification for pending approvals to users with approve permission"""
    try:
        # Get SMTP config
        smtp_settings = await db.app_settings.find_one({"setting_type": "email"}, {"_id": 0})
        smtp_host = (smtp_settings or {}).get("smtp_host") or os.environ.get("SMTP_HOST", "smtp.gmail.com")
        smtp_port = (smtp_settings or {}).get("smtp_port") or int(os.environ.get("SMTP_PORT", "587"))
        smtp_email = (smtp_settings or {}).get("smtp_email") or os.environ.get("SMTP_USER", "")
        smtp_password = (smtp_settings or {}).get("smtp_password") or os.environ.get("SMTP_PASSWORD", "")
        smtp_from = (smtp_settings or {}).get("smtp_from_email") or os.environ.get("SMTP_FROM_EMAIL", smtp_email)
        
        if not smtp_email or not smtp_password:
            return
        
        # Get users with approvals permission
        roles_with_approve = await db.roles.find({"permissions.approvals": {"$exists": True}}, {"_id": 0, "role_id": 1, "name": 1}).to_list(20)
        role_ids = [r.get("role_id") or r.get("name") for r in roles_with_approve]
        
        approvers = await db.users.find(
            {"$or": [{"role": {"$in": role_ids}}, {"role_id": {"$in": role_ids}}], "is_active": {"$ne": False}},
            {"_id": 0, "user_id": 1, "email": 1, "name": 1}
        ).to_list(50)
        
        if not approvers:
            return
        
        # Filter by notification preference (default: ON)
        approver_ids = [u["user_id"] for u in approvers]
        prefs = await db.user_preferences.find(
            {"user_id": {"$in": approver_ids}, "approval_notifications": False},
            {"_id": 0, "user_id": 1}
        ).to_list(50)
        opted_out = set(p["user_id"] for p in prefs)
        
        to_emails = [u["email"] for u in approvers if u.get("email") and u["user_id"] not in opted_out]
        if not to_emails:
            return
        
        if notification_type == "transaction":
            subject = f"Miles Capitals - New Transaction Pending Approval"
            html = f"""<div style="font-family:Arial,sans-serif;max-width:500px;margin:0 auto;padding:30px;background:#0B0C10;color:white;border-radius:8px;">
                <h2 style="color:#66FCF1;text-align:center;margin:0 0 20px;">MILES CAPITALS</h2>
                <div style="background:#1F2833;padding:20px;border-radius:8px;margin:15px 0;">
                    <h3 style="color:#fbbf24;margin:0 0 15px;">New Transaction Pending Approval</h3>
                    <table style="width:100%;font-size:13px;color:#C5C6C7;">
                        <tr><td style="padding:4px 0;color:#888;">Reference</td><td style="padding:4px 0;font-family:monospace;">{details.get('reference', '-')}</td></tr>
                        <tr><td style="padding:4px 0;color:#888;">Type</td><td style="padding:4px 0;text-transform:capitalize;color:{'#4ade80' if details.get('type') == 'deposit' else '#f87171'}">{details.get('type', '-')}</td></tr>
                        <tr><td style="padding:4px 0;color:#888;">Client</td><td style="padding:4px 0;">{details.get('client', '-')}</td></tr>
                        <tr><td style="padding:4px 0;color:#888;">Amount</td><td style="padding:4px 0;font-weight:bold;font-size:16px;">${details.get('amount', 0):,.2f} USD</td></tr>
                        {f'<tr><td style="padding:4px 0;color:#888;">Base Amount</td><td style="padding:4px 0;">{details.get("base_amount", 0):,.2f} {details.get("base_currency", "")}</td></tr>' if details.get('base_currency') and details.get('base_currency') != 'USD' else ''}
                        <tr><td style="padding:4px 0;color:#888;">Destination</td><td style="padding:4px 0;">{details.get('destination', '-')}</td></tr>
                        <tr><td style="padding:4px 0;color:#888;">Created By</td><td style="padding:4px 0;">{details.get('created_by', '-')}</td></tr>
                    </table>
                </div>
                <p style="color:#C5C6C7;text-align:center;font-size:12px;">Please review and approve/reject this transaction in the Back Office.</p></div>"""
        
        elif notification_type == "settlement":
            html = f"""<div style="font-family:Arial,sans-serif;max-width:500px;margin:0 auto;padding:30px;background:#0B0C10;color:white;border-radius:8px;">
                <h2 style="color:#66FCF1;text-align:center;margin:0 0 20px;">MILES CAPITALS</h2>
                <div style="background:#1F2833;padding:20px;border-radius:8px;margin:15px 0;">
                    <h3 style="color:#fbbf24;margin:0 0 15px;">New Settlement Pending Approval</h3>
                    <table style="width:100%;font-size:13px;color:#C5C6C7;">
                        <tr><td style="padding:4px 0;color:#888;">Settlement ID</td><td style="padding:4px 0;font-family:monospace;">{details.get('settlement_id', '-')}</td></tr>
                        <tr><td style="padding:4px 0;color:#888;">Exchanger</td><td style="padding:4px 0;">{details.get('vendor_name', '-')}</td></tr>
                        <tr><td style="padding:4px 0;color:#888;">Gross Amount</td><td style="padding:4px 0;font-weight:bold;">{details.get('gross_amount', 0):,.2f} {details.get('currency', 'USD')}</td></tr>
                        <tr><td style="padding:4px 0;color:#888;">Net Settlement</td><td style="padding:4px 0;font-weight:bold;font-size:16px;color:#66FCF1;">{details.get('net_amount', 0):,.2f} {details.get('dest_currency', 'USD')}</td></tr>
                        <tr><td style="padding:4px 0;color:#888;">Transactions</td><td style="padding:4px 0;">{details.get('tx_count', 0)} entries</td></tr>
                        <tr><td style="padding:4px 0;color:#888;">Created By</td><td style="padding:4px 0;">{details.get('created_by', '-')}</td></tr>
                    </table>
                </div>
                <p style="color:#C5C6C7;text-align:center;font-size:12px;">Please review and approve/reject this settlement in the Back Office.</p></div>"""
            subject = f"Miles Capitals - Settlement Pending Approval ({details.get('vendor_name', '')})"
        else:
            return
        
        await send_email(
            to_emails=to_emails, subject=subject, html_content=html,
            smtp_host=smtp_host, smtp_port=smtp_port,
            smtp_email=smtp_email, smtp_password=smtp_password, smtp_from_email=smtp_from
        )
        logger.info(f"Approval notification sent to {len(to_emails)} approvers for {notification_type}")
    except Exception as e:
        logger.error(f"Failed to send approval notification: {e}")



async def send_exchanger_notification(notification_type: str, vendor_id: str, details: dict):
    """Send email notification to a specific exchanger"""
    try:
        vendor = await db.vendors.find_one({"vendor_id": vendor_id}, {"_id": 0, "user_id": 1, "vendor_name": 1})
        if not vendor or not vendor.get("user_id"):
            return
        
        user_doc = await db.users.find_one({"user_id": vendor["user_id"]}, {"_id": 0, "email": 1, "user_id": 1})
        if not user_doc or not user_doc.get("email"):
            return
        
        # Check notification preference
        prefs = await db.user_preferences.find_one({"user_id": user_doc["user_id"]}, {"_id": 0})
        if prefs and prefs.get("approval_notifications") is False:
            return
        
        smtp_settings = await db.app_settings.find_one({"setting_type": "email"}, {"_id": 0})
        smtp_host = (smtp_settings or {}).get("smtp_host") or os.environ.get("SMTP_HOST", "smtp.gmail.com")
        smtp_port = (smtp_settings or {}).get("smtp_port") or int(os.environ.get("SMTP_PORT", "587"))
        smtp_email = (smtp_settings or {}).get("smtp_email") or os.environ.get("SMTP_USER", "")
        smtp_password = (smtp_settings or {}).get("smtp_password") or os.environ.get("SMTP_PASSWORD", "")
        smtp_from = (smtp_settings or {}).get("smtp_from_email") or os.environ.get("SMTP_FROM_EMAIL", smtp_email)
        
        if not smtp_email or not smtp_password:
            return
        
        vendor_name = vendor.get("vendor_name", "Exchanger")
        
        if notification_type == "transaction":
            subject = f"Miles Capitals - New Transaction Assigned"
            title = "New Transaction Assigned to You"
            color = "#4ade80" if details.get("type") == "deposit" else "#f87171"
            rows = f"""<tr><td style="padding:4px 0;color:#888;">Reference</td><td style="padding:4px 0;font-family:monospace;">{details.get('reference', '-')}</td></tr>
                <tr><td style="padding:4px 0;color:#888;">Type</td><td style="padding:4px 0;color:{color};text-transform:capitalize;">{details.get('type', '-')}</td></tr>
                <tr><td style="padding:4px 0;color:#888;">Client</td><td style="padding:4px 0;">{details.get('client', '-')}</td></tr>
                <tr><td style="padding:4px 0;color:#888;">Amount</td><td style="padding:4px 0;font-weight:bold;font-size:16px;">{details.get('amount_display', '-')}</td></tr>"""
        elif notification_type == "ie":
            subject = f"Miles Capitals - New I&E Entry Assigned"
            title = "New Income/Expense Entry Assigned"
            is_income = details.get("entry_type") == "income"
            color = "#4ade80" if is_income else "#f87171"
            rows = f"""<tr><td style="padding:4px 0;color:#888;">Type</td><td style="padding:4px 0;color:{color};">{'Income (IN)' if is_income else 'Expense (OUT)'}</td></tr>
                <tr><td style="padding:4px 0;color:#888;">Category</td><td style="padding:4px 0;">{details.get('category', '-')}</td></tr>
                <tr><td style="padding:4px 0;color:#888;">Amount</td><td style="padding:4px 0;font-weight:bold;font-size:16px;">{details.get('amount_display', '-')}</td></tr>"""
        elif notification_type == "loan":
            subject = f"Miles Capitals - Loan Transaction Assigned"
            title = "New Loan Transaction Assigned"
            is_in = details.get("loan_type") == "repayment"
            color = "#4ade80" if is_in else "#f87171"
            rows = f"""<tr><td style="padding:4px 0;color:#888;">Type</td><td style="padding:4px 0;color:{color};">{'Loan Repayment (IN)' if is_in else 'Loan Disbursement (OUT)'}</td></tr>
                <tr><td style="padding:4px 0;color:#888;">Borrower</td><td style="padding:4px 0;">{details.get('borrower', '-')}</td></tr>
                <tr><td style="padding:4px 0;color:#888;">Amount</td><td style="padding:4px 0;font-weight:bold;font-size:16px;">{details.get('amount_display', '-')}</td></tr>"""
        elif notification_type == "settlement":
            subject = f"Miles Capitals - Settlement Initiated"
            title = "Settlement Initiated for Your Account"
            rows = f"""<tr><td style="padding:4px 0;color:#888;">Settlement ID</td><td style="padding:4px 0;font-family:monospace;">{details.get('settlement_id', '-')}</td></tr>
                <tr><td style="padding:4px 0;color:#888;">Gross Amount</td><td style="padding:4px 0;font-weight:bold;">{details.get('gross_display', '-')}</td></tr>
                <tr><td style="padding:4px 0;color:#888;">Net Settlement</td><td style="padding:4px 0;font-weight:bold;font-size:16px;color:#66FCF1;">{details.get('net_display', '-')}</td></tr>
                <tr><td style="padding:4px 0;color:#888;">Entries</td><td style="padding:4px 0;">{details.get('tx_count', 0)}</td></tr>"""
        else:
            return
        
        html = f"""<div style="font-family:Arial,sans-serif;max-width:500px;margin:0 auto;padding:30px;background:#0B0C10;color:white;border-radius:8px;">
            <h2 style="color:#66FCF1;text-align:center;margin:0 0 20px;">MILES CAPITALS</h2>
            <div style="background:#1F2833;padding:20px;border-radius:8px;margin:15px 0;">
                <h3 style="color:#fbbf24;margin:0 0 15px;">{title}</h3>
                <table style="width:100%;font-size:13px;color:#C5C6C7;">{rows}
                    <tr><td style="padding:4px 0;color:#888;">Exchanger</td><td style="padding:4px 0;">{vendor_name}</td></tr>
                </table>
            </div>
            <p style="color:#C5C6C7;text-align:center;font-size:12px;">Please review in your Exchanger Portal.</p></div>"""
        
        await send_email(
            to_emails=[user_doc["email"]], subject=subject, html_content=html,
            smtp_host=smtp_host, smtp_port=smtp_port,
            smtp_email=smtp_email, smtp_password=smtp_password, smtp_from_email=smtp_from
        )
        logger.info(f"Exchanger notification sent to {vendor_name} for {notification_type}")
    except Exception as e:
        logger.error(f"Failed to send exchanger notification: {e}")



# ============== EMAIL SETTINGS & DAILY REPORTS ==============

class EmailSettingsUpdate(BaseModel):
    smtp_host: Optional[str] = None
    smtp_port: Optional[int] = None
    smtp_email: Optional[str] = None  # username for auth
    smtp_password: Optional[str] = None
    smtp_from_email: Optional[str] = None  # email to send from
    director_emails: Optional[List[str]] = None
    report_enabled: Optional[bool] = None
    report_time: Optional[str] = None  # Format: "HH:MM"
    monthly_report_enabled: Optional[bool] = None

@api_router.get("/settings/email")
async def get_email_settings(user: dict = Depends(require_permission(Modules.SETTINGS, Actions.VIEW))):
    """Get email/report settings (password masked)"""
    settings = await db.app_settings.find_one({"setting_type": "email"}, {"_id": 0})
    if not settings:
        return {
            "smtp_host": "smtp.gmail.com",
            "smtp_port": 587,
            "smtp_email": "",
            "smtp_from_email": "",
            "smtp_password_set": False,
            "director_emails": [],
            "report_enabled": False,
            "report_time": "03:00",
            "monthly_report_enabled": False
        }
    return {
        "smtp_host": settings.get("smtp_host", "smtp.gmail.com"),
        "smtp_port": settings.get("smtp_port", 587),
        "smtp_email": settings.get("smtp_email", ""),
        "smtp_from_email": settings.get("smtp_from_email", settings.get("smtp_email", "")),
        "smtp_password_set": bool(settings.get("smtp_password")),
        "director_emails": settings.get("director_emails", []),
        "report_enabled": settings.get("report_enabled", False),
        "report_time": settings.get("report_time", "03:00"),
        "monthly_report_enabled": settings.get("monthly_report_enabled", False)
    }

@api_router.put("/settings/email")
async def update_email_settings(request: Request, settings: EmailSettingsUpdate, user: dict = Depends(require_permission(Modules.SETTINGS, Actions.EDIT))):

    """Update email/report settings"""
    now = datetime.now(timezone.utc)
    
    existing = await db.app_settings.find_one({"setting_type": "email"}, {"_id": 0})
    
    updates = {"updated_at": now.isoformat(), "updated_by": user["user_id"]}
    
    if settings.smtp_host is not None:
        updates["smtp_host"] = settings.smtp_host
    if settings.smtp_port is not None:
        updates["smtp_port"] = settings.smtp_port
    if settings.smtp_email is not None:
        updates["smtp_email"] = settings.smtp_email
    if settings.smtp_from_email is not None:
        updates["smtp_from_email"] = settings.smtp_from_email
    if settings.smtp_password is not None and settings.smtp_password != "":
        updates["smtp_password"] = settings.smtp_password  # In production, encrypt this
    if settings.director_emails is not None:
        updates["director_emails"] = settings.director_emails
    if settings.report_enabled is not None:
        updates["report_enabled"] = settings.report_enabled
    if settings.report_time is not None:
        updates["report_time"] = settings.report_time
    if settings.monthly_report_enabled is not None:
        updates["monthly_report_enabled"] = settings.monthly_report_enabled
    
    if existing:
        await db.app_settings.update_one(
            {"setting_type": "email"},
            {"$set": updates}
        )
    else:
        updates["setting_type"] = "email"
        updates["created_at"] = now.isoformat()
        await db.app_settings.insert_one(updates)
    
    # Reschedule the daily report if time changed
    if settings.report_time or settings.report_enabled is not None:
        await reschedule_daily_report()
    
    await log_activity(request, user, "edit", "settings", "Updated email settings")

    return {"message": "Email settings updated successfully"}

@api_router.post("/settings/email/test")
async def test_email_settings(user: dict = Depends(require_permission(Modules.SETTINGS, Actions.EDIT))):
    """Send a test email to verify SMTP settings"""
    settings = await db.app_settings.find_one({"setting_type": "email"}, {"_id": 0})
    
    if not settings or not settings.get("smtp_email") or not settings.get("smtp_password"):
        raise HTTPException(status_code=400, detail="SMTP settings not configured")
    
    if not settings.get("director_emails"):
        raise HTTPException(status_code=400, detail="No director emails configured")
    
    try:
        await send_email(
            to_emails=settings["director_emails"],
            subject="Miles Capitals - Test Email",
            html_content=f"""
            <div style="font-family: Arial, sans-serif; padding: 20px; background-color: #0B0C10; color: white;">
                <h2 style="color: #66FCF1;">Test Email Successful!</h2>
                <p>This is a test email from Miles Capitals Back Office.</p>
                <p>If you received this, your email settings are configured correctly.</p>
                <p style="color: #C5C6C7; font-size: 12px;">Sent at: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}</p>
            </div>
            """,
            smtp_host=settings.get("smtp_host", "smtp.gmail.com"),
            smtp_port=settings.get("smtp_port", 587),
            smtp_email=settings["smtp_email"],
            smtp_password=settings["smtp_password"],
            smtp_from_email=settings.get("smtp_from_email", settings["smtp_email"])
        )
        return {"message": "Test email sent successfully"}
    except Exception as e:
        logger.error(f"Test email failed: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to send test email: {str(e)}")


@api_router.get("/reports/daily/preview")
async def preview_daily_report(user: dict = Depends(require_permission(Modules.REPORTS, Actions.VIEW))):
    """Preview the daily report HTML without sending it"""
    try:
        html_content = await generate_daily_report_html()
        return HTMLResponse(content=html_content)
    except Exception as e:
        logger.error(f"Daily report preview failed: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to generate report: {str(e)}")


@api_router.post("/settings/email/send-daily-report")
async def send_daily_report_now(user: dict = Depends(require_permission(Modules.REPORTS, Actions.EXPORT))):
    """Manually trigger and send the daily report"""
    try:
        settings = await db.app_settings.find_one({"setting_type": "email"}, {"_id": 0})
        if not settings:
            raise HTTPException(status_code=400, detail="SMTP settings not configured")
        
        smtp_host = settings.get("smtp_host")
        smtp_port = settings.get("smtp_port", 587)
        smtp_email = settings.get("smtp_email")
        smtp_password = settings.get("smtp_password")
        smtp_from_email = settings.get("smtp_from_email", smtp_email)
        director_emails = settings.get("director_emails", [])
        
        if not all([smtp_host, smtp_email, smtp_password]):
            raise HTTPException(status_code=400, detail="SMTP settings incomplete")
        
        if not director_emails:
            raise HTTPException(status_code=400, detail="No director emails configured")
        
        # Generate the daily report
        html_content = await generate_daily_report_html()
        
        # Send to all directors
        now = datetime.now(timezone.utc)
        subject = f"Miles Capitals Daily Report - {now.strftime('%Y-%m-%d')}"
        
        await send_email(
            to_emails=director_emails,
            subject=subject,
            html_content=html_content,
            smtp_host=smtp_host,
            smtp_port=smtp_port,
            smtp_email=smtp_email,
            smtp_password=smtp_password,
            smtp_from_email=smtp_from_email
        )
        
        logger.info(f"Daily report sent manually to: {director_emails}")
        return {"message": f"Daily report sent successfully to {', '.join(director_emails)}"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to send daily report: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to send daily report: {str(e)}")


async def send_email(to_emails: List[str], subject: str, html_content: str, 
                     smtp_host: str, smtp_port: int, smtp_email: str, smtp_password: str, 
                     smtp_from_email: str = None):
    """Send email via SMTP"""
    from_email = smtp_from_email or smtp_email
    
    message = MIMEMultipart("alternative")
    message["From"] = from_email
    message["To"] = ", ".join(to_emails)
    message["Subject"] = subject
    
    html_part = MIMEText(html_content, "html")
    message.attach(html_part)
    
    # Determine TLS settings based on port
    use_tls = smtp_port in [587, 25]  # STARTTLS ports
    use_ssl = smtp_port == 465  # Direct SSL port
    
    if use_ssl:
        await aiosmtplib.send(
            message,
            hostname=smtp_host,
            port=smtp_port,
            use_tls=True,
            username=smtp_email,
            password=smtp_password,
        )
    else:
        await aiosmtplib.send(
            message,
            hostname=smtp_host,
            port=smtp_port,
            start_tls=use_tls,
            username=smtp_email,
            password=smtp_password,
        )


async def generate_reconciliation_section_html():
    """Generate reconciliation section for daily report"""
    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = today_start + timedelta(days=1)
    
    # Get today's reconciliation stats
    today_items = await db.reconciliation_items.find({
        "date": {"$gte": today_start.isoformat(), "$lt": tomorrow.isoformat()}
    }, {"_id": 0}).to_list(10000)
    
    bank_entries = await db.reconciliation_entries.find({
        "created_at": {"$gte": today_start.isoformat(), "$lt": tomorrow.isoformat()}
    }, {"_id": 0}).to_list(10000)
    
    # Calculate by category
    by_category = {
        "bank": {"reconciled": 0, "pending": 0, "flagged": 0},
        "psp": {"reconciled": 0, "pending": 0, "flagged": 0},
        "transaction": {"reconciled": 0, "pending": 0, "flagged": 0},
        "income_expense": {"reconciled": 0, "pending": 0, "flagged": 0},
        "treasury": {"reconciled": 0, "pending": 0, "flagged": 0},
    }
    
    for item in today_items:
        cat = item.get("item_type", "other")
        if cat not in by_category:
            by_category[cat] = {"reconciled": 0, "pending": 0, "flagged": 0}
        
        status = item.get("status", "pending")
        if status == "reconciled":
            by_category[cat]["reconciled"] += 1
        elif status == "flagged":
            by_category[cat]["flagged"] += 1
        else:
            by_category[cat]["pending"] += 1
    
    # Add bank entries
    for entry in bank_entries:
        status = entry.get("status", "pending")
        if status == "matched":
            by_category["bank"]["reconciled"] += 1
        elif status in ["unmatched", "pending", "discrepancy"]:
            by_category["bank"]["pending"] += 1
    
    total_reconciled = sum(c["reconciled"] for c in by_category.values())
    total_pending = sum(c["pending"] for c in by_category.values())
    total_flagged = sum(c["flagged"] for c in by_category.values())
    total = total_reconciled + total_pending + total_flagged
    
    # Get flagged items details
    flagged_items = await db.reconciliation_items.find(
        {"status": "flagged"},
        {"_id": 0, "reference_id": 1, "item_type": 1, "flag_reason": 1}
    ).to_list(10)
    
    # Build category rows
    category_rows = ""
    for cat_name, stats in by_category.items():
        if stats["reconciled"] + stats["pending"] + stats["flagged"] > 0:
            cat_label = cat_name.replace("_", " ").title()
            category_rows += f"<tr><td>{cat_label}</td><td class='green'>{stats['reconciled']}</td><td class='yellow'>{stats['pending']}</td><td class='red'>{stats['flagged']}</td></tr>"
    
    # Build flagged items rows
    flagged_rows = ""
    for item in flagged_items[:5]:
        flagged_rows += f"<tr><td>{item.get('item_type', 'Unknown')}</td><td>{item.get('reference_id', 'N/A')[:20]}</td><td>{item.get('flag_reason', 'No reason')[:50]}</td></tr>"
    
    # Determine status color
    if total == 0:
        status_text = "✅ No items to reconcile"
        status_color = "green"
    elif total_pending == 0 and total_flagged == 0:
        status_text = "✅ All items reconciled"
        status_color = "green"
    elif total_flagged > 0:
        status_text = f"⚠️ {total_flagged} items need review"
        status_color = "red"
    else:
        status_text = f"⏳ {total_pending} items pending"
        status_color = "yellow"
    
    # Build category table
    category_table = ""
    if category_rows:
        category_table = f'''<table>
                        <tr><th>Category</th><th>Reconciled</th><th>Pending</th><th>Flagged</th></tr>
                        {category_rows}
                    </table>'''
    
    # Build flagged section
    flagged_section = ""
    if flagged_rows:
        flagged_section = f'''<div style="margin-top: 15px;">
                        <strong>Items Requiring Attention:</strong>
                        <table style="margin-top: 10px;">
                            <tr><th>Type</th><th>Reference</th><th>Reason</th></tr>
                            {flagged_rows}
                        </table>
                    </div>'''
    
    reconciliation_html = f'''
                <!-- Daily Reconciliation Summary -->
                <div class="section">
                    <div class="section-title">Daily Reconciliation Summary</div>
                    <div class="stat-grid">
                        <div class="stat-box">
                            <div class="stat-label">Reconciled</div>
                            <div class="stat-value green">{total_reconciled}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Pending</div>
                            <div class="stat-value yellow">{total_pending}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Flagged</div>
                            <div class="stat-value red">{total_flagged}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Status</div>
                            <div class="stat-value {status_color}">{status_text}</div>
                        </div>
                    </div>
                    {category_table}
                    {flagged_section}
                </div>
    '''
    
    return reconciliation_html


async def generate_daily_report_html():
    """Generate comprehensive daily report HTML"""
    now = datetime.now(timezone.utc)
    yesterday = now - timedelta(days=1)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday_start = yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
    today_date = now.strftime('%Y-%m-%d')
    
    # Fetch data for report
    # Today's transactions
    today_txs = await db.transactions.find({
        "created_at": {"$gte": today_start.isoformat()}
    }, {"_id": 0}).to_list(10000)
    
    today_deposits = sum(t.get("amount", 0) for t in today_txs if t.get("transaction_type") == "deposit")
    today_withdrawals = sum(t.get("amount", 0) for t in today_txs if t.get("transaction_type") == "withdrawal")
    
    # Treasury balances
    treasury_accounts = await db.treasury_accounts.find({}, {"_id": 0}).to_list(100)
    total_treasury = sum(convert_to_usd(a.get("balance", 0), a.get("currency", "USD")) for a in treasury_accounts)
    
    # PSP pending
    psps = await db.psps.find({"status": "active"}, {"_id": 0}).to_list(100)
    total_psp_pending = sum(p.get("pending_settlement", 0) for p in psps)
    
    # Pending PSP transactions
    psp_pending_txs = await db.transactions.find({
        "destination_type": "psp",
        "settled": {"$ne": True}
    }, {"_id": 0}).to_list(1000)
    
    # Outstanding accounts
    debts = await db.debts.find({"status": {"$ne": "fully_paid"}}, {"_id": 0}).to_list(1000)
    total_receivables = sum(d.get("amount", 0) - d.get("paid_amount", 0) for d in debts if d.get("debt_type") == "receivable")
    total_payables = sum(d.get("amount", 0) - d.get("paid_amount", 0) for d in debts if d.get("debt_type") == "payable")
    
    # Vendor pending
    vendors = await db.vendors.find({"status": "active"}, {"_id": 0}).to_list(100)
    
    # Pending approvals
    pending_txs = await db.transactions.find({"status": "pending"}, {"_id": 0}).to_list(1000)
    
    # ===== NEW: Daily Loan Transactions =====
    today_loan_txs = await db.loan_transactions.find({
        "created_at": {"$gte": today_start.isoformat()}
    }, {"_id": 0}).to_list(1000)
    
    today_disbursements = [lt for lt in today_loan_txs if lt.get("transaction_type") == "disbursement"]
    today_repayments = [lt for lt in today_loan_txs if lt.get("transaction_type") == "repayment"]
    total_disbursed_today = sum(lt.get("amount", 0) for lt in today_disbursements)
    total_repaid_today = sum(lt.get("amount", 0) for lt in today_repayments)
    
    # Get loan details for context
    loans = await db.loans.find({}, {"_id": 0}).to_list(1000)
    loans_dict = {l.get("loan_id"): l for l in loans}
    
    # ===== NEW: Exchangers Summary with Commissions & Settlements =====
    # Calculate vendor balances including I&E and Loans
    vendor_summaries = []
    for vendor in vendors:
        vendor_id = vendor.get("vendor_id")
        vendor_name = vendor.get("vendor_name", "Unknown")
        
        # Get pending transactions for this vendor
        vendor_txs = await db.transactions.find({
            "vendor_id": vendor_id,
            "status": "approved",
            "settled": {"$ne": True}
        }, {"_id": 0}).to_list(1000)
        
        pending_deposits = sum(t.get("amount", 0) for t in vendor_txs if t.get("transaction_type") == "deposit")
        pending_withdrawals = sum(t.get("amount", 0) for t in vendor_txs if t.get("transaction_type") == "withdrawal")
        total_commission = sum(t.get("vendor_commission_amount", 0) or 0 for t in vendor_txs)
        
        # Get I&E entries for this vendor
        vendor_ie = await db.income_expense_entries.find({
            "vendor_id": vendor_id,
            "status": "approved"
        }, {"_id": 0}).to_list(1000)
        
        ie_income = sum(e.get("base_amount", e.get("amount", 0)) for e in vendor_ie if e.get("entry_type") == "income")
        ie_expense = sum(e.get("base_amount", e.get("amount", 0)) for e in vendor_ie if e.get("entry_type") == "expense")
        
        # Get loan transactions for this vendor
        vendor_loans = await db.loan_transactions.find({
            "vendor_id": vendor_id,
            "status": "approved"
        }, {"_id": 0}).to_list(1000)
        
        loan_in = sum(lt.get("amount", 0) for lt in vendor_loans if lt.get("transaction_type") == "repayment")
        loan_out = sum(lt.get("amount", 0) for lt in vendor_loans if lt.get("transaction_type") == "disbursement")
        
        # Calculate settlement balance (what we owe vendor or vendor owes us)
        # Deposits: vendor collected money for us (we owe them)
        # Withdrawals: vendor paid out for us (they owe us)
        # I&E Income from vendor: they owe us
        # I&E Expense to vendor: we owe them
        # Loan repayments: they paid us (reduces what we owe)
        # Loan disbursements: we gave them (increases what we owe)
        settlement_balance = (pending_deposits - pending_withdrawals) - total_commission + (ie_expense - ie_income) + (loan_out - loan_in)
        
        # Get today's settlements
        today_settlements = await db.vendor_settlements.find({
            "vendor_id": vendor_id,
            "created_at": {"$gte": today_start.isoformat()}
        }, {"_id": 0}).to_list(100)
        settled_today = sum(s.get("amount", 0) for s in today_settlements)
        
        vendor_summaries.append({
            "name": vendor_name,
            "pending_deposits": pending_deposits,
            "pending_withdrawals": pending_withdrawals,
            "commission": total_commission,
            "ie_balance": ie_expense - ie_income,
            "loan_balance": loan_out - loan_in,
            "settlement_balance": settlement_balance,
            "settled_today": settled_today
        })
    
    total_vendor_settlements_today = sum(v["settled_today"] for v in vendor_summaries)
    total_pending_to_vendors = sum(v["settlement_balance"] for v in vendor_summaries if v["settlement_balance"] > 0)
    total_pending_from_vendors = sum(abs(v["settlement_balance"]) for v in vendor_summaries if v["settlement_balance"] < 0)
    
    # Dealing P&L - Get today's record and calculate
    dealing_pnl_record = await db.dealing_pnl.find_one({"date": today_date}, {"_id": 0})
    dealing_pnl_html = ""
    
    if dealing_pnl_record:
        # Get previous day's record for floating change calculation
        prev_record = await db.dealing_pnl.find_one(
            {"date": {"$lt": today_date}},
            {"_id": 0},
            sort=[("date", -1)]
        )
        
        prev_mt5_floating = prev_record.get("mt5_floating_pnl", 0) if prev_record else 0
        prev_lp_entries = {}
        if prev_record:
            for lp in prev_record.get("lp_entries", []):
                prev_lp_entries[lp.get("lp_id")] = lp.get("floating_pnl", 0)
        
        mt5_booked = dealing_pnl_record.get("mt5_booked_pnl", 0)
        mt5_floating = dealing_pnl_record.get("mt5_floating_pnl", 0)
        mt5_floating_change = mt5_floating - prev_mt5_floating
        broker_mt5_pnl = -mt5_booked - mt5_floating_change
        
        # Calculate LP P&L
        total_lp_booked = 0
        total_lp_floating = 0
        total_broker_lp_pnl = 0
        lp_rows = ""
        
        for lp in dealing_pnl_record.get("lp_entries", []):
            lp_id = lp.get("lp_id")
            lp_name = lp.get("lp_name", lp_id)
            lp_booked = lp.get("booked_pnl", 0)
            lp_floating = lp.get("floating_pnl", 0)
            prev_floating = prev_lp_entries.get(lp_id, 0)
            floating_change = lp_floating - prev_floating
            lp_pnl = lp_booked + floating_change
            
            total_lp_booked += lp_booked
            total_lp_floating += lp_floating
            total_broker_lp_pnl += lp_pnl
            
            pnl_color = "green" if lp_pnl >= 0 else "red"
            lp_rows += f"<tr><td>{lp_name}</td><td>${lp_booked:,.0f}</td><td>${lp_floating:,.0f}</td><td class='{pnl_color}'>${lp_pnl:+,.0f}</td></tr>"
        
        total_dealing_pnl = broker_mt5_pnl + total_broker_lp_pnl
        total_color = "green" if total_dealing_pnl >= 0 else "red"
        mt5_color = "green" if broker_mt5_pnl >= 0 else "red"
        lp_color = "green" if total_broker_lp_pnl >= 0 else "red"
        
        dealing_pnl_html = f'''
                <!-- Dealing P&L Section -->
                <div class="section">
                    <div class="section-title">📈 Today's Dealing P&L</div>
                    <div class="stat-grid">
                        <div class="stat-box">
                            <div class="stat-label">MT5 Broker P&L</div>
                            <div class="stat-value {mt5_color}">${broker_mt5_pnl:+,.0f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">LP Hedging P&L</div>
                            <div class="stat-value {lp_color}">${total_broker_lp_pnl:+,.0f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">MT5 Client Booked</div>
                            <div class="stat-value">${mt5_booked:+,.0f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">MT5 Floating</div>
                            <div class="stat-value">${mt5_floating:,.0f}</div>
                        </div>
                    </div>
                    <div style="background-color: #0B0C10; border-radius: 6px; padding: 20px; margin-top: 15px; text-align: center;">
                        <div class="stat-label">TOTAL DEALING P&L</div>
                        <div class="stat-value {total_color}" style="font-size: 32px;">${total_dealing_pnl:+,.0f} USD</div>
                    </div>
                    {f"""<h4 style="color: #66FCF1; margin-top: 20px; font-size: 12px;">LP BREAKDOWN</h4>
                    <table>
                        <tr><th>LP Name</th><th>Booked</th><th>Floating</th><th>P&L</th></tr>
                        {lp_rows}
                    </table>""" if lp_rows else ""}
                </div>
        '''
    else:
        dealing_pnl_html = '''
                <div class="section">
                    <div class="section-title">📈 Today's Dealing P&L</div>
                    <p style="color: #C5C6C7; text-align: center; padding: 15px;">No dealing P&L record for today</p>
                </div>
        '''
    
    # ===== Reconciliation Summary =====
    today_recon_batches = await db.reconciliation_batches.find({
        "created_at": {"$gte": today_start.isoformat()}
    }, {"_id": 0}).to_list(100)
    
    total_recon_batches = len(today_recon_batches)
    total_recon_matched = sum(b.get("matched", 0) for b in today_recon_batches)
    total_recon_unmatched = sum(b.get("unmatched", 0) for b in today_recon_batches)
    total_recon_discrepancies = sum(b.get("discrepancies", 0) for b in today_recon_batches)
    
    # Also get recent reconciliation history entries
    today_recon_history = await db.reconciliations.find({
        "created_at": {"$gte": today_start.isoformat()}
    }, {"_id": 0}).to_list(100)
    recon_history_matched = sum(r.get("matched_count", 0) for r in today_recon_history)
    recon_history_flagged = sum(r.get("flagged_count", 0) for r in today_recon_history)
    
    daily_recon_html = ""
    recon_section_html = f'''
                <div class="section">
                    <div class="section-title">🔄 Reconciliation</div>
                    <div class="stat-grid">
                        <div class="stat-box">
                            <div class="stat-label">Statements Processed</div>
                            <div class="stat-value cyan">{total_recon_batches}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Entries Matched</div>
                            <div class="stat-value green">{total_recon_matched + recon_history_matched}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Unmatched</div>
                            <div class="stat-value red">{total_recon_unmatched}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Flagged / Discrepancies</div>
                            <div class="stat-value yellow">{recon_history_flagged + total_recon_discrepancies}</div>
                        </div>
                    </div>'''
    
    if total_recon_batches > 0:
        recon_rows = ""
        for b in today_recon_batches[:10]:
            status_color = "#4ade80" if b.get("status") == "completed" else "#fbbf24"
            recon_rows += f"<tr><td>{b.get('account_name', '-')}</td><td>{b.get('filename', '-')}</td><td>{b.get('total_rows', 0)}</td><td style='color:#4ade80'>{b.get('matched', 0)}</td><td style='color:#f87171'>{b.get('unmatched', 0)}</td><td style='color:{status_color}'>{b.get('status', '-').upper()}</td></tr>"
        recon_section_html += f"""
                    <table>
                        <tr><th>Account</th><th>File</th><th>Rows</th><th>Matched</th><th>Unmatched</th><th>Status</th></tr>
                        {recon_rows}
                    </table>"""
    
    recon_section_html += "\n                </div>"
    daily_recon_html = recon_section_html
    
    # Generate HTML
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 0; padding: 0; background-color: #f5f5f5; }}
            .container {{ max-width: 700px; margin: 0 auto; background-color: #0B0C10; color: white; }}
            .header {{ background: linear-gradient(135deg, #1F2833 0%, #0B0C10 100%); padding: 30px; text-align: center; border-bottom: 3px solid #66FCF1; }}
            .header h1 {{ color: #66FCF1; margin: 0; font-size: 28px; letter-spacing: 2px; }}
            .header p {{ color: #C5C6C7; margin: 10px 0 0; font-size: 14px; }}
            .content {{ padding: 30px; }}
            .section {{ background-color: #1F2833; border-radius: 8px; padding: 20px; margin-bottom: 20px; }}
            .section-title {{ color: #66FCF1; font-size: 16px; font-weight: bold; margin-bottom: 15px; text-transform: uppercase; letter-spacing: 1px; border-bottom: 1px solid #66FCF1; padding-bottom: 10px; }}
            .stat-grid {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px; }}
            .stat-box {{ background-color: #0B0C10; border-radius: 6px; padding: 15px; }}
            .stat-label {{ color: #C5C6C7; font-size: 11px; text-transform: uppercase; letter-spacing: 1px; }}
            .stat-value {{ color: white; font-size: 24px; font-weight: bold; margin-top: 5px; }}
            .stat-value.green {{ color: #4ade80; }}
            .stat-value.red {{ color: #f87171; }}
            .stat-value.cyan {{ color: #66FCF1; }}
            .stat-value.yellow {{ color: #fbbf24; }}
            .alert {{ background-color: #7f1d1d; border-left: 4px solid #ef4444; padding: 15px; margin-bottom: 20px; border-radius: 0 8px 8px 0; }}
            .alert-title {{ color: #fca5a5; font-weight: bold; margin-bottom: 5px; }}
            .alert-text {{ color: #fecaca; font-size: 14px; }}
            .footer {{ background-color: #1F2833; padding: 20px; text-align: center; border-top: 1px solid #333; }}
            .footer p {{ color: #C5C6C7; font-size: 12px; margin: 0; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            th, td {{ padding: 10px; text-align: left; border-bottom: 1px solid #333; }}
            th {{ color: #66FCF1; font-size: 11px; text-transform: uppercase; }}
            td {{ color: white; font-size: 13px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>MILES CAPITALS</h1>
                <p>Daily Business Report - {now.strftime('%B %d, %Y')}</p>
            </div>
            
            <div class="content">
                <!-- Alerts Section -->
                {"" if len(pending_txs) == 0 else f'''
                <div class="alert">
                    <div class="alert-title">⚠️ Action Required</div>
                    <div class="alert-text">{len(pending_txs)} transactions pending approval</div>
                </div>
                '''}
                
                {dealing_pnl_html}
                
                <!-- Today's Activity -->
                <div class="section">
                    <div class="section-title">📊 Today's Activity</div>
                    <div class="stat-grid">
                        <div class="stat-box">
                            <div class="stat-label">Deposits</div>
                            <div class="stat-value green">+${today_deposits:,.2f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Withdrawals</div>
                            <div class="stat-value red">-${today_withdrawals:,.2f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Net Flow</div>
                            <div class="stat-value {'green' if today_deposits - today_withdrawals >= 0 else 'red'}">${today_deposits - today_withdrawals:,.2f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Transactions</div>
                            <div class="stat-value cyan">{len(today_txs)}</div>
                        </div>
                    </div>
                </div>
                
                <!-- Treasury Status -->
                <div class="section">
                    <div class="section-title">🏦 Treasury Status</div>
                    <div class="stat-grid">
                        <div class="stat-box">
                            <div class="stat-label">Total Balance (USD)</div>
                            <div class="stat-value cyan">${total_treasury:,.2f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Active Accounts</div>
                            <div class="stat-value">{len(treasury_accounts)}</div>
                        </div>
                    </div>
                    <table>
                        <tr><th>Account</th><th>Currency</th><th>Balance</th><th>USD Value</th></tr>
                        {''.join(f"<tr><td>{a.get('account_name', 'N/A')}</td><td>{a.get('currency', 'USD')}</td><td>{a.get('balance', 0):,.2f}</td><td>${convert_to_usd(a.get('balance', 0), a.get('currency', 'USD')):,.2f}</td></tr>" for a in treasury_accounts)}
                    </table>
                </div>
                
                <!-- PSP Status -->
                <div class="section">
                    <div class="section-title">💳 PSP Status</div>
                    <div class="stat-grid">
                        <div class="stat-box">
                            <div class="stat-label">Pending Settlements</div>
                            <div class="stat-value yellow">${total_psp_pending:,.2f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Pending Transactions</div>
                            <div class="stat-value">{len(psp_pending_txs)}</div>
                        </div>
                    </div>
                </div>
                
                <!-- Outstanding Accounts -->
                <div class="section">
                    <div class="section-title">📋 Outstanding Accounts</div>
                    <div class="stat-grid">
                        <div class="stat-box">
                            <div class="stat-label">Receivables (Owed to us)</div>
                            <div class="stat-value green">${total_receivables:,.2f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Payables (We owe)</div>
                            <div class="stat-value red">${total_payables:,.2f}</div>
                        </div>
                    </div>
                </div>
                
                <!-- Daily Loan Transactions Log -->
                <div class="section">
                    <div class="section-title">💰 Daily Loan Transactions</div>
                    <div class="stat-grid">
                        <div class="stat-box">
                            <div class="stat-label">Total Disbursed Today</div>
                            <div class="stat-value red">-${total_disbursed_today:,.2f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Total Repaid Today</div>
                            <div class="stat-value green">+${total_repaid_today:,.2f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Disbursements</div>
                            <div class="stat-value">{len(today_disbursements)}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Repayments</div>
                            <div class="stat-value">{len(today_repayments)}</div>
                        </div>
                    </div>
                    {"" if len(today_loan_txs) == 0 else f'''
                    <table>
                        <tr><th>Type</th><th>Borrower</th><th>Amount</th><th>Currency</th></tr>
                        {"".join(f"<tr><td style='color: {'#4ade80' if lt.get('transaction_type') == 'repayment' else '#f87171'}'>{lt.get('transaction_type', 'N/A').upper()}</td><td>{loans_dict.get(lt.get('loan_id'), dict()).get('borrower_name', 'N/A')}</td><td>${lt.get('amount', 0):,.2f}</td><td>{lt.get('currency', 'USD')}</td></tr>" for lt in today_loan_txs[:10])}
                    </table>
                    {f"<p style='color: #C5C6C7; font-size: 11px; margin-top: 10px;'>Showing 10 of {len(today_loan_txs)} transactions</p>" if len(today_loan_txs) > 10 else ""}
                    '''}
                </div>
                
                <!-- Exchangers Summary -->
                <div class="section">
                    <div class="section-title">🏪 Exchangers Summary</div>
                    <div class="stat-grid">
                        <div class="stat-box">
                            <div class="stat-label">Settled Today</div>
                            <div class="stat-value cyan">${total_vendor_settlements_today:,.2f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Active Exchangers</div>
                            <div class="stat-value">{len(vendors)}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">We Owe (Total)</div>
                            <div class="stat-value red">${total_pending_to_vendors:,.2f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">They Owe (Total)</div>
                            <div class="stat-value green">${total_pending_from_vendors:,.2f}</div>
                        </div>
                    </div>
                    {"" if len(vendor_summaries) == 0 else f'''
                    <table>
                        <tr><th>Exchanger</th><th>Deposits</th><th>Withdrawals</th><th>Commission</th><th>I&E</th><th>Loans</th><th>Balance</th></tr>
                        {"".join(f"<tr><td>{v['name']}</td><td style='color: #4ade80'>${v['pending_deposits']:,.0f}</td><td style='color: #f87171'>${v['pending_withdrawals']:,.0f}</td><td>${v['commission']:,.0f}</td><td>${v['ie_balance']:,.0f}</td><td>${v['loan_balance']:,.0f}</td><td style='color: {'#f87171' if v['settlement_balance'] > 0 else '#4ade80' if v['settlement_balance'] < 0 else 'white'}'>${v['settlement_balance']:+,.0f}</td></tr>" for v in vendor_summaries[:10])}
                    </table>
                    {f"<p style='color: #C5C6C7; font-size: 11px; margin-top: 10px;'>Showing 10 of {len(vendor_summaries)} exchangers</p>" if len(vendor_summaries) > 10 else ""}
                    <p style='color: #C5C6C7; font-size: 10px; margin-top: 5px;'>* Positive balance = We owe them | Negative balance = They owe us</p>
                    '''}
                </div>
                
                <!-- Pending Actions -->
                <div class="section">
                    <div class="section-title">⏳ Pending Actions</div>
                    <div class="stat-grid">
                        <div class="stat-box">
                            <div class="stat-label">Pending Approvals</div>
                            <div class="stat-value yellow">{len(pending_txs)}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Active Vendors</div>
                            <div class="stat-value">{len(vendors)}</div>
                        </div>
                    </div>
                </div>
                
                {daily_recon_html}
                
            </div>
            
            <div class="footer">
                <p>This is an automated report from Miles Capitals Back Office</p>
                <p>Generated at {now.strftime('%Y-%m-%d %H:%M:%S UTC')}</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    return html

_daily_report_lock = asyncio.Lock()

async def send_daily_report():
    """Send daily report to all directors"""
    if _daily_report_lock.locked():
        logger.warning("Daily report already in progress - skipping duplicate execution")
        return
    
    async with _daily_report_lock:
        try:
            settings = await db.app_settings.find_one({"setting_type": "email"}, {"_id": 0})
            
            if not settings or not settings.get("report_enabled"):
                logger.info("Daily report is disabled or not configured")
                return
            
            if not settings.get("smtp_email") or not settings.get("smtp_password"):
                logger.warning("SMTP settings not configured - skipping daily report")
                return
            
            if not settings.get("director_emails"):
                logger.warning("No director emails configured - skipping daily report")
                return
            
            # Dedup check: skip if report was already sent in the last 30 minutes
            recent_log = await db.email_logs.find_one({
                "type": "daily_report",
                "status": "sent",
                "sent_at": {"$gte": (datetime.now(timezone.utc) - timedelta(minutes=30)).isoformat()}
            }, {"_id": 0})
            if recent_log:
                logger.warning("Daily report already sent within last 30 minutes - skipping duplicate")
                return
            
            html_content = await generate_daily_report_html()
            
            await send_email(
                to_emails=settings["director_emails"],
                subject=f"Miles Capitals - Daily Report ({datetime.now(timezone.utc).strftime('%Y-%m-%d')})",
                html_content=html_content,
                smtp_host=settings.get("smtp_host", "smtp.gmail.com"),
                smtp_port=settings.get("smtp_port", 587),
                smtp_email=settings["smtp_email"],
                smtp_password=settings["smtp_password"],
                smtp_from_email=settings.get("smtp_from_email", settings["smtp_email"])
            )
            
            # Log successful send
            await db.email_logs.insert_one({
                "log_id": f"email_{uuid.uuid4().hex[:12]}",
                "type": "daily_report",
                "recipients": settings["director_emails"],
                "status": "sent",
                "sent_at": datetime.now(timezone.utc).isoformat()
            })
            
            logger.info(f"Daily report sent to {len(settings['director_emails'])} directors")
            
        except Exception as e:
            logger.error(f"Failed to send daily report: {e}")
            await db.email_logs.insert_one({
                "log_id": f"email_{uuid.uuid4().hex[:12]}",
                "type": "daily_report",
                "status": "failed",
                "error": str(e),
                "attempted_at": datetime.now(timezone.utc).isoformat()
            })

@api_router.post("/reports/send-now")
async def send_report_now(user: dict = Depends(require_permission(Modules.REPORTS, Actions.EXPORT))):
    """Manually trigger daily report send"""
    settings = await db.app_settings.find_one({"setting_type": "email"}, {"_id": 0})
    
    if not settings or not settings.get("smtp_email") or not settings.get("smtp_password"):
        raise HTTPException(status_code=400, detail="SMTP settings not configured")
    
    if not settings.get("director_emails"):
        raise HTTPException(status_code=400, detail="No director emails configured")
    
    try:
        html_content = await generate_daily_report_html()
        
        await send_email(
            to_emails=settings["director_emails"],
            subject=f"Miles Capitals - Daily Report ({datetime.now(timezone.utc).strftime('%Y-%m-%d')})",
            html_content=html_content,
            smtp_host=settings.get("smtp_host", "smtp.gmail.com"),
            smtp_port=settings.get("smtp_port", 587),
            smtp_email=settings["smtp_email"],
            smtp_password=settings["smtp_password"],
            smtp_from_email=settings.get("smtp_from_email", settings["smtp_email"])
        )
        
        return {"message": f"Report sent to {len(settings['director_emails'])} directors"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to send report: {str(e)}")

@api_router.get("/reports/email-logs")
async def get_email_logs(limit: int = 20, user: dict = Depends(require_permission(Modules.REPORTS, Actions.VIEW))):
    """Get email send history"""
    logs = await db.email_logs.find({}, {"_id": 0}).sort("sent_at", -1).to_list(limit)
    return logs

# Scheduler instance
scheduler = AsyncIOScheduler()
_scheduler_started = False

async def reschedule_daily_report():
    """Reschedule the daily report based on current settings"""
    global scheduler
    
    # Remove existing job if any
    try:
        scheduler.remove_job("daily_report")
    except:
        pass
    
    settings = await db.app_settings.find_one({"setting_type": "email"}, {"_id": 0})
    
    if not settings or not settings.get("report_enabled"):
        logger.info("Daily report scheduling skipped - disabled")
        return
    
    report_time = settings.get("report_time", "03:00")
    hour, minute = map(int, report_time.split(":"))
    
    scheduler.add_job(
        send_daily_report,
        CronTrigger(hour=hour, minute=minute),
        id="daily_report",
        replace_existing=True
    )
    
    logger.info(f"Daily report scheduled for {report_time}")
    
    # Schedule monthly report (last day of month at same time)
    if settings.get("monthly_report_enabled"):
        scheduler.add_job(
            send_monthly_report,
            CronTrigger(day="last", hour=hour, minute=minute),
            id="monthly_report",
            replace_existing=True
        )
        logger.info("Monthly report scheduled for last day of each month")
    else:
        try:
            scheduler.remove_job("monthly_report")
        except:
            pass


# ============== MONTHLY REPORT SYSTEM ==============

async def generate_monthly_report_html(year: int = None, month: int = None):
    """Generate comprehensive monthly summary report HTML"""
    now = datetime.now(timezone.utc)
    
    if not year or not month:
        # Default to previous month
        first_of_current = now.replace(day=1)
        last_month_end = first_of_current - timedelta(days=1)
        year = last_month_end.year
        month = last_month_end.month
    
    import calendar
    _, last_day = calendar.monthrange(year, month)
    month_start = f"{year}-{month:02d}-01"
    month_end = f"{year}-{month:02d}-{last_day:02d}"
    month_start_iso = f"{month_start}T00:00:00"
    month_end_iso = f"{month_end}T23:59:59"
    month_name = calendar.month_name[month]
    
    # --- Transactions ---
    month_txs = await db.transactions.find({
        "created_at": {"$gte": month_start_iso, "$lte": month_end_iso}
    }, {"_id": 0}).to_list(50000)
    
    approved_txs = [t for t in month_txs if t.get("status") in ["approved", "completed"]]
    total_deposits = sum(t.get("amount", 0) for t in approved_txs if t.get("transaction_type") == "deposit")
    total_withdrawals = sum(t.get("amount", 0) for t in approved_txs if t.get("transaction_type") == "withdrawal")
    total_tx_count = len(month_txs)
    approved_count = len(approved_txs)
    rejected_count = len([t for t in month_txs if t.get("status") == "rejected"])
    pending_count = len([t for t in month_txs if t.get("status") == "pending"])
    
    # --- Treasury Balances (end of month snapshot) ---
    treasury_accounts = await db.treasury_accounts.find({}, {"_id": 0}).to_list(100)
    total_treasury = sum(convert_to_usd(a.get("balance", 0), a.get("currency", "USD")) for a in treasury_accounts)
    treasury_rows = ""
    for acc in sorted(treasury_accounts, key=lambda x: -abs(x.get("balance", 0))):
        bal = acc.get("balance", 0)
        curr = acc.get("currency", "USD")
        usd_val = convert_to_usd(bal, curr)
        treasury_rows += f"<tr><td>{acc.get('account_name', '-')}</td><td>{acc.get('bank_name', '-')}</td><td>{curr}</td><td style='text-align:right'>{bal:,.2f}</td><td style='text-align:right'>${usd_val:,.2f}</td></tr>"
    
    # --- Income & Expenses ---
    month_ie = await db.income_expenses.find({
        "created_at": {"$gte": month_start_iso, "$lte": month_end_iso},
        "status": {"$in": ["approved", "completed"]}
    }, {"_id": 0}).to_list(10000)
    
    ie_income = sum(e.get("amount_usd", e.get("amount", 0)) for e in month_ie if e.get("entry_type") == "income")
    ie_expense = sum(e.get("amount_usd", e.get("amount", 0)) for e in month_ie if e.get("entry_type") == "expense")
    ie_net = ie_income - ie_expense
    
    # --- Loans ---
    month_loan_txs = await db.loan_transactions.find({
        "created_at": {"$gte": month_start_iso, "$lte": month_end_iso}
    }, {"_id": 0}).to_list(10000)
    
    total_disbursed = sum(lt.get("amount", 0) for lt in month_loan_txs if lt.get("transaction_type") == "disbursement")
    total_repaid = sum(lt.get("amount", 0) for lt in month_loan_txs if lt.get("transaction_type") == "repayment")
    
    active_loans = await db.loans.find({"status": {"$in": ["active", "partially_paid"]}}, {"_id": 0}).to_list(1000)
    total_outstanding_loans = sum(l.get("amount", 0) - l.get("total_repaid", 0) for l in active_loans)
    
    # --- Vendor/Exchanger Summary ---
    vendors = await db.vendors.find({"status": "active"}, {"_id": 0}).to_list(100)
    vendor_commission_total = sum(t.get("vendor_commission_amount", 0) or 0 for t in approved_txs if t.get("vendor_id"))
    ie_commission_total = sum(e.get("vendor_commission_amount", 0) or 0 for e in month_ie if e.get("vendor_id"))
    
    vendor_rows = ""
    for v in vendors:
        vid = v.get("vendor_id")
        v_txs = [t for t in approved_txs if t.get("vendor_id") == vid]
        v_deps = sum(t.get("amount", 0) for t in v_txs if t.get("transaction_type") == "deposit")
        v_wds = sum(t.get("amount", 0) for t in v_txs if t.get("transaction_type") == "withdrawal")
        v_comm = sum(t.get("vendor_commission_amount", 0) or 0 for t in v_txs)
        v_ie = [e for e in month_ie if e.get("vendor_id") == vid]
        v_ie_in = sum(e.get("amount_usd", e.get("amount", 0)) for e in v_ie if e.get("entry_type") == "income")
        v_ie_out = sum(e.get("amount_usd", e.get("amount", 0)) for e in v_ie if e.get("entry_type") == "expense")
        v_ie_comm = sum(e.get("vendor_commission_amount", 0) or 0 for e in v_ie)
        v_net = (v_deps + v_ie_in) - (v_wds + v_ie_out) - (v_comm + v_ie_comm)
        net_color = "green" if v_net >= 0 else "red"
        vendor_rows += f"<tr><td>{v.get('vendor_name', '-')}</td><td>${v_deps:,.0f}</td><td>${v_wds:,.0f}</td><td>${v_comm:,.2f}</td><td>${v_ie_in:,.0f}</td><td>${v_ie_out:,.0f}</td><td class='{net_color}'>${v_net:+,.0f}</td></tr>"
    
    # --- PSP Summary ---
    psps = await db.psps.find({"status": "active"}, {"_id": 0}).to_list(100)
    psp_rows = ""
    for p in psps:
        pid = p.get("psp_id")
        p_txs = [t for t in approved_txs if t.get("psp_id") == pid]
        p_vol = sum(t.get("amount", 0) for t in p_txs)
        p_comm = sum(t.get("psp_commission_amount", 0) or 0 for t in p_txs)
        psp_rows += f"<tr><td>{p.get('psp_name', '-')}</td><td>{len(p_txs)}</td><td>${p_vol:,.0f}</td><td>${p_comm:,.2f}</td><td>${p.get('pending_settlement', 0):,.0f}</td></tr>"
    
    # --- Outstanding Accounts ---
    debts = await db.debts.find({"status": {"$ne": "fully_paid"}}, {"_id": 0}).to_list(1000)
    total_receivables = sum(d.get("amount", 0) - d.get("paid_amount", 0) for d in debts if d.get("debt_type") == "receivable")
    total_payables = sum(d.get("amount", 0) - d.get("paid_amount", 0) for d in debts if d.get("debt_type") == "payable")
    
    # --- Settlements ---
    month_settlements = await db.vendor_settlements.find({
        "created_at": {"$gte": month_start_iso, "$lte": month_end_iso}
    }, {"_id": 0}).to_list(1000)
    total_settled = sum(s.get("settlement_amount", s.get("amount", 0)) for s in month_settlements)
    
    # --- Dealing P&L ---
    dealing_records = await db.dealing_pnl.find({
        "date": {"$gte": month_start, "$lte": month_end}
    }, {"_id": 0}).sort("date", 1).to_list(31)
    
    total_mt5_booked = 0
    total_lp_booked = 0
    total_broker_pnl = 0
    dealing_day_rows = ""
    prev_mt5_floating = 0
    prev_lp_floating = {}
    
    for dr in dealing_records:
        mt5_booked = dr.get("mt5_booked_pnl", 0)
        mt5_floating = dr.get("mt5_floating_pnl", 0)
        mt5_float_change = mt5_floating - prev_mt5_floating
        broker_mt5_pnl = -mt5_booked - mt5_float_change
        
        day_lp_booked = 0
        day_lp_pnl = 0
        for lp in dr.get("lp_entries", []):
            lp_booked = lp.get("booked_pnl", 0)
            lp_float = lp.get("floating_pnl", 0)
            prev_f = prev_lp_floating.get(lp.get("lp_id"), 0)
            day_lp_booked += lp_booked
            day_lp_pnl += lp_booked + (lp_float - prev_f)
            prev_lp_floating[lp.get("lp_id")] = lp_float
        
        day_total = broker_mt5_pnl + day_lp_pnl
        total_mt5_booked += mt5_booked
        total_lp_booked += day_lp_booked
        total_broker_pnl += day_total
        prev_mt5_floating = mt5_floating
        
        pnl_color = "#4ade80" if day_total >= 0 else "#f87171"
        dealing_day_rows += f"<tr><td>{dr.get('date', '-')}</td><td>${mt5_booked:+,.0f}</td><td>${broker_mt5_pnl:+,.0f}</td><td>${day_lp_pnl:+,.0f}</td><td style='color:{pnl_color}'>${day_total:+,.0f}</td></tr>"
    
    # --- Reconciliation Summary ---
    month_recon_batches = await db.reconciliation_batches.find({
        "created_at": {"$gte": month_start_iso, "$lte": month_end_iso}
    }, {"_id": 0}).to_list(500)
    
    month_recon_history = await db.reconciliations.find({
        "created_at": {"$gte": month_start_iso, "$lte": month_end_iso}
    }, {"_id": 0}).to_list(500)
    
    recon_total_batches = len(month_recon_batches)
    recon_total_matched = sum(b.get("matched", 0) for b in month_recon_batches) + sum(r.get("matched_count", 0) for r in month_recon_history)
    recon_total_unmatched = sum(b.get("unmatched", 0) for b in month_recon_batches)
    recon_total_flagged = sum(b.get("discrepancies", 0) for b in month_recon_batches) + sum(r.get("flagged_count", 0) for r in month_recon_history)
    recon_total_rows = sum(b.get("total_rows", 0) for b in month_recon_batches)
    
    # Pre-build dealing P&L table
    if dealing_day_rows:
        dealing_table_html = f'<table><tr><th>Date</th><th>MT5 Booked</th><th>Broker MT5</th><th>LP P&L</th><th>Total</th></tr>{dealing_day_rows}</table>'
    else:
        dealing_table_html = '<p style="color:#C5C6C7; text-align:center; padding:10px;">No dealing P&L records this month</p>'

    # Generate HTML
    html = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 0; padding: 0; background-color: #f5f5f5; }}
            .container {{ max-width: 750px; margin: 0 auto; background-color: #0B0C10; color: white; }}
            .header {{ background: linear-gradient(135deg, #1F2833 0%, #0B0C10 100%); padding: 30px; text-align: center; border-bottom: 3px solid #66FCF1; }}
            .header h1 {{ color: #66FCF1; margin: 0; font-size: 28px; letter-spacing: 2px; }}
            .header p {{ color: #C5C6C7; margin: 10px 0 0; font-size: 14px; }}
            .content {{ padding: 30px; }}
            .section {{ background-color: #1F2833; border-radius: 8px; padding: 20px; margin-bottom: 20px; }}
            .section-title {{ color: #66FCF1; font-size: 16px; font-weight: bold; margin-bottom: 15px; text-transform: uppercase; letter-spacing: 1px; border-bottom: 1px solid #66FCF1; padding-bottom: 10px; }}
            .stat-grid {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 15px; }}
            .stat-grid-3 {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 15px; }}
            .stat-box {{ background-color: #0B0C10; border-radius: 6px; padding: 15px; }}
            .stat-label {{ color: #C5C6C7; font-size: 11px; text-transform: uppercase; letter-spacing: 1px; }}
            .stat-value {{ color: white; font-size: 24px; font-weight: bold; margin-top: 5px; }}
            .stat-value.green {{ color: #4ade80; }}
            .stat-value.red {{ color: #f87171; }}
            .stat-value.cyan {{ color: #66FCF1; }}
            .stat-value.yellow {{ color: #fbbf24; }}
            table {{ width: 100%; border-collapse: collapse; margin: 10px 0; font-size: 12px; }}
            th {{ background-color: #0B3D91; color: white; padding: 8px; text-align: left; font-size: 10px; text-transform: uppercase; }}
            td {{ padding: 6px 8px; border-bottom: 1px solid #333; color: #C5C6C7; }}
            td.green {{ color: #4ade80; }}
            td.red {{ color: #f87171; }}
            .footer {{ background-color: #1F2833; padding: 20px; text-align: center; border-top: 1px solid #333; }}
            .footer p {{ color: #C5C6C7; font-size: 12px; margin: 0; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>MILES CAPITALS</h1>
                <p>Monthly Report - {month_name} {year}</p>
                <p style="font-size:12px; color:#66FCF1;">{month_start} to {month_end}</p>
            </div>
            <div class="content">
                <!-- Transaction Summary -->
                <div class="section">
                    <div class="section-title">Transaction Summary</div>
                    <div class="stat-grid">
                        <div class="stat-box">
                            <div class="stat-label">Total Deposits</div>
                            <div class="stat-value green">${total_deposits:,.0f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Total Withdrawals</div>
                            <div class="stat-value red">${total_withdrawals:,.0f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Net Flow</div>
                            <div class="stat-value {'green' if total_deposits - total_withdrawals >= 0 else 'red'}">${total_deposits - total_withdrawals:+,.0f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Total Transactions</div>
                            <div class="stat-value cyan">{total_tx_count}</div>
                        </div>
                    </div>
                    <div class="stat-grid" style="margin-top:15px">
                        <div class="stat-box">
                            <div class="stat-label">Approved</div>
                            <div class="stat-value green">{approved_count}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Rejected / Pending</div>
                            <div class="stat-value yellow">{rejected_count} / {pending_count}</div>
                        </div>
                    </div>
                </div>

                <!-- Treasury Balances -->
                <div class="section">
                    <div class="section-title">Treasury Balances</div>
                    <div class="stat-box" style="text-align:center; margin-bottom:15px;">
                        <div class="stat-label">Total Treasury (USD Equiv.)</div>
                        <div class="stat-value cyan" style="font-size:32px">${total_treasury:,.0f}</div>
                    </div>
                    <table>
                        <tr><th>Account</th><th>Bank</th><th>Currency</th><th style='text-align:right'>Balance</th><th style='text-align:right'>USD Equiv.</th></tr>
                        {treasury_rows}
                    </table>
                </div>

                <!-- Income & Expenses -->
                <div class="section">
                    <div class="section-title">Income & Expenses</div>
                    <div class="stat-grid-3">
                        <div class="stat-box">
                            <div class="stat-label">Income</div>
                            <div class="stat-value green">${ie_income:,.0f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Expenses</div>
                            <div class="stat-value red">${ie_expense:,.0f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Net</div>
                            <div class="stat-value {'green' if ie_net >= 0 else 'red'}">${ie_net:+,.0f}</div>
                        </div>
                    </div>
                </div>

                <!-- Loans -->
                <div class="section">
                    <div class="section-title">Loans</div>
                    <div class="stat-grid-3">
                        <div class="stat-box">
                            <div class="stat-label">Disbursed</div>
                            <div class="stat-value red">${total_disbursed:,.0f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Repaid</div>
                            <div class="stat-value green">${total_repaid:,.0f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Outstanding</div>
                            <div class="stat-value yellow">${total_outstanding_loans:,.0f}</div>
                        </div>
                    </div>
                </div>

                <!-- Exchanger Summary -->
                <div class="section">
                    <div class="section-title">Exchanger Summary</div>
                    <div class="stat-grid">
                        <div class="stat-box">
                            <div class="stat-label">Tx Commission</div>
                            <div class="stat-value yellow">${vendor_commission_total:,.2f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Settled This Month</div>
                            <div class="stat-value cyan">${total_settled:,.0f}</div>
                        </div>
                    </div>
                    {'<table><tr><th>Exchanger</th><th>Deposits</th><th>Withdrawals</th><th>Tx Comm</th><th>I&E In</th><th>I&E Out</th><th>Net</th></tr>' + vendor_rows + '</table>' if vendor_rows else '<p style="color:#C5C6C7; text-align:center; padding:10px;">No exchanger activity</p>'}
                </div>

                <!-- PSP Summary -->
                {'<div class="section"><div class="section-title">PSP Summary</div><table><tr><th>PSP</th><th>Txns</th><th>Volume</th><th>Commission</th><th>Pending</th></tr>' + psp_rows + '</table></div>' if psp_rows else ''}

                <!-- Outstanding Accounts -->
                <div class="section">
                    <div class="section-title">Outstanding Accounts</div>
                    <div class="stat-grid">
                        <div class="stat-box">
                            <div class="stat-label">Receivables</div>
                            <div class="stat-value green">${total_receivables:,.0f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Payables</div>
                            <div class="stat-value red">${total_payables:,.0f}</div>
                        </div>
                    </div>
                </div>

                <!-- Dealing P&L -->
                <div class="section">
                    <div class="section-title">Dealing P&L</div>
                    <div class="stat-grid-3">
                        <div class="stat-box">
                            <div class="stat-label">MT5 Client Booked</div>
                            <div class="stat-value">${total_mt5_booked:+,.0f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">LP Booked</div>
                            <div class="stat-value">${total_lp_booked:+,.0f}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Broker Net P&L</div>
                            <div class="stat-value {'green' if total_broker_pnl >= 0 else 'red'}" style="font-size:28px">${total_broker_pnl:+,.0f}</div>
                        </div>
                    </div>
                    <p style="color:#C5C6C7; font-size:11px; margin:10px 0 5px;">Records: {len(dealing_records)} trading days</p>
                    {dealing_table_html}
                </div>

                <!-- Reconciliation -->
                <div class="section">
                    <div class="section-title">Reconciliation</div>
                    <div class="stat-grid">
                        <div class="stat-box">
                            <div class="stat-label">Statements Processed</div>
                            <div class="stat-value cyan">{recon_total_batches}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Total Rows Parsed</div>
                            <div class="stat-value">{recon_total_rows}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Entries Matched</div>
                            <div class="stat-value green">{recon_total_matched}</div>
                        </div>
                        <div class="stat-box">
                            <div class="stat-label">Flagged / Discrepancies</div>
                            <div class="stat-value {'red' if recon_total_flagged > 0 else 'yellow'}">{recon_total_flagged}</div>
                        </div>
                    </div>
                </div>

            </div>
            <div class="footer">
                <p>Miles Capitals - Monthly Report | Generated {now.strftime('%Y-%m-%d %H:%M UTC')}</p>
                <p style="margin-top:5px; font-size:10px; color:#666;">This is an automated report. Please contact admin for any discrepancies.</p>
            </div>
        </div>
    </body>
    </html>
    """
    return html


async def send_monthly_report():
    """Send monthly report to all directors on the last day of the month"""
    try:
        settings = await db.app_settings.find_one({"setting_type": "email"}, {"_id": 0})
        
        if not settings or not settings.get("monthly_report_enabled"):
            logger.info("Monthly report is disabled or not configured")
            return
        
        if not settings.get("smtp_email") or not settings.get("smtp_password"):
            logger.warning("SMTP settings not configured - skipping monthly report")
            return
        
        if not settings.get("director_emails"):
            logger.warning("No director emails configured - skipping monthly report")
            return
        
        now = datetime.now(timezone.utc)
        html_content = await generate_monthly_report_html(now.year, now.month)
        
        import calendar
        month_name = calendar.month_name[now.month]
        
        await send_email(
            to_emails=settings["director_emails"],
            subject=f"Miles Capitals - Monthly Report ({month_name} {now.year})",
            html_content=html_content,
            smtp_host=settings.get("smtp_host", "smtp.gmail.com"),
            smtp_port=settings.get("smtp_port", 587),
            smtp_email=settings["smtp_email"],
            smtp_password=settings["smtp_password"],
            smtp_from_email=settings.get("smtp_from_email", settings["smtp_email"])
        )
        
        await db.email_logs.insert_one({
            "log_id": f"email_{uuid.uuid4().hex[:12]}",
            "type": "monthly_report",
            "recipients": settings["director_emails"],
            "status": "sent",
            "month": f"{now.year}-{now.month:02d}",
            "sent_at": datetime.now(timezone.utc).isoformat()
        })
        
        logger.info(f"Monthly report sent for {month_name} {now.year}")
        
    except Exception as e:
        logger.error(f"Failed to send monthly report: {e}")
        await db.email_logs.insert_one({
            "log_id": f"email_{uuid.uuid4().hex[:12]}",
            "type": "monthly_report",
            "status": "failed",
            "error": str(e),
            "attempted_at": datetime.now(timezone.utc).isoformat()
        })


@api_router.get("/reports/monthly/preview")
async def preview_monthly_report(
    year: Optional[int] = None, month: Optional[int] = None,
    user: dict = Depends(require_permission(Modules.REPORTS, Actions.VIEW))
):
    """Preview monthly report HTML"""
    try:
        html_content = await generate_monthly_report_html(year, month)
        from fastapi.responses import HTMLResponse
        return HTMLResponse(content=html_content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate report: {str(e)}")


@api_router.post("/reports/monthly/send-now")
async def send_monthly_report_now(
    year: Optional[int] = None, month: Optional[int] = None,
    user: dict = Depends(require_permission(Modules.REPORTS, Actions.EXPORT))
):
    """Manually trigger monthly report send"""
    settings = await db.app_settings.find_one({"setting_type": "email"}, {"_id": 0})
    
    if not settings or not settings.get("smtp_email") or not settings.get("smtp_password"):
        raise HTTPException(status_code=400, detail="SMTP settings not configured")
    
    if not settings.get("director_emails"):
        raise HTTPException(status_code=400, detail="No director emails configured")
    
    try:
        now = datetime.now(timezone.utc)
        y = year or now.year
        m = month or now.month
        html_content = await generate_monthly_report_html(y, m)
        
        import calendar
        month_name = calendar.month_name[m]
        
        await send_email(
            to_emails=settings["director_emails"],
            subject=f"Miles Capitals - Monthly Report ({month_name} {y})",
            html_content=html_content,
            smtp_host=settings.get("smtp_host", "smtp.gmail.com"),
            smtp_port=settings.get("smtp_port", 587),
            smtp_email=settings["smtp_email"],
            smtp_password=settings["smtp_password"],
            smtp_from_email=settings.get("smtp_from_email", settings["smtp_email"])
        )
        
        return {"message": f"Monthly report for {month_name} {y} sent to {len(settings['director_emails'])} directors"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to send report: {str(e)}")

# ============== AUDIT & COMPLIANCE MODULE ==============

class AuditSeverity:
    CRITICAL = "critical"
    WARNING = "warning"
    INFO = "info"

class AuditCategory:
    TRANSACTION_INTEGRITY = "transaction_integrity"
    FX_RATE_VERIFICATION = "fx_rate_verification"
    PSP_SETTLEMENT = "psp_settlement"
    ANOMALY_DETECTION = "anomaly_detection"
    TREASURY_BALANCE = "treasury_balance"

async def run_audit_checks() -> dict:
    """Run all audit checks and return findings."""
    now = datetime.now(timezone.utc)
    findings = []
    stats = {
        "total_checks": 0,
        "critical": 0,
        "warning": 0,
        "info": 0,
        "passed": 0,
    }
    
    # ---- 1. TRANSACTION INTEGRITY ----
    all_txs = await db.transactions.find({}, {"_id": 0}).to_list(50000)
    stats["total_checks"] += 1
    
    # 1a. Missing fields
    for tx in all_txs:
        issues = []
        if not tx.get("client_id"):
            issues.append("Missing client_id")
        if not tx.get("reference"):
            issues.append("Missing reference")
        if not tx.get("amount") or tx.get("amount", 0) <= 0:
            issues.append("Invalid or zero amount")
        if not tx.get("currency"):
            issues.append("Missing currency")
        if tx.get("status") == "completed" and tx.get("transaction_type") == "deposit" and not tx.get("proof_image") and not tx.get("accountant_proof_image"):
            issues.append("Completed deposit without proof")
        if issues:
            findings.append({
                "category": AuditCategory.TRANSACTION_INTEGRITY,
                "severity": AuditSeverity.WARNING,
                "title": f"Data issue: {tx.get('reference', tx['transaction_id'])}",
                "description": "; ".join(issues),
                "transaction_id": tx["transaction_id"],
                "reference": tx.get("reference"),
            })
            stats["warning"] += 1
    
    # 1b. Duplicate detection (same client, same amount, within 5 minutes)
    from collections import defaultdict
    client_txs = defaultdict(list)
    for tx in all_txs:
        key = f"{tx.get('client_id')}_{tx.get('amount')}_{tx.get('transaction_type')}"
        client_txs[key].append(tx)
    
    for key, txs in client_txs.items():
        if len(txs) < 2:
            continue
        txs_sorted = sorted(txs, key=lambda x: x.get("created_at", ""))
        for i in range(1, len(txs_sorted)):
            try:
                t1 = datetime.fromisoformat(txs_sorted[i-1].get("created_at", ""))
                t2 = datetime.fromisoformat(txs_sorted[i].get("created_at", ""))
                if abs((t2 - t1).total_seconds()) < 300:  # 5 minutes
                    findings.append({
                        "category": AuditCategory.ANOMALY_DETECTION,
                        "severity": AuditSeverity.WARNING,
                        "title": f"Potential duplicate: {txs_sorted[i].get('reference', '')}",
                        "description": f"Same client ({txs_sorted[i].get('client_name', 'N/A')}), same amount ({txs_sorted[i].get('amount', 0)} {txs_sorted[i].get('currency', 'USD')}), within 5 minutes of {txs_sorted[i-1].get('reference', '')}",
                        "transaction_id": txs_sorted[i]["transaction_id"],
                        "reference": txs_sorted[i].get("reference"),
                        "related_transaction": txs_sorted[i-1]["transaction_id"],
                    })
                    stats["warning"] += 1
            except (ValueError, TypeError):
                continue
    stats["total_checks"] += 1
    
    # ---- 2. FX RATE VERIFICATION ----
    rates = _fx_cache.get("rates") or FALLBACK_RATES_TO_USD
    
    for tx in all_txs:
        if tx.get("base_currency") and tx.get("base_amount") and tx.get("amount"):
            base_curr = tx["base_currency"]
            if base_curr != "USD" and base_curr in rates:
                expected_usd = round(tx["base_amount"] * rates.get(base_curr, 1.0), 2)
                actual_usd = tx["amount"]
                if actual_usd > 0:
                    deviation_pct = abs(expected_usd - actual_usd) / actual_usd * 100
                    if deviation_pct > 5:
                        findings.append({
                            "category": AuditCategory.FX_RATE_VERIFICATION,
                            "severity": AuditSeverity.CRITICAL if deviation_pct > 10 else AuditSeverity.WARNING,
                            "title": f"FX rate deviation: {tx.get('reference', tx['transaction_id'])}",
                            "description": f"Expected ~{expected_usd:,.2f} USD from {tx['base_amount']:,.2f} {base_curr}, got {actual_usd:,.2f} USD ({deviation_pct:.1f}% deviation)",
                            "transaction_id": tx["transaction_id"],
                            "reference": tx.get("reference"),
                            "deviation_percent": round(deviation_pct, 2),
                        })
                        if deviation_pct > 10:
                            stats["critical"] += 1
                        else:
                            stats["warning"] += 1
    stats["total_checks"] += 1
    
    # ---- 3. PSP SETTLEMENT AUDIT ----
    psp_txs = [tx for tx in all_txs if tx.get("destination_type") == "psp"]
    psps = {p["psp_id"]: p for p in await db.psps.find({}, {"_id": 0}).to_list(1000)}
    treasury_accounts = {a["account_id"]: a for a in await db.treasury_accounts.find({}, {"_id": 0}).to_list(100)}
    
    for tx in psp_txs:
        psp = psps.get(tx.get("psp_id"))
        if not psp:
            continue
        
        # 3a. Verify net amount math
        # Note: psp_net_amount is amount after commission only (reserve fund deducted at settlement)
        amount = tx.get("amount", 0)
        commission = tx.get("psp_commission_amount", 0)
        expected_net = amount - commission
        actual_net = tx.get("psp_net_amount", 0)
        
        if actual_net and abs(expected_net - actual_net) > 0.02:
            findings.append({
                "category": AuditCategory.PSP_SETTLEMENT,
                "severity": AuditSeverity.CRITICAL,
                "title": f"Net amount mismatch: {tx.get('reference', tx['transaction_id'])}",
                "description": f"Expected net {expected_net:,.2f} (amount {amount:,.2f} - commission {commission:,.2f}), actual net {actual_net:,.2f}",
                "transaction_id": tx["transaction_id"],
                "reference": tx.get("reference"),
            })
            stats["critical"] += 1
        
        # 3b. Check reserve fund rate matches PSP setting
        reserve_fund = tx.get("psp_reserve_fund_amount", 0) or 0
        if reserve_fund > 0 and amount > 0:
            expected_rate = psp.get("reserve_fund_rate", 0)
            actual_rate = round(reserve_fund / amount * 100, 2)
            if expected_rate > 0 and abs(actual_rate - expected_rate) > 0.5:
                findings.append({
                    "category": AuditCategory.PSP_SETTLEMENT,
                    "severity": AuditSeverity.WARNING,
                    "title": f"Reserve fund rate mismatch: {tx.get('reference', tx['transaction_id'])}",
                    "description": f"PSP rate is {expected_rate}%, but transaction shows {actual_rate}% ({reserve_fund:,.2f} of {amount:,.2f})",
                    "transaction_id": tx["transaction_id"],
                    "reference": tx.get("reference"),
                })
                stats["warning"] += 1
        
        # 3c. Check settled transactions have currency conversion if needed
        if tx.get("settled"):
            dest_id = psp.get("settlement_destination_id")
            dest_acct = treasury_accounts.get(dest_id) if dest_id else None
            if dest_acct and dest_acct.get("currency", "USD") != tx.get("currency", "USD"):
                # Check if treasury transaction has conversion
                ttx = await db.treasury_transactions.find_one({
                    "related_transaction_id": tx["transaction_id"],
                    "transaction_type": "psp_settlement"
                }, {"_id": 0})
                if ttx and not ttx.get("original_currency"):
                    findings.append({
                        "category": AuditCategory.PSP_SETTLEMENT,
                        "severity": AuditSeverity.CRITICAL,
                        "title": f"Missing currency conversion: {tx.get('reference', tx['transaction_id'])}",
                        "description": f"Settled {tx.get('currency','USD')} transaction to {dest_acct.get('currency','USD')} account without conversion",
                        "transaction_id": tx["transaction_id"],
                        "reference": tx.get("reference"),
                    })
                    stats["critical"] += 1
    stats["total_checks"] += 1
    
    # ---- 4. ANOMALY DETECTION ----
    # 4a. Large transactions
    anomaly_settings = await db.app_settings.find_one({"setting_type": "audit"}, {"_id": 0}) or {}
    large_threshold = anomaly_settings.get("large_transaction_threshold", 50000)
    
    for tx in all_txs:
        if tx.get("amount", 0) >= large_threshold:
            findings.append({
                "category": AuditCategory.ANOMALY_DETECTION,
                "severity": AuditSeverity.INFO,
                "title": f"Large transaction: {tx.get('reference', tx['transaction_id'])}",
                "description": f"{tx.get('transaction_type', 'N/A').upper()} of {tx.get('amount', 0):,.2f} {tx.get('currency', 'USD')} for {tx.get('client_name', 'N/A')}",
                "transaction_id": tx["transaction_id"],
                "reference": tx.get("reference"),
            })
            stats["info"] += 1
    
    # 4b. Round number detection (exact multiples of 10000)
    for tx in all_txs:
        amt = tx.get("amount", 0)
        if amt >= 10000 and amt % 10000 == 0:
            findings.append({
                "category": AuditCategory.ANOMALY_DETECTION,
                "severity": AuditSeverity.INFO,
                "title": f"Round amount: {tx.get('reference', tx['transaction_id'])}",
                "description": f"Exact round amount {amt:,.0f} {tx.get('currency', 'USD')} - {tx.get('client_name', 'N/A')}",
                "transaction_id": tx["transaction_id"],
                "reference": tx.get("reference"),
            })
            stats["info"] += 1
    stats["total_checks"] += 1
    
    # ---- 5. TREASURY BALANCE VERIFICATION ----
    for acct_id, acct in treasury_accounts.items():
        stored_balance = acct.get("balance", 0)
        
        # Sum all treasury transactions for this account
        ttxs = await db.treasury_transactions.find({"account_id": acct_id}, {"_id": 0}).to_list(50000)
        calculated_balance = sum(t.get("amount", 0) for t in ttxs)
        
        # Note: initial balance is not tracked in transactions, so we can only flag large discrepancies
        if len(ttxs) > 0 and abs(stored_balance - calculated_balance) > 1.0:
            findings.append({
                "category": AuditCategory.TREASURY_BALANCE,
                "severity": AuditSeverity.WARNING,
                "title": f"Balance discrepancy: {acct.get('account_name', acct_id)}",
                "description": f"Stored balance: {stored_balance:,.2f} {acct.get('currency', 'USD')}, Sum of transactions: {calculated_balance:,.2f} {acct.get('currency', 'USD')}. Difference may include initial balance or manual adjustments.",
                "account_id": acct_id,
                "stored_balance": round(stored_balance, 2),
                "calculated_balance": round(calculated_balance, 2),
                "difference": round(stored_balance - calculated_balance, 2),
            })
            stats["warning"] += 1
    stats["total_checks"] += 1
    
    stats["passed"] = max(0, stats["total_checks"] - stats["critical"] - stats["warning"])
    health_score = max(0, 100 - (stats["critical"] * 15) - (stats["warning"] * 3))
    
    return {
        "scan_id": f"audit_{uuid.uuid4().hex[:12]}",
        "scanned_at": now.isoformat(),
        "health_score": min(100, health_score),
        "stats": stats,
        "findings": findings,
        "summary": {
            "total_transactions": len(all_txs),
            "psp_transactions": len(psp_txs),
            "treasury_accounts": len(treasury_accounts),
            "categories_checked": 5,
        }
    }

@api_router.post("/audit/run-scan")
async def run_audit_scan(request: Request, user: dict = Depends(require_permission(Modules.AUDIT, Actions.CREATE))):

    """Run a full audit scan and save results."""
    result = await run_audit_checks()
    
    # Save to DB
    await db.audit_scans.insert_one({**result})
    
    # Remove _id before returning
    result.pop("_id", None)
    await log_activity(request, user, "create", "audit", "Ran audit scan")

    return result

@api_router.get("/audit/latest")
async def get_latest_audit(user: dict = Depends(require_permission(Modules.AUDIT, Actions.VIEW))):
    """Get the latest audit scan result."""
    scan = await db.audit_scans.find_one({}, {"_id": 0}, sort=[("scanned_at", -1)])
    if not scan:
        return {"message": "No audit scans found. Run a scan first.", "scan_id": None}
    return scan

@api_router.get("/audit/history")
async def get_audit_history(user: dict = Depends(require_permission(Modules.AUDIT, Actions.VIEW)), limit: int = 20):
    """Get audit scan history."""
    scans = await db.audit_scans.find({}, {"_id": 0, "findings": 0}).sort("scanned_at", -1).to_list(limit)
    return scans

@api_router.get("/audit/settings")
async def get_audit_settings(user: dict = Depends(require_permission(Modules.AUDIT, Actions.VIEW))):
    """Get audit module settings."""
    settings = await db.app_settings.find_one({"setting_type": "audit"}, {"_id": 0})
    if not settings:
        settings = {
            "setting_type": "audit",
            "large_transaction_threshold": 50000,
            "fx_deviation_threshold": 5,
            "auto_scan_enabled": False,
            "auto_scan_time": "02:00",
            "alert_emails": [],
        }
    return settings

class AuditSettingsUpdate(BaseModel):
    large_transaction_threshold: Optional[float] = None
    fx_deviation_threshold: Optional[float] = None
    auto_scan_enabled: Optional[bool] = None
    auto_scan_time: Optional[str] = None
    alert_emails: Optional[List[str]] = None

@api_router.put("/audit/settings")
async def update_audit_settings(request: Request, settings: AuditSettingsUpdate, user: dict = Depends(require_permission(Modules.AUDIT, Actions.EDIT))):

    """Update audit module settings."""
    updates = {k: v for k, v in settings.model_dump().items() if v is not None}
    updates["setting_type"] = "audit"
    
    await db.app_settings.update_one(
        {"setting_type": "audit"},
        {"$set": updates},
        upsert=True
    )
    
    # Reschedule auto-scan if needed
    if settings.auto_scan_enabled is not None or settings.auto_scan_time is not None:
        await reschedule_audit_scan()
    
    await log_activity(request, user, "edit", "audit", "Updated audit settings")

    return await db.app_settings.find_one({"setting_type": "audit"}, {"_id": 0})

async def send_audit_alert_email(result: dict):
    """Send audit alert email if issues found."""
    try:
        audit_settings = await db.app_settings.find_one({"setting_type": "audit"}, {"_id": 0})
        if not audit_settings or not audit_settings.get("alert_emails"):
            return
        
        smtp_settings = await db.app_settings.find_one({"setting_type": "email"}, {"_id": 0})
        if not smtp_settings or not smtp_settings.get("smtp_email"):
            return
        
        stats = result.get("stats", {})
        if stats.get("critical", 0) == 0 and stats.get("warning", 0) == 0:
            return  # No issues to report
        
        score = result.get("health_score", 100)
        color = "#ef4444" if score < 60 else "#f59e0b" if score < 80 else "#22c55e"
        
        html = f"""
        <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;background:#0B0C10;color:#C5C6C7;padding:20px;border-radius:8px;">
            <h1 style="color:#66FCF1;margin-bottom:4px;">Miles Capitals - Audit Alert</h1>
            <p style="color:#888;margin-top:0;">{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}</p>
            <div style="text-align:center;margin:20px 0;">
                <div style="display:inline-block;width:80px;height:80px;border-radius:50%;border:4px solid {color};line-height:80px;font-size:28px;font-weight:bold;color:{color};">{score}</div>
                <p style="color:#888;margin-top:8px;">Health Score</p>
            </div>
            <table style="width:100%;border-collapse:collapse;margin:16px 0;">
                <tr><td style="padding:8px;color:#ef4444;">Critical Issues</td><td style="padding:8px;text-align:right;font-weight:bold;">{stats.get('critical', 0)}</td></tr>
                <tr><td style="padding:8px;color:#f59e0b;">Warnings</td><td style="padding:8px;text-align:right;font-weight:bold;">{stats.get('warning', 0)}</td></tr>
                <tr><td style="padding:8px;color:#3b82f6;">Info</td><td style="padding:8px;text-align:right;font-weight:bold;">{stats.get('info', 0)}</td></tr>
            </table>
        """
        
        # Add critical findings
        criticals = [f for f in result.get("findings", []) if f.get("severity") == "critical"]
        if criticals:
            html += '<h3 style="color:#ef4444;">Critical Findings</h3>'
            for f in criticals[:10]:
                html += f'<div style="background:#1a1a2e;padding:10px;margin:6px 0;border-left:3px solid #ef4444;border-radius:4px;">'
                html += f'<strong>{f.get("title","")}</strong><br/><span style="color:#999;font-size:13px;">{f.get("description","")}</span></div>'
        
        html += "</div>"
        
        await send_email(
            to_emails=audit_settings["alert_emails"],
            subject=f"Miles Capitals - Audit Alert (Score: {score}/100)",
            html_content=html,
            smtp_host=smtp_settings.get("smtp_host", "smtp.gmail.com"),
            smtp_port=smtp_settings.get("smtp_port", 587),
            smtp_email=smtp_settings["smtp_email"],
            smtp_password=smtp_settings["smtp_password"],
            smtp_from_email=smtp_settings.get("smtp_from_email", smtp_settings["smtp_email"])
        )
        logger.info(f"Audit alert sent to {len(audit_settings['alert_emails'])} recipients")
    except Exception as e:
        logger.error(f"Failed to send audit alert: {e}")

async def run_scheduled_audit():
    """Run scheduled audit scan and send alerts."""
    try:
        result = await run_audit_checks()
        await db.audit_scans.insert_one({**result})
        await send_audit_alert_email(result)
        logger.info(f"Scheduled audit complete. Score: {result['health_score']}/100")
    except Exception as e:
        logger.error(f"Scheduled audit failed: {e}")

async def reschedule_audit_scan():
    """Reschedule the automated audit scan."""
    global scheduler
    try:
        scheduler.remove_job("audit_scan")
    except:
        pass
    
    settings = await db.app_settings.find_one({"setting_type": "audit"}, {"_id": 0})
    if not settings or not settings.get("auto_scan_enabled"):
        logger.info("Auto audit scan disabled")
        return
    
    scan_time = settings.get("auto_scan_time", "02:00")
    hour, minute = map(int, scan_time.split(":"))
    
    scheduler.add_job(
        run_scheduled_audit,
        CronTrigger(hour=hour, minute=minute),
        id="audit_scan",
        replace_existing=True
    )
    logger.info(f"Audit scan scheduled for {scan_time}")

# ============== FX RATE ENDPOINTS ==============
@api_router.get("/fx-rates")
async def get_fx_rates_endpoint(user: dict = Depends(get_current_user)):
    """Return current exchange rates (1 unit of currency = X USD)."""
    rates = await get_fx_rates()
    # Return a curated list of commonly used currencies
    popular = ["USD", "EUR", "GBP", "AED", "SAR", "INR", "JPY", "USDT", "CAD", "AUD", "CHF", "CNY", "SGD", "HKD", "KWD", "BHD", "OMR", "QAR", "MYR", "THB"]
    filtered = {k: v for k, v in rates.items() if k in popular}
    # Always include all rates as a secondary field
    return {
        "rates": filtered,
        "all_rates": {k: v for k, v in sorted(rates.items())},
        "source": _fx_cache.get("source", "fallback"),
        "fetched_at": _fx_cache.get("fetched_at", "").isoformat() if _fx_cache.get("fetched_at") else None,
        "cache_ttl_minutes": int(FX_CACHE_TTL.total_seconds() / 60),
    }

@api_router.post("/fx-rates/refresh")
async def refresh_fx_rates(user: dict = Depends(require_permission(Modules.SETTINGS, Actions.EDIT))):
    """Force-refresh exchange rates from the live API."""
    _fx_cache["fetched_at"] = None  # invalidate cache
    rates = await get_fx_rates()
    return {
        "message": "Rates refreshed",
        "source": _fx_cache.get("source", "fallback"),
        "fetched_at": _fx_cache.get("fetched_at", "").isoformat() if _fx_cache.get("fetched_at") else None,
        "sample_rates": {k: rates.get(k) for k in ["USD", "EUR", "GBP", "AED", "INR"] if k in rates},
    }

@api_router.get("/settings/manual-fx-rates")
async def get_manual_fx_rates(user: dict = Depends(require_permission(Modules.SETTINGS, Actions.VIEW))):
    """Get manual FX rates (1 unit of currency = X USD)"""
    settings = await db.app_settings.find_one({"setting_type": "manual_fx_rates"}, {"_id": 0})
    return {
        "rates": settings.get("rates", {}) if settings else {},
        "updated_at": settings.get("updated_at") if settings else None,
        "updated_by_name": settings.get("updated_by_name") if settings else None,
    }

@api_router.put("/settings/manual-fx-rates")
async def update_manual_fx_rates(
    request: Request,
    data: dict = Body(...),
    user: dict = Depends(require_permission(Modules.SETTINGS, Actions.EDIT))
):
    """Update manual FX rates. Body: {rates: {INR: 0.012, AED: 0.272, ...}}"""
    rates = data.get("rates", {})
    now = datetime.now(timezone.utc)
    
    await db.app_settings.update_one(
        {"setting_type": "manual_fx_rates"},
        {"$set": {
            "setting_type": "manual_fx_rates",
            "rates": rates,
            "updated_at": now.isoformat(),
            "updated_by": user["user_id"],
            "updated_by_name": user["name"]
        }},
        upsert=True
    )
    
    await log_activity(request, user, "edit", "settings", "Updated manual FX rates")
    return {"message": "Manual FX rates updated", "rates": rates}



@api_router.get("/fx-rates/convert")
async def convert_currency_endpoint(
    amount: float, from_currency: str, to_currency: str,
    user: dict = Depends(get_current_user)
):
    """Convert an amount between two currencies using live rates."""
    rates = await get_fx_rates()
    from_rate = rates.get(from_currency.upper())
    to_rate = rates.get(to_currency.upper())
    if from_rate is None or to_rate is None:
        raise HTTPException(status_code=400, detail=f"Unsupported currency: {from_currency if from_rate is None else to_currency}")
    usd_amount = amount * from_rate
    converted = usd_amount / to_rate if to_rate else usd_amount
    return {
        "from_currency": from_currency.upper(),
        "to_currency": to_currency.upper(),
        "amount": amount,
        "converted_amount": round(converted, 4),
        "usd_equivalent": round(usd_amount, 4),
        "rate": round(from_rate / to_rate, 6) if to_rate else None,
    }

# ============== COMMISSION SETTINGS ENDPOINTS ==============
@api_router.get("/settings/commission")
async def get_commission_settings(user: dict = Depends(require_permission(Modules.SETTINGS, Actions.VIEW))):
    """Get global commission settings for deposits and withdrawals."""
    settings = await db.app_settings.find_one({"setting_type": "commission"}, {"_id": 0})
    if not settings:
        return {
            "deposit_commission_rate": 0.0,
            "withdrawal_commission_rate": 0.0,
            "commission_enabled": False,
        }
    return {
        "deposit_commission_rate": settings.get("deposit_commission_rate", 0.0),
        "withdrawal_commission_rate": settings.get("withdrawal_commission_rate", 0.0),
        "commission_enabled": settings.get("commission_enabled", False),
    }

@api_router.put("/settings/commission")
async def update_commission_settings(request: Request, user: dict = Depends(require_permission(Modules.SETTINGS, Actions.EDIT))):
    """Update global commission settings."""
    data = await request.json()
    deposit_rate = float(data.get("deposit_commission_rate", 0))
    withdrawal_rate = float(data.get("withdrawal_commission_rate", 0))
    enabled = bool(data.get("commission_enabled", False))

    await db.app_settings.update_one(
        {"setting_type": "commission"},
        {"$set": {
            "setting_type": "commission",
            "deposit_commission_rate": deposit_rate,
            "withdrawal_commission_rate": withdrawal_rate,
            "commission_enabled": enabled,
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "updated_by": user["user_id"],
        }},
        upsert=True
    )
    await log_activity(request, user, "edit", "settings", "Updated commission settings")

    return {"message": "Commission settings updated", "deposit_commission_rate": deposit_rate, "withdrawal_commission_rate": withdrawal_rate, "commission_enabled": enabled}

# ============== LOGS MANAGEMENT ==============

@api_router.get("/logs")
async def get_all_logs(
    user: dict = Depends(require_permission(Modules.LOGS, Actions.VIEW)),
    log_type: Optional[str] = None,
    action: Optional[str] = None,
    module: Optional[str] = None,
    user_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    search: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
):
    """Get all logs with filtering and pagination"""
    query = {}
    
    if log_type:
        query["log_type"] = log_type
    if action:
        query["action"] = action
    if module:
        query["module"] = module
    if user_id:
        query["user_id"] = user_id
    if date_from:
        query["timestamp"] = {"$gte": date_from}
    if date_to:
        if "timestamp" in query:
            query["timestamp"]["$lte"] = date_to + "T23:59:59"
        else:
            query["timestamp"] = {"$lte": date_to + "T23:59:59"}
    if search:
        query["$or"] = [
            {"description": {"$regex": search, "$options": "i"}},
            {"user_name": {"$regex": search, "$options": "i"}},
            {"user_email": {"$regex": search, "$options": "i"}},
            {"reference_id": {"$regex": search, "$options": "i"}},
            {"module": {"$regex": search, "$options": "i"}},
        ]
    
    total = await db.system_logs.count_documents(query)
    total_pages = max(1, (total + page_size - 1) // page_size)
    skip_n = (page - 1) * page_size
    
    logs = await db.system_logs.find(query, {"_id": 0}).sort("timestamp", -1).skip(skip_n).limit(page_size).to_list(page_size)
    
    return {
        "items": logs,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages
    }

@api_router.get("/logs/stats")
async def get_logs_stats(user: dict = Depends(require_permission(Modules.LOGS, Actions.VIEW))):
    """Get logs statistics"""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    week_ago = (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y-%m-%d")
    
    total_logs = await db.system_logs.count_documents({})
    today_logs = await db.system_logs.count_documents({"timestamp": {"$gte": today}})
    
    # Logs by type
    activity_count = await db.system_logs.count_documents({"log_type": "activity"})
    auth_count = await db.system_logs.count_documents({"log_type": "auth"})
    audit_count = await db.system_logs.count_documents({"log_type": "audit"})
    error_count = await db.system_logs.count_documents({"log_type": "error"})
    
    # Failed logins in last 7 days
    failed_logins = await db.system_logs.count_documents({
        "log_type": "auth",
        "action": "login_failed",
        "timestamp": {"$gte": week_ago}
    })
    
    # Most active users
    pipeline = [
        {"$match": {"timestamp": {"$gte": week_ago}}},
        {"$group": {"_id": "$user_name", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 5}
    ]
    active_users = await db.system_logs.aggregate(pipeline).to_list(length=5)
    
    # Most common actions
    action_pipeline = [
        {"$match": {"timestamp": {"$gte": week_ago}}},
        {"$group": {"_id": "$action", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 10}
    ]
    common_actions = await db.system_logs.aggregate(action_pipeline).to_list(length=10)
    
    return {
        "total_logs": total_logs,
        "today_logs": today_logs,
        "by_type": {
            "activity": activity_count,
            "auth": auth_count,
            "audit": audit_count,
            "error": error_count
        },
        "failed_logins_7d": failed_logins,
        "active_users": [{"user": u["_id"], "count": u["count"]} for u in active_users if u["_id"]],
        "common_actions": [{"action": a["_id"], "count": a["count"]} for a in common_actions if a["_id"]]
    }

@api_router.get("/logs/activity")
async def get_activity_logs(
    user: dict = Depends(require_permission(Modules.LOGS, Actions.VIEW)),
    module: Optional[str] = None,
    user_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 100
):
    """Get activity logs"""
    query = {"log_type": "activity"}
    if module:
        query["module"] = module
    if user_id:
        query["user_id"] = user_id
    if date_from:
        query["timestamp"] = {"$gte": date_from}
    if date_to:
        if "timestamp" in query:
            query["timestamp"]["$lte"] = date_to + "T23:59:59"
        else:
            query["timestamp"] = {"$lte": date_to + "T23:59:59"}
    
    logs = await db.system_logs.find(query, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(length=limit)
    return {"logs": logs}

@api_router.get("/logs/auth")
async def get_auth_logs(
    user: dict = Depends(require_permission(Modules.LOGS, Actions.VIEW)),
    action: Optional[str] = None,
    user_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 100
):
    """Get authentication logs (login, logout, failed attempts)"""
    query = {"log_type": "auth"}
    if action:
        query["action"] = action
    if user_id:
        query["user_id"] = user_id
    if date_from:
        query["timestamp"] = {"$gte": date_from}
    if date_to:
        if "timestamp" in query:
            query["timestamp"]["$lte"] = date_to + "T23:59:59"
        else:
            query["timestamp"] = {"$lte": date_to + "T23:59:59"}
    
    logs = await db.system_logs.find(query, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(length=limit)
    return {"logs": logs}

@api_router.get("/logs/audit")
async def get_audit_logs(
    user: dict = Depends(require_permission(Modules.LOGS, Actions.VIEW)),
    module: Optional[str] = None,
    reference_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 100
):
    """Get audit logs (financial changes)"""
    query = {"log_type": "audit"}
    if module:
        query["module"] = module
    if reference_id:
        query["reference_id"] = reference_id
    if date_from:
        query["timestamp"] = {"$gte": date_from}
    if date_to:
        if "timestamp" in query:
            query["timestamp"]["$lte"] = date_to + "T23:59:59"
        else:
            query["timestamp"] = {"$lte": date_to + "T23:59:59"}
    
    logs = await db.system_logs.find(query, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(length=limit)
    return {"logs": logs}

@api_router.delete("/logs/clear")
async def clear_old_logs(
    user: dict = Depends(require_permission(Modules.LOGS, Actions.DELETE)),
    days_to_keep: int = 90
):
    """Clear logs older than specified days (default: 90 days)"""
    cutoff_date = (datetime.now(timezone.utc) - timedelta(days=days_to_keep)).isoformat()
    result = await db.system_logs.delete_many({"timestamp": {"$lt": cutoff_date}})
    
    # Log this action
    await create_log(
        log_type="activity",
        action="delete",
        module="logs",
        user_id=user["user_id"],
        user_name=user["name"],
        user_email=user["email"],
        user_role=user["role"],
        description=f"Cleared {result.deleted_count} logs older than {days_to_keep} days"
    )
    
    return {"message": f"Deleted {result.deleted_count} logs older than {days_to_keep} days"}

# ============== ADMIN IMPERSONATION ==============

@api_router.post("/admin/impersonate/{user_id}")
async def start_impersonation(user_id: str, request: Request, admin: dict = Depends(require_admin)):
    """Admin impersonates a sub-user. Returns a new JWT for the target user."""
    # Find target user
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")

    # Prevent impersonating admins
    if target.get("role") == "admin":
        raise HTTPException(status_code=403, detail="Cannot impersonate another Admin")

    if target.get("is_active") is False:
        raise HTTPException(status_code=400, detail="Cannot impersonate an inactive user")

    # Create a JWT for the target user
    impersonation_token = create_jwt_token(target["user_id"], target["email"], target["role"])

    # Create an impersonation log entry
    log_id = f"imp_{uuid.uuid4().hex[:12]}"
    ip_address = request.client.host if request.client else "unknown"
    await db.impersonation_logs.insert_one({
        "log_id": log_id,
        "admin_id": admin["user_id"],
        "admin_name": admin.get("name", ""),
        "admin_email": admin.get("email", ""),
        "user_id": target["user_id"],
        "user_name": target.get("name", ""),
        "user_email": target.get("email", ""),
        "user_role": target.get("role", ""),
        "login_time": datetime.now(timezone.utc).isoformat(),
        "logout_time": None,
        "ip_address": ip_address,
        "status": "active"
    })

    # Activity log
    await log_activity(
        request, admin, "impersonate", "users",
        f"Admin impersonated user {target.get('name')} ({target.get('email')})",
        reference_id=target["user_id"],
        details={"target_user": target.get("email"), "target_role": target.get("role")}
    )

    return {
        "access_token": impersonation_token,
        "impersonation_log_id": log_id,
        "user": {
            "user_id": target["user_id"],
            "email": target["email"],
            "name": target.get("name", ""),
            "role": target.get("role", ""),
        }
    }


@api_router.post("/admin/stop-impersonate")
async def stop_impersonation(request: Request, user: dict = Depends(get_current_user)):
    """End impersonation session. Called by the impersonated session to log the end time."""
    log_id = None
    body = {}
    try:
        body = await request.json()
        log_id = body.get("log_id")
    except Exception:
        pass

    if log_id:
        await db.impersonation_logs.update_one(
            {"log_id": log_id, "status": "active"},
            {"$set": {"logout_time": datetime.now(timezone.utc).isoformat(), "status": "ended"}}
        )
    else:
        # Fallback: close most recent active session for this user
        await db.impersonation_logs.update_one(
            {"user_id": user["user_id"], "status": "active"},
            {"$set": {"logout_time": datetime.now(timezone.utc).isoformat(), "status": "ended"}},
        )

    return {"message": "Impersonation ended"}


@api_router.get("/admin/impersonation-logs")
async def get_impersonation_logs(
    limit: int = 50,
    user: dict = Depends(require_permission(Modules.LOGS, Actions.VIEW))
):
    """Get impersonation audit logs"""
    logs = await db.impersonation_logs.find(
        {}, {"_id": 0}
    ).sort("login_time", -1).to_list(limit)
    return logs


# Include router
app.include_router(api_router)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_db_indexes():
    """Create database indexes and start scheduler"""
    try:
        # Unique index on transaction reference (sparse to allow null values)
        await db.transactions.create_index("reference", unique=True, sparse=True)
        # Index for duplicate detection queries
        await db.transactions.create_index([
            ("client_id", 1),
            ("transaction_type", 1),
            ("amount", 1),
            ("created_at", -1)
        ])
        # Index for common transaction queries
        await db.transactions.create_index([("created_at", -1)])
        await db.transactions.create_index([("status", 1), ("created_at", -1)])
        await db.transactions.create_index([("vendor_id", 1), ("status", 1)])
        await db.transactions.create_index([("destination_account_id", 1), ("status", 1)])
        
        # Vendor indexes
        await db.vendors.create_index("vendor_id", unique=True, sparse=True)
        await db.vendors.create_index([("status", 1)])
        await db.vendors.create_index([("vendor_name", "text")])
        
        # Income/Expense indexes
        await db.income_expense_entries.create_index([("created_at", -1)])
        await db.income_expense_entries.create_index([("entry_type", 1), ("created_at", -1)])
        await db.income_expense_entries.create_index([("vendor_id", 1), ("status", 1)])
        
        # Treasury transaction indexes
        await db.treasury_transactions.create_index([("account_id", 1), ("created_at", -1)])
        await db.treasury_transactions.create_index([("transaction_id", 1)])
        
        # Client indexes
        await db.clients.create_index("client_id", unique=True, sparse=True)
        await db.clients.create_index([("first_name", "text"), ("last_name", "text"), ("email", "text")])
        
        # Index for roles
        await db.roles.create_index("role_id", unique=True, sparse=True)
        await db.roles.create_index("name", unique=True, sparse=True)
        
        # CRITICAL: transaction_id index (used in 20+ find_one calls)
        try:
            await db.transactions.create_index("transaction_id", unique=True)
        except Exception:
            # Index may already exist without unique constraint
            pass
        await db.transactions.create_index("psp_id", sparse=True)
        await db.transactions.create_index("crm_reference", sparse=True)
        
        # PSP indexes
        await db.psps.create_index("psp_id", unique=True, sparse=True)
        await db.psps.create_index([("status", 1)])
        
        # PSP settlements
        await db.psp_settlements.create_index([("psp_id", 1), ("created_at", -1)])
        await db.psp_settlements.create_index([("status", 1)])
        
        # Treasury accounts
        await db.treasury_accounts.create_index("account_id", unique=True, sparse=True)
        await db.treasury_accounts.create_index([("status", 1)])
        
        # Users
        await db.users.create_index("user_id", unique=True, sparse=True)
        await db.users.create_index("email", unique=True, sparse=True)
        
        # Transaction requests
        await db.transaction_requests.create_index([("status", 1), ("created_at", -1)])
        await db.transaction_requests.create_index("request_id", unique=True, sparse=True)
        
        # Reconciliation
        await db.reconciliations.create_index([("date", -1), ("account_type", 1), ("account_id", 1)])
        await db.reconciliations.create_index("recon_id", unique=True, sparse=True)
        await db.reconciliations.create_index([("status", 1), ("created_at", -1)])
        
        # Activity log
        await db.activity_log.create_index([("created_at", -1)])
        await db.activity_log.create_index([("user_id", 1), ("created_at", -1)])
        
        # Client bank accounts
        await db.client_bank_accounts.create_index([("client_id", 1)])
        
        # Loan transactions
        await db.loan_transactions.create_index("transaction_id", unique=True, sparse=True)
        await db.loan_transactions.create_index([("vendor_id", 1), ("status", 1)])
        
        # App settings
        await db.app_settings.create_index("setting_type", unique=True, sparse=True)
        
        logger.info("Database indexes created/verified successfully")
    except Exception as e:
        logger.warning(f"Index creation warning (may already exist): {e}")
    
    # Initialize default roles
    try:
        await initialize_default_roles()
        logger.info("Default roles initialized")
    except Exception as e:
        logger.error(f"Failed to initialize default roles: {e}")
    
    # Start scheduler and schedule daily report
    try:
        global _scheduler_started
        # Prevent duplicate scheduler starts during hot-reload
        if scheduler.running:
            scheduler.shutdown(wait=False)
        scheduler.start()
        _scheduler_started = True
        await reschedule_daily_report()
        await reschedule_audit_scan()
        logger.info("Scheduler started successfully")
    except Exception as e:
        logger.error(f"Failed to start scheduler: {e}")

    # Warm FX rate cache
    try:
        await get_fx_rates()
        logger.info(f"FX rates loaded (source: {_fx_cache.get('source', 'unknown')})")
    except Exception as e:
        logger.warning(f"FX rate warm-up failed: {e}")

@app.on_event("shutdown")
async def shutdown_db_client():
    scheduler.shutdown(wait=False)
    client.close()
